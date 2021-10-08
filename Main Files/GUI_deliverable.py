# from multiple_frame import Ui_MainWindow
# import pyinstaller
from Fault_Response_main import Ui_Fault_response
from svi2_app import Ui_Form
from collections import deque
from SVID_custom_commands_main import Ui_SVID_custom_commands
from PMBus_custom_commands_main import Ui_PMBus_custom_commands
from Fault_Configuration_main import Ui_Fault_config
from PMBus_Configuration_main import Ui_PMBus_configuration_main_window
from SVI2_Configuration_main import Ui_label_main_SVI2_Configuration
from SVID_configuration_main import Ui_SVID_Configuration
from Phase_thermal_balance_main import Ui_Phase_thermal_balance
from Telemetry_senstivity_main import Ui_Telemetry_senstivity
from Telemetry_calibration_main import Ui_Telemetry_calibration
from Transient_Configuration_main import Ui_Transient_Window
from Phase_assignment_main import Ui_Phase_configuration_main_window
from Boot_voltage_main import Ui_Boot_Voltage
from Phase_add_drop_main import Ui_Phase_add_drop
from PMBus_Address_main import Ui_PMBus_address_configuration
from Start_up_frame_main import Ui_start_up
from mtp_read_main import Ui_Reading_Device_Registers
from Home_main import Ui_Home
# import clr

import sys, time, string, openpyxl, struct, os, numpy as np
from datetime import datetime

from PyQt5.QtCore import QThread, QObject, pyqtSignal, QDate
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMainWindow, QApplication, QMessageBox, QInputDialog, QLineEdit, QDesktopWidget, QFileDialog
import math
# from NI8452.dio import DIO
# from NI8452.i2c import I2C
# from NI8452.spi import SPI
# from NI8452.spistream import SPIStream
# import usb_to_gpio
from AMPS_API.TI import usb_to_gpio
from openpyxl import Workbook, load_workbook
import random
global register_database, parallel_thread, stop_thread, PMBUS_ACK_status, PMBus_send
register_database = {

    "010": {
        "size_in_bits": 8,
        "default_value": "01110000",
        "Initial_device_MTP_value": "01110000",
        "Temp_update_from_customer": "01110000",
        "Final_register_value": "01110000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "01110000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "011": {
        "size_in_bits": 8,
        "default_value": "01110000",
        "Initial_device_MTP_value": "01110000",
        "Temp_update_from_customer": "01110000",
        "Final_register_value": "01110000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "01110000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "020": {
        "size_in_bits": 8,
        "default_value": "00010110",
        "Initial_device_MTP_value": "00010110",
        "Temp_update_from_customer": "00010110",
        "Final_register_value": "00010110",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00010110",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "021": {
        "size_in_bits": 8,
        "default_value": "00010110",
        "Initial_device_MTP_value": "00010110",
        "Temp_update_from_customer": "00010110",
        "Final_register_value": "00010110",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00010110",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "1B0": {
        "size_in_bits": 32,
        "default_value": "00000000000000000000000000000000",
        "Initial_device_MTP_value": "00000000000000000000000000000000",
        "Temp_update_from_customer": "00000000000000000000000000000000",
        "Final_register_value": "00000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000000000000000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Process Call",
    }
    ,

    "1B1": {
        "size_in_bits": 32,
        "default_value": "00000000000000000000000000000000",
        "Initial_device_MTP_value": "00000000000000000000000000000000",
        "Temp_update_from_customer": "00000000000000000000000000000000",
        "Final_register_value": "00000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000000000000000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Process Call",
    }
    ,

    "200": {
        "size_in_bits": 8,
        "default_value": "00111110",
        "Initial_device_MTP_value": "00111110",
        "Temp_update_from_customer": "00111110",
        "Final_register_value": "00111110",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00111110",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "201": {
        "size_in_bits": 8,
        "default_value": "00111110",
        "Initial_device_MTP_value": "00111110",
        "Temp_update_from_customer": "00111110",
        "Final_register_value": "00111110",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00111110",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "210": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "211": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "220": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "221": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "240": {
        "size_in_bits": 16,
        "default_value": "0000011111111111",
        "Initial_device_MTP_value": "0000011111111111",
        "Temp_update_from_customer": "0000011111111111",
        "Final_register_value": "0000011111111111",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000011111111111",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "241": {
        "size_in_bits": 16,
        "default_value": "0000011111111111",
        "Initial_device_MTP_value": "0000011111111111",
        "Temp_update_from_customer": "0000011111111111",
        "Final_register_value": "0000011111111111",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000011111111111",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "250": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "251": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "260": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "261": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "270": {
        "size_in_bits": 16,
        "default_value": "1110001100000000",
        "Initial_device_MTP_value": "1110001100000000",
        "Temp_update_from_customer": "1110001100000000",
        "Final_register_value": "1110001100000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1110001100000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "271": {
        "size_in_bits": 16,
        "default_value": "1110001100000000",
        "Initial_device_MTP_value": "1110001100000000",
        "Temp_update_from_customer": "1110001100000000",
        "Final_register_value": "1110001100000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1110001100000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "290": {
        "size_in_bits": 16,
        "default_value": "Its Notsupported",
        "Initial_device_MTP_value": "Its Notsupported",
        "Temp_update_from_customer": "Its Notsupported",
        "Final_register_value": "Its Notsupported",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "Its Notsupported",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "291": {
        "size_in_bits": 16,
        "default_value": "Its Notsupported",
        "Initial_device_MTP_value": "Its Notsupported",
        "Temp_update_from_customer": "Its Notsupported",
        "Final_register_value": "Its Notsupported",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "Its Notsupported",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "2B0": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "2B1": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "330": {
        "size_in_bits": 16,
        "default_value": "1100101000100000",
        "Initial_device_MTP_value": "1100101000100000",
        "Temp_update_from_customer": "1100101000100000",
        "Final_register_value": "1100101000100000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1100101000100000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "331": {
        "size_in_bits": 16,
        "default_value": "1100101000100000",
        "Initial_device_MTP_value": "1100101000100000",
        "Temp_update_from_customer": "1100101000100000",
        "Final_register_value": "1100101000100000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1100101000100000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "35": {
        "size_in_bits": 16,
        "default_value": "1100001111000000",
        "Initial_device_MTP_value": "1100001111000000",
        "Temp_update_from_customer": "1100001111000000",
        "Final_register_value": "1100001111000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1100001111000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "380": {
        "size_in_bits": 16,
        "default_value": "0000000100000000",
        "Initial_device_MTP_value": "0000000100000000",
        "Temp_update_from_customer": "0000000100000000",
        "Final_register_value": "0000000100000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000100000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "381": {
        "size_in_bits": 16,
        "default_value": "0000000100000000",
        "Initial_device_MTP_value": "0000000100000000",
        "Temp_update_from_customer": "0000000100000000",
        "Final_register_value": "0000000100000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000100000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "390": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "391": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "400": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "401": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "410": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "411": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "440": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "441": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "450": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "451": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "470": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "471": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "500": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "501": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "55": {
        "size_in_bits": 16,
        "default_value": "1101101000000000",
        "Initial_device_MTP_value": "1101101000000000",
        "Temp_update_from_customer": "1101101000000000",
        "Final_register_value": "1101101000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1101101000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "56": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "59": {
        "size_in_bits": 16,
        "default_value": "1100001110000000",
        "Initial_device_MTP_value": "1100001110000000",
        "Temp_update_from_customer": "1100001110000000",
        "Final_register_value": "1100001110000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1100001110000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "5A": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "600": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "601": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "620": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "621": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "630": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "631": {
        "size_in_bits": 8,
        "default_value": "00000000",
        "Initial_device_MTP_value": "00000000",
        "Temp_update_from_customer": "00000000",
        "Final_register_value": "00000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "640": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "641": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "99": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "9A": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "9B": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "9C": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "9D": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "9E": {
        "size_in_bits": 16,
        "default_value": "0000000000000000",
        "Initial_device_MTP_value": "0000000000000000",
        "Temp_update_from_customer": "0000000000000000",
        "Final_register_value": "0000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "C50": {
        "size_in_bits": 88,
        "default_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Initial_device_MTP_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Temp_update_from_customer": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Final_register_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "C51": {
        "size_in_bits": 88,
        "default_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Initial_device_MTP_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Temp_update_from_customer": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Final_register_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000011001000110001010100000111000000001111000110000000101111110000000000000000000001010",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "C9": {
        "size_in_bits": 88,
        "default_value": "0000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Initial_device_MTP_value": "0000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Temp_update_from_customer": "0000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Final_register_value": "0000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "CD": {
        "size_in_bits": 56,
        "default_value": "01010101010101010101001001001000100101000000000000000000",
        "Initial_device_MTP_value": "01010101010101010101001001001000100101000000000000000000",
        "Temp_update_from_customer": "01010101010101010101001001001000100101000000000000000000",
        "Final_register_value": "01010101010101010101001001001000100101000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "01010101010101010101001001001000100101000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "D8": {
        "size_in_bits": 168,
        "default_value": "000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Initial_device_MTP_value": "000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Temp_update_from_customer": "000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Final_register_value": "000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "DF": {
        "size_in_bits": 16,
        "default_value": "0000001001010100",
        "Initial_device_MTP_value": "0000001001010100",
        "Temp_update_from_customer": "0000001001010100",
        "Final_register_value": "0000001001010100",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000001001010100",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "E10": {
        "size_in_bits": 32,
        "default_value": "00000100000101000000000000001011",
        "Initial_device_MTP_value": "00000100000101000000000000001011",
        "Temp_update_from_customer": "00000100000101000000000000001011",
        "Final_register_value": "00000100000101000000000000001011",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000100000101000000000000001011",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "E11": {
        "size_in_bits": 32,
        "default_value": "00000100000101000000000000001011",
        "Initial_device_MTP_value": "00000100000101000000000000001011",
        "Temp_update_from_customer": "00000100000101000000000000001011",
        "Final_register_value": "00000100000101000000000000001011",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000100000101000000000000001011",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "E4": {
        "size_in_bits": 16,
        "default_value": "1011001000000000",
        "Initial_device_MTP_value": "1011001000000000",
        "Temp_update_from_customer": "1011001000000000",
        "Final_register_value": "1011001000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "1011001000000000",
        "Write_command_type": "Write Word",
        "Read_command_type": "Read Word",
    }
    ,

    "E6": {
        "size_in_bits": 8,
        "default_value": "00000001",
        "Initial_device_MTP_value": "00000001",
        "Temp_update_from_customer": "00000001",
        "Final_register_value": "00000001",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000001",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "E80": {
        "size_in_bits": 128,
        "default_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Initial_device_MTP_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Temp_update_from_customer": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Final_register_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "E81": {
        "size_in_bits": 128,
        "default_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Initial_device_MTP_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Temp_update_from_customer": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Final_register_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000000000000100010000000000000000111000000000000000000000000000000000000000000000000000000011111111000000000000000001000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "E9": {
        "size_in_bits": 88,
        "default_value": "0000000001011000100000000011000000111000000001110000001100000000000000000000000000000000",
        "Initial_device_MTP_value": "0000000001011000100000000011000000111000000001110000001100000000000000000000000000000000",
        "Temp_update_from_customer": "0000000001011000100000000011000000111000000001110000001100000000000000000000000000000000",
        "Final_register_value": "0000000001011000100000000011000000111000000001110000001100000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000001011000100000000011000000111000000001110000001100000000000000000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "EF": {
        "size_in_bits": 8,
        "default_value": "11110101",
        "Initial_device_MTP_value": "11110101",
        "Temp_update_from_customer": "11110101",
        "Final_register_value": "11110101",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "11110101",
        "Write_command_type": "Write Byte",
        "Read_command_type": "Read Byte",
    }
    ,

    "F2": {
        "size_in_bits": 312,
        "default_value": "000000010110100100000010110100100000010110100100000010101111101000010101111101000010011100100000010110100100000000111100100001000000000000110111101000000001000000001000000001000000001000000001000000001000000001000000000000000000000000000000000000000000000000000000000000000000000000000011110110000000001111000010",
        "Initial_device_MTP_value": "000000010110100100000010110100100000010110100100000010101111101000010101111101000010011100100000010110100100000000111100100001000000000000110111101000000001000000001000000001000000001000000001000000001000000001000000000000000000000000000000000000000000000000000000000000000000000000000011110110000000001111000010",
        "Temp_update_from_customer": "000000010110100100000010110100100000010110100100000010101111101000010101111101000010011100100000010110100100000000111100100001000000000000110111101000000001000000001000000001000000001000000001000000001000000001000000000000000000000000000000000000000000000000000000000000000000000000000011110110000000001111000010",
        "Final_register_value": "000000010110100100000010110100100000010110100100000010101111101000010101111101000010011100100000010110100100000000111100100001000000000000110111101000000001000000001000000001000000001000000001000000001000000001000000000000000000000000000000000000000000000000000000000000000000000000000011110110000000001111000010",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "000000010110100100000010110100100000010110100100000010101111101000010101111101000010011100100000010110100100000000111100100001000000000000110111101000000001000000001000000001000000001000000001000000001000000001000000000000000000000000000000000000000000000000000000000000000000000000000011110110000000001111000010",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "F3": {
        "size_in_bits": 152,
        "default_value": "00000000000110000000001000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Initial_device_MTP_value": "00000000000110000000001000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Temp_update_from_customer": "00000000000110000000001000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Final_register_value": "00000000000110000000001000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "00000000000110000000001000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "FC0": {
        "size_in_bits": 64,
        "default_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Initial_device_MTP_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Temp_update_from_customer": "0000000000000000000000000000000000000000000000000000000000000000",
        "Final_register_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,

    "FC1": {
        "size_in_bits": 64,
        "default_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Initial_device_MTP_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Temp_update_from_customer": "0000000000000000000000000000000000000000000000000000000000000000",
        "Final_register_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Register_changed": "FALSE",
        "Customer_interaction": "NO",
        "Updated_device_MTP_value": "0000000000000000000000000000000000000000000000000000000000000000",
        "Write_command_type": "Block Write",
        "Read_command_type": "Block Read",
    }
    ,
}
stop_thread = False
PMBus_send = False
PMBUS_ACK_status = -1
initial_device_status = 1  # 1: not proper,0: proper,OR of pmbus connecton + supply is up+ dongle is connected properly

# hardware functions
def power_up_device():
    global VR_Enabled
    # here we need to turn the device , call the driver function which will turn on the device
    VR_Enabled = "ON"
    # print_log("Device turn on DLL is invoked.", "INFO")

    obj = MyThread()
    obj.power_toggle()  # it uses device_status then sets the en pin
    log_gui_interaction("Inside power_up_device()")

def power_down_device():
    global VR_Enabled
    # here we need to turn the device , call the driver function which will turn on the device
    VR_Enabled = "OFF"
    # print_log("Device turn off DLL is invoked.", "INFO")
    obj = MyThread()
    obj.power_toggle()  # it uses device_status then sets the en pin
    log_gui_interaction("Inside power_down_device()")

def is_PMBus_connected():
    global parallel_thread, PMBUS_ADDR, PAGE, PMBUS_ACK_status, queue
    log_gui_interaction("Inside is_PMBus_connected")
    # write page command and process the ACK if device is connected or not.
    # parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
    #                             pullup_enable=1, clockrate=100, writesize=2,
    #                             writedata=[0x00, 0x00])

    # to check pmbus connectivity perform a read command
    value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                       clockrate=100, writesize=1, writedata=[0x00], noofbytestoread=1)

    PAGE = 0

    # checking acknowledgement
    ack = "Pass" if PMBUS_ACK_status == 0 else "Fail"
    queue = deque()

    if ack == "Fail":
        queue.appendleft("PMBus connectivity has some issue.%ERROR")
        return "NO"

    else:
        queue.appendleft("PMBus connectivity checked and it is connected.%INFO")
        return "YES"

def usb_replace_bit(reg_address=10, reg_size=10, high_index=10, low_index=0, reg_value='110', command_format="Block"):
    if command_format == "Block":
        usb2gpio = usb_to_gpio.USB_TO_GPIO()
        value = usb2gpio.i2c_read(address=int(PMBUS_ADDR, 16), commandcode=reg_address, noofbytestoread=reg_size + 1)
        usb2gpio.close()
        # value = [10,20,30,40,50,60,70]
        print(value)
        rcvd_value = ""
        for i in range(reg_size):
            rcvd_value = bin(value[i+1]).replace("0b", "").zfill(8) + rcvd_value
        #print(rcvd_value)
        try:
            high_index_pointer = int(high_index)  # ensure input is an int
            low_index_pointer = int(low_index)  # ensure input is an int
        except:
            print("Invalid input or register key is not available")
            return "INVALID"
        if (high_index_pointer >= low_index_pointer) and (low_index_pointer >= 0):
            string_high_index = len(rcvd_value) - low_index_pointer
            string_low_index = len(rcvd_value) - high_index_pointer - 1
            # Return the binary string between two binary pointers.
            print(string_high_index)
            print(string_low_index)
            rcvd_value = rcvd_value[:string_low_index] + reg_value + rcvd_value[string_high_index:]
        else:
            return "INVALID"
        reg_data = []
        for i in range(reg_size):
            reg_data.append(int(rcvd_value[8 * i:8 * i + 8], 2))
            print(int(rcvd_value[8 * i:8 * i + 8], 2))

        reg_data = reg_data[::-1]
        # reg_data.insert(0, reg_address)
        reg_data.insert(0, reg_size)
        print(reg_address, reg_data)
        usb2gpio = usb_to_gpio.USB_TO_GPIO()
        usb2gpio.i2c_write(address=int(PMBUS_ADDR, 16), commandcode=reg_address, writesize=reg_size+1, writedata=reg_data)
        usb2gpio.close()

    else:
        usb2gpio = usb_to_gpio.USB_TO_GPIO()
        value = usb2gpio.i2c_read(address=int(PMBUS_ADDR, 16), commandcode=reg_address, noofbytestoread=reg_size)
        usb2gpio.close()
        # value = [10,20,30,40,50,60,70]
        rcvd_value = ""
        for i in range(reg_size):
            rcvd_value = bin(value[i]).replace("0b", "").zfill(8) + rcvd_value
        # print(rcvd_value)
        try:
            high_index_pointer = int(high_index)  # ensure input is an int
            low_index_pointer = int(low_index)  # ensure input is an int
        except:
            print("Invalid input or register key is not available")
            return "INVALID"
        if (high_index_pointer >= low_index_pointer) and (low_index_pointer >= 0):
            string_high_index = len(rcvd_value) - low_index_pointer
            string_low_index = len(rcvd_value) - high_index_pointer - 1
            # Return the binary string between two binary pointers.
            print(string_high_index)
            print(string_low_index)
            rcvd_value = rcvd_value[:string_low_index] + reg_value + rcvd_value[string_high_index:]
        else:
            return "INVALID"
        reg_data = []
        for i in range(reg_size):
            reg_data.append(int(rcvd_value[8 * i:8 * i + 8], 2))
            print(int(rcvd_value[8 * i:8 * i + 8], 2))

        reg_data = reg_data[::-1]
        # reg_data.insert(0, reg_address)
        print(reg_address, reg_size, reg_data)
        usb2gpio = usb_to_gpio.USB_TO_GPIO()
        usb2gpio.i2c_write(address=int(PMBUS_ADDR, 16), commandcode=reg_address, writesize=reg_size, writedata=reg_data)
        usb2gpio.close()

def read_bits(reg_address=10, reg_size=10, high_index=10, low_index=0, command_format = 'Block'):
    global parallel_thread
    if command_format == 'Block':
        value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                clockrate=100, writesize=1, writedata=[reg_address], noofbytestoread=reg_size + 1)
        print(value)
        # value = [10,20,30,40,50,60,70]
        rcvd_value = ""
        for i in range(reg_size):
            rcvd_value = bin(value[i+1]).replace("0b", "").zfill(8) + rcvd_value

        try:
            high_index_pointer = int(high_index)  # ensure input is an int
            low_index_pointer = int(low_index)  # ensure input is an int
        except:
            print("Invalid input or register key is not available")
            return "INVALID"
        if (high_index_pointer >= low_index_pointer) and (low_index_pointer >= 0):
            string_high_index = len(rcvd_value) - low_index_pointer
            string_low_index = len(rcvd_value) - high_index_pointer - 1
            # Return the binary string between two binary pointers.
            print(string_high_index)
            print(string_low_index)
            print("####", rcvd_value)
            rcvd_value = rcvd_value[string_low_index:string_high_index]
            print("@@@@@", rcvd_value)
            # rcvd_value = rcvd_value[:string_low_index] + reg_value + rcvd_value[string_high_index:]
            # print("####", rcvd_value)
            # print_log("Bit field read = "+str(rcvd_value), "INFO")
            return rcvd_value
        else:
            return "INVALID"
    else:
        value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                              clockrate=100, writesize=1, writedata=[reg_address], noofbytestoread=reg_size
                              )
        print(value)
        # value = [10,20,30,40,50,60,70]
        rcvd_value = ""
        for i in range(reg_size):
            rcvd_value = bin(value[i]).replace("0b", "").zfill(8) + rcvd_value

        try:
            high_index_pointer = int(high_index)  # ensure input is an int
            low_index_pointer = int(low_index)  # ensure input is an int
        except:
            print("Invalid input or register key is not available")
            return "INVALID"
        if (high_index_pointer >= low_index_pointer) and (low_index_pointer >= 0):
            string_high_index = len(rcvd_value) - low_index_pointer
            string_low_index = len(rcvd_value) - high_index_pointer - 1
            # Return the binary string between two binary pointers.
            print(string_high_index)
            print(string_low_index)
            print("####", rcvd_value)
            rcvd_value = rcvd_value[string_low_index:string_high_index]
            print("@@@@@", rcvd_value)
            # rcvd_value = rcvd_value[:string_low_index] + reg_value + rcvd_value[string_high_index:]
            # print("####", rcvd_value)
            # print_log("Bit field read = "+str(rcvd_value), "INFO")
            return rcvd_value
        else:
            return "INVALID"


#MTP Burn verification
def otp_test():
    global parallel_thread
    number_of_sections_written = 0
    #usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=70, low_index=69, reg_value='01')
    usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=36, low_index=36, reg_value='1')
    usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=35, low_index=24, reg_value='000000000000')
    usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=10, low_index=6, reg_value='10001')
    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                             writesize=3, writedata=[0xDD, 0x96, 0x60])
    for i in range(32):
        reg_data = bin(i).replace("0b", "").zfill(12)
        usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=35, low_index=24, reg_value=reg_data)
        parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=3, writedata=[0xDD, 0x96, 0x60])
        otp_read = read_bits(reg_address=0xDC, reg_size=8, high_index=7, low_index=0)
        # read_address = 10, read_size = 10, read_high_index = 10, read_low_index = 0, register_block_type = "Block"
        #self.trim_func.read_i2c(voltagelevel=33, address_size=0, address=PMBUS_ADDR, pullup_enable=1, clockrate=100, writesize=1, writedata=[0xDC], noofbytestoread=8)
        number_of_sections_written = i
        if(i != int(otp_read,2)):
            break
    usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=36, low_index=36, reg_value="0")
    usb_replace_bit(reg_address=int(0xFD), reg_size=9, high_index=10, low_index=10, reg_value="0")
    usb_replace_bit(reg_address=int(0xDD), reg_size=2, high_index=15, low_index=0, reg_value="0000000000000000", command_format="Non-block")
    return number_of_sections_written

# def otp_test():
#     global parallel_thread, PMBUS_ADDR
#     log_gui_interaction("Inside otp_test()")
#     number_of_sections_written = 0
#     replace_bit(reg_address=0xFD, reg_size=9, high_index=70, low_index=69, reg_value='01')
#     replace_bit(reg_address=0xFD, reg_size=9, high_index=36, low_index=36, reg_value='1')
#     parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
#                                 clockrate=100, writesize=3, writedata=[0xDD, 0x96, 0x60])
#     for i in range(32):
#         reg_data = bin(i).replace("0b", "").zfill(11)
#         replace_bit(reg_address=int(0xFD), reg_size=9, high_index=34, low_index=24,
#                     reg_value=str(reg_data))
#         otp_read = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
#                                               pullup_enable=1, clockrate=100, writesize=1, writedata=[0xDC],
#                                               noofbytestoread=9)  # [block count, 1st data byte, 2nd db, ...., 8th db]
#         number_of_sections_written = i
#         if i != otp_read[2]:
#             break
#
#     print_log("number of sections written are " + str(number_of_sections_written), "INFO")
#     return number_of_sections_written

def replace_bit(reg_address=10, reg_size=10, high_index=10, low_index=0, reg_value='110'):
    global PMBUS_ADDR, parallel_thread
    log_gui_interaction("Inside replace_bit()")
    value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                       clockrate=100, writesize=1, writedata=[reg_address], noofbytestoread=reg_size)

    rcvd_value = ""
    for i in range(reg_size):
        rcvd_value = bin(value[i]).replace("0b", "").zfill(8) + rcvd_value
    print(rcvd_value)
    try:
        high_index_pointer = int(high_index)  # ensure input is an int
        low_index_pointer = int(low_index)  # ensure input is an int
    except:
        print("Invalid input or register key is not available")
        return "INVALID"
    if (high_index_pointer >= low_index_pointer) and (low_index_pointer >= 0):
        string_high_index = len(rcvd_value) - low_index_pointer
        string_low_index = len(rcvd_value) - high_index_pointer - 1
        # Return the binary string between two binary pointers.
        rcvd_value = rcvd_value[:string_low_index] + reg_value + rcvd_value[string_high_index:]
    else:
        return "INVALID"
    reg_data = []
    for i in range(reg_size):
        reg_data.append(int(rcvd_value[8 * i:8 * i + 8], 2))
        print(int(rcvd_value[8 * i:8 * i + 8], 2))

    reg_data = reg_data[::-1]
    reg_data.insert(0, reg_address)
    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                clockrate=100, writesize=reg_size + 1, writedata=reg_data)
def testmode_entry():
    #############################################################
    # case1 : Testmode Entry                                    #
    #############################################################
    log_gui_interaction("Inside testmode_entry()")
    global parallel_thread, PMBUS_ADDR
    print(">>> Testmode Entry >>>")
    # print(input("Press Enter button if 'PROG' pin is high!!!"))

    """ Enter the Key in LOCK register : MFR_SPECIFIC_DE[15:0] - 0xBCE7 """
    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                pullup_enable=1, clockrate=100, writesize=3, writedata=[222, 231, 188])
    """ MFR_SPECIFIC_DD[2:0]  - 3'b110 to unlock OTP and Test Mode Reg """
    usb_replace_bit(reg_address=0xDD, reg_size=2, high_index=2, low_index=0, reg_value='110', command_format="Non-block")

def testmode_exit():
    #############################################################
    # case19 : Testmode Exit                                    #
    #############################################################
    log_gui_interaction("Inside testmode_exit()")
    print(">>> Testmode Exit >>>")

    usb_replace_bit(reg_address=0xDD, reg_size=2, high_index=2, low_index=0,
                reg_value='000', command_format="Non-block")  # MFR_SPECIFIC_DD[2:0]-[0x00]
    usb_replace_bit(reg_address=0xDE, reg_size=2, high_index=15,  # MFR_SPECIFIC_DE[15:0]-[0x00]
                low_index=0, reg_value='0000000000000000', command_format="Non-block")

def exec_command(vdd1, vdd2, vid_code, psi0_l, psi1_l, tfn, load_line, offset_trim, clk):
    log_gui_interaction("Inside exec_command()")
    global next_row_pointer_command_xlsx
    vdd1_s = str(vdd1)
    vdd2_s = str(vdd2)
    vid_code_s = str(bin(vid_code).replace("0b", "").zfill(8))
    psi0_l_s = str(psi0_l)
    psi1_l_s = str(psi1_l)
    tfn_s = str(tfn)
    load_line_s = str(bin(load_line).replace("0b", "").zfill(3))
    offset_trim_s = str(bin(offset_trim).replace("0b", "").zfill(2))
    svd = str("11000") + vdd1_s + vdd2_s + str("00") + psi0_l_s + vid_code_s[0:7] + str("0") + vid_code_s[7] + psi1_l_s + tfn_s + load_line_s + offset_trim_s + str("0")
    # svd = str("11000") + vdd1_s + vdd2_s + str("0") + psi0_l_s + vid_code_s[0:7] + vid_code_s[7] + psi1_l_s + tfn_s + load_line_s + offset_trim_s
    workbook = openpyxl.load_workbook('command.xlsx')
    worksheet = workbook["Sheet"]
    worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
    worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "SVI2"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = hex(int(svd[0:8], 2))
    worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = "2"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = hex(int(svd[8:], 2))
    worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = clk
    worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"

    next_row_pointer_command_xlsx += 1
    votf = decode_telemetry()  # add NI driver function here which will return ack value
    workbook.save('command.xlsx')
    # print(svd)
    # votf = 1
    return votf
def decode_telemetry():
    log_gui_interaction("Inside decode_telemetry()")
    global parallel_thread

    workbook2 = openpyxl.open('telemetry_data.xlsx')
    worksheet2 = workbook2.active

    votf_received = 0
    for x in range(8):
        # readline,telemetry = parallel_thread.svi2_telemetry(12, 0, 1, 2, 0, 21000, 0, 1, 0, 0, [], 3) #add NI driver function here which will return telemetry packets
        readline, telemetry = parallel_thread.svi2_telemetry(12, 0, 1, 2, 0, 21000, 1, 1, 65, 8,
                                                             [0xff, 0xaa, 0x55, 0xff, 0xaa, 0xaa, 0x55, 0xff], 8)
        print(telemetry)
        if (readline == 0):
            if (telemetry[0:2] == "00"):  # vdd1 voltage and current
                # print("vdd1 domain")
                vout1 = telemetry[2:11]
                c2 = worksheet2.cell(row=x + 2, column=1)
                c2.value = vout1
                iout1 = telemetry[12:20]
                c3 = worksheet2.cell(row=x + 2, column=3)
                c3.value = iout1
            elif (telemetry[0:2] == "01"):  # vdd2 voltage and current
                # print("vdd2 domain")
                vout2 = telemetry[2:11]
                c2 = worksheet2.cell(row=x + 2, column=2)
                c2.value = vout2
                iout2 = telemetry[12:20]
                c3 = worksheet2.cell(row=x + 2, column=4)
                c3.value = iout2
            elif (telemetry[0:2] == "11"):  # vdd1 voltage and vdd2 voltage
                # print("voltage domain")
                vout1 = telemetry[2:11]
                c2 = worksheet2.cell(row=x + 2, column=1)
                c2.value = vout1
                vout2 = telemetry[11:20]
                c3 = worksheet2.cell(row=x + 2, column=2)
                c3.value = vout2
            elif (telemetry[0:2] == "10"):
                # print("votf")
                votf_received = 1
                break
        else:
            # print("Timeout")

            break
    workbook2.save('telemetry_data.xlsx')
    return votf_received

def MTP_burnt_verification():
    global register_database, parallel_thread, PMBUS_ADDR, PAGE
    print("Verifying the mtp updation")
    log_gui_interaction("Inside MTP_burnt_verification()")

    testmode_entry()
    paged_0 = []
    paged_1 = []
    not_paged = []
    for i in register_database:
        if len(i) == 3:
            if i[2] == "0":
                paged_0.append(i)
            else:
                paged_1.append(i)
        else:
            not_paged.append(i)

    # 3 lists for storing the unsuccessful mtp burn command codes
    burnt_unsuccessful_for_not_paged = []
    burnt_unsuccessful_for_page_0 = []
    burnt_unsuccessful_for_page_1 = []

    # Page command  = 0 (RailA)
    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                clockrate=100, writesize=2, writedata=[0, 0])
    PAGE = 0

    for i in not_paged:
        if register_database[i]["Read_command_type"] == "Read Byte":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=100, writesize=1,
                                               writedata=[int(i, base=16)], noofbytestoread=1)

            register_database[i]["Updated_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)

        elif register_database[i]["Read_command_type"] == "Read Word":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=100, writesize=1,
                                               writedata=[int(i, base=16)], noofbytestoread=2)

            register_database[i]["Updated_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))

        elif register_database[i]["Read_command_type"] == "Block Read":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=100, writesize=1,
                                               writedata=[int(i, base=16)], noofbytestoread
                                               =(int(register_database[i]["size_in_bits"] / 8) + 1))
            value = value[1:]
            register_database[i]["Updated_device_MTP_value"] = ''.join(
                format(x, '08b') for x in reversed(value))

        if register_database[i]["Updated_device_MTP_value"] != register_database[i]["Final_register_value"]:
            burnt_unsuccessful_for_not_paged.append("0x" + i)
        else:
            pass

    for i in paged_0:
        if register_database[i]["Read_command_type"] == "Read Byte":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=100, writesize=1,
                                               writedata=[int(i[0:-1], base=16)], noofbytestoread=1)

            register_database[i]["Updated_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)

        elif register_database[i]["Read_command_type"] == "Read Word":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=100, writesize=1,
                                               writedata=[int(i[0:-1], base=16)], noofbytestoread=2)

            register_database[i]["Updated_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))

        elif register_database[i]["Read_command_type"] == "Block Read":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=100, writesize=1,
                                               writedata=[int(i[0:-1], base=16)], noofbytestoread
                                               =(int(register_database[i]["size_in_bits"] / 8) + 1))
            value = value[1:]
            register_database[i]["Updated_device_MTP_value"] = ''.join(
                format(x, '08b') for x in reversed(value))

        else:
            pass

        if register_database[i]["Updated_device_MTP_value"] != register_database[i]["Final_register_value"]:
            burnt_unsuccessful_for_page_0.append("0x" + i[:2])
        else:
            pass

    # Page command = 1 (Rail B)
    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                writesize=2, writedata=[0, 1])
    PAGE = 1
    for i in paged_1:
        if (i == "331") or (i == "621") or (i == "601") or (i == "641") or (i == "271"):
            # changing from 0x01 to 0xff (for freq switch problem)
            parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                        pullup_enable=1, clockrate=100,
                                        writesize=2, writedata=[0, 255])

        if register_database[i]["Read_command_type"] == "Read Byte":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                               clockrate=100, writesize=1,
                                               writedata=[int(i[0:-1], base=16)], noofbytestoread=1)

            register_database[i]["Updated_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)

        elif register_database[i]["Read_command_type"] == "Read Word":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                               clockrate=100, writesize=1,
                                               writedata=[int(i[0:-1], base=16)], noofbytestoread=2)

            register_database[i]["Updated_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))
            if (i == "331") or (i == "621") or (i == "601") or (i == "641") or (i == "271"):
                # changing from 0xff to 0x01 (for freq switch problem)
                parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=100,
                                            writesize=2, writedata=[0, 1])

        elif register_database[i]["Read_command_type"] == "Block Read":
            value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                               clockrate=100, writesize=1,
                                               writedata=[int(i[0:-1], base=16)], noofbytestoread
                                               =(int(register_database[i]["size_in_bits"] / 8) + 1))
            value = value[1:]
            register_database[i]["Updated_device_MTP_value"] = ''.join(
                format(x, '08b') for x in reversed(value))

        else:
            pass

        if register_database[i]["Updated_device_MTP_value"] != register_database[i]["Final_register_value"]:
            burnt_unsuccessful_for_page_1.append("0x" + i[:2])
        else:
            pass

    if len(burnt_unsuccessful_for_not_paged) != 0:
        print_log("MTP burn unsuccessful for following non-paged commands " + str(burnt_unsuccessful_for_not_paged),
                  "ERROR")
        burnt_unsuccessful_for_not_paged.clear()

    if len(burnt_unsuccessful_for_page_0) != 0:
        print_log("MTP burn unsuccessful for following non-paged commands " + str(burnt_unsuccessful_for_page_0),
                  "ERROR")
        burnt_unsuccessful_for_page_0.clear()

    if len(burnt_unsuccessful_for_page_1) != 0:
        print_log("MTP burn unsuccessful for following non-paged commands " + str(burnt_unsuccessful_for_page_1),
                  "ERROR")
        burnt_unsuccessful_for_page_1.clear()

    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                clockrate=100,
                                writesize=2, writedata=[0, 0])
    PAGE = 0

#Register database functions
def write_PMBUS_entry_in_command_xlsx(register_key):
    global register_database, next_row_pointer_command_xlsx, PMBus_freq, PMBus_parity, PAGE, stop_thread, parallel_thread
    log_gui_interaction("Inside write_PMBUS_entry_in_command_xlsx() and register key is: "+register_key)

    update_entry_in_master_command_xlsx(register_key)
    # Need to stop the parallel thread here.
    # stop_thread = True
    # time.sleep(1)
    workbook = openpyxl.load_workbook('command.xlsx')
    worksheet = workbook["Sheet"]
    if len(register_key) == 2:  # It is non paged register, so need not set the page
        pass
    else:
        if int(register_key[2]) == PAGE:  # PAGE is already set, so no need to change the PAGE
            pass
        else:
            PAGE = (PAGE + 1) % 2  # toggle the page
            worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
            worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = "0x00"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = "W"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = "1"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = "0x0"+str(PAGE)
            worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
            worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
            worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = "Write Byte"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
            worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
            next_row_pointer_command_xlsx += 1

    worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
    worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = "0x" + register_key[:2]
    worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = "W"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = str(int(register_database[register_key]["size_in_bits"] / 8))
    worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = "0x" + hex(int(register_database[register_key]["Final_register_value"], 2)).split("0x")[1].zfill(int(register_database[register_key]["size_in_bits"] / 4))
    worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
    worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
    worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = register_database[register_key]["Write_command_type"]
    worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
    workbook.save('command.xlsx')
    next_row_pointer_command_xlsx += 1
    # Need to start the thread here
    # time.sleep(0.5)
    # stop_thread = False
    # parallel_thread.start()

def update_entry_in_master_command_xlsx(register_key):
    global register_database, next_row_pointer_command_xlsx, PMBus_freq, PMBus_parity, PAGE, stop_thread, parallel_thread
    log_gui_interaction("Inside update_entry_in_master_command_xlsx() and register key is: " + register_key)
    is_paged = 'No'
    page_number = "0"
    if len(register_key)==2:
        #non paged command
        is_paged = 'No'
    else:
        is_paged = 'Yes'
        page_number=register_key[2]

    master_xlsx = openpyxl.open('master_command.xlsx')
    worksheet = master_xlsx["Sheet"]
    register_key_lower="0x"+register_key[:2].lower()
    register_key_upper = "0x"+register_key[:2].upper()
    # print(is_paged)
    for row_num in range(1,worksheet.max_row):

        if worksheet.cell(row=row_num, column=3).value == register_key_lower or worksheet.cell(row=row_num, column=3).value == register_key_upper:
            page_info_in_master = str(worksheet.cell(row=row_num, column=15).value)
            # print(worksheet.cell(row=row_num, column=3).value,type(page_info_in_master))
            if is_paged == 'No':
                worksheet.cell(row=row_num, column=6).value = "0x" + hex(int(register_database[register_key]["Final_register_value"], 2)).split("0x")[1].zfill(int(register_database[register_key]["size_in_bits"] / 4))
                worksheet.cell(row=row_num, column=15).value = "NO"
                worksheet.cell(row=row_num, column=8).value = "YES"
                # print("found")

            else:

                if page_number == "0" and page_info_in_master == "0":
                    worksheet.cell(row=row_num, column=6).value = "0x" + hex(int(register_database[register_key]["Final_register_value"], 2)).split("0x")[1].zfill(int(register_database[register_key]["size_in_bits"] / 4))
                    worksheet.cell(row=row_num, column=8).value = "YES"
                    # worksheet.cell(row=row, column=5).value = 2
                    # print("found")
                elif page_number == "1" and page_info_in_master == "1":
                    worksheet.cell(row=row_num, column=7).value = "0x" + hex(int(register_database[register_key]["Final_register_value"], 2)).split("0x")[1].zfill(int(register_database[register_key]["size_in_bits"] / 4))
                    worksheet.cell(row=row_num, column=8).value = "YES"
                    # worksheet.cell(row=row, column=5).value = 2
                    # print("found")



    master_xlsx.save('master_command.xlsx')
    return
def initialize_feature_variable(register_key, high_index_pointer, low_index_pointer):
    # inputs: string,int,int
    # register_key: it is key from register database which is dict of dict
    # high_index_pointer: higher index binary bit location
    # low_index_pointer : lower_index binary bit location
    ###################### EXAMPLE ############################################
    # we want to modify a register with register_key = "DF", it is 16 bit register and we want to
    # initialize a feature variable with binary bits stored between 8:4 bits i.e. DF[8:4] high_index_pointer = 8 low_index_pointer=4
    log_gui_interaction("Inside initialize_feature_variable() and register key is: " + register_key+"  high_index_pointer:"+str(high_index_pointer)+"  low_index_pointer:"+str(low_index_pointer))

    global register_database
    # input error handling
    try:
        register_len = register_database[register_key]["size_in_bits"]  # register length in bits # checks if register is available
        high_index_pointer = int(high_index_pointer)  # ensure input is an int
        low_index_pointer = int(low_index_pointer)  # ensure input is an int
    except:
        print("Invalid input or register key is not available")
        return "INVALID"
    if register_len > 0 and high_index_pointer >= low_index_pointer and register_len >= high_index_pointer and low_index_pointer >= 0:
        string_high_index = register_len - low_index_pointer
        string_low_index = register_len - high_index_pointer - 1
        # Return the binary string between two binary pointers.
        return register_database[register_key]["Final_register_value"][string_low_index:string_high_index]
    else:
        return "INVALID"
def update_database_with_temp_customer_input(register_key, high_index_pointer, low_index_pointer, customer_input):
    # inputs: string,int,int,string
    # register_key: it is key from register database which is dict of dict
    # high_index_pointer: higher index binary bit location
    # low_index_pointer : lower_index binary bit location
    # customer_input : binary string of length (high_index_pointer - low_index_pointer + 1)

    ###################### EXAMPLE ############################################
    # we want to modify a register with register_key = "DF", it is 16 bit register and we want to
    # update binary bits stored between 8:4 bits with "10101" i.e. DF[8:4] high_index_pointer = 8 low_index_pointer=4
    log_gui_interaction("Inside update_database_with_temp_customer_input() and register key is: " + register_key+"  high_index_pointer:"+str(high_index_pointer)+"  low_index_pointer:"+str(low_index_pointer)+"  customer_input:"+str(customer_input))

    global register_database
    # input error handling
    try:
        register_len = register_database[register_key]["size_in_bits"]  # register length in bits # checks if register is available
        high_index_pointer = int(high_index_pointer)  # ensure input is an int
        low_index_pointer = int(low_index_pointer)  # ensure input is an int
        temp_input = int(customer_input, 2)  # ensure input is a valid binary string
        customer_input_len = len(customer_input)
    except:
        print("Invalid input or register key is not available")
        return "INVALID"
    if register_len > 0 and high_index_pointer >= low_index_pointer and register_len >= high_index_pointer and low_index_pointer >= 0 and customer_input_len == (high_index_pointer - low_index_pointer + 1):
        string_high_index = register_len - low_index_pointer
        string_low_index = register_len - high_index_pointer - 1
        # Note: We are updating value stored in "Temp_update_from_customer" here because its possible that we have two features present
        # inside same frame which deal with same register key
        # Hence  during frame initialization we do,register_database[register_key]["Temp_update_from_customer"]= register_database[register_key]["Final_register_value"]

        # string left to string_low_index.
        left_string = register_database[register_key]["Temp_update_from_customer"][:string_low_index]
        # print("Left string is:", left_string)
        # string in between two indices
        # mid_string = register_database[register_key]["Final_register_value"][string_low_index:string_high_index]
        mid_string = customer_input
        # print("mid_string is:", mid_string)
        # string right to the string_high_index
        right_string = register_database[register_key]["Temp_update_from_customer"][string_high_index:]
        # print("Right_string is: ", right_string)
        final_string = left_string + mid_string + right_string
        ## Cross checking before updating the database
        if len(final_string) == int(register_database[register_key]["size_in_bits"]):
            register_database[register_key]["Temp_update_from_customer"] = final_string

        return final_string
    else:
        return "INVALID"
def initialize_Temp_update_from_customer(key_list):
    log_gui_interaction("Inside initialize_Temp_update_from_customer() and key list is: " + str(key_list))

    global register_database
    for i in key_list:
        register_database[i]["Temp_update_from_customer"] = register_database[i]["Final_register_value"]

def create_new_command_xlsx():
    log_gui_interaction("Inside create_new_command_xlsx()")

    global next_row_pointer_command_xlsx, parallel_thread
    new_xlsx = openpyxl.Workbook()
    file = 'command.xlsx'
    sheet_name = 'Sheet'
    # new_xlsx.create_sheet(sheet_name)
    new_xlsx.save(file)
    workbook = openpyxl.open(file)
    worksheet = workbook[sheet_name]
    # Create Header
    next_row_pointer_command_xlsx = 1
    parallel_thread.temp_row_number = 1
    worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = "SL.No."
    worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "Protocol_identifier"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = "Command_code"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = "Access"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = "Command_size (Bytes)"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = "command_data"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = "FREQ"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = "Parity"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "SVID address (hex)"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "Protocol_command[English]"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = "command_type"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "Tstart (ns)"
    worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "Tstop (ns)"
    next_row_pointer_command_xlsx = 2
    workbook.save(file)

def update_command_xlsx_from_master_command_xlsx():
    global parallel_thread, PMBus_freq, PMBus_parity, PMBUS_ADDR, next_row_pointer_command_xlsx
    log_gui_interaction("Inside update_command_xlsx_from_master_command_xlsx()")
    print("master_command xlsx content has been written to device")
    workbook2 = openpyxl.open('command.xlsx')
    worksheet2 = workbook2.active
    workbook3 = openpyxl.open('master_command.xlsx')
    worksheet3 = workbook3.active
    # sl_no = 1
    next_row_pointer_command_xlsx = 1
    parallel_thread.temp_row_number = 1
    # page command = 0x00 (Rail A)
    command = [str(next_row_pointer_command_xlsx), "PMBUS", "0x00", "W", "1", "0x00", PMBus_freq, PMBus_parity, "NA",
               "NA", "Write Byte", "NA", "NA"]
    worksheet2.append(command)
    next_row_pointer_command_xlsx += 1

    abc = []
    for k in worksheet3.iter_rows(min_row=2, max_row=worksheet3.max_row, values_only=True):
        if k[14] == "NO":
            abc.append(k)
    for j in abc:
        command = [str(next_row_pointer_command_xlsx), "PMBUS", j[2], "W", str(j[4]), j[5], PMBus_freq, PMBus_parity,
                   "NA", "NA", j[3], "NA", "NA"]
        next_row_pointer_command_xlsx += 1
        worksheet2.append(command)

    abc = []
    for k in worksheet3.iter_rows(min_row=2, max_row=worksheet3.max_row, values_only=True):
        if k[14] == "0":
            abc.append(k)
    for j in abc:
        if j[5] != "Its Not supported":
            command = [str(next_row_pointer_command_xlsx), "PMBUS", j[2], "W", str(j[4]), j[5], PMBus_freq,
                       PMBus_parity, "NA", "NA", j[3], "NA", "NA"]
            next_row_pointer_command_xlsx += 1
            worksheet2.append(command)

    # page command = 0x01 (Rail B)
    command = [str(next_row_pointer_command_xlsx), "PMBUS", "0x00", "W", "1", "0x01", PMBus_freq, PMBus_parity, "NA",
               "NA", "Write Byte", "NA", "NA"]
    worksheet2.append(command)
    next_row_pointer_command_xlsx += 1
    abc = []
    for k in worksheet3.iter_rows(min_row=2, max_row=worksheet3.max_row, values_only=True):
        if k[14] == "1":
            abc.append(k)
    for j in abc:
        if j[6] != "Its Not supported":
            command = [str(next_row_pointer_command_xlsx), "PMBUS", j[2], "W", str(j[4]), j[6], PMBus_freq,
                       PMBus_parity, "NA", "NA", j[3], "NA", "NA"]
            next_row_pointer_command_xlsx += 1
            worksheet2.append(command)

    workbook2.save('command.xlsx')

def update_master_command_xlsx_from_register_database():
    global register_database
    log_gui_interaction("Inside update_master_command_xlsx_from_register_database()")
    print("master_database has been updated.")
    workbook = openpyxl.open('master_command.xlsx')
    worksheet = workbook.active

    sl_no = 1

    for i in register_database:
        if len(i) == 2:
            command_code = "0x" + i[:2]
            data = [str(sl_no), "PMBUS", command_code, register_database[i]["Write_command_type"],
                    str(int(register_database[i]["size_in_bits"] / 8)),
                    "0x" + format(int(register_database[i]["Final_register_value"], 2), "02x").
                        zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                    register_database[i]["Customer_interaction"],
                    "0x" + format(int(register_database[i]["default_value"], 2), "02x").
                        zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                    "0x" + format(int(register_database[i]["Initial_device_MTP_value"], 2), "02x").
                        zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                    "0x" + format(int(register_database[i]["Final_register_value"], 2), "02x").
                        zfill(int(register_database[i]["size_in_bits"] / 4)), "", "NO"]
            worksheet.append(data)
            sl_no += 1

        else:
            if i[2] == "0":
                command_code = "0x" + i[:2]
                if i == "290":
                    data = [str(sl_no), "PMBUS", command_code, register_database[i]["Write_command_type"],
                            str(int(register_database[i]["size_in_bits"] / 8)),
                            "Its Not supported", "", register_database[i]["Customer_interaction"],
                            "Its Not supported", "", "Its Not supported", "", "Its Not supported", "", "0"]
                else:
                    data = [str(sl_no), "PMBUS", command_code, register_database[i]["Write_command_type"],
                            str(int(register_database[i]["size_in_bits"] / 8)),
                            "0x" + format(int(register_database[i]["Final_register_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                            register_database[i]["Customer_interaction"],
                            "0x" + format(int(register_database[i]["default_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                            "0x" + format(int(register_database[i]["Initial_device_MTP_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                            "0x" + format(int(register_database[i]["Final_register_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "", "0"]
                worksheet.append(data)
                sl_no += 1
            else:
                command_code = "0x" + i[:2]
                if i == "291":
                    data = [str(sl_no), "PMBUS", command_code, register_database[i]["Write_command_type"],
                            str(int(register_database[i]["size_in_bits"] / 8)), "",
                            "Its Not supported", register_database[i]["Customer_interaction"], "",
                            "Its Not supported", "", "Its Not supported", "", "Its Not supported", "1"]
                else:
                    data = [str(sl_no), "PMBUS", command_code, register_database[i]["Write_command_type"],
                            str(int(register_database[i]["size_in_bits"] / 8)), "",
                            "0x" + format(int(register_database[i]["Final_register_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)),
                            register_database[i]["Customer_interaction"],
                            "", "0x" + format(int(register_database[i]["default_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                            "0x" + format(int(register_database[i]["Initial_device_MTP_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "",
                            "0x" + format(int(register_database[i]["Final_register_value"], 2), "02x").
                                zfill(int(register_database[i]["size_in_bits"] / 4)), "1"]
                worksheet.append(data)
                sl_no += 1
        workbook.save('master_command.xlsx')

def update_register_database_from_master_command_xlsx():
    log_gui_interaction("Inside update_register_database_from_master_command_xlsx()")

    print("Updating register database from master command xlsx")
    workbook5 = openpyxl.open('master_command.xlsx')
    worksheet5 = workbook5.active

    for k in worksheet5.iter_rows(min_row=2, max_row=worksheet5.max_row, values_only=True):
        if k[14] == "0":
            if k[5] != "Its Not supported":
                register_database[k[2][2:] + "0"]["Final_register_value"] = bin(int(k[5], 16))[2:].zfill(int(int(k[4]) * 8))
                register_database[k[2][2:] + "0"]["Initial_device_MTP_value"] = bin(int(k[5], 16))[2:].zfill(int(int(k[4]) * 8))
        elif k[14] == "1":
            if k[6] != "Its Not supported":
                register_database[k[2][2:] + "1"]["Final_register_value"] = bin(int(k[6], 16))[2:].zfill(int(int(k[4]) * 8))
                register_database[k[2][2:] + "1"]["Initial_device_MTP_value"] = bin(int(k[6], 16))[2:].zfill(int(int(k[4]) * 8))

        else:
            register_database[k[2][2:]]["Final_register_value"] = bin(int(k[5], 16))[2:].zfill(int(int(k[4]) * 8))
            register_database[k[2][2:]]["Initial_device_MTP_value"] = bin(int(k[5], 16))[2:].zfill(int(int(k[4]) * 8))

def create_new_master_command_xlsx():
    # global next_row_pointer_command_xlsx
    log_gui_interaction("Inside create_new_master_command_xlsx()")

    new_xlsx = openpyxl.Workbook()
    file = 'master_command.xlsx'
    sheet_name = 'Sheet'
    # new_xlsx.create_sheet(sheet_name)
    new_xlsx.save(file)
    workbook = openpyxl.open(file)
    worksheet = workbook[sheet_name]
    # Create Header
    # next_row_pointer_command_xlsx = 1
    worksheet.cell(row=1, column=1).value = "SL.No."
    worksheet.cell(row=1, column=2).value = "Protocol_identifier"
    worksheet.cell(row=1, column=3).value = "Command_code"
    worksheet.cell(row=1, column=4).value = "Write_command_type"
    worksheet.cell(row=1, column=5).value = "Command_size (Bytes)"

    worksheet.cell(row=1, column=6).value = "command_data[Page0/non-paged]"  # this is the latest content in present in register database
    worksheet.cell(row=1, column=7).value = "command_data[Page1]"  # this is the latest content in present in register database
    worksheet.cell(row=1, column=8).value = "Customer Interaction"  # make yes if customer has changed it.
    worksheet.cell(row=1, column=9).value = "MFR_default_command_data[Page0/non-paged]"  # MFR default value, it is constant
    worksheet.cell(row=1, column=10).value = "MFR_defaultcommand_data[Page1]"  # MFR default value, it is constant
    worksheet.cell(row=1, column=11).value = "Initial_Stored_device_MTP_command_data[Page0/non-paged]"  # contains initial mtp settings stored on device
    worksheet.cell(row=1, column=12).value = "Initial_Stored_device_MTP_command_data[Page1]"
    worksheet.cell(row=1, column=13).value = "Final_Stored_device_MTP_command_data[Page0/non-paged]"  # once we burn the MTP we verify it , by default make it all zero
    worksheet.cell(row=1, column=14).value = "Final_Stored_device_MTP_command_data[Page1]"  # continued: it will be used to verify successful MTP burning.
    worksheet.cell(row=1, column=15).value = "PAGED?"
    workbook.save(file)

#Other
def print_log(string_to_be_printed, info_type):
    global homeWin_obj
    # Logger colors
    COLORS = {"DEBUG": 'blue', "INFO": 'black', "WARNING": 'orange', "ERROR": 'red'}
    string_balancing = {"DEBUG": 'DEBUG  ', "INFO": 'INFO   ', "WARNING": 'WARNING', "ERROR": 'ERROR  '}
    if info_type in COLORS.keys():
        # loggerq
        string_to_be_printed = "    " + datetime.now().strftime("%H:%M:%S") + " - " + string_balancing[info_type] + " - " + string_to_be_printed
        color = COLORS[info_type]
        s = '<pre><font color="%s">%s</font></pre>' % (color, string_to_be_printed)
        homeWin_obj.Logger_frame.appendHtml(s)

def log_gui_interaction(input_log_string):
    global GUI_log_file_name
    input_log_string = str(input_log_string)
    try:
        GUI_log_file = open('GUI_LOG\\'+GUI_log_file_name, "a")
        GUI_log_file.write(str(datetime.now())+" : "+input_log_string + "\n")
        GUI_log_file.close()
    except :
        print("gui log interaction invoked before main()")

# frame specific custom functions
def write_feature_variable(register_key, high_index_pointer, low_index_pointer):
    # inputs: string,int,int
    # register_key: it is key from register database which is dict of dict
    # high_index_pointer: higher index binary bit location
    # low_index_pointer : lower_index binary bit location
    ###################### EXAMPLE ############################################
    # we want to modify a register with register_key = "DF", it is 16 bit register and we want to
    # initialize a feature variable with binary bits stored between 8:4 bits i.e. DF[8:4] high_index_pointer = 8 low_index_pointer=4
    global register_database
    log_gui_interaction("Inside write_feature_variable() and register key is: " + register_key+"  high_index_pointer:"+str(high_index_pointer)+"  low_index_pointer:"+str(low_index_pointer))

    # input error handling
    try:
        register_len = register_database[register_key]["size_in_bits"]  # register length in bits # checks if register is available
        high_index_pointer = int(high_index_pointer)  # ensure input is an int
        low_index_pointer = int(low_index_pointer)  # ensure input is an int
    except:
        print_log("Invalid input or register key is not available", "ERROR")
        return "INVALID"
    if register_len > 0 and high_index_pointer >= low_index_pointer and register_len >= high_index_pointer and low_index_pointer >= 0:
        string_high_index = register_len - low_index_pointer
        string_low_index = register_len - high_index_pointer - 1
        # Return the binary string between two binary pointers.
        return register_database[register_key]["Temp_update_from_customer"][string_low_index:string_high_index]
    else:
        return "INVALID"

def resolution_calculation(temp_update):
    # for SVID frame only temp_update should be 1 for all other cases temp_update should be 0
    global resolutionA, resolutionB, parallel_thread
    log_gui_interaction("Inside resolution_calculation()")

    if temp_update == 1:
        if int(write_feature_variable('010', 5, 4), 2) == 0:  # pmbus override
            if int(write_feature_variable('200', 4, 0), 2) == 30:
                resolutionA = '6.25'
            elif int(write_feature_variable('E4', 3, 3), 2) == 1:
                resolutionA = '10'
            else:
                resolutionA = '5'
        elif int(write_feature_variable('E10', 25, 25), 2) == 1:
            if int(write_feature_variable('E10', 24, 24), 2) == 1:
                resolutionA = '10'
            else:
                resolutionA = '5'
        else:
            if parallel_thread.VID_SEL() == 1:
                resolutionA = '10'
            else:
                resolutionA = '5'

        if int(write_feature_variable('011', 5, 4), 2) == 0:
            if int(write_feature_variable('201', 4, 0), 2) == 30:
                resolutionB = '6.25'
            elif int(write_feature_variable('E4', 3, 3), 2) == 1:
                resolutionB = '10'
            else:
                resolutionB = '5'
        elif int(write_feature_variable('E11', 25, 25), 2) == 1:
            if int(write_feature_variable('E11', 24, 24), 2) == 1:
                resolutionB = '10'
            else:
                resolutionB = '5'
        else:
            if parallel_thread.VID_SEL() == 1:
                resolutionB = '10'
            else:
                resolutionB = '5'

    else:
        if int(initialize_feature_variable('010', 5, 4), 2) == 0:
            if int(initialize_feature_variable('200', 4, 0), 2) == 30:
                resolutionA = '6.25'
            elif int(initialize_feature_variable('E4', 3, 3), 2) == 1:
                resolutionA = '10'
            else:
                resolutionA = '5'
        elif int(initialize_feature_variable('E10', 25, 25), 2) == 1:
            if int(initialize_feature_variable('E10', 24, 24), 2) == 1:
                resolutionA = '10'
            else:
                resolutionA = '5'
        else:
            if parallel_thread.VID_SEL() == 1:
                resolutionA = '10'
            else:
                resolutionA = '5'

        if int(initialize_feature_variable('011', 5, 4), 2) == 0:
            if int(initialize_feature_variable('201', 4, 0), 2) == 30:
                resolutionB = '6.25'
            elif int(initialize_feature_variable('E4', 3, 3), 2) == 1:
                resolutionB = '10'
            else:
                resolutionB = '5'
        elif int(initialize_feature_variable('E11', 25, 25), 2) == 1:
            if int(initialize_feature_variable('E11', 24, 24), 2) == 1:
                resolutionB = '10'
            else:
                resolutionB = '5'
        else:
            if parallel_thread.VID_SEL() == 1:
                resolutionB = '10'
            else:
                resolutionB = '5'
    return

# user input function on master command xlsx loading or not
def master_command_xlsx_user_input():
    global stop_thread, initial_device_status, mtp_load_obj
    log_gui_interaction("Inside master_command_xlsx_user_input()")

    if os.path.isfile("master_command.xlsx"):
        master_command_popup = QMessageBox()
        master_command_popup.setWindowTitle("Information")
        master_command_popup.setIcon(QMessageBox.Information)
        master_command_popup.setText("Last device settings found.\n Do you want to use it?")
        master_command_popup.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        master_command_popup.setDefaultButton(QMessageBox.No)
        master_command_result = master_command_popup.exec_()

        if master_command_result == QMessageBox.No:
            create_new_command_xlsx()
            create_new_master_command_xlsx()
            wrong_combo_popup = QMessageBox()
            wrong_combo_popup.setWindowTitle("MTP load")
            wrong_combo_popup.setText("Loading MTP, Please wait while MTP loads...")
            wrong_combo_popup.exec_()
            # wrong_combo_popup.about("Loading MTP", "Please wait while MTP loads...")

            mtp_load_obj = MTPLoadingWindow()
            mtp_load_obj.show()
            # start the MTP initialization
            mtp_init_obj = MTP_register_database()
            mtp_init_obj.start()
            time.sleep(3)
            # update_master_command_xlsx_from_register_database()

        elif master_command_result == QMessageBox.Yes:
            print("Copy the master_command_xlsx content to command.xlsx and run the parallel thread.")
            stop_thread = True
            time.sleep(1)
            update_command_xlsx_from_master_command_xlsx()  # call the parallel thread when all master_command xlsx has been copied.
            stop_thread = False
            parallel_thread.start()
            # Need to read the device and refresh the register database.
            # No need to call the MTPLoadingWindow() class, updated the register database
            # from master_command_xlsx
            update_register_database_from_master_command_xlsx()

    else:
        print("master_command.xlsx not found.")
        create_new_command_xlsx()
        create_new_master_command_xlsx()
        wrong_combo_popup = QMessageBox()
        wrong_combo_popup.setWindowTitle("MTP load")
        wrong_combo_popup.setText("Loading MTP, Please wait while MTP loads...")
        wrong_combo_popup.exec_()

        mtp_load_obj = MTPLoadingWindow()
        mtp_load_obj.show()
        # start the MTP initialization
        mtp_init_obj = MTP_register_database()
        mtp_init_obj.start()

        print("update master command xlsx")
        update_master_command_xlsx_from_register_database()

nsleep = lambda x: time.sleep(x/1000000000.0)

# internal GUI related global variables
global RailA_name, RailB_name, next_row_pointer_command_xlsx, PMBus_freq, PMBus_parity, PARTNAME, counter, PMBUS_ADDR, list_of_registers_used_in_this_frame,GUI_log_file_name
global slew_rate, offset, homeWin_obj, device_status, vid_max_value, vid_min_value, svid_protocol, VR_Enabled, PAGE, svid_address, load_settings_done
RailA_name = "Rail-A"  # It is internal to GUi and used for display only.
RailB_name = "Rail-B"
PMBus_freq = "100k"  #
PMBus_parity = "DISABLE"  #
next_row_pointer_command_xlsx = 1
PARTNAME = "Not defined"
counter = 0
PMBUS_ADDR = "75"
load_settings_done = "NO"
list_of_registers_used_in_this_frame = []
offset = {
    "5": 245, "6.25": 0, "10": 190
}
slew_rate = {
    "2.5": 0xC280,
    "5": 0xCA80,
    "7.5": 0xCBC0,
    "10": 0xD280,
    "12.5": 0xD320,
    "15": 0xD3C0,
    "17.5": 0xDA30,
    "20": 0xDA80,
    "22.5": 0xDAD0,
    "25": 0xDB20,
    "27.5": 0xDB70,
    "30": 0xDBC0,
    "32.5": 0xE208,
    "35": 0xE230,
    "37.5": 0xE258,
    "40": 0xE280,
    "48": 0xE300,
    "60": 0xE3C0,
    "80": 0xEA80,
    "96": 0xEB00,
    "125": 0xEBE8,
    "0.12": 0x9BD7,
    "0.5": 0xB200,
    "1": 0xBA00
}
device_status = "OFF"  # ON or OFF
svid_protocol = {
    "0": 5, "1": 8, "2": 14
}
vid_max_value = {
    '6.25': 1.55,
    '5': 1.52,
    '10': 2.74
}
vid_min_value = {
    '6.25': 0,
    '5': 0.25,
    '10': 0.2
}
VR_Enabled = "OFF"  # ON or OFF
PAGE = 0
svid_address = "0"
# GUI_log_file_name = 'gui_log.txt'
# feature related global variables # these needs to be initialized with values stored on device during MTP read inside each frame class
global RailA_phase_count_arg, RailB_phase_count_arg, PS1_active_phases_argA, PS1_active_phases_argB, PS2_active_phases_argA, PS2_active_phases_argB
global resolutionA, resolutionB
RailA_phase_count_arg = ""  # "5"
RailB_phase_count_arg = ""  # "4"
PS1_active_phases_argA = ""
PS1_active_phases_argB = ""
PS2_active_phases_argA = ""
PS2_active_phases_argB = ""
resolutionA = ""
resolutionB = ""

# feature related error handling global variables # these are facts associated with device
# They are hard coded here itself
global RailA_phase_count_max, RailB_phase_count_max, RailAB_total_phase
RailA_phase_count_max = "9"
RailB_phase_count_max = "4"
RailAB_total_phase = "9"

# MTP read function will be called here .. MTP_read() # Now all the database registers have been updated.
# For now assume all entries in database registers have been updated.
# register_database["DF"]["Final_register_value"] = register_database["DF"]["Initial_device_MTP_value"]

# Initializing feature related global variables.
# write a generic function to read a selected bits from a register
# call this function to get the string bits between two bit locations..for example DF[8:4] in 16 bit DF register
# Map this binary number to the actual global variable and assign it. For ex: ""00001" from DF[8:4]
# which means RailA_phase_count_arg = str(bin(DF[8:4]))
# Example RailA_phase_count_arg # 5 bit string between DF[8:4]
# temp_val = register_database["DF"]["Final_register_value"]  # Extract 8:4 bits from DF
# for given feature, we want to change the bits from 8th to 4th out of 16 bit. # changing 8:4 bits comes from xlsx

# few of the variables assigning it temperorily as they are being assigned in one frame but being used in other frame without initialization
resolutionA = "6.25"  # in mv
resolutionB = "6.25"  # in mV
RailA_phase_count_arg = "5"  # "5"
RailB_phase_count_arg = "4"
svid_address = str(hex(int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", ""))


# Parallel thread class
class MyThread(QThread):
    global myWin, PARTNAME, homeWin_obj, stop_thread, PMBUS_ADDR

    def __init__(self, parent=None, enable_dio=True):
        super(MyThread, self).__init__(parent)

        # self.enable_i2c = enable_i2c
        # self.enable_spi = enable_spi
        self.enable_dio = enable_dio


        """ Initialize i2c, spi, dio modules from NI8452 package
        """
        '''
        if enable_i2c:
            self.i2c = I2C()
        if enable_spi:
            self.spi = SPI()
        if enable_dio:
            self.dio = DIO()

        """ Opens and sets the VoltageLevel of NI8452 device
        """

        if enable_i2c:
            self.i2c.ni845xOpen()
            #self.i2c.ni845xSetIoVoltageLevel(i2c_voltagelevel)
        if enable_spi:
            self.spi.ni845xOpen()
            #self.spi.ni845xSetIoVoltageLevel(spi_voltagelevel)
        if enable_dio:
            self.dio.ni845xOpen()
            self.dio.ni845xSetIoVoltageLevel(dio_voltagelevel)
        '''
        """ Initializes Protocol Configurations
        """
        """
        if self.enable_i2c:
            self.i2c.ni845xI2cConfigurationOpen()
        if self.enable_spi:
            self.spi.ni845xSpiConfigurationOpen()
        if self.enable_dio:
            self.dio.ni845xDioConfigurationOpen()
        """
        """ Initialize the slave address for PMBUS device
        """
        self.slave_addr = int(PMBUS_ADDR, 16) << 1
        self.temp_row_number = 1

        """ Initialize variable for part name of the device
        """

        """ Initialize variable for telemetry row entries
        """
        self.entries = 0

        # """ Initializes the command.xlsx workbook
        # """
        # self.wb = load_workbook("C:\\Users\\3329\\Desktop\\AMPS\\Command_1.xlsx")
        # self.ws = self.wb.active
        # self.totalnoofrow = self.ws.max_row

        """ Initializes the response.xlsx workbook
        """
        self.workbook = Workbook()
        self.save_sheet = self.workbook.active

        header = [["SL.NO.", "Protocol_identifier", "Command_code", "data_size(in bytes)", "ACK", "Data",
                   "Process call command code", "Parity_status", "Transaction status"]]
        for i in header:
            self.save_sheet.append(i)

        """ Initializes the Display_response.xlsx workbook
        """
        self.workbook2 = Workbook()
        self.sheet = self.workbook2["Sheet"]
        self.save_sheet2 = self.workbook2.active
        row = 1
        self.save_sheet2.cell(row, 1).value = "SL.NO / PARTNAME"

        header2 = [[PARTNAME, "VIN", "PSYS", "IMON AUX", "VOUTA", "IOUTA", "TMONA", "VOUTB", "IOUTB", "TMONB",
                    "POUTA", "POUTB"]]
        for i in header2:
            self.save_sheet2.append(i)

        header3 = [["Register details", "READ_VIN (88h)", "READ_PIN (97h)", "svid read (15h)", "READ_VOUT (8Bh)",
                    "READ_IOUT (8Ch)", "READ_TEMPERATURE_1 (8Dh)", "READ_VOUT (8Bh)", "READ_IOUT (8Ch)",
                    "READ_TEMPERATURE_1 (8Dh)", "READ_POUT (96h)", "READ_POUT (96h)"]]
        for i in header3:
            self.save_sheet2.append(i)

        self.save_sheet2.cell(row=28, column=1).value = "Fault:" + PARTNAME

        header4 = [["Fault name", "VIN Under voltage", "VIN over voltage", "VOUTA fixed Overvoltage",
                    "VOUTA tracking Overvoltage", "VOUTB fixed Overvoltage", "VOUTB tracking Overvoltage",
                    "VOUTA fixed undervoltage", "VOUTA tracking undervoltage", "VOUTB fixed undervoltage",
                    "VOUTB tracking undervoltage", "Over Temp A", "Over Temp B", "IOUT OverCurrent A",
                    "IOUT Over Current B", "PSYS Critical", "SPS  VDD Under voltage A", "SPS  VDD Under voltage B",
                    "SPS VIN Under voltage", "SPS Over current", "SPS Over Temperature A", "SPS Over Temperature B",
                    "SPS High Side FET short", "Controller Over temperature"]]
        for i in header4:
            self.save_sheet2.append(i)

    def crc(self, msg, div='100000111', code='00000000'):
        """
        Cyclic Redundancy Check generates an error detecting code based on an inputted message
        and divisor in the form of a polynomial representation.
        Arguments:
            msg: The input message of which to generate the output code.
            div: The divisor in polynomial form. For example, if the polynomial
                 of x^8 + x^2 + x + 1 is given, this should be represented as '100000111' in the div argument.
            code: This is an option argument where a previously generated code may
                  be passed in. This can be used to check validity. If the input
                  code produces an output code of all zeros, then the message has no errors.
        Returns:
            An error-detecting code generated by the message and the given divisor.


        crc_8 calculation method:
        Polynomial = x^8 + x^2 + x + 1  --> 100000111
        message = 0101 1100 = 0x5c

        # 0101110000000000 =CRC
        #  100000111 XOR polynomial
        #  001110111000000 =CRC
        #    100000111 XOR polynomial
        #    0110110110000 =CRC
        #     100000111 XOR polynomial
        #     010110001000 =CRC
        #      100000111 XOR polynomial
        #      00110010100 =CRC
        #        100000111 XOR polynomial
        #        010010011 =CRC
        """

        msg = msg + code  # Append the code to the message. If no code is given, default to '00000000'

        msg = list(msg)  # Convert msg and div into list form for easier handling
        div = list(div)

        for i in range(len(msg) - len(code)):  # Loop over every message bit (minus the appended code)
            if msg[i] == '1':  # If that message bit is 1, perform xor operation
                for j in range(len(div)):
                    msg[i + j] = str(
                        (int(msg[i + j]) ^ int(div[j])))  # Perform xor operation on each index of the divisor

        return ''.join(msg[-len(code):])  # Output the last error-checking code portion of the message generated

    def i2c_open(self, voltagelevel=33):
        self.usb_to_gpio = usb_to_gpio.USB_TO_GPIO()
        self.usb_to_gpio.configure(pec_enabled=False)


    def spi_open(self, voltagelevel=12):
        return
        # self.spi = SPI()
        # self.spi.ni845xOpen()
        # self.spi.ni845xSetIoVoltageLevel(voltagelevel)
        # self.spi.ni845xSpiConfigurationOpen()

    def i2c_close(self):
        self.usb_to_gpio.close()

    def spi_close(self):
        return
        # self.spi.ni845xSpiConfigurationClose()
        # self.spi.ni845xClose()

    def usb_to_gpio_send_byte(self, voltagelevel=33, address_size=0, address=0x00, pullup_enable=1,
                    clockrate=100, writesize=0, writedata=[0x00]):
        self.i2c_open(voltagelevel)
        self.usb_to_gpio.send_byte(address=address, commandcode=writedata[0])
        PMBUS_ACK_status = self.usb_to_gpio.status  # returns 0 if i2c transaction is passed, else a non-zero value
        self.i2c_close()

    def write_PMBus(self, voltagelevel=33, address_size=0, address=0x00, pullup_enable=1,
                    clockrate=100, writesize=0, writedata=[0x00]):
        global PMBUS_ACK_status
        """ Configures required parameters and writes data into specified register address
        """
        self.i2c_open(voltagelevel)
        # self.i2c.ni845xI2cConfigurationOpen()
        # self.i2c.ni845xSetIoVoltageLevel(voltagelevel)
        # self.i2c.ni845xI2cConfigurationSetAddressSize(address_size)  # Sets the configuration address size
        # self.i2c.ni845xI2cConfigurationSetAddress(address)  # Sets the configuration address
        # self.i2c.ni845xI2cSetPullupEnable(pullup_enable)
        # self.i2c.ni845xI2cConfigurationSetClockRate(clockrate)  # Sets the configuration clock rate in kilohertz
        # self.i2c.ni845xI2cWrite(writesize, writedata)  # Write an array of data into an I2C slave device
        # self.i2c.ni845xI2cConfigurationClose()
        self.usb_to_gpio.i2c_write(address=address, commandcode=writedata[0], writesize=writesize-1, writedata=writedata[1:])
        PMBUS_ACK_status = self.usb_to_gpio.status  # returns 0 if i2c transaction is passed, else a non-zero value
        self.i2c_close()

    """def write_read_PMBus(self, voltagelevel=33, address_size=0, address=0x00, pullup_enable=1, clockrate=100, 
    writesize=0, writedata=[0x00], noofbytestoread=1): #Configures required parameters and writes data into specified 
    register address 

        self.i2c_open(voltagelevel) self.i2c.ni845xI2cConfigurationSetAddressSize(address_size)                 # 
        Sets the configuration address size self.i2c.ni845xI2cConfigurationSetAddress(address)                        
          # Sets the configuration address self.i2c.ni845xI2cSetPullupEnable(pullup_enable) 
          self.i2c.ni845xI2cConfigurationSetClockRate(clockrate)                      # Sets the configuration clock 
          rate in kilohertz print("!!!!!", self.i2c.ni845xI2cConfigurationGetClockRate()) val = 
          self.i2c.ni845xI2cWriteRead(writesize, writedata, noofbytestoread) self.i2c_close() return val """

    def read_PMBus(self, voltagelevel=33, address_size=0, address=0x00, pullup_enable=1, clockrate=100,
                   writesize=1, writedata=[0x00], noofbytestoread=1):
        global PMBUS_ACK_status
        try:
            self.i2c_open(voltagelevel)
            # self.i2c.ni845xI2cConfigurationOpen()
            # self.i2c.ni845xSetIoVoltageLevel(voltagelevel)
            # self.i2c.ni845xI2cConfigurationSetAddressSize(address_size)  # Sets the configuration address size
            # self.i2c.ni845xI2cConfigurationSetAddress(address)  # Sets the configuration address
            # self.i2c.ni845xI2cSetPullupEnable(pullup_enable)
            # self.i2c.ni845xI2cConfigurationSetClockRate(clockrate)  # Sets the configuration clock rate in kilohertz
            read_pmbus = self.usb_to_gpio.i2c_read (address=address, commandcode=writedata[0],  noofbytestoread=noofbytestoread)        # Reads an array of data from an I2C slave device
            PMBUS_ACK_status =self.usb_to_gpio.status # this needs to be updated . It has issue. As of now , it has been bypassed.
            self.i2c_close()
            # self.i2c.ni845xI2cConfigurationClose()
            # print("%%%%%%%%%%%%%", read_pmbus)
            return read_pmbus

            # self.response_excel_write(read_pmbus)

        except Exception as error:
            self.update_error(error)

    def i2c_protocol(self):
        global PMBUS_ADDR, PMBus_custom_obj, PMBus_send
        # PMBus_custom_obj = PMBus_custom_commands()
        """ Configures the PMBus formats for read and write commands
        """
        config_clockrate = int(''.join(filter(str.isdigit, self.row[6])))

        if config_clockrate == 1:
            config_clockrate = 1000
        if self.row[3] == ('R' or "r"):
            if self.row[10] == 'Read Byte':  # Read Byte
                commandcode, commandsize = self.row[2], int(self.row[4])  # Command code and Command size
                commandcode = int(commandcode, base=16)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=1,
                                            writedata=[commandcode], noofbytestoread=int(commandsize + 1))
                    m = [self.slave_addr, commandcode, self.slave_addr + 1]
                    m.extend(value[0:-1])
                    print(value[0:-1])
                    print(m)
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break

                    print(pec)  # remove
                    if pec == value[-1]:
                        pec_status = "Pass"
                    else:
                        pec_status = "Fail"

                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA',
                              ('0x' + ''.join(format(value[0], '02x'))), 'NA', pec_status, ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)

                else:
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=1,
                                            writedata=[commandcode], noofbytestoread=int(commandsize))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA',
                              ('0x' + ''.join(format(value[0], '02x'))), 'NA', self.row[7], ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)

                # Passing read byte data to pmbus custom command frame
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(False)
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText(''.join(format(value[0], '02x')))

            elif self.row[10] == 'Read Word':  # Read Word
                commandcode, commandsize = self.row[2], int(self.row[4])  # Command code and Command size
                commandcode = int(commandcode, base=16)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=1,
                                            writedata=[commandcode], noofbytestoread=int(commandsize + 1))
                    m = [self.slave_addr, commandcode, self.slave_addr + 1]
                    m.extend(value[0:-1])
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break

                    if pec == value[-1]:
                        pec_status = "Pass"
                    else:
                        pec_status = "Fail"
                    value = value[:-1]
                    value = ''.join(format(x, '02x') for x in reversed(value))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', ('0x' + value), 'NA',
                              pec_status, ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)

                else:
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                            clockrate=config_clockrate, writesize=1, writedata=[commandcode],
                                            noofbytestoread=int(commandsize))
                    value = ''.join(format(x, '02x') for x in reversed(value))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', ('0x' + value), 'NA',
                              self.row[7], ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)
                # print(value)

                # Passing read byte data to pmbus custom command frame
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(False)
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText(value)

            elif self.row[10] == 'Block Read':  # Block Read
                commandcode, commandsize = self.row[2], int(self.row[4])
                commandcode = int(commandcode, base=16)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=1,
                                            writedata=[commandcode], noofbytestoread=int(commandsize + 2)
                                            # Command code and Command size + no of bytes to read + read pec byte
                                            )

                    m = [self.slave_addr, commandcode, self.slave_addr + 1]
                    m.extend(value[0:-1])
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break

                    if pec == value[-1]:
                        pec_status = "Pass"
                    else:
                        pec_status = "Fail"
                    value = value[1:-1]
                    value = ''.join(format(x, '02x') for x in reversed(value))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', ('0x' + value), 'NA',
                              pec_status, ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)

                else:
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=1,
                                            writedata=[commandcode], noofbytestoread=int(commandsize + 1)
                                            # Command code and Command size + no of bytes to read
                                            )
                    value = value[1:]
                    value = ''.join(format(x, '02x') for x in reversed(value))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', ('0x' + value), 'NA',
                              self.row[7], ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)

                # Passing read byte data to pmbus custom command frame
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(False)
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText(value)

            elif self.row[10] == 'Process Call':  # Process Call read
                commandcode, commandsize, commanddata = (self.row[2], int(self.row[4]), self.row[5])
                commandcode = int(commandcode, base=16)
                commanddata = ''.join(commanddata[2:])
                size = int(len(commanddata) / 2)
                data = []
                for i in range(size):
                    byte = commanddata[-2:]
                    byte = int(byte, base=16)
                    data.append(byte)
                    commanddata = commanddata[: len(commanddata) - 2]
                writevalue = [commandcode, commandsize]
                writevalue.extend(data)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize + 2),
                                            writedata=writevalue, noofbytestoread=int(commandsize + 2)
                                            )
                    m = [self.slave_addr, commandcode, commandsize]
                    m.extend(data)
                    m.extend([self.slave_addr + 1])
                    m.extend(value[0:-1])
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break

                    if pec == value[-1]:
                        pec_status = "Pass"
                    else:
                        pec_status = "Fail"
                    value = value[1:-1]
                    value = ''.join(format(x, '02x') for x in reversed(value))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', ('0x' + value), self.row[5],
                              pec_status, ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)

                else:
                    value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize + 2),
                                            writedata=writevalue, noofbytestoread=int(commandsize + 1)
                                            )
                    value = value[1:]
                    value = ''.join(format(x, '02x') for x in reversed(value))
                    result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', ('0x' + value), self.row[5],
                              self.row[7], ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                    self.response_excel_write(save_data=result)
                    # print(value)

                # Passing read byte data to pmbus custom command frame
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(False)
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText(value)

        else:
            if self.row[10] == 'Write Byte':  # Write Byte
                commandcode, commandsize, commanddata = (self.row[2], int(self.row[4]) + 1, self.row[5])
                commandcode = int(commandcode, base=16)
                commanddata = int(commanddata, base=16)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    m = [self.slave_addr, commandcode, commanddata]
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break

                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize + 1),
                                     writedata=[commandcode, commanddata, pec])

                else:
                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize),
                                     writedata=[commandcode, commanddata])

                result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', self.row[5], 'NA', self.row[7],
                          ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))

                # Passing ack and received data to PMBus custom command
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(True)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText("--")
                self.response_excel_write(save_data=result)

            elif self.row[10] == 'Send Byte':  # Send Byte
                commandsize, commandcode = (int(self.row[4]), self.row[2])
                commandcode = int(commandcode, base=16)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    m = [self.slave_addr, commandcode]
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break

                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize + 1),
                                     writedata=[commandcode, pec])

                else:
                    # self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                    #                  pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize),
                    #                  writedata=[commandcode])

                    # send byte function in usb2any adapter has a separate driver function
                    self.usb_to_gpio_send_byte(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                               pullup_enable=1, clockrate=config_clockrate, writesize=0,
                                               writedata=[commandcode])

                result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', self.row[5], 'NA', self.row[7],
                          ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                self.response_excel_write(save_data=result)

            elif self.row[10] == 'Write Word':  # Write Word
                commandcode, commandsize, commanddata = (self.row[2], int(self.row[4]) + 1, self.row[5])
                commandcode = int(commandcode, base=16)
                y = int(commanddata, base=16)
                y1 = y & 0xff
                y2 = (y & 0xff00) >> 8
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    m = [self.slave_addr, commandcode, y1, y2]
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break
                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize + 1),
                                     writedata=[commandcode, y1, y2, pec])
                else:
                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate, writesize=int(commandsize),
                                     writedata=[commandcode, y1, y2])

                result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', self.row[5], 'NA', self.row[7],
                          ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))

                # Passing ack and received data to PMBus custom command
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(True)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText("--")
                self.response_excel_write(save_data=result)

            elif self.row[10] == 'Block Write':  # Block Write
                commandcode, commandsize, commanddata = (self.row[2], int(self.row[4]), self.row[5])
                commandcode = int(commandcode, base=16)
                commanddata = ''.join(commanddata[2:])
                size = int(len(commanddata) / 2)
                data = []
                for i in range(size):
                    byte = commanddata[-2:]
                    byte = int(byte, base=16)
                    data.append(byte)
                    commanddata = commanddata[: (len(commanddata) - 2)]
                writevalue = [commandcode, commandsize]
                writevalue.extend(data)
                if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
                    m = [self.slave_addr, commandcode, commandsize]
                    m.extend(data)
                    new_msg = m[0]

                    for k in range(len(m)):
                        new_msg = bin(new_msg)
                        n_msg = ''.join(new_msg[2:])
                        pec = self.crc(msg=n_msg)
                        pec = int(pec, base=2)
                        if k < (len(m) - 1):
                            new_msg = pec ^ m[k + 1]
                        else:
                            break
                    writevalue.extend([pec])
                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate,
                                     writesize=int(commandsize + 3),
                                     writedata=writevalue)
                else:
                    self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                     pullup_enable=1, clockrate=config_clockrate,
                                     writesize=int(commandsize + 2),
                                     writedata=writevalue)

                result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', self.row[5], 'NA', self.row[7],
                          ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))

                # Passing ack and received data to PMBus custom command
                if PMBus_send:
                    ack = "Pass" if self.usb_to_gpio.status == 0 else "Fail"
                    PMBus_custom_obj.label_ACK.setText("--")
                    time.sleep(1)
                    PMBus_custom_obj.label_ACK.setText(ack)
                    PMBus_custom_obj.lineEdit_Received_DATA.setDisabled(True)
                    PMBus_custom_obj.lineEdit_Received_DATA.setText("--")
                self.response_excel_write(save_data=result)

            # elif self.row[10] == 'Process call Write':  # Process call Write
            #     commandcode, commandsize, commanddata = (self.row[2], self.row[4] + 1, self.row[5])
            #     commandcode = int(commandcode, base=16)
            #     y = int(commanddata, base=16)
            #     y1 = y & 0xff
            #     y2 = (y & 0xff00) >> 8
            #     if self.row[7] in ('EN', 'ENABLE', 'en', 'enable'):
            #         m = [self.slave_addr, commandcode, y1, y2]
            #         new_msg = m[0]
            #
            #         for k in range(len(m)):
            #             new_msg = bin(new_msg)
            #             n_msg = ''.join(new_msg[2:])
            #             pec = self.crc(msg=n_msg)
            #             pec = int(pec, base=2)
            #             if k < (len(m) - 1):
            #                 new_msg = pec ^ m[k + 1]
            #             else:
            #                 break
            #         self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
            #                          pullup_enable=1, clockrate=config_clockrate,
            #                          writesize=int(commandsize + 1),  # +1 to add pec
            #                          writedata=[commandcode, y1, y2, pec])
            #     else:
            #         self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
            #                          pullup_enable=1, clockrate=config_clockrate,
            #                          writesize=int(commandsize),
            #                          writedata=[commandcode, y1, y2])
            #
            #     result = (self.row[0], self.row[1], self.row[2], self.row[4], 'NA', self.row[5], 'NA', self.row[7],
            #               ("Pass" if self.usb_to_gpio.status == 0 else "Fail"))
            #     self.response_excel_write(save_data=result)

    def spi_protocol(self):

        return
        global SVID_custom_obj
        """ Identifies protocol name from command.xlsx file, based on protocol name this function executes specified
        state.
        """
        clockrate = int(''.join(filter(str.isdigit, self.row[6])))

        if self.row[1] in ("SVID", "Svid", "svid"):
            if len(str(clockrate)) in (1, 2):
                clockrate = clockrate * 1000  # converting to kilohertz
            if self.row[3] in ('WR', "RW", "wr", "rw", "W/R", "R/W", "r/w", "w/r"):
                start = format(2, "03b")  # formatting 3bits of binary data (010)
                address = format(int(self.row[8], base=16), "04b")  # formatting 4bits of binary data (0000)
                commandcode = format(int(self.row[2], base=16),
                                     "05b")  # Converting string into binary format followed by 5 bits of binary data
                commanddata = format(int(self.row[5], base=16),
                                     "08b")  # Converting string into binary format followed by 8 bits of binary data
                parity_cal = format(int(start + address + commandcode + commanddata, 2), "020b")
                self.p = 0  # Defining parity as p
                for i in range(len(parity_cal)):  # Calculating parity bit with (start+add+c.code+c.data) by Ex-or gate
                    if i == 0:
                        self.p = i ^ (int(parity_cal[i + 1], 2))
                    else:
                        try:
                            self.p = self.p ^ (int(parity_cal[i + 1], 2))
                        except IndexError:
                            pass
                parity = format(self.p, "01b")  # Assigning 1bit of binary data from output of "p"
                end = format(3, "03b")  # formatting 3bits of binary data (011)
                writedata = []
                svid_write = "0" + start + address + commandcode + commanddata + parity + end + "110011111111011"
                for i in range(0, len(svid_write), 8):  # Finding length of data and offsetting by 8bits format
                    writedata.append(int(svid_write[i:i + 8], 2))  # Appending in int data type
                    i += 8
                # writedata.append(85)  # to be removed, (when tested on device)
                # writedata.append(85)  # to be removed, (when tested on device)
                # read_spi = self.write_read_SPI(voltagelevel=12, chipselect=0, clockrate=clockrate, clockpolarity=1,
                #                                clockphase=0, numbitspersample=40,
                #                                # Ex: 8 * 5 = 40 bits per sample
                #                                writesize=len(writedata), writedata=writedata,
                #                                readsize=5  # read_size = 2 (with device) (last 2 bytes)
                #                                )
                # # print(read_spi)
                # savedata = format(read_spi[0][3], "08b") + format(read_spi[0][4],
                #                                                   "08b")  # Converts dec data type to bin data type
                # savedata = format(read_spi[0][0], "08b") + format(read_spi[0][1], "08b")  # with device
                savedata='0101010101010101'
                # Passing SVID ACK and Received data to SVID_custom_command_frame
                if self.spi.status <= 0:
                    if savedata[3:5] == "00":
                        SVID_custom_obj.label_ACK.setText("MIXED")
                    elif savedata[3:5] == "01":
                        SVID_custom_obj.label_ACK.setText("NAK")
                    elif savedata[3:5] == "10":
                        SVID_custom_obj.label_ACK.setText("ACK")
                    else:
                        SVID_custom_obj.label_ACK.setText("REJ")
                    SVID_custom_obj.lineEdit_Received_DATA.setText(format(int(savedata[5:13], 2), "02x").zfill(2))
                else:
                    SVID_custom_obj.label_ACK.setText("Fail")
                    SVID_custom_obj.lineEdit_Received_DATA.setText("")

                # # received byte parity calculation
                # p = 0
                # for i in range(8):  # Calculating parity bit with received data by Ex-or gate
                #     if i == 0:
                #         p = i ^ (int(savedata[5:13][i + 1], 2))
                #     else:
                #         try:
                #             p = p ^ (int(savedata[5:13][i + 1], 2))
                #         except IndexError:
                #             pass

                # print(savedata, read_spi)
                savedata = (self.row[0], "SVID", (self.row[2]), "1",
                            ("NAK (01)" if savedata[3:5] == "01" else
                             "ACK (10)" if savedata[3:5] == "10" else
                             "MIXED (00)" if savedata[3:5] == "00" else "REJ (11)"),
                            "0x" + format(int(savedata[5:13], 2), "02x").zfill(2),
                            "NA", str(savedata[13:14] == self.p),
                            ("Pass" if self.spi.status <= 0 else "Fail")
                            )
                # self.spi.ni845xSpiConfigurationClose()
                self.response_excel_write(save_data=savedata)  # Calls response_excel_write function to save the datum

    def write_read_SPI(self, voltagelevel=12, chipselect=0, clockrate=1000, clockpolarity=0, clockphase=0,
                       numbitspersample=8, writesize=0, writedata=0, readsize=5):
        """ Configures required parameters and writes & reads data into/from specified register address
        """
        return
        try:
            # print("write:", chipselect, clockrate, clockpolarity, clockphase, numbitspersample, writesize, writedata)
            self.spi_open(voltagelevel)
            # self.spi.ni845xSpiConfigurationOpen()
            self.spi.ni845xSetIoVoltageLevel(voltagelevel)
            self.spi.ni845xSpiConfigurationSetChipSelect(
                chipselect)  # Sets chip select where the SPI slave device resides.
            self.spi.ni845xSpiConfigurationSetClockRate(clockrate)  # Sets the SPI configuration clock rate in kilohertz
            self.spi.ni845xSpiConfigurationSetClockPolarity(
                clockpolarity)  # Sets clock polarity to use when communicating with the SPI slave device.
            self.spi.ni845xSpiConfigurationSetClockPhase(
                clockphase)  # Sets the clock phase on the first/second edge of the clock period
            self.spi.ni845xSpiConfigurationSetNumBitsPerSample(
                numbitspersample)  # Sets the number of bits per sample for an SPI transmission
            read_spi = self.spi.ni845xSpiWriteRead(writesize, writedata,
                                                   readsize)  # Writes defined SVID data through through SPI

            # transmission
            '''
            savedata = format(read_spi[0][3], "08b") + format(read_spi[0][4],
                                                              "08b")  # Converts dec data type to bin data type
            savedata = (self.row[0], "SVID", (self.row[2][2:]), "1",
                        ("Not-Acknowledge (01)" if savedata[2:4] == "01" else
                         "Acknowledge (10)" if savedata[2:4] == "10" else "Reject (11)"),
                        format(int(savedata[4:12], 2), "#01x"),
                        "NA", str(savedata[12:13] == self.p),
                        ("Pass" if self.spi.status <= 0 else "Fail")
                        )
            # self.spi.ni845xSpiConfigurationClose()
            self.response_excel_write(save_data=savedata)  # Calls response_excel_write function to save the datum
            '''
            self.spi_close()
            # return read_spi, savedata
            return read_spi
        except Exception as error:
            self.update_error("SPI Protocol Error:", error)

    def svi2_clk_data(self, voltagelevel=15, numsamples=1, numbits=28, clockpolarity=1, clockphase=1,
                      timingparameter1=0, ParameterValue1=2, timingparameter2=1, ParameterValue2=2,
                      timingparameter3=8, ParameterValue3=7, timingparameter4=12, ParameterValue4=1,
                      dataarray=[], arraysize=4, pinnumber=2, mode=1):

        """ Configures the SVI2 formats for read and write commands
        """
        return
        clockrate = (1 / int(self.row[6]) * 1000) / 2

        if (self.row[1] in ("SVI2", "svi2")) and (self.row[3] in ("W", "w")):
            iter_data = (format(0, "04b") +  # (27 clock cycles)4 bits will be ignored from MSB
                         format(int(self.row[2][2:4], base=16), "08b") +
                         format(0, "01b") +  # Matching with 0 to make 27 clock cycles
                         format(int(self.row[5][2:4], base=16), "08b") +
                         format(0, "01b") +  # Matching with 0 to make 27 clock cycles
                         format(int(self.row[5][4:6], base=16), "08b") +
                         format(0, "02b")  # Matching with 00 to make 27 clock cycles
                         )
            print("***********", iter_data, int(self.row[2][2:4], 16), int(self.row[5][2:4], 16),
                  int(self.row[5][4:6], 16))
            dataarray = []
            for i in range(0, len(iter_data), 8):
                dataarray.append(int(iter_data[i:i + 8], base=2))
                i += 8

            print("SVD", dataarray)
            self.spistream = SPIStream()
            self.spistream.ni845xOpen()
            self.spistream.ni845xSetIoVoltageLevel(voltagelevel)
            self.spistream.ni845xSpiStreamConfigurationOpen()
            self.spistream.ni845xSpiStreamConfigurationSetNumSamples(numsamples)
            self.spistream.ni845xSpiStreamConfigurationSetNumBits(numbits)
            self.spistream.ni845xSpiStreamConfigurationSetClockPolarity(clockpolarity)
            self.spistream.ni845xSpiStreamConfigurationSetClockPhase(clockphase)
            self.spistream.ni845xSpiStreamConfigurationWave1SetTimingParam(timingparameter1,
                                                                           math.floor(
                                                                               clockrate // 10))  # Rounds to smallest number and dividing by 10ns
            self.spistream.ni845xSpiStreamConfigurationWave1SetTimingParam(timingparameter2,
                                                                           math.ceil(
                                                                               clockrate // 10))  # Rounds to biggest number and dividing by 10ns
            self.spistream.ni845xSpiStreamConfigurationWave1SetTimingParam(timingparameter3,
                                                                           math.ceil(
                                                                               self.row[11] // 10))  # dividing by 10ns
            self.spistream.ni845xSpiStreamConfigurationWave1SetTimingParam(timingparameter4,
                                                                           math.ceil(
                                                                               self.row[12] // 10))  # dividing by 10ns
            self.spistream.ni845xSpiStreamConfigurationWave1SetMosiData(dataarray, arraysize)
            self.spistream.ni845xSpiStreamConfigurationWave1SetPinConfig(pinnumber, mode)
            self.spistream.ni845xSpiStreamStart()
            self.spistream.ni845xSpiStreamStop()
            self.spistream.ni845xSpiStreamConfigurationClose()
            self.spistream.ni845xClose()
            savedata = (self.row[0], "SVI2", (self.row[2]), "2",
                        "NA", format(int((iter_data[4:12] + iter_data[13:21]), 2), "#02x"),
                        "NA", "NA", ("Pass" if self.spistream.status <= 0 else "Fail")
                        )
            self.response_excel_write(save_data=savedata)  # Calls response_excel_write function to save the datum

    def svi2_telemetry(self, voltagelevel=15, port=0, type=1, line=2,
                       chipselect=0, clockrate=20000, clockpolarity=1,
                       clockphase=1, numbitspersample=65,  # Ex: 8 * 5 = 40 bits per sample
                       writesize=8, writedata=[0xff, 0xaa, 0x55, 0xff, 0xaa, 0xaa, 0x55, 0xff], readsize=8
                       ):
        return
        if self.enable_dio:
            """ Read SVI2 telemetry line"""
            l = list(format(0, "08b"))  # 8 bits format for DIO - 0000 0000
            l[line] = str(1)  # l[2] = 0010 0000
            l = l[::-1]  # l = 0000 0100 (reversed)

            self.dio = DIO()  # Initializes DIO class
            self.dio.ni845xOpen()  # Opens the device reference with DIO
            self.dio.ni845xSetIoVoltageLevel(voltagelevel)  # Sets the I/O Voltage Level
            self.dio.ni845xDioSetDriverType(port, type)  # Sets the DIO driver type as open-drain/push-pull
            self.dio.ni845xDioSetPortLineDirectionMap(port,
                                                      int("".join(l), 2))  # Sets the line direction map as input/output
            loop_count = 1000
            for i in range(loop_count):
                readline = self.dio.ni845xDioReadLine(port, line)  # Reads the digital line
                if readline == 0:
                    break
                time.sleep(0.000001)  # 1 microsencond

            self.dio.ni845xClose()  # Closes the device reference with DIO

        self.spi_open(voltagelevel)
        # self.spi.ni845xOpen()
        # self.spi.ni845xSetIoVoltageLevel(voltagelevel)
        # self.spi.ni845xSpiConfigurationOpen()
        self.spi.ni845xSpiConfigurationSetChipSelect(chipselect)  # Sets chip select where the SPI slave device resides.
        self.spi.ni845xSpiConfigurationSetClockRate(clockrate)  # Sets the SPI configuration clock rate in kilohertz
        self.spi.ni845xSpiConfigurationSetClockPolarity(
            clockpolarity)  # Sets clock polarity to use when communicating with the SPI slave device.
        self.spi.ni845xSpiConfigurationSetClockPhase(
            clockphase)  # Sets the clock phase on the first/second edge of the clock period.
        self.spi.ni845xSpiConfigurationSetNumBitsPerSample(
            numbitspersample)  # Sets the number of bits per sample for an SPI transmission
        print(numbitspersample, clockrate, writesize, writedata)
        read_spi = self.spi.ni845xSpiWriteRead(writesize, writedata,
                                               readsize)  # Writes defined SVID data through through SPI transmission
        read_spi = (format(read_spi[0][0], "08b") +
                    format(read_spi[0][1], "08b") +
                    format(read_spi[0][2], "08b")
                    )[:20]  # Ignore 4 bits from LSB
        print("SVT", read_spi)
        # self.spi.ni845xSpiConfigurationClose()
        # self.spi.ni845xClose()
        self.spi_close()
        return readline, read_spi

    def power_toggle(self, port=0, type=0, line=3):
        #third binary location bit, or 4th from right side.
        global device_status, VR_Enabled
        """ Writes specified dio line to high if power button has toggled to True else False from UI """

        a = usb_to_gpio.USB_TO_GPIO()
        a.configure(pec_enabled=False)
        # a.set_control(control_line_number=3, control_on=True)             # enable
        # print(a.get_control(control_line_number=3))

        if VR_Enabled == "ON":  # Reads an instance from UI
            a.set_control(control_line_number=3, control_on=True)         # Writes 1 to Enable pin
            # pass
        else:
            a.set_control(control_line_number=3, control_on=False)        # Writes 0 to Enable pin
            # pass
        a.close()


        # if self.enable_dio:
        #     l = list(format(0, "08b"))  # 8 bits format for DIO - 0000 0000
        #     l[line] = str(1)  # l[3] = 0001 0000
        #     l = l[::-1]  # l = 0000 1000 (reversed from MSB)
        #
        #     # self.dio = DIO()  # Initializes DIO class
        #     # self.dio.ni845xOpen()  # Opens the device resource with DIO
        #     # # self.dio.ni845xSetIoVoltageLevel(voltagelevel)  # Sets the I/O Voltage Level
        #     # self.dio.ni845xSetIoVoltageLevel(33)  # Sets the I/O Voltage Level
        #     # self.dio.ni845xDioSetDriverType(port, type)  # Sets the DIO driver type as open-drain/push-pull
        #     # self.dio.ni845xDioSetPortLineDirectionMap(port,
        #     #                                           int("".join(l), 2))  # Sets the line direction map as input/output
        #
        #     # if device_status == "ON":  # Reads an instance from UI
        #     #     self.dio.ni845xDioWriteLine(port, line, writedata=1)  # Writes the Digital Line
        #     # else:
        #     #     self.dio.ni845xDioWriteLine(port, line, writedata=0)
        #     # self.dio.ni845xClose()  # Closes the device resource with DIO

    def VID_SEL(self, voltagelevel=33, port=0, type=1, line=2):
        global initial_device_status

        if initial_device_status == 1:
            return 0
        a = usb_to_gpio.USB_TO_GPIO()
        a.configure(pec_enabled=False)
        # a.set_control(control_line_number=1, control_on=True)             # VID_SEL
        if str(a.get_control(control_line_number=1)) == "ACK: High":
            readline = 1
        else:
            readline = 0
        return readline

        # l = list(format(0, "08b"))  # 8 bits format for DIO
        # l[line] = str(1)
        # l = l[::-1]

        # Port line direction ??

        # self.dio = DIO()  # Initializes DIO class
        # self.dio.ni845xOpen()  # Opens the device reference with DIO
        # self.dio.ni845xSetIoVoltageLevel(voltagelevel)  # Sets the I/O Voltage Level
        # self.dio.ni845xDioSetDriverType(port, type)  # Sets the DIO driver type as open-drain/push-pull
        # self.dio.ni845xDioSetPortLineDirectionMap(port,
        #                                           int("".join(l), 2))  # Sets the line direction map as input/output

        # readline =self.usb_to_gpio.gpio_read_write(readmask=1, writedata=1)  # reads the Digital Line
        # readline = self.dio.ni845xDioReadLine(port, line)  # Reads the digital line
        # self.dio.ni845xClose()


    def V3p3_comparator(self, voltagelevel=33, port=0, type=1, line=4):
        return 0
        # self.dio = DIO()  # Initializes DIO class
        # self.dio.ni845xOpen()  # Opens the device reference with DIO
        # self.dio.ni845xSetIoVoltageLevel(voltagelevel)  # Sets the I/O Voltage Level
        # self.dio.ni845xDioSetDriverType(port, type)  # Sets the DIO driver type as open-drain/push-pull
        # self.dio.ni845xDioSetPortLineDirectionMap(port, 2 ** line)  # Sets the line direction map as input/output
        # readline = self.dio.ni845xDioReadLine(port, line)  # Reads the digital line
        # self.dio.ni845xClose()
        return readline

    def pmbus_alert(self, voltagelevel=33, port=0, type=1, line=0):
        l = list(format(0, "08b"))  # 8 bits format for DIO
        l[line] = str(1)
        l = l[::-1]

        # self.dio = DIO()  # Initializes DIO class
        # self.dio.ni845xOpen()  # Opens the device reference with DIO
        # self.dio.ni845xSetIoVoltageLevel(voltagelevel)  # Sets the I/O Voltage Level
        # self.dio.ni845xDioSetDriverType(port, type)  # Sets the DIO driver type as open-drain/push-pull
        # self.dio.ni845xDioSetPortLineDirectionMap(port,
        #                                           int("".join(l), 2))  # Sets the line direction map as input/output
        readline = 1
        # readline =self.usb_to_gpio.gpio_read_write(readmask=1, writedata=1)  # reads the Digital Line
        # self.dio.ni845xClose()
        return readline

    def svid_alert(self, voltagelevel=33, port=0, type=1, line=1):
        l = list(format(0, "08b"))  # 8 bits format for DIO
        l[line] = str(1)
        l = l[::-1]

        return 0
        # self.dio = DIO()  # Initializes DIO class
        # self.dio.ni845xOpen()  # Opens the device reference with DIO
        # self.dio.ni845xSetIoVoltageLevel(voltagelevel)  # Sets the I/O Voltage Level
        # self.dio.ni845xDioSetDriverType(port, type)  # Sets the DIO driver type as open-drain/push-pull
        # self.dio.ni845xDioSetPortLineDirectionMap(port,
        #                                           int("".join(l), 2))  # Sets the line direction map as input/output
        #
        # readline = self.dio.ni845xDioReadLine(port, line)  # Reads the digital line
        # self.dio.ni845xClose()
        # return readline

    def response_excel_write(self, save_data,
                             filename: str = "response.xlsx"):
        """ Saves device data in an excel file """
        if save_data is not None:
            try:
                self.save_sheet.append(save_data)
                self.workbook.save(filename=filename)
                # print(f"PMBus read bytes {save_data} has been saved successfully.")
            except Exception as error:
                self.update_error(error)

    def display_response_excel_write(self, save_data,
                                     filename: str = "Display_response.xlsx"):
        """ Saves device data in an excel file """
        if save_data is not None:
            try:
                for i in range(len(self.result)):
                    self.save_sheet2.cell(row=self.entries + 3, column=i + 1).value = self.result[i]
                for j in range(len(self.result)):
                    self.save_sheet2.cell(row=25, column=j + 1).value = self.result[j]
                if self.entries == 20:
                    self.sheet.delete_rows(idx=4, amount=23)
                    self.entries = 0
                self.workbook2.save(filename=filename)
                # print(f"PMBus read bytes {save_data} has been saved successfully.")

            except Exception as error:
                self.update_error(error)

    def close(self):
        """ Closes the device port and Protocol Configuration ports of NI8452
        """
        """
        if self.enable_i2c: self.i2c.ni845xI2cConfigurationClose()
        elif self.enable_spi: self.spi.ni845xSpiConfigurationClose()
        elif self.enable_dio: self.dio.ni845xDioConfigurationClose()
        # """
        return
        # try:
        #     if self.enable_i2c:
        #         self.i2c.ni845xClose()
        # except AttributeError:
        #     try:
        #         if self.enable_spi:
        #             self.spi.ni845xClose()
        #     except AttributeError:
        #         try:
        #             if self.enable_dio:
        #                 self.dio.ni845xClose()
        #         except AttributeError as error:
        #             print("No module connected:", error)

    ###########################################
    # Error Handler
    ###########################################

    def update_error(self, error):  # Update an exception error
        print("Error Occurred:", error)

    def linear_to_float(self, val):
        """
        :param val: list output of read word with the format [LS_byte, MS_byte]
        :return: returns the float converted value of X = Y.2^N
        """
        value = "0x" + ''.join(format(x, '02x') for x in reversed(val))
        val = format(val[1], "08b") + format(val[0], "08b")
        N = val[0:5]
        Y = val[5:]
        if N[0] == "1":
            N = ''.join([str((int(i) ^ 1)) for i in N])
            N = int(N, base=2)
            N = (N + 1) * -1
        else:
            N = int(N, base=2)

        if Y[0] == "1":
            Y = ''.join([str((int(i) ^ 1)) for i in Y])
            Y = int(Y, base=2)
            Y = (Y + 1) * -1
        else:
            Y = int(Y, base=2)
        X = Y * (2 ** N)

        return value, X

    def get_vid_max_railA(self):
        if int(initialize_feature_variable("010", 5, 4), 2) == 0:
            # PMBUS_override enabled
            if resolutionA == "6.25":
                max = round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125, 3)
                if max > 1.55:
                    max = 1.55
            elif resolutionA == "5":
                max = round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125 + 0.245, 3)
                if max > 1.52:
                    max = 1.52
            else:
                max = round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125 + 0.190, 3)
                if max > 2.74:
                    max = 2.74
        else:
            max = float((int(initialize_feature_variable('E80', 31, 24), 2) * float(resolutionA) + offset[
                str(resolutionA)]) / 1000) if int(initialize_feature_variable('E80', 31, 24), 2) != 0 else int(0)
        return max

    def get_vid_max_railB(self):
        if int(initialize_feature_variable("011", 5, 4), 2) == 0:
            # PMBUS_override enabled
            if resolutionB == "6.25":
                max = round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125, 3)
                if max > 1.55:
                    max = 1.55
            elif resolutionB == "5":
                max = round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125 + 0.245, 3)
                if max > 1.52:
                    max = 1.52
            else:
                max = round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125 + 0.190, 3)
                if max > 2.74:
                    max = 2.74
        else:
            max = float((int(initialize_feature_variable('E81', 31, 24), 2) * float(resolutionB) + offset[
                str(resolutionB)]) / 1000) if int(initialize_feature_variable('E81', 31, 24), 2) != 0 else int(0)
        return max

    def Telemetry(self):
        global homeWin_obj, PARTNAME, PMBUS_ADDR, PAGE
        # return
        # print(PMBUS_ADDR)
        self.entries += 1
        time.sleep(0.5)
        # There is issue with IMONA PMBUs register, so we have to write 80 to register 04.
        parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100, writesize=2, writedata=[4, 128])
        self.result = [self.entries]
        if PARTNAME == "AMP4592":
            # GUI display initialization
            homeWin_obj.frame_VIN.hide()
            homeWin_obj.frame_EFFICIENCY.hide()
            homeWin_obj.label.setText(PARTNAME)

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 0])  # Page command, value to 0 (Channel A)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[136],
                                    noofbytestoread=2
                                    )  # READ_VIN

            print(value)

            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-V")
            # updating VIN in the Telemetry GUI display
            homeWin_obj.label_slew_A.setText(str(round(returnvalue2, 3)) + " V")
            homeWin_obj.label_slew_B.setText(str(round(returnvalue2, 3)) + " V")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[151],
                                    noofbytestoread=2
                                    )  # READ_PIN
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-W")
            # updating PIN in the Telemetry GUI display
            homeWin_obj.label_absolute7.setText(
                str(int(round((returnvalue2 * ((int(initialize_feature_variable("E9", 33, 26), 2)) / 255)), 2)))
                + "/" + str(int(initialize_feature_variable("E9", 33, 26), 2)) + " W")
            if int(initialize_feature_variable("E9", 33, 26), 2) == 0:
                homeWin_obj.label_per7.setText(
                    "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                    "vertical-align:super;\">%</span></p>")
            else:
                homeWin_obj.label_per7.setText(
                    "<p><span style=\" font-size:28pt;\">" + str(round(int(
                        round((returnvalue2 * ((int(initialize_feature_variable("E9", 33, 26), 2)) / 255)),
                              2)) * 100 / int(
                        initialize_feature_variable("E9", 33, 26), 2), 2))
                    + "</span><span style=\" font-size:18pt; vertical-align:center;\">%</span></p>")

            # read_svid = self.write_read_SPI(voltagelevel=12, chipselect=0, clockrate=25000, clockpolarity=0,
            #                                 clockphase=1, numbitspersample=40,
            #                                 # Ex: 8 * 5 = 40 bits per sample
            #                                 writesize=5, writedata=[90, 113, 91, 255, 255],
            #                                 readsize=5
            #                                 )  # svid read 0Dh
            #Temp Manual override
            # savedata='0101010101010101'
            # # savedata = format(read_svid[0][3], "08b") + format(read_svid[0][4], "08b")
            # savedata1 = format(int(savedata[4:12], 2), "#01x")
            # savedata2 = int(savedata1, base=16)
            # self.result.append(savedata1 + "-" + str(savedata2) + "-A")
            # updating IMON_AUX, in the GUI display
            self.result.append("NA")
            # homeWin_obj.label_absolute8.setText(
            #     str(int(round((savedata2 * ((int(initialize_feature_variable("E9", 61, 54), 2)) / 255)))))
            #     + "/" + str(int(initialize_feature_variable("E9", 61, 54), 2)) + " W")
            # if int(initialize_feature_variable("E9", 61, 54), 2) == 0:
            #     homeWin_obj.label_per8.setText(
            #         "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
            #         "vertical-align:super;\">%</span></p>")
            # else:
            #     homeWin_obj.label_per8.setText(
            #         "<p><span style=\" font-size:28pt;\">" + str(round(int(
            #             round((savedata2 * ((int(initialize_feature_variable("E9", 61, 54), 2)) / 255)),
            #                   2)) * 100 / int(
            #             initialize_feature_variable("E9", 61, 54), 2), 2))
            #         + "</span><span style=\" font-size:18pt; vertical-align:center;\">%</span></p>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[139],
                                    noofbytestoread=2
                                    )  # READ_VOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-V")
            # updating VOUTA max to Telemetry GUI display
            max_val = self.get_vid_max_railA()
            homeWin_obj.label_absolute3.setText("Max:" + str(max_val) + " V")
            # updating VOUTA to Telemetry GUI display
            homeWin_obj.label_per3.setText("<html><head/><body><p><span style=\" font-size:28pt;\">" +
                                           str(round(returnvalue2, 3)) + "</span><span style=\" font-size:20pt;\">"
                                                                         "V</span></p></body></html>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[140],
                                    noofbytestoread=2
                                    )  # READ_IOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-A")
            # updating IOUTA, in the GUI display
            homeWin_obj.label_absolute5.setText("Max:" + str(int(initialize_feature_variable("E80", 55, 48), 2)) + " A")
            homeWin_obj.label_per5.setText("<html><head/><body><p><span style=\" font-size:21pt;\">"
                                           + str(round(returnvalue2, 3)) + "</span><span style=\" font-size:18pt;\">"
                                                                           "A</span></p></body></html>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[141],
                                    noofbytestoread=2
                                    )  # READ_TEMPERATURE_1
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-C")
            # updating Temp of RAIL A, in the Telemetry GUI display
            Rail_B_temp = returnvalue2

            font = QtGui.QFont()
            font.setFamily("Arial")
            font.setBold(False)
            font.setWeight(50)
            if len(str(round(returnvalue2, 1))) == 7:
                font.setPointSize(11)
            elif len(str(round(returnvalue2, 1))) == 6:
                font.setPointSize(13)                       # -147.6
            elif len(str(round(returnvalue2, 1))) == 5:
                font.setPointSize(16)                       # -70.5
            elif len(str(round(returnvalue2, 1))) == 4:
                font.setPointSize(19)                       # 17.5
            else:
                font.setPointSize(20)                       # 25
            homeWin_obj.label_temp_A.setFont(font)
            homeWin_obj.label_temp_A.setText(str(round(returnvalue2, 1)))

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 1])  # Page command, value to 1 (Channel B)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[139],
                                    noofbytestoread=2
                                    )  # READ_VOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-V")
            # updating VOUTB max to Telemetry GUI display
            max_val = self.get_vid_max_railB()
            homeWin_obj.label_absolute4.setText("Max:" + str(max_val) + " V")
            # updating VOUTB to Telemetry GUI display
            homeWin_obj.label_per4.setText("<html><head/><body><p><span style=\" font-size:28pt;\">" +
                                           str(round(returnvalue2, 3)) + "</span><span style=\" font-size:20pt;\">"
                                                                         "V</span></p></body></html>")

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 255])  # Page command modified to support 8Ch register read for RailB

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[140],
                                    noofbytestoread=2
                                    )  # READ_IOUT

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 1])

            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-A")
            # updating IOUTB, in the GUI display
            homeWin_obj.label_absolute6.setText("Max:" + str(int(initialize_feature_variable("E81", 55, 48), 2)) + " A")
            # homeWin_obj.label_per6.setText("<html><head/><body><p><span style=\" font-size:21pt;\">"
            #                                + str(round(returnvalue2, 3)) + "</span><span style=\" font-size:18pt;\">"
            #                                                                "A</span></p></body></html>")

            # IMONB is hardcoded to 0A, as its reading the same value of IMONA
            homeWin_obj.label_per6.setText("<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" font-size:18pt;\">"
                                                                           "A</span></p></body></html>")



            self.result.append("NA")  # "NA" for READ_TEMPERATURE_1 command for channel B
            # updating Temp of RAIL B, in the Telemetry GUI display --> (value = RailA temp)
            font = QtGui.QFont()
            font.setFamily("Arial")
            font.setBold(False)
            font.setWeight(50)
            if len(str(round(returnvalue2, 1))) == 7:
                font.setPointSize(11)
            elif len(str(round(Rail_B_temp, 1))) == 6:
                font.setPointSize(13)  # -147.6
            elif len(str(round(Rail_B_temp, 1))) == 5:
                font.setPointSize(16)  # -70.5
            elif len(str(round(Rail_B_temp, 1))) == 4:
                font.setPointSize(19)  # 17.5
            else:
                font.setPointSize(20)  # 25

            homeWin_obj.label_temp_B.setFont(font)
            homeWin_obj.label_temp_B.setText(str(round(Rail_B_temp, 1)))

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 0])  # Page command, value to 0 (Channel A)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[150],
                                    noofbytestoread=2
                                    )  # READ_POUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-W")

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 1])  # Page command, value to 1 (Channel B)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[150],
                                    noofbytestoread=2
                                    )  # READ_POUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-W")
            self.display_response_excel_write(save_data=self.result)

            if PAGE == 0:
                self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                 pullup_enable=1, clockrate=100, writesize=2,
                                 writedata=[0, 0])  # Page command, value to 0 (Channel A)
            else:
                self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                 pullup_enable=1, clockrate=100, writesize=2,
                                 writedata=[0, 1])  # Page command, value to 1 (Channel B)

        elif PARTNAME == "AMP4692":
            # GUI display initialization
            homeWin_obj.label.setText(PARTNAME)
            homeWin_obj.frame_VIN.hide()
            homeWin_obj.frame_EFFICIENCY.hide()
            homeWin_obj.frame_IMONA.setGeometry(QtCore.QRect(40, 300, 150, 150))
            homeWin_obj.frame_IMONB.setGeometry(QtCore.QRect(230, 300, 150, 150))
            homeWin_obj.frame_VOUTA.setGeometry(QtCore.QRect(40, 100, 150, 150))
            homeWin_obj.frame_VOUTB.setGeometry(QtCore.QRect(230, 100, 150, 150))

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 0])  # Page command, value to 0 (Channel A)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[136],
                                    noofbytestoread=2
                                    )  # READ_VIN
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-V")
            # updating VIN in the Telemetry GUI display
            homeWin_obj.label_slew_A.setText(str(round(returnvalue2, 3)) + " V")
            homeWin_obj.label_slew_B.setText(str(round(returnvalue2, 3)) + " V")
            self.result.append("NA")
            self.result.append("NA")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[139],
                                    noofbytestoread=2
                                    )  # READ_VOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-V")
            # updating VOUTA max to GUI display
            homeWin_obj.label_absolute3.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
            # updating VOUTA to GUI display
            homeWin_obj.label_per3.setText("<html><head/><body><p><span style=\" font-size:28pt;\">" +
                                           str(round(returnvalue2, 3)) + "</span><span style=\" font-size:20pt;\">"
                                                                         "V</span></p></body></html>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[140],
                                    noofbytestoread=2
                                    )  # READ_IOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-A")
            # updating IOUTA, in the GUI display
            homeWin_obj.label_absolute5.setText("Max:" + str(int(initialize_feature_variable("E80", 55, 48), 2)) + " A")
            homeWin_obj.label_per5.setText("<html><head/><body><p><span style=\" font-size:21pt;\">"
                                           + str(round(returnvalue2, 3)) + "</span><span style=\" font-size:18pt;\">"
                                                                           "A</span></p></body></html>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[141],
                                    noofbytestoread=2
                                    )  # READ_TEMPERATURE_1
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-C")
            # updating Temp of RAIL A, in the GUI display
            homeWin_obj.label_temp_A.setText(str(round(returnvalue2, 2)))

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                             pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 1])  # Page command, value to 1 (Channel B)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[139],
                                    noofbytestoread=2
                                    )  # READ_VOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-V")
            # updating VOUTB max to GUI display
            homeWin_obj.label_absolute4.setText("Max:" + str(vid_max_value[resolutionB]) + " V")
            # updating VOUTB to GUI display
            homeWin_obj.label_per4.setText("<html><head/><body><p><span style=\" font-size:28pt;\">" +
                                           str(round(returnvalue2, 3)) + "</span><span style=\" font-size:20pt;\">"
                                                                         "V</span></p></body></html>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[140],
                                    noofbytestoread=2
                                    )  # READ_IOUT
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-A")
            # updating IOUTB, in the GUI display
            homeWin_obj.label_absolute6.setText("Max:" + str(int(initialize_feature_variable("E81", 55, 48), 2)) + " A")
            homeWin_obj.label_per6.setText("<html><head/><body><p><span style=\" font-size:21pt;\">"
                                           + str(round(returnvalue2, 3)) + "</span><span style=\" font-size:18pt;\">"
                                                                           "A</span></p></body></html>")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100,
                                    writesize=1, writedata=[141],
                                    noofbytestoread=2
                                    )  # READ_TEMPERATURE_1 (Rail B)
            returnvalue1, returnvalue2 = self.linear_to_float(val=value)
            self.result.append(returnvalue1 + "-" + str(returnvalue2) + "-A")
            # updating Temp of RAIL B, in the GUI display
            homeWin_obj.label_temp_B.setText(str(round(returnvalue2, 2)))
            self.result.append("NA")
            self.result.append("NA")
            self.display_response_excel_write(save_data=self.result)

        elif PARTNAME == "AMP4792":
            print("Telemetry", PARTNAME)
        elif PARTNAME == "AMP4291":
            print("Telemetry", PARTNAME)
        else:
            print("Telemetry", "unknown")

    def Fault(self):
        global homeWin_obj
        # return
        fault_value = 0
        fault_status = ["Bitwise fault status"]
        header2 = ["Register details"]
        header1 = ["Bit mapping", "STATUS_INPUT[4]", "STATUS_INPUT[7]", "MFR_SPECIFIC_FA[0]",
                   "STATUS_VOUT_RAIL_A[7]", "MFR_SPECIFIC_FA[1]", "STATUS_VOUT_RAIL_B[7]", "NA",
                   "STATUS_VOUT_RAIL_A[4]", "NA", "STATUS_VOUT_RAIL_B[4]", "STATUS_TEMPERATURE_RAIL_A[7]",
                   "STATUS_TEMPERATURE_RAIL_B[7]", "STATUS_IOUT_RAIL_A[7]", "STATUS_IOUT_RAIL_B[7]",
                   "STATUS_MFR_SPECIFIC[4]", "STATUS_PWR_STAGE_FAULTS[32]", "STATUS_PWR_STAGE_FAULTS[33]",
                   "STATUS_PWR_STAGE_FAULTS[34]", "STATUS_PWR_STAGE_FAULTS[8:0]", "STATUS_PWR_STAGE_FAULTS[36]",
                   "STATUS_PWR_STAGE_FAULTS[37]", "STATUS_PWR_STAGE_FAULTS[24:16]", "MFR_SPECIFIC_FA[2]"]

        if PARTNAME == "AMP4592":
            print("Fault occurred for", PARTNAME)

            for i in range(len(header1)):
                self.save_sheet2.cell(row=30, column=i + 1).value = header1[i]

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 0])  # Page command, value to 0 (Channel A)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[124], noofbytestoread=1)
            header2.insert(1, ("7Ch-" + ('0x' + ''.join(format(value[0], '02x')))))
            header2.insert(2, ("7Ch-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(1, value[3])  # VIN Under voltage
            fault_value = fault_value | int(value[3])
            # Passing VIN Under voltage fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_1.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_1.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(2, value[0])  # VIN Over voltage
            fault_value = fault_value | int(value[0])
            # Passing VIN Over voltage fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_2.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_2.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[250], noofbytestoread=5)
            header2.insert(3, ("FAh-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))
            header2.insert(5, ("FAh-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))
            header2.insert(23, ("FAh-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))
            value = format(value[1], "08b")

            fault_status.insert(3, value[7])  # VOUT A fixed Overvoltage
            fault_value = fault_value | int(value[7])
            # Passing VOUT A fixed Overvoltage fault status to the GUI Fault display
            if value[7] == "1":
                homeWin_obj.label_Fault_3.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_3.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(5, value[6])  # VOUT B fixed Overvoltage
            fault_value = fault_value | int(value[6])
            # Passing VOUT B fixed Overvoltage fault status to the GUI Fault display
            if value[6] == "1":
                homeWin_obj.label_Fault_7.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_7.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(23, value[5])  # Controller Over temperature
            # fault_value = fault_value | int(value[5])
            # # Passing Controller Over temperature fault status to the GUI Fault display
            # if value[5] == "1":
            #     homeWin_obj.label_Fault_22.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_22.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[122], noofbytestoread=1)
            header2.insert(4, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            header2.insert(8, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(4, value[0])  # VOUT A tracking Overvoltage
            fault_value = fault_value | int(value[0])
            # Passing VOUT A tracking Overvoltage fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_4.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_4.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(8, value[3])  # VOUT A tracking undervoltage
            fault_value = fault_value | int(value[3])
            # Passing VOUT A tracking undervoltage fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_6.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_6.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[125], noofbytestoread=1)
            header2.insert(11, ("7Dh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")
            fault_status.insert(11, value[0])  # Over Temp A
            fault_value = fault_value | int(value[0])
            # Passing Over Temp A fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_11.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_11.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[123], noofbytestoread=1)
            header2.insert(13, ("7Bh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")
            fault_status.insert(13, value[0])  # IOUT OverCurrent A
            fault_value = fault_value | int(value[0])
            # Passing IOUT OverCurrent A fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_13.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_13.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")


            #PSYS Fault GUI
            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[128], noofbytestoread=1)
            header2.insert(15, ("80h-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")
            fault_status.insert(15, value[3])  # PSYS Critical
            fault_value = fault_value | int(value[3])
            # Passing IOUT OverCurrent A fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_23.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_23.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[249], noofbytestoread=6)
            for i in range(16, 23):
                header2.insert(i, ("F9h-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))

            value1 = format(value[-1], "08b")
            fault_status.insert(16, value1[7])  # SPS VDD Under voltage A
            # fault_value = fault_value | int(value1[7])
            # Passing SPS VDD Under voltage A fault status to the GUI Fault display
            # if value1[7] == "1":
            #     homeWin_obj.label_Fault_15.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_15.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            fault_status.insert(17, value1[6])  # SPS VDD Under voltage B
            # fault_value = fault_value | int(value1[6])
            # # Passing SPS VDD Under voltage B fault status to the GUI Fault display
            # if value1[6] == "1":
            #     homeWin_obj.label_Fault_16.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_16.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            fault_status.insert(18, str((int(value1[5]) | int(value1[4]))))  # SPS VIN Under voltage
            # fault_value = fault_value | (int(value1[5]) | int(value1[4]))
            # # Passing SPS VIN Under voltage fault status to the GUI Fault display
            # if str((int(value1[5]) | int(value1[4]))) == "1":
            #     homeWin_obj.label_Fault_17.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_17.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            fault_status.insert(19, value1[3])  # SPS Over Temperature A
            fault_value = fault_value | int(value1[3])
            # Passing SPS Over Temperature A fault status to the GUI Fault display
            if value1[3] == "1":
                homeWin_obj.label_Fault_19.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_19.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            fault_status.insert(20, value1[2])  # SPS Over Temperature B
            # fault_value = fault_value | int(value1[2])
            # # Passing SPS Over Temperature B fault status to the GUI Fault display
            # if value1[3] == "1":
            #     homeWin_obj.label_Fault_20.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_20.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            value2 = format(value[1], "08b")
            value3 = format(value[2], "08b")
            value3 = value3[7]
            value4 = format(value[3], "08b")
            value5 = format(value[4], "08b")
            value5 = value5[7]
            p = 0
            for i in range(len(value2)):
                p = p | int(value2[i])

            fault_status.insert(21, str(p | int(value3)))  # SPS Over current
            # fault_value = fault_value | (p | int(value3))
            # # Passing SPS Over current fault status to the GUI Fault display
            # if (p | int(value3)) == 1:
            #     homeWin_obj.label_Fault_18.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_18.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            p = 0
            for i in range(len(value4)):
                p = p | int(value4[i])
            fault_status.insert(22, str(p | int(value5)))  # SPS High Side FET short
            # fault_value = fault_value | (p | int(value5))
            # # Passing SPS High Side FET short fault status to the GUI Fault display
            # if (p | int(value5)) == 1:
            #     homeWin_obj.label_Fault_21.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_21.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 1])  # Page command, value to 1 (Channel B)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[122], noofbytestoread=1)
            header2.insert(6, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            header2.insert(7, "NA")
            header2.insert(9, "NA")
            header2.insert(10, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(6, value[0])  # VOUT B tracking Overvoltage
            fault_value = fault_value | int(value[0])
            # Passing VOUT B tracking Overvoltage fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_8.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_8.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(7, "NA")  # VOUT A fixed undervoltage
            fault_status.insert(9, "NA")  # VOUT B fixed undervoltage
            fault_status.insert(10, value[3])  # VOUT B tracking undervoltage
            fault_value = fault_value | int(value[3])
            # Passing VOUT B tracking undervoltage fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_10.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_10.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[125], noofbytestoread=1)
            header2.insert(12, ("7Dh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(12, value[0])  # Over Temp B
            # fault_value = fault_value | int(value[0])
            # # Passing Over Temp B fault status to the GUI Fault display
            # if value[0] == "1":
            #     homeWin_obj.label_Fault_12.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_12.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[123], noofbytestoread=1)
            header2.insert(14, ("7Bh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(14, value[0])  # IOUT Over Current B
            fault_value = fault_value | int(value[0])
            # Passing IOUT Over Current B fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_14.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_14.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")


            for i in range(len(header2)):
                self.save_sheet2.cell(row=31, column=i + 1).value = header2[i]

            for i in range(len(fault_status)):
                self.save_sheet2.cell(row=32, column=i + 1).value = fault_status[i]

            self.workbook2.save(filename="Display_response.xlsx")
            if fault_value == 1:
                homeWin_obj.pushButton_Fault.setStyleSheet("QPushButton {\n"
                                                           "background-color: red;\n"
                                                           "border-radius:5px;\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-style:outset;\n"
                                                           "border-width:2px;\n"
                                                           "border-color:black;\n"
                                                           "}\n"
                                                           "QPushButton::hover {\n"
                                                           "background-color:red;\n"
                                                           "}\n"
                                                           "\n"
                                                           "")
            else:
                homeWin_obj.pushButton_Fault.setStyleSheet("QPushButton {\n"
                                                           "background-color: grey;\n"
                                                           "border-radius:5px;\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-style:outset;\n"
                                                           "border-width:2px;\n"
                                                           "border-color:black;\n"
                                                           "}\n"
                                                           "QPushButton::hover {\n"
                                                           "background-color:darkgrey;\n"
                                                           "}\n"
                                                           "\n"
                                                           "")

        elif PARTNAME == "AMP4692":
            print("Fault occurred for", PARTNAME)

            for i in range(len(header1)):
                self.save_sheet2.cell(row=30, column=i + 1).value = header1[i]

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 0])  # Page command, value to 0 (Channel A)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[124], noofbytestoread=1)
            header2.insert(1, ("7Ch-" + ('0x' + ''.join(format(value[0], '02x')))))
            header2.insert(2, ("7Ch-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(1, value[3])  # VIN Under voltage
            fault_value = fault_value | int(value[3])
            # Passing VIN Under voltage fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_1.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_1.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(2, value[0])  # VIN Over voltage
            fault_value = fault_value | int(value[0])
            # Passing VIN Over voltage fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_2.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_2.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[250], noofbytestoread=5)
            header2.insert(3, ("FAh-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))
            header2.insert(5, ("FAh-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))
            header2.insert(23, ("FAh-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))
            value = format(value[1], "08b")

            fault_status.insert(3, value[7])  # VOUT A fixed Overvoltage
            fault_value = fault_value | int(value[7])
            # Passing VOUT A fixed Overvoltage fault status to the GUI Fault display
            if value[7] == "1":
                homeWin_obj.label_Fault_3.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_3.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(5, value[6])  # VOUT B fixed Overvoltage
            fault_value = fault_value | int(value[6])
            # Passing VOUT B fixed Overvoltage fault status to the GUI Fault display
            if value[6] == "1":
                homeWin_obj.label_Fault_7.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_7.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(23, value[5])  # Controller Over temperature
            # fault_value = fault_value | int(value[5])
            # # Passing Controller Over temperature fault status to the GUI Fault display
            # if value[5] == "1":
            #     homeWin_obj.label_Fault_22.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_22.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[122], noofbytestoread=1)
            header2.insert(4, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            header2.insert(8, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(4, value[0])  # VOUT A tracking Overvoltage
            fault_value = fault_value | int(value[0])
            # Passing VOUT A tracking Overvoltage fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_4.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_4.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(8, value[3])  # VOUT A tracking undervoltage
            fault_value = fault_value | int(value[3])
            # Passing VOUT A tracking undervoltage fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_6.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_6.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[125], noofbytestoread=1)
            header2.insert(11, ("7Dh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")
            fault_status.insert(11, value[0])  # Over Temp A
            fault_value = fault_value | int(value[0])
            # Passing Over Temp A fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_11.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_11.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[123], noofbytestoread=1)
            header2.insert(13, ("7Bh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")
            fault_status.insert(13, value[0])  # IOUT OverCurrent A
            fault_value = fault_value | int(value[0])
            # Passing IOUT OverCurrent A fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_13.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_13.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            header2.insert(15, "NA")
            fault_status.insert(15, "NA")  # PSYS Critical NA for AMP4792 (SVI2 part)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[249], noofbytestoread=6)
            for i in range(16, 23):
                header2.insert(i, ("F9h-" + (''.join(format(x, '02x') for x in reversed(value[1:])))))

            value1 = format(value[-1], "08b")
            fault_status.insert(16, value1[7])  # SPS VDD Under voltage A
            # fault_value = fault_value | int(value1[7])
            # # Passing SPS VDD Under voltage A fault status to the GUI Fault display
            # if value1[7] == "1":
            #     homeWin_obj.label_Fault_15.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_15.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            fault_status.insert(17, value1[6])  # SPS VDD Under voltage B
            # fault_value = fault_value | int(value1[6])
            # # Passing SPS VDD Under voltage B fault status to the GUI Fault display
            # if value1[6] == "1":
            #     homeWin_obj.label_Fault_16.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_16.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            fault_status.insert(18, str((int(value1[5]) | int(value1[4]))))  # SPS VIN Under voltage
            # fault_value = fault_value | (int(value1[5]) | int(value1[4]))
            # # Passing SPS VIN Under voltage fault status to the GUI Fault display
            # if str((int(value1[5]) | int(value1[4]))) == "1":
            #     homeWin_obj.label_Fault_17.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_17.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            fault_status.insert(19, value1[3])  # SPS Over Temperature A
            fault_value = fault_value | int(value1[3])
            # Passing SPS Over Temperature A fault status to the GUI Fault display
            if value1[3] == "1":
                homeWin_obj.label_Fault_19.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_19.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            fault_status.insert(20, value1[2])  # SPS Over Temperature B
            # fault_value = fault_value | int(value1[2])
            # # Passing SPS Over Temperature B fault status to the GUI Fault display
            # if value1[3] == "1":
            #     homeWin_obj.label_Fault_20.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_20.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            value2 = format(value[1], "08b")
            value3 = format(value[2], "08b")
            value3 = value3[7]
            value4 = format(value[3], "08b")
            value5 = format(value[4], "08b")
            value5 = value5[7]
            p = 0
            for i in range(len(value2)):
                p = p | int(value2[i])

            fault_status.insert(21, str(p | int(value3)))  # SPS Over current
            # fault_value = fault_value | (p | int(value3))
            # # Passing SPS Over current fault status to the GUI Fault display
            # if (p | int(value3)) == 1:
            #     homeWin_obj.label_Fault_18.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_18.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            p = 0
            for i in range(len(value4)):
                p = p | int(value4[i])
            fault_status.insert(22, str(p | int(value5)))  # SPS High Side FET short
            # fault_value = fault_value | (p | int(value5))
            # # Passing SPS High Side FET short fault status to the GUI Fault display
            # if (p | int(value5)) == 1:
            #     homeWin_obj.label_Fault_21.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_21.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            self.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100, writesize=2,
                             writedata=[0, 1])  # Page command, value to 1 (Channel B)

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[122], noofbytestoread=1)
            header2.insert(6, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            header2.insert(7, "NA")
            header2.insert(9, "NA")
            header2.insert(10, ("7Ah-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(6, value[0])  # VOUT B tracking Overvoltage
            fault_value = fault_value | int(value[0])
            # Passing VOUT B tracking Overvoltage fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_8.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_8.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:2px;")

            fault_status.insert(7, "NA")  # VOUT A fixed undervoltage
            fault_status.insert(9, "NA")  # VOUT B fixed undervoltage
            fault_status.insert(10, value[3])  # VOUT B tracking undervoltage
            fault_value = fault_value | int(value[3])
            # Passing VOUT B tracking undervoltage fault status to the GUI Fault display
            if value[3] == "1":
                homeWin_obj.label_Fault_10.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_10.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[125], noofbytestoread=1)
            header2.insert(12, ("7Dh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(12, value[0])  # Over Temp B
            # fault_value = fault_value | int(value[0])
            # # Passing Over Temp B fault status to the GUI Fault display
            # if value[0] == "1":
            #     homeWin_obj.label_Fault_12.setStyleSheet("background-color: rgb(255, 0, 0);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")
            # else:
            #     homeWin_obj.label_Fault_12.setStyleSheet("background-color: rgb(172, 172, 172);\n"
            #                                              "color: rgb(255, 255, 255);\n"
            #                                              "border-radius:2px;")

            value = self.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=1, writedata=[123], noofbytestoread=1)
            header2.insert(14, ("7Bh-" + ('0x' + ''.join(format(value[0], '02x')))))
            value = format(value[0], "08b")

            fault_status.insert(14, value[0])  # IOUT Over Current B
            fault_value = fault_value | int(value[0])
            # Passing IOUT Over Current B fault status to the GUI Fault display
            if value[0] == "1":
                homeWin_obj.label_Fault_14.setStyleSheet("background-color: rgb(255, 0, 0);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")
            else:
                homeWin_obj.label_Fault_14.setStyleSheet("background-color: rgb(172, 172, 172);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:2px;")

            for i in range(len(header2)):
                self.save_sheet2.cell(row=31, column=i + 1).value = header2[i]

            for i in range(len(fault_status)):
                self.save_sheet2.cell(row=32, column=i + 1).value = fault_status[i]

            self.workbook2.save(filename="Display_response.xlsx")
            if fault_value == 1:
                homeWin_obj.pushButton_Fault.setStyleSheet("QPushButton {\n"
                                                           "background-color: red;\n"
                                                           "border-radius:5px;\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-style:outset;\n"
                                                           "border-width:2px;\n"
                                                           "border-color:black;\n"
                                                           "}\n"
                                                           "QPushButton::hover {\n"
                                                           "background-color:red;\n"
                                                           "}\n"
                                                           "\n"
                                                           "")
            else:
                homeWin_obj.pushButton_Fault.setStyleSheet("QPushButton {\n"
                                                           "background-color: grey;\n"
                                                           "border-radius:5px;\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-style:outset;\n"
                                                           "border-width:2px;\n"
                                                           "border-color:black;\n"
                                                           "}\n"
                                                           "QPushButton::hover {\n"
                                                           "background-color:darkgrey;\n"
                                                           "}\n"
                                                           "\n"
                                                           "")

        elif PARTNAME == "AMP4792":
            print("Fault occurred for", PARTNAME)
        else:  # AMP4291
            print("Fault occurred for", PARTNAME)

    def run(self):
        global stop_thread, VR_Enabled, load_settings_done, PMBUS_ADDR, PMBus_send
        """ Start of Command.xlsx execution and updating to response.xlsx
        """
        abc = []

        time.sleep(0.1)
        # self.wb = load_workbook("command.xlsx")
        # self.wb = openpyxl.load_workbook('command.xlsx')
        # self.ws = self.wb.active
        """
        for k in self.ws.iter_rows(values_only=True):
            if 'PMBUS' in k[1]:
                self.row = k
                self.i2c_protocol()

            elif 'SVID' in k[1]:
                self.row = k
                self.spi_protocol()

            elif 'SVI2' in k[1]:
                self.row = k
                self.svi2_clk_data()
        """

        while True:
            if stop_thread:
                stop_thread = False
                # time.sleep(2)
                break
            if self.temp_row_number < next_row_pointer_command_xlsx - 1:
                self.wb = openpyxl.load_workbook('command.xlsx')
                self.ws = self.wb.active

                for k in self.ws.iter_rows(min_row=self.temp_row_number + 1, max_row=next_row_pointer_command_xlsx - 1,
                                           values_only=True):
                    time.sleep(0.05)
                    abc.append(list(k))

                # print(abc)
                # print(self.temp_row_number)
                # print(next_row_pointer_command_xlsx)

                for j in abc:
                    if 'PMBUS' in j[1]:
                        self.row = j
                        time.sleep(0.1)
                        self.i2c_protocol()
                    elif 'SVID' in j[1]:
                        pass
                        # self.row = j
                        # self.spi_protocol()
                    elif 'SVI2' in j[1]:
                            pass
                        # self.row = j
                        # self.svi2_clk_data()
                self.temp_row_number = next_row_pointer_command_xlsx - 1
            abc.clear()

            if PMBus_send:
                PMBus_send = False

            if PMBUS_ADDR != (''.join(format(int(initialize_feature_variable("EF", 6, 0), 2), '02x'))):
                PMBUS_ADDR = ''.join(format(int(initialize_feature_variable("EF", 6, 0), 2), '02x'))
                time.sleep(0.1)

            if load_settings_done == "YES":
                print_log("Previously saved MTP settings have been loaded on the device.", "INFO")
                load_settings_done = "NO"

            # if stop_thread:
            #     stop_thread = False
            #     # time.sleep(2)
            #     break

            """ Start of Telemetry commands execution and updating to Display_response.xlsx
            """
            if VR_Enabled != "OFF":
                self.Telemetry()
            """ Calling the pmbus and svid alert function
            """
            pmb_alert = self.pmbus_alert()
            pmb_alert = 0
            if pmb_alert == 0:  # to be changed to high for fault function check (default_alert = 0)
                self.Fault()

            # svid_alert = self.svid_alert()
            # print(pmb_alert, svid_alert)

# MTP Initialization class
class MTP_register_database(QThread):
    def __init__(self, parent=None):
        global register_database, parallel_thread, PMBUS_ADDR, homeWin_obj, initial_device_status, RailA_phase_count_arg, RailB_phase_count_arg
        super(MTP_register_database, self).__init__(parent)

    def run(self):

        global register_database, parallel_thread, PMBUS_ADDR, initial_device_status,stop_thread

        # return # temperory thing
        if initial_device_status == 1:
            return
        testmode_entry()                # To read E4 register testmode entry is required
        paged_0 = []
        paged_1 = []
        not_paged = []
        for i in register_database:
            if len(i) == 3:
                if i[2] == "0":
                    paged_0.append(i)
                else:
                    paged_1.append(i)
            else:
                not_paged.append(i)

        # Page command  = 0 (RailA)
        parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100, writesize=2, writedata=[0, 0])

        for i in not_paged:
            if register_database[i]["Read_command_type"] == "Read Byte":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1,
                                                   writedata=[int(i, base=16)], noofbytestoread=1)

                # register_database[i]["Initial_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)
                register_database[i]["Temp_update_from_customer"] = bin(value[0]).split("0b")[1].zfill(8)
                register_database[i]["Final_register_value"] = bin(value[0]).split("0b")[1].zfill(8)
                # register_database[i]["Updated_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)

            elif register_database[i]["Read_command_type"] == "Read Word":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1,
                                                   writedata=[int(i, base=16)], noofbytestoread=2)

                # register_database[i]["Initial_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))
                register_database[i]["Temp_update_from_customer"] = ''.join(format(x, '08b') for x in reversed(value))
                register_database[i]["Final_register_value"] = ''.join(format(x, '08b') for x in reversed(value))
                # register_database[i]["Updated_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))

            elif register_database[i]["Read_command_type"] == "Block Read":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1,
                                                   writedata=[int(i, base=16)], noofbytestoread
                                                   =(int(register_database[i]["size_in_bits"] / 8) + 1))
                value = value[1:]
                # register_database[i]["Initial_device_MTP_value"] = ''.join(
                #     format(x, '08b') for x in reversed(value))
                register_database[i]["Temp_update_from_customer"] = ''.join(
                    format(x, '08b') for x in reversed(value))
                register_database[i]["Final_register_value"] = ''.join(
                    format(x, '08b') for x in reversed(value))
                # register_database[i]["Updated_device_MTP_value"] = ''.join(
                #     format(x, '08b') for x in reversed(value))

        for i in paged_0:
            if register_database[i]["Read_command_type"] == "Read Byte":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1,
                                                   writedata=[int(i[0:-1], base=16)], noofbytestoread=1)

                # register_database[i]["Initial_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)
                register_database[i]["Temp_update_from_customer"] = bin(value[0]).split("0b")[1].zfill(8)
                register_database[i]["Final_register_value"] = bin(value[0]).split("0b")[1].zfill(8)
                # register_database[i]["Updated_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)

            elif register_database[i]["Read_command_type"] == "Read Word":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1,
                                                   writedata=[int(i[0:-1], base=16)], noofbytestoread=2)

                # register_database[i]["Initial_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))
                register_database[i]["Temp_update_from_customer"] = ''.join(format(x, '08b') for x in reversed(value))
                register_database[i]["Final_register_value"] = ''.join(format(x, '08b') for x in reversed(value))
                # register_database[i]["Updated_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))

            elif register_database[i]["Read_command_type"] == "Block Read":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1,
                                                   writedata=[int(i[0:-1], base=16)], noofbytestoread
                                                   =(int(register_database[i]["size_in_bits"] / 8) + 1))
                value = value[1:]
                # register_database[i]["Initial_device_MTP_value"] = ''.join(
                #     format(x, '08b') for x in reversed(value))
                register_database[i]["Temp_update_from_customer"] = ''.join(
                    format(x, '08b') for x in reversed(value))
                register_database[i]["Final_register_value"] = ''.join(
                    format(x, '08b') for x in reversed(value))
                # register_database[i]["Updated_device_MTP_value"] = ''.join(
                #     format(x, '08b') for x in reversed(value))

            else:
                print_log("Process call for " + i[0:-1] + ", not implemented in GUI", "INFO")

        # Page command = 1 (Rail B)
        parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1, clockrate=100,
                                    writesize=2, writedata=[0, 1])

        for i in paged_1:
            if (i == "331") or (i == "621") or (i == "601") or (i == "641") or (i == "271"):
                # changing from 0x01 to 0xff (for freq switch problem)
                parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                            pullup_enable=1, clockrate=100,
                                            writesize=2, writedata=[0, 255])

            if register_database[i]["Read_command_type"] == "Read Byte":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                                   clockrate=100, writesize=1,
                                                   writedata=[int(i[0:-1], base=16)], noofbytestoread=1)

                # register_database[i]["Initial_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)
                register_database[i]["Temp_update_from_customer"] = bin(value[0]).split("0b")[1].zfill(8)
                register_database[i]["Final_register_value"] = bin(value[0]).split("0b")[1].zfill(8)
                # register_database[i]["Updated_device_MTP_value"] = bin(value[0]).split("0b")[1].zfill(8)

            elif register_database[i]["Read_command_type"] == "Read Word":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                                   clockrate=100, writesize=1,
                                                   writedata=[int(i[0:-1], base=16)], noofbytestoread=2)

                if (i == "331") or (i == "621") or (i == "601") or (i == "641") or (i == "271"):
                    # changing from 0xff to 0x01 (for freq switch problem)
                    parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                pullup_enable=1, clockrate=100,
                                                writesize=2, writedata=[0, 1])

                # register_database[i]["Initial_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))
                register_database[i]["Temp_update_from_customer"] = ''.join(format(x, '08b') for x in reversed(value))
                register_database[i]["Final_register_value"] = ''.join(format(x, '08b') for x in reversed(value))
                # register_database[i]["Updated_device_MTP_value"] = ''.join(format(x, '08b') for x in reversed(value))

            elif register_database[i]["Read_command_type"] == "Block Read":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                                   clockrate=100, writesize=1,
                                                   writedata=[int(i[0:-1], base=16)], noofbytestoread
                                                   =(int(register_database[i]["size_in_bits"] / 8) + 1))
                value = value[1:]
                # register_database[i]["Initial_device_MTP_value"] = ''.join(
                #     format(x, '08b') for x in reversed(value))
                register_database[i]["Temp_update_from_customer"] = ''.join(
                    format(x, '08b') for x in reversed(value))
                register_database[i]["Final_register_value"] = ''.join(
                    format(x, '08b') for x in reversed(value))
                # register_database[i]["Updated_device_MTP_value"] = ''.join(
                #     format(x, '08b') for x in reversed(value))

            else:
                print_log("Process call for " + i[0:-1] + ", not implemented in GUI", "INFO")

        # Page = 0 (Rail A)
        parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16), pullup_enable=1,
                                    clockrate=100, writesize=2, writedata=[0, 0])

        update_master_command_xlsx_from_register_database()

        resolution_calculation(0)

        # Updating the home page labels when refresh condition is selected
        RailA_phase_count_arg = str(int(initialize_feature_variable("DF", 8, 4), 2))
        RailB_phase_count_arg = str(int(initialize_feature_variable("DF", 3, 0), 2))
        homeWin_obj.label_phase_A.setText(RailA_phase_count_arg)
        homeWin_obj.label_phase_B.setText(RailB_phase_count_arg)
        homeWin_obj.label_PMBus_addressA.setText(PMBUS_ADDR + "h")
        homeWin_obj.label_PMBus_addressB.setText(PMBUS_ADDR + "h")
        homeWin_obj.label_prot_address_A.setText(
            str(hex(int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
        homeWin_obj.label_prot_address_B.setText(
            str(hex(1 + int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
        homeWin_obj.label_boot_vol_A.setText(str(round(float((int(initialize_feature_variable('E10', 23, 13), 2) * float(
            resolutionA) + offset[str(resolutionA)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13),
                                                                    2) != 0 else int(0)) + "V")
        homeWin_obj.label_boot_vol_B.setText(str(round(float((int(initialize_feature_variable('E11', 23, 13), 2) * float(
            resolutionB) + offset[str(resolutionB)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13),
                                                                    2) != 0 else int(0)) + "V")

        if homeWin_obj.pushButton_Enable_VR.isChecked():
            stop_thread = False
            parallel_thread.start()

# Device Configuration class
class Transient_Configuration(QMainWindow, Ui_Transient_Window):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(Transient_Configuration, self).__init__(parent)
        self.setupUi(self)
        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["C50", "C51"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        # GUI default value look initialization
        # Non feature related initializaton on GUI_ display
        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        # Hiding Active Phases in UVR2 event for RAILB as no of phases is only 4, while the dropdown feature has 7 and
        # 9 phase options
        self.comboBox_UVR2_phase_RailB.hide()
        self.comboBox_UVR1_phase_RailB.hide()
        update_database_with_temp_customer_input("C51", 24, 23, "00")

        # Feature related initialization on GUI display
        self.comboBox_isum_gain_railA.setCurrentIndex(int(initialize_feature_variable("C50", 4, 0), 2))
        self.comboBox_isum_gain_railB.setCurrentIndex(int(initialize_feature_variable("C51", 4, 0), 2))
        self.comboBox_dynamic_ramp_height_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 75, 74), 2))
        self.comboBox_dynamic_ramp_height_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 75, 74), 2))
        if int(initialize_feature_variable("C50", 49, 49), 2) == 0:
            self.comboBox_loop_ac_gain_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 56, 55), 2))
        else:
            self.comboBox_loop_ac_gain_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 56, 55), 2) + 4)

        if int(initialize_feature_variable("C51", 49, 49), 2) == 0:
            self.comboBox_loop_ac_gain_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 56, 55), 2))
        else:
            self.comboBox_loop_ac_gain_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 56, 55), 2) + 4)
        self.comboBox_ramp_height_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 33, 31), 2))
        self.comboBox_ramp_height_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 33, 31), 2))
        self.comboBox_integ_time_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 52, 50), 2))
        self.comboBox_integ_time_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 52, 50), 2))
        self.comboBox_OVR_threshold_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 42, 40), 2))
        self.comboBox_OVR_threshold_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 42, 40), 2))
        self.comboBox_UVR1_threshold_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 22, 20), 2))
        self.comboBox_UVR1_threshold_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 22, 20), 2))
        self.comboBox_UVR2_threshold_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 17, 15), 2))
        self.comboBox_UVR2_threshold_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 17, 15), 2))
        self.comboBox_UVR1_phase_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 24, 23), 2))
        self.comboBox_UVR1_phase_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 24, 23), 2))
        self.comboBox_UVR2_phase_RailA.setCurrentIndex(int(initialize_feature_variable("C50", 19, 18), 2))
        self.comboBox_UVR2_phase_RailB.setCurrentIndex(int(initialize_feature_variable("C51", 19, 18), 2))

        # Customer GUI interaction related function mapping
        self.comboBox_isum_gain_railA.activated.connect(self.isum_gain_changed_A)
        self.comboBox_isum_gain_railB.activated.connect(self.isum_gain_changed_B)
        self.comboBox_dynamic_ramp_height_RailA.activated.connect(self.dynamic_ramp_height_changed_A)
        self.comboBox_dynamic_ramp_height_RailB.activated.connect(self.dynamic_ramp_height_changed_B)
        self.comboBox_loop_ac_gain_RailA.activated.connect(self.loop_ac_gain_changed_A)
        self.comboBox_loop_ac_gain_RailB.activated.connect(self.loop_ac_gain_changed_B)
        self.comboBox_ramp_height_RailA.activated.connect(self.ramp_height_changed_A)
        self.comboBox_ramp_height_RailB.activated.connect(self.ramp_height_changed_B)
        self.comboBox_integ_time_RailA.activated.connect(self.integ_time_changed_A)
        self.comboBox_integ_time_RailB.activated.connect(self.integ_time_changed_B)
        self.comboBox_OVR_threshold_RailA.activated.connect(self.OVR_threshold_changed_A)
        self.comboBox_OVR_threshold_RailB.activated.connect(self.OVR_threshold_changed_B)
        self.comboBox_UVR1_threshold_RailA.activated.connect(self.UVR1_threshold_changed_A)
        self.comboBox_UVR1_threshold_RailB.activated.connect(self.UVR1_threshold_changed_B)
        self.comboBox_UVR2_threshold_RailA.activated.connect(self.UVR2_threshold_changed_A)
        self.comboBox_UVR2_threshold_RailB.activated.connect(self.UVR2_threshold_changed_B)
        self.comboBox_UVR1_phase_RailA.activated.connect(self.UVR1_phase_changed_A)
        self.comboBox_UVR1_phase_RailB.activated.connect(self.UVR1_phase_changed_B)
        self.comboBox_UVR2_phase_RailA.activated.connect(self.UVR2_phase_changed_A)
        self.comboBox_UVR2_phase_RailB.activated.connect(self.UVR2_phase_changed_B)
        self.pushButton_Discard.clicked.connect(self.Discard)
        self.pushButton_Save.clicked.connect(self.Save)

    def isum_gain_changed_A(self):
        if self.comboBox_isum_gain_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 4, 0, "00000")
        elif self.comboBox_isum_gain_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 4, 0, "00001")
        elif self.comboBox_isum_gain_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 4, 0, "00010")
        elif self.comboBox_isum_gain_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 4, 0, "00011")
        elif self.comboBox_isum_gain_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 4, 0, "00100")
        elif self.comboBox_isum_gain_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 4, 0, "00101")
        elif self.comboBox_isum_gain_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 4, 0, "00110")
        elif self.comboBox_isum_gain_railA.currentIndex() == 7:
            update_database_with_temp_customer_input("C50", 4, 0, "00111")
        elif self.comboBox_isum_gain_railA.currentIndex() == 8:
            update_database_with_temp_customer_input("C50", 4, 0, "01000")
        elif self.comboBox_isum_gain_railA.currentIndex() == 9:
            update_database_with_temp_customer_input("C50", 4, 0, "01001")
        elif self.comboBox_isum_gain_railA.currentIndex() == 10:
            update_database_with_temp_customer_input("C50", 4, 0, "01010")
        elif self.comboBox_isum_gain_railA.currentIndex() == 11:
            update_database_with_temp_customer_input("C50", 4, 0, "01011")
        elif self.comboBox_isum_gain_railA.currentIndex() == 12:
            update_database_with_temp_customer_input("C50", 4, 0, "01100")
        elif self.comboBox_isum_gain_railA.currentIndex() == 13:
            update_database_with_temp_customer_input("C50", 4, 0, "01101")
        elif self.comboBox_isum_gain_railA.currentIndex() == 14:
            update_database_with_temp_customer_input("C50", 4, 0, "01110")
        elif self.comboBox_isum_gain_railA.currentIndex() == 15:
            update_database_with_temp_customer_input("C50", 4, 0, "01111")
        elif self.comboBox_isum_gain_railA.currentIndex() == 16:
            update_database_with_temp_customer_input("C50", 4, 0, "10000")
        elif self.comboBox_isum_gain_railA.currentIndex() == 17:
            update_database_with_temp_customer_input("C50", 4, 0, "10001")
        elif self.comboBox_isum_gain_railA.currentIndex() == 18:
            update_database_with_temp_customer_input("C50", 4, 0, "10010")
        elif self.comboBox_isum_gain_railA.currentIndex() == 19:
            update_database_with_temp_customer_input("C50", 4, 0, "10011")
        else:
            update_database_with_temp_customer_input("C50", 4, 0, "10100")

    def isum_gain_changed_B(self):
        if self.comboBox_isum_gain_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 4, 0, "00000")
        elif self.comboBox_isum_gain_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 4, 0, "00001")
        elif self.comboBox_isum_gain_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 4, 0, "00010")
        elif self.comboBox_isum_gain_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 4, 0, "00011")
        elif self.comboBox_isum_gain_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 4, 0, "00100")
        elif self.comboBox_isum_gain_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 4, 0, "00101")
        elif self.comboBox_isum_gain_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 4, 0, "00110")
        elif self.comboBox_isum_gain_railB.currentIndex() == 7:
            update_database_with_temp_customer_input("C51", 4, 0, "00111")
        elif self.comboBox_isum_gain_railB.currentIndex() == 8:
            update_database_with_temp_customer_input("C51", 4, 0, "01000")
        elif self.comboBox_isum_gain_railB.currentIndex() == 9:
            update_database_with_temp_customer_input("C51", 4, 0, "01001")
        elif self.comboBox_isum_gain_railB.currentIndex() == 10:
            update_database_with_temp_customer_input("C51", 4, 0, "01010")
        elif self.comboBox_isum_gain_railB.currentIndex() == 11:
            update_database_with_temp_customer_input("C51", 4, 0, "01011")
        elif self.comboBox_isum_gain_railB.currentIndex() == 12:
            update_database_with_temp_customer_input("C51", 4, 0, "01100")
        elif self.comboBox_isum_gain_railB.currentIndex() == 13:
            update_database_with_temp_customer_input("C51", 4, 0, "01101")
        elif self.comboBox_isum_gain_railB.currentIndex() == 14:
            update_database_with_temp_customer_input("C51", 4, 0, "01110")
        elif self.comboBox_isum_gain_railB.currentIndex() == 15:
            update_database_with_temp_customer_input("C51", 4, 0, "01111")
        elif self.comboBox_isum_gain_railB.currentIndex() == 16:
            update_database_with_temp_customer_input("C51", 4, 0, "10000")
        elif self.comboBox_isum_gain_railB.currentIndex() == 17:
            update_database_with_temp_customer_input("C51", 4, 0, "10001")
        elif self.comboBox_isum_gain_railB.currentIndex() == 18:
            update_database_with_temp_customer_input("C51", 4, 0, "10010")
        elif self.comboBox_isum_gain_railB.currentIndex() == 19:
            update_database_with_temp_customer_input("C51", 4, 0, "10011")
        else:
            update_database_with_temp_customer_input("C51", 4, 0, "10100")

    def dynamic_ramp_height_changed_A(self):
        if self.comboBox_dynamic_ramp_height_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 75, 74, "00")
        elif self.comboBox_dynamic_ramp_height_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 75, 74, "01")
        elif self.comboBox_dynamic_ramp_height_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 75, 74, "10")
        else:
            update_database_with_temp_customer_input("C50", 75, 74, "11")

    def dynamic_ramp_height_changed_B(self):
        if self.comboBox_dynamic_ramp_height_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 75, 74, "00")
        elif self.comboBox_dynamic_ramp_height_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 75, 74, "01")
        elif self.comboBox_dynamic_ramp_height_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 75, 74, "10")
        else:
            update_database_with_temp_customer_input("C51", 75, 74, "11")

    def loop_ac_gain_changed_A(self):
        if self.comboBox_loop_ac_gain_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 56, 55, "00")
            update_database_with_temp_customer_input("C50", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 56, 55, "01")
            update_database_with_temp_customer_input("C50", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 56, 55, "10")
            update_database_with_temp_customer_input("C50", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 56, 55, "11")
            update_database_with_temp_customer_input("C50", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 56, 55, "00")
            update_database_with_temp_customer_input("C50", 49, 49, "1")

        elif self.comboBox_loop_ac_gain_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 56, 55, "01")
            update_database_with_temp_customer_input("C50", 49, 49, "1")

        elif self.comboBox_loop_ac_gain_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 56, 55, "10")
            update_database_with_temp_customer_input("C50", 49, 49, "1")

        else:
            update_database_with_temp_customer_input("C50", 56, 55, "11")
            update_database_with_temp_customer_input("C50", 49, 49, "1")

    def loop_ac_gain_changed_B(self):
        if self.comboBox_loop_ac_gain_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 56, 55, "00")
            update_database_with_temp_customer_input("C51", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 56, 55, "01")
            update_database_with_temp_customer_input("C51", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 56, 55, "10")
            update_database_with_temp_customer_input("C51", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 56, 55, "11")
            update_database_with_temp_customer_input("C51", 49, 49, "0")

        elif self.comboBox_loop_ac_gain_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 56, 55, "00")
            update_database_with_temp_customer_input("C51", 49, 49, "1")

        elif self.comboBox_loop_ac_gain_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 56, 55, "01")
            update_database_with_temp_customer_input("C51", 49, 49, "1")

        elif self.comboBox_loop_ac_gain_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 56, 55, "10")
            update_database_with_temp_customer_input("C51", 49, 49, "1")

        else:
            update_database_with_temp_customer_input("C51", 56, 55, "11")
            update_database_with_temp_customer_input("C51", 49, 49, "1")

    def ramp_height_changed_A(self):
        if self.comboBox_ramp_height_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 33, 31, "000")
        elif self.comboBox_ramp_height_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 33, 31, "001")
        elif self.comboBox_ramp_height_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 33, 31, "010")
        elif self.comboBox_ramp_height_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 33, 31, "011")
        elif self.comboBox_ramp_height_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 33, 31, "100")
        elif self.comboBox_ramp_height_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 33, 31, "101")
        elif self.comboBox_ramp_height_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 33, 31, "110")
        else:
            update_database_with_temp_customer_input("C50", 33, 31, "111")

    def ramp_height_changed_B(self):
        if self.comboBox_ramp_height_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 33, 31, "000")
        elif self.comboBox_ramp_height_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 33, 31, "001")
        elif self.comboBox_ramp_height_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 33, 31, "010")
        elif self.comboBox_ramp_height_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 33, 31, "011")
        elif self.comboBox_ramp_height_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 33, 31, "100")
        elif self.comboBox_ramp_height_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 33, 31, "101")
        elif self.comboBox_ramp_height_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 33, 31, "110")
        else:
            update_database_with_temp_customer_input("C51", 33, 31, "111")

    def integ_time_changed_A(self):
        if self.comboBox_integ_time_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 52, 50, "000")
        elif self.comboBox_integ_time_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 52, 50, "001")
        elif self.comboBox_integ_time_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 52, 50, "010")
        elif self.comboBox_integ_time_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 52, 50, "011")
        elif self.comboBox_integ_time_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 52, 50, "100")
        elif self.comboBox_integ_time_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 52, 50, "101")
        elif self.comboBox_integ_time_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 52, 50, "110")
        else:
            update_database_with_temp_customer_input("C50", 52, 50, "111")

    def integ_time_changed_B(self):
        if self.comboBox_integ_time_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 52, 50, "000")
        elif self.comboBox_integ_time_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 52, 50, "001")
        elif self.comboBox_integ_time_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 52, 50, "010")
        elif self.comboBox_integ_time_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 52, 50, "011")
        elif self.comboBox_integ_time_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 52, 50, "100")
        elif self.comboBox_integ_time_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 52, 50, "101")
        elif self.comboBox_integ_time_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 52, 50, "110")
        else:
            update_database_with_temp_customer_input("C51", 52, 50, "111")

    def OVR_threshold_changed_A(self):
        if self.comboBox_OVR_threshold_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 42, 40, "000")
        elif self.comboBox_OVR_threshold_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 42, 40, "001")
        elif self.comboBox_OVR_threshold_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 42, 40, "010")
        elif self.comboBox_OVR_threshold_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 42, 40, "011")
        elif self.comboBox_OVR_threshold_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 42, 40, "100")
        elif self.comboBox_OVR_threshold_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 42, 40, "101")
        elif self.comboBox_OVR_threshold_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 42, 40, "110")
        else:
            update_database_with_temp_customer_input("C50", 42, 40, "111")

    def OVR_threshold_changed_B(self):
        if self.comboBox_OVR_threshold_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 42, 40, "000")
        elif self.comboBox_OVR_threshold_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 42, 40, "001")
        elif self.comboBox_OVR_threshold_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 42, 40, "010")
        elif self.comboBox_OVR_threshold_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 42, 40, "011")
        elif self.comboBox_OVR_threshold_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 42, 40, "100")
        elif self.comboBox_OVR_threshold_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 42, 40, "101")
        elif self.comboBox_OVR_threshold_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 42, 40, "110")
        else:
            update_database_with_temp_customer_input("C51", 42, 40, "111")

    def UVR1_threshold_changed_A(self):
        if self.comboBox_UVR1_threshold_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 22, 20, "000")
        elif self.comboBox_UVR1_threshold_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 22, 20, "001")
        elif self.comboBox_UVR1_threshold_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 22, 20, "010")
        elif self.comboBox_UVR1_threshold_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 22, 20, "011")
        elif self.comboBox_UVR1_threshold_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 22, 20, "100")
        elif self.comboBox_UVR1_threshold_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 22, 20, "101")
        elif self.comboBox_UVR1_threshold_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 22, 20, "110")
        else:
            update_database_with_temp_customer_input("C50", 22, 20, "111")

    def UVR1_threshold_changed_B(self):
        if self.comboBox_UVR1_threshold_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 22, 20, "000")
        elif self.comboBox_UVR1_threshold_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 22, 20, "001")
        elif self.comboBox_UVR1_threshold_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 22, 20, "010")
        elif self.comboBox_UVR1_threshold_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 22, 20, "011")
        elif self.comboBox_UVR1_threshold_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 22, 20, "100")
        elif self.comboBox_UVR1_threshold_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 22, 20, "101")
        elif self.comboBox_UVR1_threshold_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 22, 20, "110")
        else:
            update_database_with_temp_customer_input("C51", 22, 20, "111")

    def UVR2_threshold_changed_A(self):
        if self.comboBox_UVR2_threshold_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 17, 15, "000")
        elif self.comboBox_UVR2_threshold_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 17, 15, "001")
        elif self.comboBox_UVR2_threshold_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 17, 15, "010")
        elif self.comboBox_UVR2_threshold_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("C50", 17, 15, "011")
        elif self.comboBox_UVR2_threshold_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("C50", 17, 15, "100")
        elif self.comboBox_UVR2_threshold_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("C50", 17, 15, "101")
        elif self.comboBox_UVR2_threshold_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("C50", 17, 15, "110")
        else:
            update_database_with_temp_customer_input("C50", 17, 15, "111")

    def UVR2_threshold_changed_B(self):
        if self.comboBox_UVR2_threshold_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 17, 15, "000")
        elif self.comboBox_UVR2_threshold_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 17, 15, "001")
        elif self.comboBox_UVR2_threshold_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 17, 15, "010")
        elif self.comboBox_UVR2_threshold_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("C51", 17, 15, "011")
        elif self.comboBox_UVR2_threshold_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("C51", 17, 15, "100")
        elif self.comboBox_UVR2_threshold_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("C51", 17, 15, "101")
        elif self.comboBox_UVR2_threshold_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("C51", 17, 15, "110")
        else:
            update_database_with_temp_customer_input("C51", 17, 15, "111")

    def UVR1_phase_changed_A(self):
        if self.comboBox_UVR1_phase_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 24, 23, "00")
        else:
            update_database_with_temp_customer_input("C50", 24, 23, "01")

    def UVR1_phase_changed_B(self):
        if self.comboBox_UVR1_phase_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 24, 23, "00")
        else:
            update_database_with_temp_customer_input("C51", 24, 23, "01")

    def UVR2_phase_changed_A(self):
        if self.comboBox_UVR2_phase_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 19, 18, "00")
        else:
            update_database_with_temp_customer_input("C50", 19, 18, "01")

    def UVR2_phase_changed_B(self):
        if self.comboBox_UVR2_phase_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 19, 18, "00")
        else:
            update_database_with_temp_customer_input("C51", 19, 18, "01")

    def Save(self):
        global initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                print_log("Transient Configuration Settings saved.", "INFO")
                global list_of_registers_used_in_this_frame
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def Discard(self):
        print_log("Transient Configuration Settings discarded.", "INFO")
        self.main = Transient_Configuration()
        self.main.show()
        self.close()


class PMBus_Address(QMainWindow, Ui_PMBus_address_configuration):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(PMBus_Address, self).__init__(parent)
        self.setupUi(self)
        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["EF", "99", "9E", "9C", "9A", "9B", "9D"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        # GUI default value look initialization
        # Non feature related initializaton on GUI_ display
        # self.lineEdit_RailA.setText(RailA_name)
        # self.lineEdit_RailB.setText(RailB_name)

        # Feature related initialization on GUI display
        self.label_part_name.setText(PARTNAME)
        self.lineEdit_PMBus_address.setText(''.join(format(int(initialize_feature_variable("EF", 6, 0), 2), '02x')))
        self.lineEdit_MFR_ID.setText(''.join(format(int(initialize_feature_variable("99", 15, 0), 2), '04x')))
        self.lineEdit_MFR_location.setText(''.join(format(int(initialize_feature_variable("9C", 15, 0), 2), '04x')))
        self.lineEdit_MFR_model.setText(''.join(format(int(initialize_feature_variable("9A", 15, 0), 2), '04x')))
        self.lineEdit_MFR_revision.setText(''.join(format(int(initialize_feature_variable("9B", 15, 0), 2), '04x')))
        self.lineEdit_MFR_date.setDate(QDate((int(initialize_feature_variable("9D", 6, 0), 2) + 2000),
                                             (int(initialize_feature_variable("9D", 10, 7), 2) + 1),
                                             (int(initialize_feature_variable("9D", 15, 11), 2) + 1)))
        self.lineEdit_device_serial_number.setText(
            ''.join(format(int(initialize_feature_variable("9E", 15, 0), 2), '04x')))

        # Changing all the PMBus address frame features to read only, except PMBus address change feature
        self.lineEdit_MFR_ID.setReadOnly(True)
        self.lineEdit_MFR_model.setReadOnly(True)
        self.lineEdit_MFR_location.setReadOnly(True)
        self.lineEdit_MFR_revision.setReadOnly(True)
        self.lineEdit_MFR_date.setReadOnly(True)
        self.lineEdit_device_serial_number.setReadOnly(True)
        self.lineEdit_MFR_ID.setStyleSheet("QLineEdit{\n"
                                           "    background-color: rgb(160, 188, 173);\n"
                                           "    color: rgb(0, 0, 0);\n"
                                           "border:1px solid white;\n"
                                           "border-radius:3px;\n"
                                           "}\n"
                                           "\n"
                                           "QLineEdit::hover{\n"
                                           "border:1px solid  rgb(160, 188, 173);\n"
                                           "}\n"
                                           "\n"
                                           "QLineEdit::focus{\n"
                                           "border:2px solid grey;\n"
                                           "    background-color: rgb(160, 188, 173);\n"
                                           "}")
        self.lineEdit_MFR_date.setStyleSheet("QDateEdit{\n"
                                             "    background-color: rgb(160, 188, 173);\n"
                                             "    color: rgb(0, 0, 0);\n"
                                             "border:1px solid grey;\n"
                                             "border-radius:3px;\n"
                                             "}\n"
                                             "\n"
                                             "QDateEdit::hover{\n"
                                             "border:2px solid  rgb(160, 188, 173);\n"
                                             "    color: rgb(0, 0, 0);\n"
                                             "}\n"
                                             "\n"
                                             "QDateEdit::focus{\n"
                                             "border:2px solid grey;\n"
                                             "    background-color: rgb(160, 188, 173);\n"
                                             "    color: rgb(0, 0, 0);\n"
                                             "}")
        self.lineEdit_MFR_location.setStyleSheet("QLineEdit{\n"
                                                 "    background-color: rgb(160, 188, 173);\n"
                                                 "    color: rgb(0, 0, 0);\n"
                                                 "border:1px solid white;\n"
                                                 "border-radius:3px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QLineEdit::hover{\n"
                                                 "border:1px solid  rgb(160, 188, 173);\n"
                                                 "}\n"
                                                 "\n"
                                                 "QLineEdit::focus{\n"
                                                 "border:2px solid grey;\n"
                                                 "    background-color: rgb(160, 188, 173);\n"
                                                 "}")
        self.lineEdit_MFR_model.setStyleSheet("QLineEdit{\n"
                                              "    background-color: rgb(160, 188, 173);\n"
                                              "    color: rgb(0, 0, 0);\n"
                                              "border:1px solid white;\n"
                                              "border-radius:3px;\n"
                                              "}\n"
                                              "\n"
                                              "QLineEdit::hover{\n"
                                              "border:1px solid  rgb(160, 188, 173);\n"
                                              "}\n"
                                              "\n"
                                              "QLineEdit::focus{\n"
                                              "border:2px solid grey;\n"
                                              "    background-color: rgb(160, 188, 173);\n"
                                              "}")
        self.lineEdit_MFR_revision.setStyleSheet("QLineEdit{\n"
                                                 "    background-color: rgb(160, 188, 173);\n"
                                                 "    color: rgb(0, 0, 0);\n"
                                                 "border:1px solid white;\n"
                                                 "border-radius:3px;\n"
                                                 "}\n"
                                                 "\n"
                                                 "QLineEdit::hover{\n"
                                                 "border:1px solid  rgb(160, 188, 173);\n"
                                                 "}\n"
                                                 "\n"
                                                 "QLineEdit::focus{\n"
                                                 "border:2px solid grey;\n"
                                                 "    background-color: rgb(160, 188, 173);\n"
                                                 "}")
        self.lineEdit_device_serial_number.setStyleSheet("QLineEdit{\n"
                                                         "    background-color: rgb(160, 188, 173);\n"
                                                         "    color: rgb(0, 0, 0);\n"
                                                         "border:1px solid white;\n"
                                                         "border-radius:3px;\n"
                                                         "}\n"
                                                         "\n"
                                                         "QLineEdit::hover{\n"
                                                         "border:1px solid  rgb(160, 188, 173);\n"
                                                         "}\n"
                                                         "\n"
                                                         "QLineEdit::focus{\n"
                                                         "border:2px solid grey;\n"
                                                         "    background-color: rgb(160, 188, 173);\n"
                                                         "}")

        # Customer GUI interaction related function mapping
        self.pushButton_Discard.clicked.connect(self.Discard)
        self.pushButton_Save.clicked.connect(self.Save)
        self.lineEdit_PMBus_address.textChanged.connect(self.PMBus_address_change)
        # self.lineEdit_MFR_ID.textChanged.connect(self.MFR_ID_change)
        # self.lineEdit_MFR_location.textChanged.connect(self.MFR_location_change)
        # self.lineEdit_MFR_model.textChanged.connect(self.MFR_model_change)
        # self.lineEdit_MFR_revision.textChanged.connect(self.MFR_revision_change)
        # self.lineEdit_device_serial_number.textChanged.connect(self.device_serial_number_change)
        # self.lineEdit_MFR_date.dateChanged.connect(self.MFR_date_change)

    def PMBus_address_change(self):
        if self.lineEdit_PMBus_address.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_PMBus_address.text()) and int(
                    self.lineEdit_PMBus_address.text(), base=16) <= 127:
                update_database_with_temp_customer_input("EF", 6, 0,
                                                         bin(int(self.lineEdit_PMBus_address.text(), base=16)).split(
                                                             "0b")[1].zfill(7))
            else:
                print_log("Entered " + self.lineEdit_PMBus_address.text() + ", give PMBus address ranging (0x00 - 0x7F)", "ERROR")
                time.sleep(2)
                self.lineEdit_PMBus_address.setText(
                    ''.join(format(int(initialize_feature_variable("EF", 6, 0), 2), '02x')))

    def MFR_ID_change(self):
        if self.lineEdit_MFR_ID.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_MFR_ID.text()) and int(
                    self.lineEdit_MFR_ID.text(), base=16) <= 65535:
                update_database_with_temp_customer_input("99", 15, 0,
                                                         bin(int(self.lineEdit_MFR_ID.text(), base=16)).split("0b")[
                                                             1].zfill(16))
            else:
                print_log("Entered " + self.lineEdit_MFR_ID.text() + ", give MFR ID ranging (0x0000 - 0xffff)", "ERROR")
                time.sleep(2)
                self.lineEdit_MFR_ID.setText(''.join(format(int(initialize_feature_variable("99", 15, 0), 2), '04x')))

    def MFR_location_change(self):
        if self.lineEdit_MFR_location.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_MFR_location.text()) and int(
                    self.lineEdit_MFR_location.text(), base=16) <= 65535:
                update_database_with_temp_customer_input("9C", 15, 0,
                                                         bin(int(self.lineEdit_MFR_location.text(), base=16)).split(
                                                             "0b")[1].zfill(16))
            else:
                print_log("Entered " + self.lineEdit_MFR_location.text() + ", give MFR location ranging (0x0000 - 0xffff)", "ERROR")
                time.sleep(2)
                self.lineEdit_MFR_location.setText(
                    ''.join(format(int(initialize_feature_variable("9C", 15, 0), 2), '04x')))

    def MFR_model_change(self):
        if self.lineEdit_MFR_model.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_MFR_model.text()) and int(
                    self.lineEdit_MFR_model.text(), base=16) <= 65535:
                update_database_with_temp_customer_input("9A", 15, 0,
                                                         bin(int(self.lineEdit_MFR_model.text(), base=16)).split("0b")[
                                                             1].zfill(16))
            else:
                print_log("Entered " + self.lineEdit_MFR_model.text() + ", give MFR model ranging (0x0000 - 0xffff)", "ERROR")
                time.sleep(2)
                self.lineEdit_MFR_model.setText(
                    ''.join(format(int(initialize_feature_variable("9A", 15, 0), 2), '04x')))

    def MFR_revision_change(self):
        if self.lineEdit_MFR_revision.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_MFR_revision.text()) and int(
                    self.lineEdit_MFR_revision.text(), base=16) <= 65535:
                update_database_with_temp_customer_input("9B", 15, 0,
                                                         bin(int(self.lineEdit_MFR_revision.text(), base=16)).split(
                                                             "0b")[1].zfill(16))
            else:
                print_log("Entered " + self.lineEdit_MFR_revision.text() + ", give MFR revision ranging (0x0000 - 0xffff)", "ERROR")
                time.sleep(2)
                self.lineEdit_MFR_revision.setText(
                    ''.join(format(int(initialize_feature_variable("9B", 15, 0), 2), '04x')))

    def device_serial_number_change(self):
        if self.lineEdit_device_serial_number.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_device_serial_number.text()) and int(
                    self.lineEdit_device_serial_number.text(), base=16) <= 65535:
                update_database_with_temp_customer_input("9E", 15, 0, bin(
                    int(self.lineEdit_device_serial_number.text(), base=16)).split("0b")[1].zfill(16))
            else:
                print_log("Entered " + self.lineEdit_device_serial_number.text() + ", give device serial number ranging (0x0000 - 0xffff)", "ERROR")
                time.sleep(2)
                self.lineEdit_device_serial_number.setText(
                    ''.join(format(int(initialize_feature_variable("9E", 15, 0), 2), '04x')))

    def MFR_date_change(self):
        update_database_with_temp_customer_input("9D", 15, 0,
                                                 (bin(self.lineEdit_MFR_date.date().day() - 1).split("0b")[1].zfill(
                                                     5)) +
                                                 (bin(self.lineEdit_MFR_date.date().month() - 1).split("0b")[
                                                      1].zfill(4)) +
                                                 (bin(self.lineEdit_MFR_date.date().year() - 2000).split("0b")[
                                                      1].zfill(7)))

    def Save(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, PMBUS_ADDR, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("PMBus Address frame settings saved", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

        self.lineEdit_PMBus_address.setText(self.lineEdit_PMBus_address.text().zfill(2))
        self.lineEdit_MFR_ID.setText(hex(int(self.lineEdit_MFR_ID.text(), base=16)).split("0x")[1].zfill(4))
        self.lineEdit_MFR_model.setText(hex(int(self.lineEdit_MFR_model.text(), base=16)).split("0x")[1].zfill(4))
        self.lineEdit_MFR_location.setText(hex(int(self.lineEdit_MFR_location.text(), base=16)).split("0x")[1].zfill(4))
        self.lineEdit_MFR_revision.setText(hex(int(self.lineEdit_MFR_revision.text(), base=16)).split("0x")[1].zfill(4))
        self.lineEdit_device_serial_number.setText(hex(int(self.lineEdit_device_serial_number.text(), base=16)).split("0x")[1].zfill(4))

    def Discard(self):
        print_log("PMBus address settings discarded.", "INFO")
        self.main = PMBus_Address()
        self.main.show()
        self.close()

class frame_phase_configuration(QMainWindow, Ui_Phase_configuration_main_window):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, RailA_phase_count_arg, RailB_phase_count_arg, list_of_registers_used_in_this_frame
        super(frame_phase_configuration, self).__init__(parent)
        self.setupUi(self)
        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["DF", "C9", "E80", "E81"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        RailA_phase_count_arg = str(int(initialize_feature_variable("DF", 8, 4), 2))  # binary->decimal->string because it will be shown on GUI
        RailB_phase_count_arg = str(int(initialize_feature_variable("DF", 3, 0), 2))
        current_sense_resistor_arg = int(initialize_feature_variable("C9", 32, 32), 2)  # binary->decimal->decimal index used for combo_box in gui
        PS1_active_phases_argA = int(initialize_feature_variable("E80", 109, 106), 2)  # binary->decimal->decimal index used for combo_box in gui
        PS1_active_phases_argB = int(initialize_feature_variable("E81", 109, 106), 2)  # binary->decimal->decimal index used for combo_box in gui
        PS2_active_phases_argA = int(initialize_feature_variable("E80", 113, 110), 2)  # binary->decimal->decimal index used for combo_box in gui
        PS2_active_phases_argB = int(initialize_feature_variable("E81", 113, 110), 2)  # binary->decimal->decimal index used for combo_box in gui

        # GUI default value look initialization
        # Non feature related initializaton on GUI_ dispaly
        self.lineEdit_RailA.setText(RailA_name)
        self.lineEdit_RailB.setText(RailB_name)

        self.label_display_RailA.setText(self.lineEdit_RailA.text())
        self.label_display_RailB.setText(self.lineEdit_RailB.text())

        # Feature related initialization on GUI display
        self.lineEdit_RailA_count.setText(RailA_phase_count_arg)
        self.lineEdit_RailB_count.setText(RailB_phase_count_arg)

        self.phase_image_update()

        self.comboBox_Current_Sense_Resistor.setCurrentIndex(current_sense_resistor_arg)
        self.comboBox_PS1_railA.setCurrentIndex(PS1_active_phases_argA)
        self.comboBox_PS1_railB.setCurrentIndex(PS1_active_phases_argB)
        self.comboBox_PS2_railA.setCurrentIndex(PS2_active_phases_argA)
        self.comboBox_PS2_railB.setCurrentIndex(PS2_active_phases_argB)

        # Customer GUI interaction related function mapping
        self.lineEdit_RailA.textEdited.connect(self.railA)
        self.lineEdit_RailB.textEdited.connect(self.railB)
        self.lineEdit_RailA_count.textEdited.connect(self.RailA_count)
        self.lineEdit_RailB_count.textEdited.connect(self.RailB_count)
        self.pushButton_Save.clicked.connect(self.save)
        self.pushButton_Discard.clicked.connect(self.Discard)
        self.pushButton_RailA_plus.clicked.connect(self.RailA_plus)
        self.pushButton_RailA_minus.clicked.connect(self.RailA_minus)
        self.pushButton_RailB_plus.clicked.connect(self.RailB_plus)
        self.pushButton_RailB_minus.clicked.connect(self.RailB_minus)
        self.comboBox_PS1_railB.activated.connect(self.PS1_railB_phase_number)
        self.comboBox_PS1_railA.activated.connect(self.PS1_railA_phase_number)
        self.comboBox_PS2_railB.activated.connect(self.PS2_railB_phase_number)
        self.comboBox_PS2_railA.activated.connect(self.PS2_railA_phase_number)
        self.comboBox_Current_Sense_Resistor.activated.connect(self.current_sense_resistor)

    def phase_image_update(self):
        total_phase_count = int(RailAB_total_phase)
        phase_mapping = "A" * int(RailA_phase_count_arg) + "X" * (total_phase_count - int(RailA_phase_count_arg) - int(RailB_phase_count_arg)) + "B" * int(RailB_phase_count_arg)
        # print(phase_mapping)

        # First phase
        if phase_mapping[0] == "A":
            self.label_PhaseA1_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA1_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[0] == "X":
            self.label_PhaseUA1_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA1_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
        # Second phase
        if phase_mapping[1] == "A":
            self.label_PhaseA2_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA2_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[1] == "X":
            self.label_PhaseUA2_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA2_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
        # Third Phase
        if phase_mapping[2] == "A":
            self.label_PhaseA3_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA3_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[2] == "X":
            self.label_PhaseUA3_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA3_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
        # Fourth Phase
        if phase_mapping[3] == "A":
            self.label_PhaseA4_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA4_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[3] == "X":
            self.label_PhaseUA4_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA4_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        # Fifth Phase
        if phase_mapping[4] == "A":
            self.label_PhaseA5_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA5_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[4] == "X":
            self.label_PhaseUA5_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA5_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        # Sixth Phase
        if phase_mapping[5] == "A":
            self.label_PhaseA6_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA6_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB4_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[5] == "X":
            self.label_PhaseUA6_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA6_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB4_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[5] == "B":
            self.label_PhaseUA6_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseA6_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB4_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")

        # Seventh Phase
        if phase_mapping[6] == "A":
            self.label_PhaseA7_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA7_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB3_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[6] == "X":
            self.label_PhaseUA7_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA7_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB3_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[6] == "B":
            self.label_PhaseUA7_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseA7_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB3_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")

        # Eigth Phase
        if phase_mapping[7] == "A":
            self.label_PhaseA8_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA8_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB2_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[7] == "X":
            self.label_PhaseUA8_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA8_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB2_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[7] == "B":
            self.label_PhaseUA8_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseA8_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB2_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")

        # Ninth Phase
        if phase_mapping[8] == "A":
            self.label_PhaseA9_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")
            self.label_PhaseUA9_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB1_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[8] == "X":
            self.label_PhaseUA9_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase.png);\n""border:none;")
            self.label_PhaseA9_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB1_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")

        elif phase_mapping[8] == "B":
            self.label_PhaseUA9_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseA9_image.setStyleSheet("image: url(GUI_IMAGE/Unassigned_phase_blank.png);\n""border:none;")
            self.label_PhaseB1_image.setStyleSheet("image: url(GUI_IMAGE/Assigned_phase.png);\n""border:none;")

    def railA(self):
        # print(self.lineEdit_RailA.text()) # Print the RailA text
        self.label_display_RailA.setText(self.lineEdit_RailA.text())  # Set the RailA entered text as Label

    def railB(self):
        # print(self.lineEdit_RailB.text()) # Print the RailB text
        self.label_display_RailB.setText(self.lineEdit_RailB.text())  # Set the RailB entered text as Label

    def RailA_count(self):
        global RailA_phase_count_arg, RailB_phase_count_arg, RailA_name, RailB_name

        # print_log(RailA_name+" entered count is: "+ self.lineEdit_RailA_count.text(),"INFO")
        try:
            val = int(self.lineEdit_RailA_count.text())
        except ValueError:
            print_log("Please enter a valid integer number.", "ERROR")
            self.lineEdit_RailA_count.setText(RailA_phase_count_arg)
            return

        if int(RailB_phase_count_arg) + int(self.lineEdit_RailA_count.text()) > int(RailAB_total_phase):
            print_log("Both Rails phase sum can`t exceed: " + RailAB_total_phase, "ERROR")
            self.lineEdit_RailA_count.setText(RailA_phase_count_arg)
            return 0
        elif int(RailA_phase_count_arg) > int(RailA_phase_count_max):
            print_log(RailA_name + " has exceeded max phase count limit", "ERROR")
            self.lineEdit_RailA_count.setText(RailA_phase_count_arg)
            return 0
        else:
            print_log(RailA_name + " entered count is: " + self.lineEdit_RailA_count.text(), "INFO")
            RailA_phase_count_arg = self.lineEdit_RailA_count.text()
            update_database_with_temp_customer_input("DF", 8, 4, bin(int(RailA_phase_count_arg)).split("0b")[1].zfill(5))
            self.phase_image_update()
            self.PS2_railA_phase_number()
            self.PS1_railA_phase_number()

    def RailB_count(self):
        global RailA_phase_count_arg, RailB_phase_count_arg, RailA_name, RailB_name, RailAB_total_phase

        # print_log(RailB_name+" entered count is: "+ self.lineEdit_RailB_count.text(),"INFO")
        try:
            val = int(self.lineEdit_RailB_count.text())
        except ValueError:
            print_log("Please enter a valid integer number.", "ERROR")
            self.lineEdit_RailB_count.setText(RailB_phase_count_arg)
            return

        if int(RailA_phase_count_arg) + int(self.lineEdit_RailB_count.text()) > int(RailAB_total_phase):
            print_log("Both Rails phase sum can`t exceed: " + RailAB_total_phase, "ERROR")
            self.lineEdit_RailB_count.setText(RailB_phase_count_arg)
            return 0
        elif int(RailB_phase_count_arg) > int(RailB_phase_count_max):
            print_log(RailB_name + " has exceeded max phase count limit", "ERROR")
            self.lineEdit_RailB_count.setText(RailB_phase_count_arg)
            return 0
        else:
            print_log(RailB_name + " entered count is: " + self.lineEdit_RailB_count.text(), "INFO")
            RailB_phase_count_arg = self.lineEdit_RailB_count.text()
            update_database_with_temp_customer_input("DF", 3, 0, bin(int(RailB_phase_count_arg)).split("0b")[1].zfill(4))
            self.phase_image_update()
            self.PS2_railB_phase_number()
            self.PS1_railB_phase_number()

    def RailA_plus(self):
        global RailA_phase_count_arg, RailB_phase_count_arg, RailAB_total_phase, RailA_name, RailB_name

        if int(RailA_phase_count_arg) + int(RailB_phase_count_arg) > int(RailAB_total_phase) - 1:
            print_log("Both Rails phase sum can`t exceed: " + RailAB_total_phase, "ERROR")
            return
        elif int(RailA_phase_count_arg) > int(RailA_phase_count_max) - 1:
            print_log(RailA_name + " has exceeded max phase count limit", "ERROR")
        else:
            print_log(RailA_name + " phase count is: " + self.lineEdit_RailA_count.text(), "INFO")
            self.lineEdit_RailA_count.setText(str(int(RailA_phase_count_arg) + 1))
            RailA_phase_count_arg = str(int(RailA_phase_count_arg) + 1)
            update_database_with_temp_customer_input("DF", 8, 4, bin(int(RailA_phase_count_arg)).split("0b")[1].zfill(5))
            self.phase_image_update()
            self.PS2_railA_phase_number()
            self.PS1_railA_phase_number()

    def RailB_plus(self):
        global RailA_phase_count_arg, RailB_phase_count_arg, RailAB_total_phase, RailA_name, RailB_name
        if int(RailA_phase_count_arg) + int(RailB_phase_count_arg) > int(RailAB_total_phase) - 1:
            print_log("Both Rails phase sum can`t exceed: " + RailAB_total_phase, "ERROR")
            return 0
        elif int(RailB_phase_count_arg) > int(RailB_phase_count_max) - 1:
            print_log(RailB_name + " has exceeded max phase count limit", "ERROR")
            return 0
        else:
            print_log(RailB_name + " phase count is: " + self.lineEdit_RailB_count.text(), "INFO")
            self.lineEdit_RailB_count.setText(str(int(RailB_phase_count_arg) + 1))
            RailB_phase_count_arg = str(int(RailB_phase_count_arg) + 1)
            update_database_with_temp_customer_input("DF", 3, 0, bin(int(RailB_phase_count_arg)).split("0b")[1].zfill(4))
            self.phase_image_update()
            self.PS2_railB_phase_number()
            self.PS1_railB_phase_number()

    def RailA_minus(self):
        global RailA_phase_count_arg, RailB_phase_count_arg, RailAB_total_phase, RailA_name, RailB_name

        if int(RailA_phase_count_arg) + int(RailB_phase_count_arg) == 0:
            print_log("Both Rails phase sum can`t be negative.", "ERROR")
            return 0
        elif int(RailA_phase_count_arg) <= 0:
            print_log(RailA_name + " phase count Cant be negative.", "ERROR")
            return 0
        else:
            print_log(RailA_name + " phase count is: " + self.lineEdit_RailA_count.text(), "INFO")
            self.lineEdit_RailA_count.setText(str(int(RailA_phase_count_arg) - 1))
            RailA_phase_count_arg = str(int(RailA_phase_count_arg) - 1)
            update_database_with_temp_customer_input("DF", 8, 4, bin(int(RailA_phase_count_arg)).split("0b")[1].zfill(5))
            self.phase_image_update()

            self.PS2_railA_phase_number()
            self.PS1_railA_phase_number()

    def RailB_minus(self):
        global RailA_phase_count_arg, RailB_phase_count_arg, RailAB_total_phase, RailA_name, RailB_name

        if int(RailA_phase_count_arg) + int(RailB_phase_count_arg) == 0:
            print_log("Both Rails phase sum can`t be negative.", "ERROR")
            return
        elif int(RailB_phase_count_arg) <= 0:
            print_log(RailB_name + " phase count Cant be negative.", "ERROR")
            return
        else:
            print_log(RailB_name + " phase count is: " + self.lineEdit_RailB_count.text(), "INFO")
            self.lineEdit_RailB_count.setText(str(int(RailB_phase_count_arg) - 1))
            RailB_phase_count_arg = str(int(RailB_phase_count_arg) - 1)
            update_database_with_temp_customer_input("DF", 3, 0, bin(int(RailB_phase_count_arg)).split("0b")[1].zfill(4))
            self.phase_image_update()
            self.PS2_railB_phase_number()
            self.PS1_railB_phase_number()

    def PS1_railB_phase_number(self):
        num_phase = self.comboBox_PS1_railB.currentIndex()
        if (num_phase > int(RailB_phase_count_arg)):
            num_phase = int(RailB_phase_count_arg)
            self.comboBox_PS1_railB.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E81", 109, 106, bin(int(num_phase)).split("0b")[1].zfill(4))
        else:

            self.comboBox_PS1_railB.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E81", 109, 106, bin(int(num_phase)).split("0b")[1].zfill(4))

    def PS1_railA_phase_number(self):
        num_phase = self.comboBox_PS1_railA.currentIndex()
        if (num_phase > int(RailA_phase_count_arg)):
            num_phase = int(RailA_phase_count_arg)
            self.comboBox_PS1_railA.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E80", 109, 106, bin(int(num_phase)).split("0b")[1].zfill(4))
        else:

            self.comboBox_PS1_railA.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E80", 109, 106, bin(int(num_phase)).split("0b")[1].zfill(4))

    def PS2_railB_phase_number(self):
        num_phase = self.comboBox_PS2_railB.currentIndex()
        if (num_phase > int(RailB_phase_count_arg)):
            num_phase = int(RailB_phase_count_arg)
            self.comboBox_PS2_railB.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E81", 113, 110, bin(int(num_phase)).split("0b")[1].zfill(4))
        else:

            self.comboBox_PS2_railB.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E81", 113, 110, bin(int(num_phase)).split("0b")[1].zfill(4))

    def PS2_railA_phase_number(self):
        num_phase = self.comboBox_PS2_railA.currentIndex()
        if (num_phase > int(RailA_phase_count_arg)):
            num_phase = int(RailA_phase_count_arg)
            self.comboBox_PS2_railA.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E80", 113, 110, bin(int(num_phase)).split("0b")[1].zfill(4))
        else:

            self.comboBox_PS2_railA.setCurrentIndex(num_phase)
            update_database_with_temp_customer_input("E80", 113, 110, bin(int(num_phase)).split("0b")[1].zfill(4))

    def current_sense_resistor(self):
        if self.comboBox_Current_Sense_Resistor.currentText() == "External":

            print_log("External resistor has been chosen for sensing current but this feature is not supported.", "INFO")
        elif self.comboBox_Current_Sense_Resistor.currentText() == "Internal":
            print_log("Internal resistor has been chosen for sensing current but this feature is not supported.", "INFO")

    def save(self):
        global list_of_registers_used_in_this_frame, RailA_name, RailB_name, RailB_phase_count_arg, \
            RailA_phase_count_arg, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        # make the temporary values as final values in the relevant register entries in database.
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)

                print_log(self.lineEdit_RailA.text() + " has " + self.lineEdit_RailA_count.text() + " active phases. " + self.lineEdit_RailB.text() + " has " + self.lineEdit_RailB_count.text() + " active phases.", "INFO")
                RailA_name = self.lineEdit_RailA.text()
                RailB_name = self.lineEdit_RailB.text()
                RailA_phase_count_arg = self.lineEdit_RailA_count.text()
                RailB_phase_count_arg = self.lineEdit_RailB_count.text()
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def Discard(self):
        print_log("Phase manager Settings discarded.", "WARNING")

        self.main = frame_phase_configuration()
        self.main.show()
        self.close()
        # initializing the global variables just in case they got updated during temp interaction
        RailA_phase_count_arg = self.lineEdit_RailA_count.text()
        RailB_phase_count_arg = self.lineEdit_RailB_count.text()
        # self.close()


class Boot_voltage_configuration(QMainWindow, Ui_Boot_Voltage):

    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame, resolutionA, resolutionB,PARTNAME
        super(Boot_voltage_configuration, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["E4", "E10", "210", "E11", "211"]

        resolution_calculation(0)

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        initial_vid_raila = float((int(initialize_feature_variable('E10', 23, 13), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('E10', 23, 13), 2) != 0 else int(0)
        self.lineEdit_RailA_bootVoltage.setText(str(initial_vid_raila))

        if (int(initialize_feature_variable("010", 5, 4), 2) == 0) and (int(initialize_feature_variable("011", 5, 4), 2) == 0):
            self.radioButton_PIN_railA.setDisabled(True)
        else:
            self.radioButton_PIN_railA.setDisabled(False)


        # if int(initialize_feature_variable("201", 4, 0), 2) == 30:
        #     resolutionB = "6.25"
        # else:
        #     if int(initialize_feature_variable("E4", 3, 3), 2) == 0:
        #         resolutionB = "5"
        #     else:
        #         resolutionB = "10"
        #
        # if int(initialize_feature_variable("200", 4, 0), 2) == 30:
        #     resolutionA = "6.25"
        # else:
        #     if int(initialize_feature_variable("E4", 3, 3), 2) == 0:
        #         resolutionA = "5"
        #     else:
        #         resolutionA = "10"

        initial_vid_railb = float((int(initialize_feature_variable('E11', 23, 13), 2) * float(resolutionB) + offset[
            str(resolutionB)]) / 1000) if int(initialize_feature_variable('E11', 23, 13), 2) != 0 else int(0)
        self.lineEdit_RailB_bootVoltage.setText(str(initial_vid_railb))

        self.comboBox_slewrate_RailA.setCurrentIndex(int(initialize_feature_variable('E10', 4, 2), 2))
        self.comboBox_slewrate_RailB.setCurrentIndex(int(initialize_feature_variable('E11', 4, 2), 2))

        if str(initialize_feature_variable('E4', 12, 12)) == '0':
            self.radioButton_PIN_railA.setChecked(True)
            self.radioButton_MTP_railA.setChecked(False)
            self.lineEdit_RailA_bootVoltage.setHidden(True)
            self.lineEdit_RailB_bootVoltage.setHidden(True)
        else:
            self.radioButton_PIN_railA.setChecked(False)
            self.radioButton_MTP_railA.setChecked(True)
            self.lineEdit_RailA_bootVoltage.setHidden(False)
            self.lineEdit_RailB_bootVoltage.setHidden(False)

        # Other initialization
        self.label_3.setText(RailA_name + " :Boot Voltage")
        self.label_4.setText(RailB_name + " :Boot Voltage")

        self.radioButton_PIN_railA.toggled.connect(self.boot_onClicked)
        self.radioButton_MTP_railA.toggled.connect(self.boot_onClicked)

        self.pushButton_save_bootVoltage.clicked.connect(self.boot_save_railA)

        self.lineEdit_RailA_bootVoltage.textEdited.connect(self.railA_boot_update)
        self.lineEdit_RailB_bootVoltage.textEdited.connect(self.railB_boot_update)

        self.pushButton_discard_bootVoltage.clicked.connect(self.boot_discard_railA)

        self.comboBox_slewrate_RailA.activated.connect(self.railA_boot_slew)
        self.comboBox_slewrate_RailB.activated.connect(self.railB_boot_slew)

    def railA_boot_slew(self):
        if self.comboBox_slewrate_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("E10", 4, 2, "000")
        elif self.comboBox_slewrate_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("E10", 4, 2, "001")
        elif self.comboBox_slewrate_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("E10", 4, 2, "010")
        elif self.comboBox_slewrate_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("E10", 4, 2, "011")
        elif self.comboBox_slewrate_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("E10", 4, 2, "100")
        elif self.comboBox_slewrate_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("E10", 4, 2, "101")
        elif self.comboBox_slewrate_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("E10", 4, 2, "110")
        else:
            update_database_with_temp_customer_input("E10", 4, 2, "111")

    def railB_boot_slew(self):
        if self.comboBox_slewrate_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("E11", 4, 2, "000")
        elif self.comboBox_slewrate_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("E11", 4, 2, "001")
        elif self.comboBox_slewrate_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("E11", 4, 2, "010")
        elif self.comboBox_slewrate_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("E11", 4, 2, "011")
        elif self.comboBox_slewrate_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("E11", 4, 2, "100")
        elif self.comboBox_slewrate_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("E11", 4, 2, "101")
        elif self.comboBox_slewrate_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("E11", 4, 2, "110")
        else:
            update_database_with_temp_customer_input("E11", 4, 2, "111")

    def railA_boot_update(self):
        if self.lineEdit_RailA_bootVoltage.text() == "":
            return
        try:
            float(self.lineEdit_RailA_bootVoltage.text()).is_integer()
        except:
            QMessageBox.about(self, "Error", "Text box value should be number")
            return

        if self.radioButton_MTP_railA.isChecked() == True:
            if self.lineEdit_RailA_bootVoltage.text() == "":
                QMessageBox.about(self, "Error", "Text box should not be empty")
            elif float(self.lineEdit_RailA_bootVoltage.text()) > 2.8 or float(self.lineEdit_RailA_bootVoltage.text()) < 0.25 and float(self.lineEdit_RailA_bootVoltage.text()) != 0:
                QMessageBox.about(self, "Error", "Input voltage is out of range")
            else:
                boot_voltage = int((float(self.lineEdit_RailA_bootVoltage.text()) * 1000 - float(offset[resolutionA])) / float(resolutionA)) if float(self.lineEdit_RailA_bootVoltage.text()) != 0 else int(0)
                update_database_with_temp_customer_input("E10", 23, 13, bin(boot_voltage).split("0b")[1].zfill(11))
                if int(initialize_feature_variable('010', 5, 4), 2) == 0:
                    update_database_with_temp_customer_input("210", 10, 0, bin(boot_voltage).split("0b")[1].zfill(11))

    def railB_boot_update(self):
        if self.lineEdit_RailB_bootVoltage.text() == "":
            return;
        try:
            float(self.lineEdit_RailB_bootVoltage.text()).is_integer()
        except:
            QMessageBox.about(self, "Error", "Text box value should be number")
            return

        if self.radioButton_MTP_railA.isChecked() == True:
            boot_voltage_sel = 1
            if self.lineEdit_RailB_bootVoltage.text() == "":
                QMessageBox.about(self, "Error", "Text box should not be empty")
            elif float(self.lineEdit_RailB_bootVoltage.text()) > 2.8 or float(self.lineEdit_RailB_bootVoltage.text()) < 0.25 and float(self.lineEdit_RailB_bootVoltage.text()) != 0:
                QMessageBox.about(self, "Error", "Input voltage is out of range")
            else:
                boot_voltage = int((float(self.lineEdit_RailB_bootVoltage.text()) * 1000 - float(offset[resolutionB])) / float(resolutionB)) if float(self.lineEdit_RailB_bootVoltage.text()) != 0 else int(0)
                update_database_with_temp_customer_input("E11", 23, 13, bin(boot_voltage).split("0b")[1].zfill(11))
                if int(initialize_feature_variable('011', 5, 4), 2) == 0:
                    update_database_with_temp_customer_input("211", 10, 0, bin(boot_voltage).split("0b")[1].zfill(11))


    def boot_onClicked(self):
        if self.radioButton_PIN_railA.isChecked() == True:
            self.lineEdit_RailA_bootVoltage.setHidden(True)
            self.lineEdit_RailB_bootVoltage.setHidden(True)
            update_database_with_temp_customer_input("E4", 12, 12, "0")
        elif self.radioButton_MTP_railA.isChecked() == True:
            self.lineEdit_RailA_bootVoltage.setHidden(False)
            self.lineEdit_RailB_bootVoltage.setHidden(False)
            update_database_with_temp_customer_input("E4", 12, 12, "1")

    def boot_save_railA(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("Boot Voltage settings saved", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def boot_discard_railA(self):
        print_log("Boot manager Settings discarded.", "INFO")
        self.main = Boot_voltage_configuration()
        self.main.show()
        self.close()

class MTPLoadingWindow(QMainWindow, Ui_Reading_Device_Registers):

    def __init__(self, parent=None):
        global PARTNAME
        super(MTPLoadingWindow, self).__init__(parent)
        self.setupUi(self)
        self.label_device.setText(PARTNAME)
        ## Remove title bar
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)  # Remove the window thing
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)  # make non-frame stuff transparent

        ## Timer Start
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.progress)
        self.timer.start(35)
        # self.timer.start(2)

    def progress(self):
        global counter, homeWin_obj, parallel_thread

        # Set value to Progress BAR
        self.progressBar.setValue(counter)
        if counter > 100:
            self.timer.stop()

            self.main = homeWin_obj
            # self.label.setText(PARTNAME)
            self.main.show()
            self.close()
            print("completed MTP load window task")
            counter = 0
            # stop_thread = False
            # parallel_thread.start()

        counter += 1

class Phase_add_drop(QMainWindow, Ui_Phase_add_drop):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(Phase_add_drop, self).__init__(parent)
        self.setupUi(self)
        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["C9", "C50", "E80", "C51", "E81"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        # GUI default value look initialization
        # Non feature related initializaton on GUI_ dispaly
        # self.lineEdit_RailA.setText(RailA_name)
        # self.lineEdit_RailB.setText(RailB_name)

        # Feature related initialization on GUI display
        self.label_railA_main.setText("<html><head/><body><p><span style=\" font-weight:600;\">" + RailA_name + "</span>: <span style=\" font-size:14pt;\">Autonomous Phase manager</span></p></body></html>")
        self.label_railB_main.setText("<html><head/><body><p><span style=\" font-weight:600;\">" + RailB_name + "</span>: <span style=\" font-size:14pt;\">Autonomous Phase manager</span></p></body></html>")
        self.comboBox_railA_coarse12.setCurrentIndex(int(initialize_feature_variable("C9", 68, 64), 2))
        self.comboBox_railA_coarse23.setCurrentIndex(int(initialize_feature_variable("C9", 63, 61), 2))
        self.comboBox_railA_coarse34.setCurrentIndex(int(initialize_feature_variable("C9", 60, 58), 2))
        self.comboBox_railA_coarse45.setCurrentIndex(int(initialize_feature_variable("C9", 57, 55), 2))
        self.comboBox_railA_coarse56.setCurrentIndex(int(initialize_feature_variable("C9", 54, 52), 2))
        self.comboBox_railA_coarse67.setCurrentIndex(int(initialize_feature_variable("C9", 51, 49), 2))
        self.comboBox_railA_coarse78.setCurrentIndex(int(initialize_feature_variable("C9", 48, 46), 2))
        self.comboBox_railA_coarse89.setCurrentIndex(int(initialize_feature_variable("C9", 45, 43), 2))
        self.comboBox_railB_coarse12.setCurrentIndex(int(initialize_feature_variable("C9", 42, 40), 2))
        self.comboBox_railB_coarse23.setCurrentIndex(int(initialize_feature_variable("C9", 39, 37), 2))
        self.comboBox_railB_coarse34.setCurrentIndex(int(initialize_feature_variable("C9", 35, 33), 2))
        self.comboBox_offset_railA.setCurrentIndex(int(initialize_feature_variable("C50", 68, 67), 2))
        self.comboBox_offset_railB.setCurrentIndex(int(initialize_feature_variable("C51", 68, 67), 2))
        self.comboBox_hysteresis_railA.setCurrentIndex(int(initialize_feature_variable("C50", 66, 65), 2))
        self.comboBox_hysteresis_railB.setCurrentIndex(int(initialize_feature_variable("C51", 66, 65), 2))

        self.pushButton_Enable_APD_railA.setCheckable(True)
        self.pushButton_Enable_APD_railA.setChecked(not (int(initialize_feature_variable("E80", 114, 114), 2)))

        if self.pushButton_Enable_APD_railA.isChecked():
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but1.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but1_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")
        else:
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but0.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but0_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")

        self.pushButton_Enable_APD_railB.setChecked(not (int(initialize_feature_variable("E81", 114, 114), 2)))
        self.pushButton_Enable_APD_railB.setCheckable(True)
        if self.pushButton_Enable_APD_railB.isChecked():
            self.pushButton_Enable_APD_railB.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but1.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but1_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")
        else:
            self.pushButton_Enable_APD_railB.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but0.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but0_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")

        self.initial_RailA()
        self.initial_RailB()

        # Customer GUI interaction related function mapping
        self.pushButton_discard_railB.clicked.connect(self.Discard_RailB)
        self.pushButton_Enable_APD_railA.clicked.connect(self.Enable_RailA)
        self.pushButton_Enable_APD_railB.clicked.connect(self.Enable_RailB)
        self.comboBox_railA_coarse12.activated.connect(self.changevalue_phase12_A)
        self.comboBox_railA_coarse23.activated.connect(self.changevalue_phase23_A)
        self.comboBox_railA_coarse34.activated.connect(self.changevalue_phase34_A)
        self.comboBox_railA_coarse45.activated.connect(self.changevalue_phase45_A)
        self.comboBox_railA_coarse56.activated.connect(self.changevalue_phase56_A)
        self.comboBox_railA_coarse67.activated.connect(self.changevalue_phase67_A)
        self.comboBox_railA_coarse78.activated.connect(self.changevalue_phase78_A)
        self.comboBox_railA_coarse89.activated.connect(self.changevalue_phase89_A)
        self.comboBox_railB_coarse12.activated.connect(self.changevalue_phase12_B)
        self.comboBox_railB_coarse23.activated.connect(self.changevalue_phase23_B)
        self.comboBox_railB_coarse34.activated.connect(self.changevalue_phase34_B)
        self.pushButton_save_railB.clicked.connect(self.save_RailB)
        self.comboBox_offset_railA.activated.connect(self.changevalue_offset_A)
        self.comboBox_offset_railB.activated.connect(self.changevalue_offset_B)
        self.comboBox_hysteresis_railA.activated.connect(self.changevalue_hysteresis_A)
        self.comboBox_hysteresis_railB.activated.connect(self.changevalue_hysteresis_B)

    def initial_RailA(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse12.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display12_railA.setText(str(val) + "A")
        self.label_displayH12_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse23.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display23_railA.setText(str(val) + "A")
        self.label_displayH23_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse34.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display34_railA.setText(str(val) + "A")
        self.label_displayH34_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse45.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display45_railA.setText(str(val) + "A")
        self.label_displayH45_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse56.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display56_railA.setText(str(val) + "A")
        self.label_displayH56_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse67.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display67_railA.setText(str(val) + "A")
        self.label_displayH67_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse78.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display78_railA.setText(str(val) + "A")
        self.label_displayH78_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse89.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display89_railA.setText(str(val) + "A")
        self.label_displayH89_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")

    def initial_RailB(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railB_coarse12.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railB.currentText())))
        self.label_display12_railB.setText(str(val) + "A")
        self.label_displayH12_railB.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railB.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railB_coarse23.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railB.currentText())))
        self.label_display23_railB.setText(str(val) + "A")
        self.label_displayH23_railB.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railB.currentText())))) + "A")

        val = int(''.join(filter(str.isdigit, self.comboBox_railB_coarse34.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railB.currentText())))
        self.label_display34_railB.setText(str(val) + "A")
        self.label_displayH34_railB.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railB.currentText())))) + "A")

    def changevalue_phase12_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse12.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display12_railA.setText(str(val) + "A")
        self.label_displayH12_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 66, 64, bin(self.comboBox_railA_coarse12.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase23_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse23.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display23_railA.setText(str(val) + "A")
        self.label_displayH23_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 63, 61, bin(self.comboBox_railA_coarse23.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase34_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse34.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display34_railA.setText(str(val) + "A")
        self.label_displayH34_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 60, 58, bin(self.comboBox_railA_coarse34.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase45_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse45.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display45_railA.setText(str(val) + "A")
        self.label_displayH45_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 57, 55, bin(self.comboBox_railA_coarse45.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase56_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse56.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display56_railA.setText(str(val) + "A")
        self.label_displayH56_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 54, 52, bin(self.comboBox_railA_coarse56.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase67_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse67.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display67_railA.setText(str(val) + "A")
        self.label_displayH67_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 51, 49, bin(self.comboBox_railA_coarse67.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase78_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse78.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display78_railA.setText(str(val) + "A")
        self.label_displayH78_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 48, 46, bin(self.comboBox_railA_coarse78.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase89_A(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railA_coarse89.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railA.currentText())))
        self.label_display89_railA.setText(str(val) + "A")
        self.label_displayH89_railA.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railA.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 45, 43, bin(self.comboBox_railA_coarse89.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase12_B(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railB_coarse12.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railB.currentText())))
        self.label_display12_railB.setText(str(val) + "A")
        self.label_displayH12_railB.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railB.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 42, 40, bin(self.comboBox_railB_coarse12.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase23_B(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railB_coarse23.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railB.currentText())))
        self.label_display23_railB.setText(str(val) + "A")
        self.label_displayH23_railB.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railB.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 39, 37, bin(self.comboBox_railB_coarse23.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_phase34_B(self):
        val = int(''.join(filter(str.isdigit, self.comboBox_railB_coarse34.currentText()))) + \
              int(''.join(filter(str.isdigit, self.comboBox_offset_railB.currentText())))
        self.label_display34_railB.setText(str(val) + "A")
        self.label_displayH34_railB.setText(
            str(val - int(''.join(filter(str.isdigit, self.comboBox_hysteresis_railB.currentText())))) + "A")
        update_database_with_temp_customer_input("C9", 35, 33, bin(self.comboBox_railB_coarse34.currentIndex()).split("0b")[1].zfill(3))

    def changevalue_offset_A(self):
        update_database_with_temp_customer_input("C50", 68, 67, bin(self.comboBox_offset_railA.currentIndex()).split("0b")[1].zfill(2))
        self.initial_RailA()

    def changevalue_offset_B(self):
        update_database_with_temp_customer_input("C51", 68, 67, bin(self.comboBox_offset_railB.currentIndex()).split("0b")[1].zfill(2))
        self.initial_RailB()

    def changevalue_hysteresis_A(self):
        update_database_with_temp_customer_input("C50", 66, 65, bin(self.comboBox_hysteresis_railA.currentIndex()).split("0b")[1].zfill(2))
        self.initial_RailA()

    def changevalue_hysteresis_B(self):
        update_database_with_temp_customer_input("C51", 66, 65, bin(self.comboBox_hysteresis_railB.currentIndex()).split("0b")[1].zfill(2))
        self.initial_RailB()

    def Enable_RailA(self):
        # if button is checked
        if self.pushButton_Enable_APD_railA.isChecked():
            print_log("Selected Auto mode for RailA", "INFO")
            update_database_with_temp_customer_input("E80", 114, 114, "0")
            # setting image to button1
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but1.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but1_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")

            # if it is unchecked
        else:
            print_log("Selected Manual mode for RailA", "INFO")
            update_database_with_temp_customer_input("E80", 114, 114, "1")
            # setting image to button0
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but0.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but0_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")

    def Enable_RailB(self):
        # if button is checked
        if self.pushButton_Enable_APD_railB.isChecked():
            print_log("Selected Auto mode for RailB", "INFO")
            update_database_with_temp_customer_input("E81", 114, 114, "0")
            # setting image to button1
            self.pushButton_Enable_APD_railB.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but1.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but1_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")

            # if it is unchecked
        else:
            print_log("Selected Manual mode for RailB", "INFO")
            update_database_with_temp_customer_input("E81", 114, 114, "1")
            # setting image to button0
            self.pushButton_Enable_APD_railB.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but0.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but0_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")

    def save_RailB(self):
        global initial_device_status, homeWin_obj, parallel_thread, stop_thread
        phase_list1a = [int(''.join(filter(str.isdigit, self.comboBox_railA_coarse12.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse23.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse34.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse45.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse56.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse67.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse78.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railA_coarse89.currentText())))]

        phase_list1 = phase_list1a[:]
        phase_list1.sort()

        phase_list1b = [int(''.join(filter(str.isdigit, self.comboBox_railB_coarse12.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railB_coarse23.currentText()))),
                        int(''.join(filter(str.isdigit, self.comboBox_railB_coarse34.currentText())))]

        phase_list2 = phase_list1b[:]
        phase_list2.sort()
        if phase_list1 == phase_list1a and phase_list2 == phase_list1b:
            if initial_device_status == 0:
                if homeWin_obj.pushButton_Enable_VR.isChecked():
                    stop_thread = True
                    time.sleep(1)
                    print_log("Autonomous phase manager settings have been saved.", "INFO")
                    global list_of_registers_used_in_this_frame
                    for i in list_of_registers_used_in_this_frame:
                        if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                            # update the final register
                            register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                            register_database[i]['Customer_interaction'] = "YES"
                            write_PMBUS_entry_in_command_xlsx(i)
                    time.sleep(0.1)
                    stop_thread = False
                    parallel_thread.start()
                else:
                    print_log("Enable VR to write into the device", "WARNING")
            else:
                print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

        else:
            if phase_list1 != phase_list1a:
                print_log("Rail A phase order " + str(phase_list1a), "ERROR")
            if phase_list2 != phase_list1b:
                print_log("Rail B phase order " + str(phase_list1b), "ERROR")
            print_log("Please choose valid phase currents for RailA and RailB", "ERROR")

    def Discard_RailB(self):
        print_log("Autonomous Phase Manager Settings discarded.", "INFO")
        self.main = Phase_add_drop()
        self.main.show()
        self.close()


class Phase_thermal_balance(QMainWindow, Ui_Phase_thermal_balance):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(Phase_thermal_balance, self).__init__(parent)
        self.setupUi(self)
        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["C9", "C50", "C51"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        # GUI default value look initialization
        # Non feature related initializaton on GUI_ display
        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        # Feature related initialization on GUI display
        self.comboBox_thermal_offset1.setCurrentIndex(int(initialize_feature_variable("C9", 26, 24), 2))
        self.comboBox_thermal_offset2.setCurrentIndex(int(initialize_feature_variable("C9", 23, 21), 2))
        self.comboBox_thermal_offset3.setCurrentIndex(int(initialize_feature_variable("C9", 20, 18), 2))
        self.comboBox_thermal_offset4.setCurrentIndex(int(initialize_feature_variable("C9", 17, 15), 2))
        self.comboBox_thermal_offset5.setCurrentIndex(int(initialize_feature_variable("C9", 14, 12), 2))
        self.comboBox_thermal_offset6.setCurrentIndex(int(initialize_feature_variable("C9", 11, 9), 2))
        self.comboBox_thermal_offset7.setCurrentIndex(int(initialize_feature_variable("C9", 8, 6), 2))
        self.comboBox_thermal_offset8.setCurrentIndex(int(initialize_feature_variable("C9", 5, 3), 2))
        self.comboBox_thermal_offset9.setCurrentIndex(int(initialize_feature_variable("C9", 2, 0), 2))

        if initialize_feature_variable("C50", 79, 78) == "11":
            self.comboBox_freq_switch_RailA.setCurrentIndex(0)
        elif initialize_feature_variable("C50", 79, 78) == "10":
            self.comboBox_freq_switch_RailA.setCurrentIndex(1)
        elif initialize_feature_variable("C50", 79, 78) == "01":
            self.comboBox_freq_switch_RailA.setCurrentIndex(2)
        else:
            self.comboBox_freq_switch_RailA.setCurrentIndex(3)

        if initialize_feature_variable("C51", 79, 78) == "11":
            self.comboBox_freq_switch_RailB.setCurrentIndex(0)
        elif initialize_feature_variable("C51", 79, 78) == "10":
            self.comboBox_freq_switch_RailB.setCurrentIndex(1)
        elif initialize_feature_variable("C51", 79, 78) == "01":
            self.comboBox_freq_switch_RailB.setCurrentIndex(2)
        else:
            self.comboBox_freq_switch_RailB.setCurrentIndex(3)

        # Customer GUI interaction related function mapping
        self.comboBox_thermal_offset1.activated.connect(self.thermal_offset1_changed)
        self.comboBox_thermal_offset2.activated.connect(self.thermal_offset2_changed)
        self.comboBox_thermal_offset3.activated.connect(self.thermal_offset3_changed)
        self.comboBox_thermal_offset4.activated.connect(self.thermal_offset4_changed)
        self.comboBox_thermal_offset5.activated.connect(self.thermal_offset5_changed)
        self.comboBox_thermal_offset6.activated.connect(self.thermal_offset6_changed)
        self.comboBox_thermal_offset7.activated.connect(self.thermal_offset7_changed)
        self.comboBox_thermal_offset8.activated.connect(self.thermal_offset8_changed)
        self.comboBox_thermal_offset9.activated.connect(self.thermal_offset9_changed)
        self.comboBox_freq_switch_RailA.activated.connect(self.freq_switch_changed_A)
        self.comboBox_freq_switch_RailB.activated.connect(self.freq_switch_changed_B)
        self.pushButton_Discard.clicked.connect(self.Discard)
        self.pushButton_Save.clicked.connect(self.Save)

    def thermal_offset1_changed(self):
        if self.comboBox_thermal_offset1.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 26, 24, "000")
        elif self.comboBox_thermal_offset1.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 26, 24, "001")
        elif self.comboBox_thermal_offset1.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 26, 24, "010")
        elif self.comboBox_thermal_offset1.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 26, 24, "011")
        elif self.comboBox_thermal_offset1.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 26, 24, "100")
        elif self.comboBox_thermal_offset1.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 26, 24, "101")
        elif self.comboBox_thermal_offset1.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 26, 24, "110")
        else:
            update_database_with_temp_customer_input("C9", 26, 24, "111")

    def thermal_offset2_changed(self):
        if self.comboBox_thermal_offset2.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 23, 21, "000")
        elif self.comboBox_thermal_offset2.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 23, 21, "001")
        elif self.comboBox_thermal_offset2.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 23, 21, "010")
        elif self.comboBox_thermal_offset2.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 23, 21, "011")
        elif self.comboBox_thermal_offset2.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 23, 21, "100")
        elif self.comboBox_thermal_offset2.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 23, 21, "101")
        elif self.comboBox_thermal_offset2.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 23, 21, "110")
        else:
            update_database_with_temp_customer_input("C9", 23, 21, "111")

    def thermal_offset3_changed(self):
        if self.comboBox_thermal_offset3.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 20, 18, "000")
        elif self.comboBox_thermal_offset3.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 20, 18, "001")
        elif self.comboBox_thermal_offset3.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 20, 18, "010")
        elif self.comboBox_thermal_offset3.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 20, 18, "011")
        elif self.comboBox_thermal_offset3.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 20, 18, "100")
        elif self.comboBox_thermal_offset3.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 20, 18, "101")
        elif self.comboBox_thermal_offset3.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 20, 18, "110")
        else:
            update_database_with_temp_customer_input("C9", 20, 18, "111")

    def thermal_offset4_changed(self):
        if self.comboBox_thermal_offset4.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 17, 15, "000")
        elif self.comboBox_thermal_offset4.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 17, 15, "001")
        elif self.comboBox_thermal_offset4.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 17, 15, "010")
        elif self.comboBox_thermal_offset4.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 17, 15, "011")
        elif self.comboBox_thermal_offset4.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 17, 15, "100")
        elif self.comboBox_thermal_offset4.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 17, 15, "101")
        elif self.comboBox_thermal_offset4.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 17, 15, "110")
        else:
            update_database_with_temp_customer_input("C9", 17, 15, "111")

    def thermal_offset5_changed(self):
        if self.comboBox_thermal_offset5.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 14, 12, "000")
        elif self.comboBox_thermal_offset5.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 14, 12, "001")
        elif self.comboBox_thermal_offset5.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 14, 12, "010")
        elif self.comboBox_thermal_offset5.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 14, 12, "011")
        elif self.comboBox_thermal_offset5.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 14, 12, "100")
        elif self.comboBox_thermal_offset5.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 14, 12, "101")
        elif self.comboBox_thermal_offset5.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 14, 12, "110")
        else:
            update_database_with_temp_customer_input("C9", 14, 12, "111")

    def thermal_offset6_changed(self):
        if self.comboBox_thermal_offset6.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 11, 9, "000")
        elif self.comboBox_thermal_offset6.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 11, 9, "001")
        elif self.comboBox_thermal_offset6.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 11, 9, "010")
        elif self.comboBox_thermal_offset6.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 11, 9, "011")
        elif self.comboBox_thermal_offset6.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 11, 9, "100")
        elif self.comboBox_thermal_offset6.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 11, 9, "101")
        elif self.comboBox_thermal_offset6.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 11, 9, "110")
        else:
            update_database_with_temp_customer_input("C9", 11, 9, "111")

    def thermal_offset7_changed(self):
        if self.comboBox_thermal_offset7.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 8, 6, "000")
        elif self.comboBox_thermal_offset7.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 8, 6, "001")
        elif self.comboBox_thermal_offset7.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 8, 6, "010")
        elif self.comboBox_thermal_offset7.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 8, 6, "011")
        elif self.comboBox_thermal_offset7.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 8, 6, "100")
        elif self.comboBox_thermal_offset7.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 8, 6, "101")
        elif self.comboBox_thermal_offset7.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 8, 6, "110")
        else:
            update_database_with_temp_customer_input("C9", 8, 6, "111")

    def thermal_offset8_changed(self):
        if self.comboBox_thermal_offset8.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 5, 3, "000")
        elif self.comboBox_thermal_offset8.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 5, 3, "001")
        elif self.comboBox_thermal_offset8.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 5, 3, "010")
        elif self.comboBox_thermal_offset8.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 5, 3, "011")
        elif self.comboBox_thermal_offset8.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 5, 3, "100")
        elif self.comboBox_thermal_offset8.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 5, 3, "101")
        elif self.comboBox_thermal_offset8.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("C9", 5, 3, "111")

    def thermal_offset9_changed(self):
        if self.comboBox_thermal_offset9.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 2, 0, "000")
        elif self.comboBox_thermal_offset9.currentIndex() == 1:
            update_database_with_temp_customer_input("C9", 2, 0, "001")
        elif self.comboBox_thermal_offset9.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 2, 0, "010")
        elif self.comboBox_thermal_offset9.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 2, 0, "011")
        elif self.comboBox_thermal_offset9.currentIndex() == 4:
            update_database_with_temp_customer_input("C9", 2, 0, "100")
        elif self.comboBox_thermal_offset9.currentIndex() == 5:
            update_database_with_temp_customer_input("C9", 2, 0, "101")
        elif self.comboBox_thermal_offset9.currentIndex() == 6:
            update_database_with_temp_customer_input("C9", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("C9", 2, 0, "111")

    def freq_switch_changed_A(self):
        if self.comboBox_freq_switch_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("C50", 79, 78, "11")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("C50", 79, 78, "10")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("C50", 79, 78, "01")
        else:
            update_database_with_temp_customer_input("C50", 79, 78, "00")

    def freq_switch_changed_B(self):
        if self.comboBox_freq_switch_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("C51", 79, 78, "11")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("C51", 79, 78, "10")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("C51", 79, 78, "01")
        else:
            update_database_with_temp_customer_input("C51", 79, 78, "00")

    def Save(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("Phase thermal balance settings saved.", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def Discard(self):
        print_log("Phase thermal balance settings discarded.", "INFO")
        self.main = Phase_thermal_balance()
        self.main.show()
        self.close()


class frame_svid(QMainWindow, Ui_SVID_Configuration):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame, homeWin_obj
        super(frame_svid, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["E6", "E9", "E4", "DF", "F2", "010", "E80", "E10", "FC0", "200", "270", "011", "E81", "E11", "FC1", "201", "271"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        resolution_calculation(0)
        homeWin_obj.label_prot_address_A.setText(str(hex(int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
        homeWin_obj.label_prot_address_B.setText(str(hex(1 + int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
        self.lineEdit_SVID_address.setText(str(hex(int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")))
        self.lineEdit_Pin_max.setText(str(int(initialize_feature_variable('E9', 33, 26), 2) * 2))
        self.lineEdit_icc_max_railA.setText(str(int(initialize_feature_variable('E80', 55, 48), 2)))
        self.lineEdit_icc_max_railB.setText(str(int(initialize_feature_variable('E81', 55, 48), 2)))
        initial_vid_raila = float((int(initialize_feature_variable('E80', 31, 24), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('E80', 31, 24), 2) != 0 else int(0)
        self.lineEdit_vid_max_railA.setText(str(initial_vid_raila))
        self.raila_vidmax = initial_vid_raila
        initial_vid_railb = float((int(initialize_feature_variable('E81', 31, 24), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('E81', 31, 24), 2) != 0 else int(0)
        self.lineEdit_vid_max_railB.setText(str(initial_vid_railb))
        self.railb_vidmax = initial_vid_railb
        self.lineEdit_temp_max.setText(str(int(initialize_feature_variable('E80', 47, 40), 2)))
        self.lineEdit_imon_aux.setText(str(int(initialize_feature_variable('E9', 61, 54), 2)))
        initial_pvid_0 = float((int(initialize_feature_variable('FC0', 7, 0), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 7, 0), 2) != 0 else int(0)
        self.lineEdit_PVID_A_0.setText(str(initial_pvid_0))
        self.lineEdit_PVID_A_1.setText(str(float((int(initialize_feature_variable('FC0', 15, 8), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 15, 8), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_2.setText(str(float((int(initialize_feature_variable('FC0', 23, 16), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 23, 16), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_3.setText(str(float((int(initialize_feature_variable('FC0', 31, 24), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 31, 24), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_4.setText(str(float((int(initialize_feature_variable('FC0', 39, 32), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 39, 32), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_5.setText(str(float((int(initialize_feature_variable('FC0', 47, 40), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 47, 40), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_6.setText(str(float((int(initialize_feature_variable('FC0', 55, 48), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 55, 48), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_7.setText(str(float((int(initialize_feature_variable('FC0', 63, 56), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 63, 56), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_0.setText(str(float((int(initialize_feature_variable('FC1', 7, 0), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 7, 0), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_1.setText(str(float((int(initialize_feature_variable('FC1', 15, 8), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 15, 8), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_2.setText(str(float((int(initialize_feature_variable('FC1', 23, 16), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 23, 16), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_3.setText(str(float((int(initialize_feature_variable('FC1', 31, 24), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 31, 24), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_4.setText(str(float((int(initialize_feature_variable('FC1', 39, 32), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 39, 32), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_5.setText(str(float((int(initialize_feature_variable('FC1', 47, 40), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 47, 40), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_6.setText(str(float((int(initialize_feature_variable('FC1', 55, 48), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 55, 48), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_7.setText(str(float((int(initialize_feature_variable('FC1', 63, 56), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 63, 56), 2) != 0 else int(0)))

        if str(initialize_feature_variable('DF', 10, 10)) == '1':
            self.radioButton_pline.setChecked(True)
            self.radioButton_sline.setChecked(False)
        else:
            self.radioButton_pline.setChecked(False)
            self.radioButton_sline.setChecked(True)

        if int(initialize_feature_variable('E10', 25, 25), 2) == 1:
            self.radioButton_internal_registerA.setChecked(True)
            self.radioButton_vid_sel_pinA.setChecked(False)
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(False)
        else:
            self.radioButton_internal_registerA.setChecked(False)
            self.radioButton_vid_sel_pinA.setChecked(True)
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)

        if int(initialize_feature_variable('E11', 25, 25), 2) == 1:
            self.radioButton_internal_registerB.setChecked(True)
            self.radioButton_vid_sel_pinB.setChecked(False)
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(False)
        else:
            self.radioButton_internal_registerB.setChecked(False)
            self.radioButton_vid_sel_pinB.setChecked(True)
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)

        def get_key_svid(val):
            for key, value in svid_protocol.items():
                if val == value:
                    return key

        def get_key_slew(val):
            for key, value in slew_rate.items():
                if val == value:
                    return key

        self.comboBox_svid_protocol_RailA.setCurrentIndex(int(get_key_svid(int(initialize_feature_variable('E80', 89, 86), 2))))
        self.comboBox_svid_protocol_RailB.setCurrentIndex(int(get_key_svid(int(initialize_feature_variable('E81', 89, 86), 2))))
        self.comboBox_vid_resolution_RailA.setCurrentIndex(not(int(initialize_feature_variable('E10', 25, 25), 2)))
        self.comboBox_vid_resolution_RailB.setCurrentIndex(not(int(initialize_feature_variable('E11', 25, 25), 2)))
        self.comboBox_vid_sel_pin_resolution_railA.setCurrentIndex((int(initialize_feature_variable('E10', 24, 24), 2)))
        self.comboBox_vid_sel_pin_resolution_railB.setCurrentIndex((int(initialize_feature_variable('E11', 24, 24), 2)))
        if int(initialize_feature_variable('200', 4, 0), 2) == 0x1E:
            self.comboBox_vid_resolution_RailA.setCurrentIndex(1)
        else:
            if int(initialize_feature_variable('E4', 3, 3), 2) == 1:
                self.comboBox_vid_resolution_RailA.setCurrentIndex(2)
            else:
                self.comboBox_vid_resolution_RailA.setCurrentIndex(0)

        if int(initialize_feature_variable('201', 4, 0), 2) == 0x1E:
            self.comboBox_vid_resolution_RailB.setCurrentIndex(1)
        else:
            if int(initialize_feature_variable('E4', 3, 3), 2) == 1:
                self.comboBox_vid_resolution_RailB.setCurrentIndex(2)
            else:
                self.comboBox_vid_resolution_RailB.setCurrentIndex(0)

        if int(initialize_feature_variable('010', 5, 4), 2) == 0:
            self.checkBox_PMBus_override_railA.setChecked(True)
            self.comboBox_vid_resolution_RailA.setDisabled(False)
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)
            self.radioButton_vid_sel_pinA.setDisabled(True)
            self.radioButton_internal_registerA.setDisabled(True)
        else:
            self.checkBox_PMBus_override_railA.setChecked(False)
            self.comboBox_vid_resolution_RailA.setDisabled(True)
            self.radioButton_vid_sel_pinA.setDisabled(False)
            self.radioButton_internal_registerA.setDisabled(False)


        if int(initialize_feature_variable('011', 5, 4), 2) == 0:
            self.checkBox_PMBus_override_railA_2.setChecked(True)
            self.comboBox_vid_resolution_RailB.setDisabled(False)
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(True)
            self.radioButton_internal_registerB.setDisabled(True)
        else:
            self.checkBox_PMBus_override_railA_2.setChecked(False)
            self.comboBox_vid_resolution_RailB.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(False)
            self.radioButton_internal_registerB.setDisabled(False)

        if initialize_feature_variable("270", 15, 0) == "1001101111010111":         # 0x9BD7
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(0)
        elif initialize_feature_variable("270", 15, 0) == "1011001000000000":       # 0xB200
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(1)
        elif initialize_feature_variable("270", 15, 0) == "1011101000000000":       # 0xBA00
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(2)
        elif initialize_feature_variable("270", 15, 0) == "1100001010000000":       # 0xC280
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(3)
        elif initialize_feature_variable("270", 15, 0) == "1100101010000000":       # 0xCA80
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(4)
        elif initialize_feature_variable("270", 15, 0) == "1100101111000000":       # 0xCBC0
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(5)
        elif initialize_feature_variable("270", 15, 0) == "1101001010000000":       # 0xD280
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(6)
        elif initialize_feature_variable("270", 15, 0) == "1101001100100000":       # 0xD320
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(7)
        elif initialize_feature_variable("270", 15, 0) == "1101001111000000":       # 0xD3C0
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(8)
        elif initialize_feature_variable("270", 15, 0) == "1101101000110000":       # 0xDA30
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(9)
        elif initialize_feature_variable("270", 15, 0) == "1101101010000000":       # 0xDA80
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(10)
        elif initialize_feature_variable("270", 15, 0) == "1101101011010000":       # 0xDAD0
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(11)
        elif initialize_feature_variable("270", 15, 0) == "1101101100100000":       # 0xDB20
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(12)
        elif initialize_feature_variable("270", 15, 0) == "1101101101110000":       # 0xDB70
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(13)
        elif initialize_feature_variable("270", 15, 0) == "1101101111000000":       # 0xDBC0
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(14)
        elif initialize_feature_variable("270", 15, 0) == "1110001000001000":       # 0xE208
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(15)
        elif initialize_feature_variable("270", 15, 0) == "1110001000110000":       # 0xE230
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(16)
        elif initialize_feature_variable("270", 15, 0) == "1110001001011000":       # 0xE258
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(17)
        elif initialize_feature_variable("270", 15, 0) == "1110001010000000":       # 0xE280
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(18)
        elif initialize_feature_variable("270", 15, 0) == "1110001100000000":       # 0xE300
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(19)
        elif initialize_feature_variable("270", 15, 0) == "1110001111000000":       # 0xE3C0
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(20)
        elif initialize_feature_variable("270", 15, 0) == "1110101010000000":       # 0xEA80
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(21)
        elif initialize_feature_variable("270", 15, 0) == "1110101100000000":       # 0xEB00
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(22)
        else:                                                                       # 0xEBE8
            self.comboBox_slewrate_fast_RailA.setCurrentIndex(23)

        if initialize_feature_variable("271", 15, 0) == "1001101111010111":         # 0x9BD7
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(0)
        elif initialize_feature_variable("271", 15, 0) == "1011001000000000":       # 0xB200
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(1)
        elif initialize_feature_variable("271", 15, 0) == "1011101000000000":       # 0xBA00
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(2)
        elif initialize_feature_variable("271", 15, 0) == "1100001010000000":       # 0xC280
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(3)
        elif initialize_feature_variable("271", 15, 0) == "1100101010000000":       # 0xCA80
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(4)
        elif initialize_feature_variable("271", 15, 0) == "1100101111000000":       # 0xCBC0
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(5)
        elif initialize_feature_variable("271", 15, 0) == "1101001010000000":       # 0xD280
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(6)
        elif initialize_feature_variable("271", 15, 0) == "1101001100100000":       # 0xD320
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(7)
        elif initialize_feature_variable("271", 15, 0) == "1101001111000000":       # 0xD3C0
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(8)
        elif initialize_feature_variable("271", 15, 0) == "1101101000110000":       # 0xDA30
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(9)
        elif initialize_feature_variable("271", 15, 0) == "1101101010000000":       # 0xDA80
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(10)
        elif initialize_feature_variable("271", 15, 0) == "1101101011010000":       # 0xDAD0
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(11)
        elif initialize_feature_variable("271", 15, 0) == "1101101100100000":       # 0xDB20
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(12)
        elif initialize_feature_variable("271", 15, 0) == "1101101101110000":       # 0xDB70
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(13)
        elif initialize_feature_variable("271", 15, 0) == "1101101111000000":       # 0xDBC0
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(14)
        elif initialize_feature_variable("271", 15, 0) == "1110001000001000":       # 0xE208
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(15)
        elif initialize_feature_variable("271", 15, 0) == "1110001000110000":       # 0xE230
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(16)
        elif initialize_feature_variable("271", 15, 0) == "1110001001011000":       # 0xE258
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(17)
        elif initialize_feature_variable("271", 15, 0) == "1110001010000000":       # 0xE280
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(18)
        elif initialize_feature_variable("271", 15, 0) == "1110001100000000":       # 0xE300
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(19)
        elif initialize_feature_variable("271", 15, 0) == "1110001111000000":       # 0xE3C0
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(20)
        elif initialize_feature_variable("271", 15, 0) == "1110101010000000":       # 0xEA80
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(21)
        elif initialize_feature_variable("271", 15, 0) == "1110101100000000":       # 0xEB00
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(22)
        else:                                                                       # 0xEBE8
            self.comboBox_slewrate_fast_RailB.setCurrentIndex(23)


        # self.comboBox_slewrate_fast_RailA.setCurrentText(str(get_key_slew(int(initialize_feature_variable('270', 15, 0),2)))+"mV/μs")
        # self.comboBox_slewrate_fast_RailB.setCurrentText(str(get_key_slew(int(initialize_feature_variable('271', 15, 0),2)))+"mV/μs")
        slow_slew_sel = {0: 8, 1: 4, 2: 2, 3: 0}

        def get_key_slow_slew(val):
            for key, value in slow_slew_sel.items():
                if val >= value:
                    return key

        # print(int(initialize_feature_variable('E80', 4, 0),2))
        self.comboBox_slew_rate_slow_railA.setCurrentIndex(int(get_key_slow_slew(int(initialize_feature_variable('E80', 4, 0), 2))))
        self.comboBox_slew_rate_slow_railB.setCurrentIndex(int(get_key_slow_slew(int(initialize_feature_variable('E81', 4, 0), 2))))

        # Other initialization
        if int(initialize_feature_variable('F2', 197, 197)) == 1:
            self.pushButton_Enable_imon_aux_PS3.setChecked(True)
            self.imon_aux_update()
        else:
            self.pushButton_Enable_imon_aux_PS3.setChecked(False)
            self.imon_aux_update()

        if int(initialize_feature_variable('DF', 11, 11)) == 1:
            self.pushButton_Enable_APD_railA.setChecked(True)
            self.pvid_enable()
        else:
            self.pushButton_Enable_APD_railA.setChecked(False)
            self.pvid_enable()

        self.label_display_RailA_2.setText(RailA_name)
        self.label_display_RailB_2.setText(RailB_name)
        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        # extra added for raila, pmbus override handling
        if int(initialize_feature_variable('010', 5, 4), 2) == 0:
            self.checkBox_PMBus_override_railA.setChecked(True)
            self.comboBox_vid_resolution_RailA.setDisabled(False)
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)
            self.radioButton_vid_sel_pinA.setDisabled(True)
            self.radioButton_internal_registerA.setDisabled(True)
        else:
            self.checkBox_PMBus_override_railA.setChecked(False)
            self.comboBox_vid_resolution_RailA.setDisabled(True)
            self.radioButton_vid_sel_pinA.setDisabled(False)
            self.radioButton_internal_registerA.setDisabled(False)

        # extra added for railb, pmbus override handling
        if int(initialize_feature_variable('011', 5, 4), 2) == 0:
            self.checkBox_PMBus_override_railA_2.setChecked(True)
            self.comboBox_vid_resolution_RailB.setDisabled(False)
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(True)
            self.radioButton_internal_registerB.setDisabled(True)
        else:
            self.checkBox_PMBus_override_railA_2.setChecked(False)
            self.comboBox_vid_resolution_RailB.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(False)
            self.radioButton_internal_registerB.setDisabled(False)

        ## code starts
        self.lineEdit_SVID_address.textEdited.connect(self.svid_address)
        self.lineEdit_Pin_max.textEdited.connect(self.svid_pin_max)
        self.lineEdit_icc_max_railA.textEdited.connect(self.svid_iccmax_railA)
        self.lineEdit_icc_max_railB.textEdited.connect(self.svid_iccmax_railB)
        self.lineEdit_vid_max_railA.textEdited.connect(self.svid_vidmax_railA)
        self.lineEdit_vid_max_railB.textEdited.connect(self.svid_vidmax_railB)
        self.lineEdit_temp_max.textEdited.connect(self.svid_tempmax)
        self.lineEdit_imon_aux.textEdited.connect(self.svid_imonaux)
        self.lineEdit_PVID_A_0.textEdited.connect(self.svid_pvid_a0)
        self.lineEdit_PVID_A_1.textEdited.connect(self.svid_pvid_a1)
        self.lineEdit_PVID_A_2.textEdited.connect(self.svid_pvid_a2)
        self.lineEdit_PVID_A_3.textEdited.connect(self.svid_pvid_a3)
        self.lineEdit_PVID_A_4.textEdited.connect(self.svid_pvid_a4)
        self.lineEdit_PVID_A_5.textEdited.connect(self.svid_pvid_a5)
        self.lineEdit_PVID_A_6.textEdited.connect(self.svid_pvid_a6)
        self.lineEdit_PVID_A_7.textEdited.connect(self.svid_pvid_a7)
        self.lineEdit_PVID_B_0.textEdited.connect(self.svid_pvid_railB_a0)
        self.lineEdit_PVID_B_1.textEdited.connect(self.svid_pvid_railB_a1)
        self.lineEdit_PVID_B_2.textEdited.connect(self.svid_pvid_railB_a2)
        self.lineEdit_PVID_B_3.textEdited.connect(self.svid_pvid_railB_a3)
        self.lineEdit_PVID_B_4.textEdited.connect(self.svid_pvid_railB_a4)
        self.lineEdit_PVID_B_5.textEdited.connect(self.svid_pvid_railB_a5)
        self.lineEdit_PVID_B_6.textEdited.connect(self.svid_pvid_railB_a6)
        self.lineEdit_PVID_B_7.textEdited.connect(self.svid_pvid_railB_a7)

        self.radioButton_pline.toggled.connect(self.boot_onClicked)
        self.radioButton_sline.toggled.connect(self.boot_onClicked)
        self.radioButton_internal_registerA.toggled.connect(self.railA_resolution_sel_onClicked)
        self.radioButton_vid_sel_pinA.toggled.connect(self.railA_resolution_sel_onClicked)
        self.radioButton_internal_registerB.toggled.connect(self.railB_resolution_sel_onClicked)
        self.radioButton_vid_sel_pinB.toggled.connect(self.railB_resolution_sel_onClicked)

        self.comboBox_svid_protocol_RailA.activated.connect(self.railA_svid_protocol)
        self.comboBox_svid_protocol_RailB.activated.connect(self.railB_svid_protocol)
        self.comboBox_vid_resolution_RailA.activated.connect(self.railA_vid_resolution)
        self.comboBox_vid_resolution_RailB.activated.connect(self.railB_vid_resolution)
        self.comboBox_vid_sel_pin_resolution_railA.activated.connect(self.railA_vid_sel)
        self.comboBox_vid_sel_pin_resolution_railB.activated.connect(self.railB_vid_sel)
        self.comboBox_all_call_address_RailA.activated.connect(self.all_call_railA)
        self.comboBox_all_call_address_RailB.activated.connect(self.all_call_railB)
        self.checkBox_PMBus_override_railA.toggled.connect(self.pmbus_override_railA)
        self.checkBox_PMBus_override_railA_2.toggled.connect(self.pmbus_override_railB)

        self.comboBox_slewrate_fast_RailA.activated.connect(self.railA_slew_fast)
        self.comboBox_slewrate_fast_RailB.activated.connect(self.railB_slew_fast)
        self.comboBox_slew_rate_slow_railA.activated.connect(self.railA_slew_slow)
        self.comboBox_slew_rate_slow_railB.activated.connect(self.railB_slew_slow)
        self.pushButton_Enable_imon_aux_PS3.toggled.connect(self.imon_aux_update)
        self.pushButton_Enable_APD_railA.toggled.connect(self.pvid_enable)

        self.pushButton_Save.clicked.connect(self.svid_save_register)
        self.pushButton_Discard.clicked.connect(self.svid_discard_changes)
        self.pushButton_Enable_APD_railA.clicked.connect(self.pvid_enable_reg)

    def imon_aux_update(self):
        if self.pushButton_Enable_imon_aux_PS3.isChecked():
            self.pushButton_Enable_imon_aux_PS3.setStyleSheet("QPushButton{\n"
                                                              "border-image: url(GUI_IMAGE/but1.png);\n"
                                                              "}\n"
                                                              "\n"
                                                              "\n"
                                                              "QPushButton::hover {\n"
                                                              "border-image: url(GUI_IMAGE/but1_hover"
                                                              ".png);\n "
                                                              "    }\n"
                                                              "")
            update_database_with_temp_customer_input("F2", 197, 197, bin(1).split("0b")[1].zfill(1))
            # if it is unchecked
        else:
            # setting image to button0
            self.pushButton_Enable_imon_aux_PS3.setStyleSheet("QPushButton{\n"
                                                              "border-image: url(GUI_IMAGE/but0.png);\n"
                                                              "}\n"
                                                              "\n"
                                                              "\n"
                                                              "QPushButton::hover {\n"
                                                              "border-image: url(GUI_IMAGE/but0_hover"
                                                              ".png);\n "
                                                              "    }\n"
                                                              "")
            update_database_with_temp_customer_input("F2", 197, 197, bin(0).split("0b")[1].zfill(1))

    def pvid_enable(self):
        if self.pushButton_Enable_APD_railA.isChecked():
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but1.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but1_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")
            update_database_with_temp_customer_input("DF", 11, 11, bin(1).split("0b")[1].zfill(1))
            self.lineEdit_PVID_A_0.setDisabled(False)
            self.lineEdit_PVID_A_1.setDisabled(False)
            self.lineEdit_PVID_A_2.setDisabled(False)
            self.lineEdit_PVID_A_3.setDisabled(False)
            self.lineEdit_PVID_A_4.setDisabled(False)
            self.lineEdit_PVID_A_5.setDisabled(False)
            self.lineEdit_PVID_A_6.setDisabled(False)
            self.lineEdit_PVID_A_7.setDisabled(False)
            self.lineEdit_PVID_B_0.setDisabled(False)
            self.lineEdit_PVID_B_1.setDisabled(False)
            self.lineEdit_PVID_B_2.setDisabled(False)
            self.lineEdit_PVID_B_3.setDisabled(False)
            self.lineEdit_PVID_B_4.setDisabled(False)
            self.lineEdit_PVID_B_5.setDisabled(False)
            self.lineEdit_PVID_B_6.setDisabled(False)
            self.lineEdit_PVID_B_7.setDisabled(False)
            self.lineEdit_SVID_address.setDisabled(True)
            self.radioButton_pline.setDisabled(True)
            self.radioButton_sline.setDisabled(True)
            self.comboBox_all_call_address_RailA.setDisabled(True)
            self.comboBox_svid_protocol_RailA.setDisabled(True)
            self.comboBox_svid_protocol_RailB.setDisabled(True)
            self.comboBox_all_call_address_RailB.setDisabled(True)
            self.comboBox_slew_rate_slow_railA.setDisabled(True)
            self.comboBox_slew_rate_slow_railB.setDisabled(True)
            self.radioButton_vid_sel_pinA.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(True)
            self.radioButton_internal_registerA.setDisabled(True)
            self.radioButton_internal_registerB.setDisabled(True)
            self.lineEdit_imon_aux.setDisabled(True)
            self.pushButton_Enable_imon_aux_PS3.setDisabled(True)
            self.lineEdit_PVID_A_0.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_1.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_2.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_3.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_4.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_5.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_6.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_7.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_0.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_1.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_2.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_3.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_4.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_5.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_6.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_7.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_0.setText(str(float((int(initialize_feature_variable('FC0', 7, 0), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 7, 0),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_1.setText(str(float((int(initialize_feature_variable('FC0', 15, 8), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 15, 8),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_2.setText(str(float((int(initialize_feature_variable('FC0', 23, 16), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 23, 16),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_3.setText(str(float((int(initialize_feature_variable('FC0', 31, 24), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 31, 24),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_4.setText(str(float((int(initialize_feature_variable('FC0', 39, 32), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 39, 32),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_5.setText(str(float((int(initialize_feature_variable('FC0', 47, 40), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 47, 40),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_6.setText(str(float((int(initialize_feature_variable('FC0', 55, 48), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 55, 48),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_7.setText(str(float((int(initialize_feature_variable('FC0', 63, 56), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 63, 56),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_0.setText(str(float((int(initialize_feature_variable('FC1', 7, 0), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 7, 0),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_1.setText(str(float((int(initialize_feature_variable('FC1', 15, 8), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 15, 8),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_2.setText(str(float((int(initialize_feature_variable('FC1', 23, 16), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 23, 16),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_3.setText(str(float((int(initialize_feature_variable('FC1', 31, 24), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 31, 24),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_4.setText(str(float((int(initialize_feature_variable('FC1', 39, 32), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 39, 32),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_5.setText(str(float((int(initialize_feature_variable('FC1', 47, 40), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 47, 40),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_6.setText(str(float((int(initialize_feature_variable('FC1', 55, 48), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 55, 48),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_7.setText(str(float((int(initialize_feature_variable('FC1', 63, 56), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 63, 56),
                                                                        2) != 0 else int(0)))
            # if it is unchecked
        else:
            # setting image to button0
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but0.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but0_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")
            update_database_with_temp_customer_input("DF", 11, 11, bin(0).split("0b")[1].zfill(1))
            self.lineEdit_PVID_A_0.setDisabled(True)
            self.lineEdit_PVID_A_1.setDisabled(True)
            self.lineEdit_PVID_A_2.setDisabled(True)
            self.lineEdit_PVID_A_3.setDisabled(True)
            self.lineEdit_PVID_A_4.setDisabled(True)
            self.lineEdit_PVID_A_5.setDisabled(True)
            self.lineEdit_PVID_A_6.setDisabled(True)
            self.lineEdit_PVID_A_7.setDisabled(True)
            self.lineEdit_PVID_B_0.setDisabled(True)
            self.lineEdit_PVID_B_1.setDisabled(True)
            self.lineEdit_PVID_B_2.setDisabled(True)
            self.lineEdit_PVID_B_3.setDisabled(True)
            self.lineEdit_PVID_B_4.setDisabled(True)
            self.lineEdit_PVID_B_5.setDisabled(True)
            self.lineEdit_PVID_B_6.setDisabled(True)
            self.lineEdit_PVID_B_7.setDisabled(True)
            self.lineEdit_SVID_address.setDisabled(False)
            self.radioButton_pline.setDisabled(False)
            self.radioButton_sline.setDisabled(False)
            self.comboBox_all_call_address_RailA.setDisabled(False)
            self.comboBox_svid_protocol_RailA.setDisabled(False)
            self.comboBox_svid_protocol_RailB.setDisabled(False)
            self.comboBox_all_call_address_RailB.setDisabled(False)
            self.comboBox_slew_rate_slow_railA.setDisabled(False)
            self.comboBox_slew_rate_slow_railB.setDisabled(False)
            self.radioButton_vid_sel_pinA.setDisabled(False)
            self.radioButton_vid_sel_pinB.setDisabled(False)
            self.radioButton_internal_registerA.setDisabled(False)
            self.radioButton_internal_registerB.setDisabled(False)
            self.lineEdit_imon_aux.setDisabled(False)
            self.pushButton_Enable_imon_aux_PS3.setDisabled(False)
            self.lineEdit_PVID_A_0.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_0.setText("--")
            self.lineEdit_PVID_A_1.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_1.setText("--")
            self.lineEdit_PVID_A_2.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_2.setText("--")
            self.lineEdit_PVID_A_3.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_3.setText("--")
            self.lineEdit_PVID_A_4.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_4.setText("--")
            self.lineEdit_PVID_A_5.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_5.setText("--")
            self.lineEdit_PVID_A_6.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_6.setText("--")
            self.lineEdit_PVID_A_7.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_7.setText("--")
            self.lineEdit_PVID_B_0.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_0.setText("--")
            self.lineEdit_PVID_B_1.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_1.setText("--")
            self.lineEdit_PVID_B_2.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_2.setText("--")
            self.lineEdit_PVID_B_3.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_3.setText("--")
            self.lineEdit_PVID_B_4.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_4.setText("--")
            self.lineEdit_PVID_B_5.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_5.setText("--")
            self.lineEdit_PVID_B_6.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_6.setText("--")
            self.lineEdit_PVID_B_7.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_7.setText("--")

    def pvid_enable_reg(self):
        if self.pushButton_Enable_APD_railA.isChecked():
            update_database_with_temp_customer_input("DF", 11, 11, bin(1))
        else:
            update_database_with_temp_customer_input("DF", 11, 11, bin(0))

    def svid_address(self):
        if self.lineEdit_SVID_address.text() == "":
            return

        if (len(self.lineEdit_SVID_address.text()) > 1):
            print_log("Only 1 digit is allowed", "ERROR")
            nsleep(5)
            self.lineEdit_SVID_address.setText(str(hex(int(write_feature_variable('E6', 7, 4), 2)).replace("0x", "")))
            return

        try:
            svid_address = int(self.lineEdit_SVID_address.text(), 16)
        except ValueError:
            print_log("Text box value should be hex number", "ERROR")
            nsleep(5)
            self.lineEdit_SVID_address.setText(str(hex(int(write_feature_variable('E6', 7, 4), 2)).replace("0x", "")))
            return

        update_database_with_temp_customer_input("E6", 7, 4, bin(svid_address).split("0b")[1].zfill(4))
        update_database_with_temp_customer_input("E6", 3, 0, bin(svid_address + 1).split("0b")[1].zfill(4))

    def svid_pin_max(self):
        if self.lineEdit_Pin_max.text() == "":
            return
        try:
            pin_max = int(self.lineEdit_Pin_max.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_Pin_max.setText(str(int(write_feature_variable('E9', 33, 26), 2) * 2))
            return

        if (pin_max > 510):
            print_log("Text box value should be less than 511", "ERROR")
            nsleep(5)
            self.lineEdit_Pin_max.setText(str(int(write_feature_variable('E9', 33, 26), 2) * 2))
            return
        else:
            update_database_with_temp_customer_input("E9", 33, 26, bin(int(pin_max / 2)).split("0b")[1].zfill(8))

    def svid_iccmax_railA(self):
        if self.lineEdit_icc_max_railA.text() == "":
            return
        try:
            icc_max = int(self.lineEdit_icc_max_railA.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railA.setText(str(int(write_feature_variable('E80', 55, 48), 2)))
            return

        if (icc_max > 255):
            print_log("Text box value should be less than 256", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railA.setText(str(int(write_feature_variable('E80', 55, 48), 2)))
            return
        else:
            update_database_with_temp_customer_input("E80", 55, 48, bin(icc_max).split("0b")[1].zfill(8))

    def svid_iccmax_railB(self):
        if self.lineEdit_icc_max_railB.text() == "":
            return
        try:
            icc_max = int(self.lineEdit_icc_max_railB.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railB.setText(str(int(write_feature_variable('E81', 55, 48), 2)))
            return

        if (icc_max > 255):
            print_log("Text box value should be less than 256", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railB.setText(str(int(write_feature_variable('E81', 55, 48), 2)))
            return
        else:
            update_database_with_temp_customer_input("E81", 55, 48, bin(icc_max).split("0b")[1].zfill(8))

    def svid_vidmax_railA(self):
        resolution_calculation(1)
        if self.lineEdit_vid_max_railA.text() == "":
            return
        try:
            float(self.lineEdit_vid_max_railA.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            initial_vid_raila = float((int(write_feature_variable('E80', 31, 24), 2) * float(resolutionA) + offset[
                str(resolutionA)]) / 1000) if int(write_feature_variable('E80', 31, 24), 2) != 0 else int(0)
            self.lineEdit_vid_max_railA.setText(str(initial_vid_raila))
            return
        if (float(self.lineEdit_vid_max_railA.text()) > vid_max_value[resolutionA] or float(self.lineEdit_vid_max_railA.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_vid_max_railA.text()) != 0:
            # string = "Please enter input in range" + vid_min_value[resolutionA] + "to" + vid_max_value[resolutionA]
            print_log("Input vid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_vid_max_railA.text()) > vid_max_value[resolutionA]:
                self.lineEdit_vid_max_railA.setText(str(vid_max_value[resolutionA]))
            else:
                self.lineEdit_vid_max_railA.setText(str(vid_min_value[resolutionA]))
        self.raila_vidmax = float(self.lineEdit_vid_max_railA.text())
        vid_max = int((float(self.lineEdit_vid_max_railA.text()) * 1000 - float(offset[resolutionA])) / float(resolutionA))
        update_database_with_temp_customer_input("E80", 31, 24, bin(vid_max).split("0b")[1].zfill(8))

    def svid_vidmax_railB(self):
        resolution_calculation(1)
        if self.lineEdit_vid_max_railB.text() == "":
            return
        try:
            float(self.lineEdit_vid_max_railB.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            initial_vid_railb = float((int(write_feature_variable('E81', 31, 24), 2) * float(resolutionB) + offset[
                str(resolutionB)]) / 1000) if int(write_feature_variable('E81', 31, 24), 2) != 0 else int(0)
            self.lineEdit_vid_max_railB.setText(str(initial_vid_railb))
            return

        if (float(self.lineEdit_vid_max_railB.text()) > vid_max_value[resolutionB] or float(self.lineEdit_vid_max_railB.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_vid_max_railB.text()) != 0:
            # string = "Please enter input in range" + vid_min_value[resolutionA] + "to" + vid_max_value[resolutionA]
            print_log("Input vid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            if float(self.lineEdit_vid_max_railB.text()) > vid_max_value[resolutionB]:
                self.lineEdit_vid_max_railB.setText(str(vid_max_value[resolutionB]))
            else:
                self.lineEdit_vid_max_railB.setText(str(vid_min_value[resolutionB]))

        self.railb_vidmax = float(self.lineEdit_vid_max_railB.text())
        vid_max = int((float(self.lineEdit_vid_max_railB.text()) * 1000 - float(offset[resolutionB])) / float(resolutionB))
        update_database_with_temp_customer_input("E81", 31, 24, bin(vid_max).split("0b")[1].zfill(8))

    def svid_tempmax(self):
        if self.lineEdit_temp_max.text() == "":
            return
        try:
            temp_max = int(self.lineEdit_temp_max.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_temp_max.setText(str(int(write_feature_variable('E80', 47, 40), 2)))
            return

        if (temp_max > 255):
            print_log("Text box value should be less than 255", "ERROR")
            nsleep(5)
            self.lineEdit_temp_max.setText(str(int(write_feature_variable('E80', 47, 40), 2)))
            return
        else:
            update_database_with_temp_customer_input("E80", 47, 40, bin(temp_max).split("0b")[1].zfill(8))

    def svid_imonaux(self):
        if self.lineEdit_imon_aux.text() == "":
            return
        try:
            imonaux = int(self.lineEdit_imon_aux.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_imon_aux.setText(str(int(write_feature_variable('E9', 61, 54), 2)))
            return

        if (imonaux > 255):
            print_log("Text box value should be less than 255", "ERROR")
            nsleep(5)
            self.lineEdit_imon_aux.setText(str(int(write_feature_variable('E9', 61, 54), 2)))
            return
        else:
            update_database_with_temp_customer_input("E9", 61, 54, bin(imonaux).split("0b")[1].zfill(8))

    def svid_pvid_a0(self):
        if self.lineEdit_PVID_A_0.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_0.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            initial_pvid_0 = float((int(write_feature_variable('FC0', 7, 0), 2) * float(resolutionA) + offset[
                str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 7, 0), 2) != 0 else int(0)
            self.lineEdit_PVID_A_0.setText(str(initial_pvid_0))
            return

        if (float(self.lineEdit_PVID_A_0.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_0.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_0.text()) != 0:
            # string = "Please enter input in range" + vid_min_value[resolutionA] + "to" + vid_max_value[resolutionA]
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_0.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_0.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_0.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_0.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_0.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 7, 0, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a1(self):
        if self.lineEdit_PVID_A_1.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_1.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_1.setText(str(float((int(write_feature_variable('FC0', 15, 8), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 15, 8),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_1.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_1.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_1.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_1.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_1.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_1.setText(str(vid_min_value[resolutionA]))

        pvid_a1 = int(
            (float(self.lineEdit_PVID_A_1.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_1.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 15, 8, bin(pvid_a1).split("0b")[1].zfill(8))

    def svid_pvid_a2(self):
        if self.lineEdit_PVID_A_2.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_2.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point  number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_2.setText(str(float((int(write_feature_variable('FC0', 23, 16), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 23, 16),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_0.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_2.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_2.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_2.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_2.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_2.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_2.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_2.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 23, 16, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a3(self):
        if self.lineEdit_PVID_A_3.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_3.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_3.setText(str(float((int(write_feature_variable('FC0', 31, 24), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 31, 24),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_3.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_3.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_3.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_3.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_3.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_3.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_3.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_3.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 31, 24, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a4(self):
        if self.lineEdit_PVID_A_4.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_4.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_4.setText(str(float((int(write_feature_variable('FC0', 39, 32), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 39, 32),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_4.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_4.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_4.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_4.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_4.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_4.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_4.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_4.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 39, 32, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a5(self):
        if self.lineEdit_PVID_A_5.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_5.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_5.setText(str(float((int(write_feature_variable('FC0', 47, 40), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 47, 40),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_5.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_5.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_5.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_5.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_5.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_5.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_5.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_5.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 47, 40, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a6(self):
        if self.lineEdit_PVID_A_6.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_6.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_6.setText(str(float((int(write_feature_variable('FC0', 55, 48), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 55, 48),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_6.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_6.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_6.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_6.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_6.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_6.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_6.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_6.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 55, 48, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a7(self):
        if self.lineEdit_PVID_A_7.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_7.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_7.setText(str(float((int(write_feature_variable('FC0', 63, 56), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 63, 56),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_7.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_7.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_7.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_7.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_7.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_7.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_7.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_7.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 63, 56, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a0(self):
        if self.lineEdit_PVID_B_0.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_0.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_0.setText(str(float((int(write_feature_variable('FC1', 7, 0), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 7, 0),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_0.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_0.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_0.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_0.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_0.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_0.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_0.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_0.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 7, 0, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a1(self):
        if self.lineEdit_PVID_B_1.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_1.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_1.setText(str(float((int(write_feature_variable('FC1', 15, 8), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 15, 8),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_1.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_1.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_1.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_1.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_1.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_1.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_1.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_1.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 15, 8, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a2(self):
        if self.lineEdit_PVID_B_2.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_2.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_2.setText(str(float((int(write_feature_variable('FC1', 23, 16), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 23, 16),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_2.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_2.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_2.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_2.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_2.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_2.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_2.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_2.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 23, 16, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a3(self):
        if self.lineEdit_PVID_B_3.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_3.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_3.setText(str(float((int(write_feature_variable('FC1', 31, 24), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 23, 16),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_3.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_3.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_3.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_3.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_3.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_3.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_3.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_3.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 31, 24, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a4(self):
        if self.lineEdit_PVID_B_4.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_4.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            self.lineEdit_PVID_B_4.setText(str(float((int(write_feature_variable('FC1', 39, 32), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 39, 32),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_4.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_4.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_4.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            nsleep(5)
            if float(self.lineEdit_PVID_B_4.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_4.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_4.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_4.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_4.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 39, 32, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a5(self):
        if self.lineEdit_PVID_B_5.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_5.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_5.setText(str(float((int(write_feature_variable('FC1', 47, 40), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 47, 40),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_5.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_5.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_5.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_5.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_5.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_5.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_5.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_5.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 47, 40, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a6(self):
        if self.lineEdit_PVID_B_6.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_6.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_6.setText(str(float((int(write_feature_variable('FC1', 55, 48), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 55, 48),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_6.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_6.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_6.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_6.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_6.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_6.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_6.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_6.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 55, 48, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a7(self):
        if self.lineEdit_PVID_B_7.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_7.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_7.setText(str(float((int(write_feature_variable('FC1', 63, 56), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 63, 56),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_7.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_7.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_7.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_7.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_7.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_7.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int((float(self.lineEdit_PVID_B_7.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(self.lineEdit_PVID_B_7.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 63, 56, bin(pvid_b0).split("0b")[1].zfill(8))

    def boot_onClicked(self):
        if self.radioButton_pline.isChecked() == True:
            update_database_with_temp_customer_input("DF", 10, 10, "1")
        elif self.radioButton_sline.isChecked() == True:
            update_database_with_temp_customer_input("DF", 10, 10, "0")

    def railA_svid_protocol(self):
        raila_svid = svid_protocol[str(self.comboBox_svid_protocol_RailA.currentIndex())]
        update_database_with_temp_customer_input("E80", 89, 86, bin(int(raila_svid)).split("0b")[1].zfill(4))

    def railB_svid_protocol(self):
        railb_svid = svid_protocol[str(self.comboBox_svid_protocol_RailB.currentIndex())]
        update_database_with_temp_customer_input("E81", 89, 86, bin(int(railb_svid)).split("0b")[1].zfill(4))

    def railA_resolution_sel_onClicked(self):
        if self.radioButton_internal_registerA.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(False)
            update_database_with_temp_customer_input("E10", 25, 25, bin(1).split("0b")[1].zfill(1))
        elif self.radioButton_vid_sel_pinA.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)
            update_database_with_temp_customer_input("E10", 25, 25, bin(0).split("0b")[1].zfill(1))
        resolution_calculation(1)
        if resolutionA == "5":
            self.comboBox_vid_sel_pin_resolution_railA.setCurrentIndex(0)
        else:
            self.comboBox_vid_sel_pin_resolution_railA.setCurrentIndex(1)
        self.svid_vidmax_railA()
        self.svid_vidmax_railB()
        self.svid_pvid_a0()
        self.svid_pvid_a1()
        self.svid_pvid_a2()
        self.svid_pvid_a3()
        self.svid_pvid_a4()
        self.svid_pvid_a5()
        self.svid_pvid_a6()
        self.svid_pvid_a7()

    def railB_resolution_sel_onClicked(self):
        if self.radioButton_internal_registerB.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(False)
            update_database_with_temp_customer_input("E11", 25, 25, bin(1).split("0b")[1].zfill(1))
        elif self.radioButton_vid_sel_pinB.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)
            update_database_with_temp_customer_input("E11", 25, 25, bin(0).split("0b")[1].zfill(1))
        resolution_calculation(1)
        if resolutionB == "5":
            self.comboBox_vid_sel_pin_resolution_railB.setCurrentIndex(0)
        else:
            self.comboBox_vid_sel_pin_resolution_railB.setCurrentIndex(1)
        self.svid_vidmax_railA()
        self.svid_vidmax_railB()
        self.svid_pvid_railB_a0()
        self.svid_pvid_railB_a1()
        self.svid_pvid_railB_a2()
        self.svid_pvid_railB_a3()
        self.svid_pvid_railB_a4()
        self.svid_pvid_railB_a5()
        self.svid_pvid_railB_a6()
        self.svid_pvid_railB_a7()

    def pmbus_override_railA(self):
        if self.checkBox_PMBus_override_railA.isChecked():
            self.comboBox_vid_resolution_RailA.setDisabled(False)
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)
            self.radioButton_vid_sel_pinA.setDisabled(True)
            self.radioButton_internal_registerA.setDisabled(True)
            update_database_with_temp_customer_input("010", 5, 4, "00")
        else:
            self.comboBox_vid_resolution_RailA.setDisabled(True)
            self.radioButton_internal_registerA.setDisabled(False)
            self.radioButton_vid_sel_pinA.setDisabled(False)
            update_database_with_temp_customer_input("010", 5, 4, "11")
            if not (self.pushButton_Enable_APD_railA.isChecked()):
                if int(initialize_feature_variable('E10', 25, 25), 2) == 1:
                    self.radioButton_internal_registerA.setChecked(True)
                    self.radioButton_vid_sel_pinA.setChecked(False)
                    self.comboBox_vid_sel_pin_resolution_railA.setDisabled(False)
                    self.railA_vid_sel()
                else:
                    self.radioButton_internal_registerA.setChecked(False)
                    self.radioButton_vid_sel_pinA.setChecked(True)
                    self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)
                    self.railA_resolution_sel_onClicked()
        resolution_calculation(1)
        self.svid_vidmax_railA()
        self.svid_pvid_a0()
        self.svid_pvid_a1()
        self.svid_pvid_a2()
        self.svid_pvid_a3()
        self.svid_pvid_a4()
        self.svid_pvid_a5()
        self.svid_pvid_a6()
        self.svid_pvid_a7()

    def pmbus_override_railB(self):
        if self.checkBox_PMBus_override_railA_2.isChecked():
            self.comboBox_vid_resolution_RailB.setDisabled(False)
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(True)
            self.radioButton_internal_registerB.setDisabled(True)
            update_database_with_temp_customer_input("011", 5, 4, "00")
        else:
            self.comboBox_vid_resolution_RailB.setDisabled(True)
            self.radioButton_vid_sel_pinB.setDisabled(False)
            self.radioButton_internal_registerB.setDisabled(False)
            update_database_with_temp_customer_input("011", 5, 4, "11")
            if not(self.pushButton_Enable_APD_railA.isChecked()):
                if int(initialize_feature_variable('E11', 25, 25), 2) == 1:
                    self.radioButton_internal_registerB.setChecked(True)
                    self.radioButton_vid_sel_pinB.setChecked(False)
                    self.comboBox_vid_sel_pin_resolution_railB.setDisabled(False)
                    self.railB_vid_sel()
                else:
                    self.radioButton_internal_registerB.setChecked(False)
                    self.radioButton_vid_sel_pinB.setChecked(True)
                    self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)
                    self.railB_resolution_sel_onClicked()
        resolution_calculation(1)
        self.svid_vidmax_railB()
        self.svid_pvid_railB_a0()
        self.svid_pvid_railB_a1()
        self.svid_pvid_railB_a2()
        self.svid_pvid_railB_a3()
        self.svid_pvid_railB_a4()
        self.svid_pvid_railB_a5()
        self.svid_pvid_railB_a6()
        self.svid_pvid_railB_a7()

    def railA_vid_resolution(self):
        if (self.comboBox_vid_resolution_RailA.currentText() == '5mV'):
            update_database_with_temp_customer_input("E4", 3, 3, "0")
            update_database_with_temp_customer_input("200", 4, 0, "11111")
            resolution_calculation(1)
            self.svid_vidmax_railA()
            self.svid_vidmax_railB()
            self.svid_pvid_railB_a0()
            self.svid_pvid_railB_a1()
            self.svid_pvid_railB_a2()
            self.svid_pvid_railB_a3()
            self.svid_pvid_railB_a4()
            self.svid_pvid_railB_a5()
            self.svid_pvid_railB_a6()
            self.svid_pvid_railB_a7()

        elif (self.comboBox_vid_resolution_RailA.currentText() == '10mV'):
            update_database_with_temp_customer_input("E4", 3, 3, "1")
            update_database_with_temp_customer_input("200", 4, 0, "11111")
            resolution_calculation(1)
            self.svid_vidmax_railA()
            self.svid_vidmax_railB()
            self.svid_pvid_railB_a0()
            self.svid_pvid_railB_a1()
            self.svid_pvid_railB_a2()
            self.svid_pvid_railB_a3()
            self.svid_pvid_railB_a4()
            self.svid_pvid_railB_a5()
            self.svid_pvid_railB_a6()
            self.svid_pvid_railB_a7()
        else:
            update_database_with_temp_customer_input("200", 4, 0, "11110")
            resolution_calculation(1)
            self.svid_vidmax_railA()
        self.svid_pvid_a0()
        self.svid_pvid_a1()
        self.svid_pvid_a2()
        self.svid_pvid_a3()
        self.svid_pvid_a4()
        self.svid_pvid_a5()
        self.svid_pvid_a6()
        self.svid_pvid_a7()

    def railB_vid_resolution(self):
        if (self.comboBox_vid_resolution_RailB.currentText() == '5mV'):
            update_database_with_temp_customer_input("E4", 3, 3, "0")
            update_database_with_temp_customer_input("201", 4, 0, "11111")
            resolution_calculation(1)
            self.svid_vidmax_railB()
            self.svid_vidmax_railA()
            self.svid_pvid_a0()
            self.svid_pvid_a1()
            self.svid_pvid_a2()
            self.svid_pvid_a3()
            self.svid_pvid_a4()
            self.svid_pvid_a5()
            self.svid_pvid_a6()
            self.svid_pvid_a7()
        elif (self.comboBox_vid_resolution_RailB.currentText() == '10mV'):
            update_database_with_temp_customer_input("E4", 3, 3, "1")
            update_database_with_temp_customer_input("201", 4, 0, "11111")
            resolution_calculation(1)
            self.svid_vidmax_railB()
            self.svid_vidmax_railA()
            self.svid_pvid_a0()
            self.svid_pvid_a1()
            self.svid_pvid_a2()
            self.svid_pvid_a3()
            self.svid_pvid_a4()
            self.svid_pvid_a5()
            self.svid_pvid_a6()
            self.svid_pvid_a7()
        else:
            update_database_with_temp_customer_input("201", 4, 0, "11110")
            resolution_calculation(1)
            self.svid_vidmax_railB()
        self.svid_pvid_railB_a0()
        self.svid_pvid_railB_a1()
        self.svid_pvid_railB_a2()
        self.svid_pvid_railB_a3()
        self.svid_pvid_railB_a4()
        self.svid_pvid_railB_a5()
        self.svid_pvid_railB_a6()
        self.svid_pvid_railB_a7()

    def railA_vid_sel(self):
        raila_svid_vid = self.comboBox_vid_sel_pin_resolution_railA.currentIndex()
        update_database_with_temp_customer_input("E10", 24, 24, bin(raila_svid_vid).split("0b")[1].zfill(1))
        resolution_calculation(1)
        self.svid_vidmax_railA()

    def railB_vid_sel(self):
        railb_svid_vid = self.comboBox_vid_sel_pin_resolution_railB.currentIndex()
        update_database_with_temp_customer_input("E11", 24, 24, bin(railb_svid_vid).split("0b")[1].zfill(1))
        resolution_calculation(1)
        self.svid_vidmax_railB()

    def railA_slew_fast(self):
        if self.comboBox_slewrate_fast_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("270", 15, 0, "1001101111010111")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("270", 15, 0, "1011001000000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("270", 15, 0, "1011101000000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("270", 15, 0, "1100001010000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("270", 15, 0, "1100101010000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("270", 15, 0, "1100101111000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("270", 15, 0, "1101001010000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 7:
            update_database_with_temp_customer_input("270", 15, 0, "1101001100100000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 8:
            update_database_with_temp_customer_input("270", 15, 0, "1101001111000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 9:
            update_database_with_temp_customer_input("270", 15, 0, "1101101000110000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 10:
            update_database_with_temp_customer_input("270", 15, 0, "1101101010000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 11:
            update_database_with_temp_customer_input("270", 15, 0, "1101101011010000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 12:
            update_database_with_temp_customer_input("270", 15, 0, "1101101100100000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 13:
            update_database_with_temp_customer_input("270", 15, 0, "1101101101110000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 14:
            update_database_with_temp_customer_input("270", 15, 0, "1101101111000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 15:
            update_database_with_temp_customer_input("270", 15, 0, "1110001000001000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 16:
            update_database_with_temp_customer_input("270", 15, 0, "1110001000110000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 17:
            update_database_with_temp_customer_input("270", 15, 0, "1110001001011000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 18:
            update_database_with_temp_customer_input("270", 15, 0, "1110001010000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 19:
            update_database_with_temp_customer_input("270", 15, 0, "1110001100000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 20:
            update_database_with_temp_customer_input("270", 15, 0, "1110001111000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 21:
            update_database_with_temp_customer_input("270", 15, 0, "1110101010000000")
        elif self.comboBox_slewrate_fast_RailA.currentIndex() == 22:
            update_database_with_temp_customer_input("270", 15, 0, "1110101100000000")
        else:
            update_database_with_temp_customer_input("270", 15, 0, "1110101111101000")

        # raila_slew_fast = slew_rate[str(self.comboBox_slewrate_fast_RailA.currentText().replace("mV/μs", ""))]
        # print(raila_slew_fast)
        # update_database_with_temp_customer_input("270", 15, 0, bin(int(raila_slew_fast)).split("0b")[1].zfill(16))

    def railB_slew_fast(self):
        if self.comboBox_slewrate_fast_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("271", 15, 0, "1001101111010111")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("271", 15, 0, "1011001000000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("271", 15, 0, "1011101000000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("271", 15, 0, "1100001010000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("271", 15, 0, "1100101010000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("271", 15, 0, "1100101111000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("271", 15, 0, "1101001010000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 7:
            update_database_with_temp_customer_input("271", 15, 0, "1101001100100000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 8:
            update_database_with_temp_customer_input("271", 15, 0, "1101001111000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 9:
            update_database_with_temp_customer_input("271", 15, 0, "1101101000110000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 10:
            update_database_with_temp_customer_input("271", 15, 0, "1101101010000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 11:
            update_database_with_temp_customer_input("271", 15, 0, "1101101011010000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 12:
            update_database_with_temp_customer_input("271", 15, 0, "1101101100100000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 13:
            update_database_with_temp_customer_input("271", 15, 0, "1101101101110000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 14:
            update_database_with_temp_customer_input("271", 15, 0, "1101101111000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 15:
            update_database_with_temp_customer_input("271", 15, 0, "1110001000001000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 16:
            update_database_with_temp_customer_input("271", 15, 0, "1110001000110000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 17:
            update_database_with_temp_customer_input("271", 15, 0, "1110001001011000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 18:
            update_database_with_temp_customer_input("271", 15, 0, "1110001010000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 19:
            update_database_with_temp_customer_input("271", 15, 0, "1110001100000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 20:
            update_database_with_temp_customer_input("271", 15, 0, "1110001111000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 21:
            update_database_with_temp_customer_input("271", 15, 0, "1110101010000000")
        elif self.comboBox_slewrate_fast_RailB.currentIndex() == 22:
            update_database_with_temp_customer_input("271", 15, 0, "1110101100000000")
        else:
            update_database_with_temp_customer_input("271", 15, 0, "1110101111101000")

        # railb_slew_fast = slew_rate[str(self.comboBox_slewrate_fast_RailB.currentText().replace("mV/μs", ""))]
        # update_database_with_temp_customer_input("271", 15, 0, bin(int(railb_slew_fast)).split("0b")[1].zfill(16))

    def railA_slew_slow(self):
        slow_slew_sel = {0: bin(8), 1: bin(4), 2: bin(2), 3: bin(1)}
        raila_slew_slow = self.comboBox_slew_rate_slow_railA.currentIndex()
        update_database_with_temp_customer_input("E80", 4, 0, slow_slew_sel[raila_slew_slow].replace("0b", "").zfill(5))

    def railB_slew_slow(self):
        slow_slew_sel = {0: bin(8), 1: bin(4), 2: bin(2), 3: bin(1)}
        railb_slew_slow = self.comboBox_slew_rate_slow_railB.currentIndex()
        update_database_with_temp_customer_input("E81", 4, 0, slow_slew_sel[railb_slew_slow].replace("0b", "").zfill(5))

    def all_call_railA(self):
        all_call = {0: bin(0), 1: bin(2), 2: bin(1), 3: bin(3)}
        raila_all_call = self.comboBox_all_call_address_RailA.currentIndex()
        update_database_with_temp_customer_input("E80", 65, 64, all_call[raila_all_call].replace("0b", "").zfill(2))

    def all_call_railB(self):
        all_call = {0: bin(0), 1: bin(2), 2: bin(1), 3: bin(3)}
        railb_all_call = self.comboBox_all_call_address_RailB.currentIndex()
        update_database_with_temp_customer_input("E81", 65, 64, all_call[railb_all_call].replace("0b", "").zfill(2))

    def svid_save_register(self):
        global list_of_registers_used_in_this_frame, resolutionA, resolutionB, initial_device_status, homeWin_obj, \
            parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                        homeWin_obj.label_prot_address_A.setText(str(hex(int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
                        homeWin_obj.label_prot_address_B.setText(str(hex(1 + int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")

                resolution_calculation(1)
                print_log("resolution of RAILA is " + resolutionA, "INFO")
                print_log("resolution of RAILB is " + resolutionB, "INFO")

                initial_vid_raila = float(((int(initialize_feature_variable('E80', 31, 24), 2) * float(resolutionA)) + offset[
                    str(resolutionA)]) / 1000) if int(initialize_feature_variable('E80', 31, 24), 2) != 0 else int(0)
                print(initial_vid_raila)
                self.lineEdit_vid_max_railA.setText(str(initial_vid_raila))
                initial_vid_railb = float(
                    ((int(initialize_feature_variable('E81', 31, 24), 2) * float(resolutionB)) + offset[
                        str(resolutionB)]) / 1000) if int(initialize_feature_variable('E81', 31, 24), 2) != 0 else int(
                    0)
                self.lineEdit_vid_max_railB.setText(str(initial_vid_railb))

                initial_pvid_0 = float((int(initialize_feature_variable('FC0', 7, 0), 2) * float(resolutionA) + offset[
                    str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 7, 0), 2) != 0 else int(0)
                self.lineEdit_PVID_A_0.setText(str(initial_pvid_0))
                self.lineEdit_PVID_A_1.setText(str(float((int(initialize_feature_variable('FC0', 15, 8), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 15, 8),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_2.setText(str(float((int(initialize_feature_variable('FC0', 23, 16), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 23, 16),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_3.setText(str(float((int(initialize_feature_variable('FC0', 31, 24), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 31, 24),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_4.setText(str(float((int(initialize_feature_variable('FC0', 39, 32), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 39, 32),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_5.setText(str(float((int(initialize_feature_variable('FC0', 47, 40), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 47, 40),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_6.setText(str(float((int(initialize_feature_variable('FC0', 55, 48), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 55, 48),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_7.setText(str(float((int(initialize_feature_variable('FC0', 63, 56), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 63, 56),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_0.setText(str(float((int(initialize_feature_variable('FC1', 7, 0), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 7, 0),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_1.setText(str(float((int(initialize_feature_variable('FC1', 15, 8), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 15, 8),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_2.setText(str(float((int(initialize_feature_variable('FC1', 23, 16), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 23, 16),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_3.setText(str(float((int(initialize_feature_variable('FC1', 31, 24), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 31, 24),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_4.setText(str(float((int(initialize_feature_variable('FC1', 39, 32), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 39, 32),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_5.setText(str(float((int(initialize_feature_variable('FC1', 47, 40), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 47, 40),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_6.setText(str(float((int(initialize_feature_variable('FC1', 55, 48), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 55, 48),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_7.setText(str(float((int(initialize_feature_variable('FC1', 63, 56), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 63, 56),
                                                                            2) != 0 else int(0)))

                print_log("SVID configuration settings are saved", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def svid_discard_changes(self):
        print_log("SVID manager Settings discarded.", "INFO")
        resolution_calculation(0)
        self.main = frame_svid()
        self.main.show()
        self.close()

class frame_svi2(QMainWindow, Ui_label_main_SVI2_Configuration):

    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(frame_svi2, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["E6", "E9", "E4", "DF", "F2", "010", "E80", "E10", "FC0", "200", "011", "E81", "E11", "FC1", "201"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        resolution_calculation(0)

        self.lineEdit_icc_max_railA.setText(str(int(initialize_feature_variable('E80', 55, 48), 2)))
        self.lineEdit_icc_max_railB.setText(str(int(initialize_feature_variable('E81', 55, 48), 2)))
        initial_vid_raila = float((int(initialize_feature_variable('E80', 31, 24), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('E80', 31, 24), 2) != 0 else int(0)
        self.lineEdit_vid_max_railA.setText(str(initial_vid_raila))
        print_log("vid max " + str(float(resolutionA)), "INFO")
        self.raila_vidmax = initial_vid_raila
        initial_vid_railb = float((int(initialize_feature_variable('E81', 31, 24), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('E81', 31, 24), 2) != 0 else int(0)
        self.lineEdit_vid_max_railB.setText(str(initial_vid_railb))
        self.railb_vidmax = initial_vid_railb
        self.lineEdit_temp_max.setText(str(int(initialize_feature_variable('E80', 47, 40), 2)))
        initial_pvid_0 = float((int(initialize_feature_variable('FC0', 7, 0), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 7, 0), 2) != 0 else int(0)
        self.lineEdit_PVID_A_0.setText(str(initial_pvid_0))
        self.lineEdit_PVID_A_1.setText(str(float((int(initialize_feature_variable('FC0', 15, 8), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 15, 8), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_2.setText(str(float((int(initialize_feature_variable('FC0', 23, 16), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 23, 16), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_3.setText(str(float((int(initialize_feature_variable('FC0', 31, 24), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 31, 24), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_4.setText(str(float((int(initialize_feature_variable('FC0', 39, 32), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 39, 32), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_5.setText(str(float((int(initialize_feature_variable('FC0', 47, 40), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 47, 40), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_6.setText(str(float((int(initialize_feature_variable('FC0', 55, 48), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 55, 48), 2) != 0 else int(0)))
        self.lineEdit_PVID_A_7.setText(str(float((int(initialize_feature_variable('FC0', 63, 56), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 63, 56), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_0.setText(str(float((int(initialize_feature_variable('FC1', 7, 0), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 7, 0), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_1.setText(str(float((int(initialize_feature_variable('FC1', 15, 8), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 15, 8), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_2.setText(str(float((int(initialize_feature_variable('FC1', 23, 16), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 23, 16), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_3.setText(str(float((int(initialize_feature_variable('FC1', 31, 24), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 31, 24), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_4.setText(str(float((int(initialize_feature_variable('FC1', 39, 32), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 39, 32), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_5.setText(str(float((int(initialize_feature_variable('FC1', 47, 40), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 47, 40), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_6.setText(str(float((int(initialize_feature_variable('FC1', 55, 48), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 55, 48), 2) != 0 else int(0)))
        self.lineEdit_PVID_B_7.setText(str(float((int(initialize_feature_variable('FC1', 63, 56), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 63, 56), 2) != 0 else int(0)))

        def get_key_slew(val):
            for key, value in slew_rate.items():
                if val == value:
                    return key

        self.comboBox_vid_resolution_RailA.setCurrentIndex(not (int(initialize_feature_variable('E10', 25, 25), 2)))
        self.comboBox_vid_resolution_RailB.setCurrentIndex(not (int(initialize_feature_variable('E11', 25, 25), 2)))

        if int(initialize_feature_variable('200', 4, 0), 2) == 0x1E:
            self.comboBox_vid_resolution_RailA.setCurrentIndex(1)
        else:
            if int(initialize_feature_variable('E4', 3, 3), 2) == 1:
                self.comboBox_vid_resolution_RailA.setCurrentIndex(2)
            else:
                self.comboBox_vid_resolution_RailA.setCurrentIndex(0)

        if int(initialize_feature_variable('201', 4, 0), 2) == 0x1E:
            self.comboBox_vid_resolution_RailB.setCurrentIndex(1)
        else:
            if int(initialize_feature_variable('E4', 3, 3), 2) == 1:
                self.comboBox_vid_resolution_RailB.setCurrentIndex(2)
            else:
                self.comboBox_vid_resolution_RailB.setCurrentIndex(0)

        if int(initialize_feature_variable('010', 5, 4), 2) == 0:
            self.checkBox_PMBus_override_railA.setChecked(True)
            self.comboBox_vid_resolution_RailA.setDisabled(False)
        else:
            self.checkBox_PMBus_override_railA.setChecked(False)
            self.comboBox_vid_resolution_RailA.setDisabled(True)

        if int(initialize_feature_variable('011', 5, 4), 2) == 0:
            self.checkBox_PMBus_override_railA_2.setChecked(True)
            self.comboBox_vid_resolution_RailB.setDisabled(False)
        else:
            self.checkBox_PMBus_override_railA_2.setChecked(False)
            self.comboBox_vid_resolution_RailB.setDisabled(True)

        # self.comboBox_slewrate_fast_RailA.setCurrentText(str(get_key_slew(int(initialize_feature_variable('270', 15, 0),2)))+"mV/μs")
        # self.comboBox_slewrate_fast_RailB.setCurrentText(str(get_key_slew(int(initialize_feature_variable('271', 15, 0),2)))+"mV/μs")

        # Other initialization

        if int(initialize_feature_variable('DF', 11, 11)) == 1:
            self.pushButton_Enable_APD_railA.setChecked(True)
            self.pvid_enable()
        else:
            self.pushButton_Enable_APD_railA.setChecked(False)
            self.pvid_enable()

        self.label_display_RailA_2.setText(RailA_name)
        self.label_display_RailB_2.setText(RailB_name)
        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        ## code starts
        self.lineEdit_icc_max_railA.textEdited.connect(self.svid_iccmax_railA)
        self.lineEdit_icc_max_railB.textEdited.connect(self.svid_iccmax_railB)
        self.lineEdit_vid_max_railA.textEdited.connect(self.svid_vidmax_railA)
        self.lineEdit_vid_max_railB.textEdited.connect(self.svid_vidmax_railB)
        self.lineEdit_temp_max.textEdited.connect(self.svid_tempmax)

        self.lineEdit_PVID_A_0.textEdited.connect(self.svid_pvid_a0)
        self.lineEdit_PVID_A_1.textEdited.connect(self.svid_pvid_a1)
        self.lineEdit_PVID_A_2.textEdited.connect(self.svid_pvid_a2)
        self.lineEdit_PVID_A_3.textEdited.connect(self.svid_pvid_a3)
        self.lineEdit_PVID_A_4.textEdited.connect(self.svid_pvid_a4)
        self.lineEdit_PVID_A_5.textEdited.connect(self.svid_pvid_a5)
        self.lineEdit_PVID_A_6.textEdited.connect(self.svid_pvid_a6)
        self.lineEdit_PVID_A_7.textEdited.connect(self.svid_pvid_a7)
        self.lineEdit_PVID_B_0.textEdited.connect(self.svid_pvid_railB_a0)
        self.lineEdit_PVID_B_1.textEdited.connect(self.svid_pvid_railB_a1)
        self.lineEdit_PVID_B_2.textEdited.connect(self.svid_pvid_railB_a2)
        self.lineEdit_PVID_B_3.textEdited.connect(self.svid_pvid_railB_a3)
        self.lineEdit_PVID_B_4.textEdited.connect(self.svid_pvid_railB_a4)
        self.lineEdit_PVID_B_5.textEdited.connect(self.svid_pvid_railB_a5)
        self.lineEdit_PVID_B_6.textEdited.connect(self.svid_pvid_railB_a6)
        self.lineEdit_PVID_B_7.textEdited.connect(self.svid_pvid_railB_a7)

        self.comboBox_vid_resolution_RailA.activated.connect(self.railA_vid_resolution)
        self.comboBox_vid_resolution_RailB.activated.connect(self.railB_vid_resolution)

        self.checkBox_PMBus_override_railA.toggled.connect(self.pmbus_override_railA)
        self.checkBox_PMBus_override_railA_2.toggled.connect(self.pmbus_override_railB)

        self.comboBox_slewrate_fast_RailA.activated.connect(self.railA_slew_fast)
        self.comboBox_slewrate_fast_RailB.activated.connect(self.railB_slew_fast)
        self.pushButton_Enable_APD_railA.toggled.connect(self.pvid_enable)

        self.pushButton_Save.clicked.connect(self.svid_save_register)
        self.pushButton_Discard.clicked.connect(self.svid_discard_changes)
        self.pushButton_Enable_APD_railA.clicked.connect(self.pvid_enable_reg)

    def pvid_enable(self):
        if self.pushButton_Enable_APD_railA.isChecked():
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but1.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but1_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")
            update_database_with_temp_customer_input("DF", 11, 11, bin(1).split("0b")[1].zfill(1))
            self.lineEdit_PVID_A_0.setDisabled(False)
            self.lineEdit_PVID_A_1.setDisabled(False)
            self.lineEdit_PVID_A_2.setDisabled(False)
            self.lineEdit_PVID_A_3.setDisabled(False)
            self.lineEdit_PVID_A_4.setDisabled(False)
            self.lineEdit_PVID_A_5.setDisabled(False)
            self.lineEdit_PVID_A_6.setDisabled(False)
            self.lineEdit_PVID_A_7.setDisabled(False)
            self.lineEdit_PVID_B_0.setDisabled(False)
            self.lineEdit_PVID_B_1.setDisabled(False)
            self.lineEdit_PVID_B_2.setDisabled(False)
            self.lineEdit_PVID_B_3.setDisabled(False)
            self.lineEdit_PVID_B_4.setDisabled(False)
            self.lineEdit_PVID_B_5.setDisabled(False)
            self.lineEdit_PVID_B_6.setDisabled(False)
            self.lineEdit_PVID_B_7.setDisabled(False)
            self.lineEdit_PVID_A_0.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_1.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_2.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_3.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_4.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_5.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_6.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_7.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_0.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_1.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_2.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_3.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_4.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_5.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_6.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_B_7.setStyleSheet("background-color: rgb(255, 255, 255);")
            self.lineEdit_PVID_A_0.setText(str(float((int(initialize_feature_variable('FC0', 7, 0), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 7, 0),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_1.setText(str(float((int(initialize_feature_variable('FC0', 15, 8), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 15, 8),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_2.setText(str(float((int(initialize_feature_variable('FC0', 23, 16), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 23, 16),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_3.setText(str(float((int(initialize_feature_variable('FC0', 31, 24), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 31, 24),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_4.setText(str(float((int(initialize_feature_variable('FC0', 39, 32), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 39, 32),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_5.setText(str(float((int(initialize_feature_variable('FC0', 47, 40), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 47, 40),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_6.setText(str(float((int(initialize_feature_variable('FC0', 55, 48), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 55, 48),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_A_7.setText(str(float((int(initialize_feature_variable('FC0', 63, 56), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 63, 56),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_0.setText(str(float((int(initialize_feature_variable('FC1', 7, 0), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 7, 0),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_1.setText(str(float((int(initialize_feature_variable('FC1', 15, 8), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 15, 8),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_2.setText(str(float((int(initialize_feature_variable('FC1', 23, 16), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 23, 16),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_3.setText(str(float((int(initialize_feature_variable('FC1', 31, 24), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 31, 24),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_4.setText(str(float((int(initialize_feature_variable('FC1', 39, 32), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 39, 32),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_5.setText(str(float((int(initialize_feature_variable('FC1', 47, 40), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 47, 40),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_6.setText(str(float((int(initialize_feature_variable('FC1', 55, 48), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 55, 48),
                                                                        2) != 0 else int(0)))
            self.lineEdit_PVID_B_7.setText(str(float((int(initialize_feature_variable('FC1', 63, 56), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 63, 56),
                                                                        2) != 0 else int(0)))

            # if it is unchecked
        else:
            # setting image to button0
            self.pushButton_Enable_APD_railA.setStyleSheet("QPushButton{\n"
                                                           "border-image: url(GUI_IMAGE/but0.png);\n"
                                                           "}\n"
                                                           "\n"
                                                           "\n"
                                                           "QPushButton::hover {\n"
                                                           "border-image: url(GUI_IMAGE/but0_hover"
                                                           ".png);\n "
                                                           "    }\n"
                                                           "")
            update_database_with_temp_customer_input("DF", 11, 11, bin(0).split("0b")[1].zfill(1))
            self.lineEdit_PVID_A_0.setDisabled(True)
            self.lineEdit_PVID_A_1.setDisabled(True)
            self.lineEdit_PVID_A_2.setDisabled(True)
            self.lineEdit_PVID_A_3.setDisabled(True)
            self.lineEdit_PVID_A_4.setDisabled(True)
            self.lineEdit_PVID_A_5.setDisabled(True)
            self.lineEdit_PVID_A_6.setDisabled(True)
            self.lineEdit_PVID_A_7.setDisabled(True)
            self.lineEdit_PVID_B_0.setDisabled(True)
            self.lineEdit_PVID_B_1.setDisabled(True)
            self.lineEdit_PVID_B_2.setDisabled(True)
            self.lineEdit_PVID_B_3.setDisabled(True)
            self.lineEdit_PVID_B_4.setDisabled(True)
            self.lineEdit_PVID_B_5.setDisabled(True)
            self.lineEdit_PVID_B_6.setDisabled(True)
            self.lineEdit_PVID_B_7.setDisabled(True)
            self.lineEdit_PVID_A_0.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_0.setText("--")
            self.lineEdit_PVID_A_1.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_1.setText("--")
            self.lineEdit_PVID_A_2.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_2.setText("--")
            self.lineEdit_PVID_A_3.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_3.setText("--")
            self.lineEdit_PVID_A_4.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_4.setText("--")
            self.lineEdit_PVID_A_5.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_5.setText("--")
            self.lineEdit_PVID_A_6.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_6.setText("--")
            self.lineEdit_PVID_A_7.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_A_7.setText("--")
            self.lineEdit_PVID_B_0.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_0.setText("--")
            self.lineEdit_PVID_B_1.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_1.setText("--")
            self.lineEdit_PVID_B_2.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_2.setText("--")
            self.lineEdit_PVID_B_3.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_3.setText("--")
            self.lineEdit_PVID_B_4.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_4.setText("--")
            self.lineEdit_PVID_B_5.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_5.setText("--")
            self.lineEdit_PVID_B_6.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_6.setText("--")
            self.lineEdit_PVID_B_7.setStyleSheet("background-color: rgb(200, 200, 200);")
            self.lineEdit_PVID_B_7.setText("--")

    def pvid_enable_reg(self):
        if self.pushButton_Enable_APD_railA.isChecked():
            update_database_with_temp_customer_input("DF", 11, 11, bin(1))
        else:
            update_database_with_temp_customer_input("DF", 11, 11, bin(0))

    def svid_iccmax_railA(self):
        if self.lineEdit_icc_max_railA.text() == "":
            return
        try:
            icc_max = int(self.lineEdit_icc_max_railA.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railA.setText(str(int(write_feature_variable('E80', 55, 48), 2)))
            return

        if (icc_max > 255):
            print_log("Text box value should be less than 256", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railA.setText(str(int(write_feature_variable('E80', 55, 48), 2)))
            return
        else:
            update_database_with_temp_customer_input("E80", 55, 48, bin(icc_max).split("0b")[1].zfill(8))

    def svid_iccmax_railB(self):
        if self.lineEdit_icc_max_railB.text() == "":
            return
        try:
            icc_max = int(self.lineEdit_icc_max_railB.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railB.setText(str(int(write_feature_variable('E81', 55, 48), 2)))
            return

        if (icc_max > 255):
            print_log("Text box value should be less than 256", "ERROR")
            nsleep(5)
            self.lineEdit_icc_max_railB.setText(str(int(write_feature_variable('E81', 55, 48), 2)))
            return
        else:
            update_database_with_temp_customer_input("E81", 55, 48, bin(icc_max).split("0b")[1].zfill(8))

    def svid_vidmax_railA(self):
        resolution_calculation(1)
        if self.lineEdit_vid_max_railA.text() == "":
            return
        try:
            float(self.lineEdit_vid_max_railA.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            initial_vid_raila = float((int(write_feature_variable('E80', 31, 24), 2) * float(resolutionA) + offset[
                str(resolutionA)]) / 1000) if int(write_feature_variable('E80', 31, 24), 2) != 0 else int(0)
            self.lineEdit_vid_max_railA.setText(str(initial_vid_raila))
            return
        if (float(self.lineEdit_vid_max_railA.text()) > vid_max_value[resolutionA] or float(self.lineEdit_vid_max_railA.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_vid_max_railA.text()) != 0:
            # string = "Please enter input in range" + vid_min_value[resolutionA] + "to" + vid_max_value[resolutionA]
            print_log("Input vid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_vid_max_railA.text()) > vid_max_value[resolutionA]:
                self.lineEdit_vid_max_railA.setText(str(vid_max_value[resolutionA]))
            else:
                self.lineEdit_vid_max_railA.setText(str(vid_min_value[resolutionA]))

        self.raila_vidmax = float(self.lineEdit_vid_max_railA.text())
        vid_max = int((float(self.lineEdit_vid_max_railA.text()) * 1000 - float(offset[resolutionA])) / float(resolutionA))
        update_database_with_temp_customer_input("E80", 31, 24, bin(vid_max).split("0b")[1].zfill(8))

    def svid_vidmax_railB(self):
        resolution_calculation(1)
        if self.lineEdit_vid_max_railB.text() == "":
            return
        try:
            float(self.lineEdit_vid_max_railB.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            initial_vid_railb = float((int(write_feature_variable('E81', 31, 24), 2) * float(resolutionB) + offset[
                str(resolutionB)]) / 1000) if int(write_feature_variable('E81', 31, 24), 2) != 0 else int(0)
            self.lineEdit_vid_max_railB.setText(str(initial_vid_railb))
            return

        if (float(self.lineEdit_vid_max_railB.text()) > vid_max_value[resolutionB] or float(self.lineEdit_vid_max_railB.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_vid_max_railB.text()) != 0:
            # string = "Please enter input in range" + vid_min_value[resolutionA] + "to" + vid_max_value[resolutionA]
            print_log("Input vid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            if float(self.lineEdit_vid_max_railB.text()) > vid_max_value[resolutionB]:
                self.lineEdit_vid_max_railB.setText(str(vid_max_value[resolutionB]))
            else:
                self.lineEdit_vid_max_railB.setText(str(vid_min_value[resolutionB]))

        self.railb_vidmax = float(self.lineEdit_vid_max_railB.text())
        vid_max = int((float(self.lineEdit_vid_max_railB.text()) * 1000 - float(offset[resolutionB])) / float(resolutionB))
        update_database_with_temp_customer_input("E81", 31, 24, bin(vid_max).split("0b")[1].zfill(8))

    def svid_tempmax(self):
        if self.lineEdit_temp_max.text() == "":
            return
        try:
            temp_max = int(self.lineEdit_temp_max.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_temp_max.setText(str(int(write_feature_variable('E80', 47, 40), 2)))
            return

        if (temp_max > 255):
            print_log("Text box value should be less than 255", "ERROR")
            nsleep(5)
            self.lineEdit_temp_max.setText(str(int(write_feature_variable('E80', 47, 40), 2)))
            return
        else:
            update_database_with_temp_customer_input("E80", 47, 40, bin(temp_max).split("0b")[1].zfill(8))

    def svid_imonaux(self):
        if self.lineEdit_imon_aux.text() == "":
            return
        try:
            imonaux = int(self.lineEdit_imon_aux.text(), 10)
        except ValueError:
            print_log("Text box value should be decimal number", "ERROR")
            nsleep(5)
            self.lineEdit_imon_aux.setText(str(int(write_feature_variable('E9', 61, 54), 2)))
            return

        if (imonaux > 255):
            print_log("Text box value should be less than 255", "ERROR")
            nsleep(5)
            self.lineEdit_imon_aux.setText(str(int(write_feature_variable('E9', 61, 54), 2)))
            return
        else:
            update_database_with_temp_customer_input("E9", 61, 54, bin(imonaux).split("0b")[1].zfill(8))

    def svid_pvid_a0(self):
        if self.lineEdit_PVID_A_0.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_0.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            initial_pvid_0 = float((int(write_feature_variable('FC0', 7, 0), 2) * float(resolutionA) + offset[
                str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 7, 0), 2) != 0 else int(0)
            self.lineEdit_PVID_A_0.setText(str(initial_pvid_0))
            return

        if (float(self.lineEdit_PVID_A_0.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_0.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_0.text()) != 0:
            # string = "Please enter input in range" + vid_min_value[resolutionA] + "to" + vid_max_value[resolutionA]
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_0.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_0.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_0.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_0.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_0.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 7, 0, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a1(self):
        if self.lineEdit_PVID_A_1.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_1.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_1.setText(str(float((int(write_feature_variable('FC0', 15, 8), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 15, 8),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_1.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_1.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_1.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_1.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_1.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_1.setText(str(vid_min_value[resolutionA]))

        pvid_a1 = int(
            (float(self.lineEdit_PVID_A_1.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_1.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 15, 8, bin(pvid_a1).split("0b")[1].zfill(8))

    def svid_pvid_a2(self):
        if self.lineEdit_PVID_A_2.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_2.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point  number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_2.setText(str(float((int(write_feature_variable('FC0', 23, 16), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 23, 16),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_0.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_2.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_2.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_2.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_2.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_2.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_2.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_2.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 23, 16, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a3(self):
        if self.lineEdit_PVID_A_3.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_3.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_3.setText(str(float((int(write_feature_variable('FC0', 31, 24), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 31, 24),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_3.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_3.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_3.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_3.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_3.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_3.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_3.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_3.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 31, 24, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a4(self):
        if self.lineEdit_PVID_A_4.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_4.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_4.setText(str(float((int(write_feature_variable('FC0', 39, 32), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 39, 32),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_4.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_4.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_4.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_4.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_4.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_4.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_4.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_4.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 39, 32, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a5(self):
        if self.lineEdit_PVID_A_5.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_5.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_5.setText(str(float((int(write_feature_variable('FC0', 47, 40), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 47, 40),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_5.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_5.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_5.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_5.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_5.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_5.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_5.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_5.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 47, 40, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a6(self):
        if self.lineEdit_PVID_A_6.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_6.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_6.setText(str(float((int(write_feature_variable('FC0', 55, 48), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 55, 48),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_6.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_6.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_6.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_6.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_6.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_6.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_6.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_6.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 55, 48, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_a7(self):
        if self.lineEdit_PVID_A_7.text() == "":
            return
        try:
            float(self.lineEdit_PVID_A_7.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_A_7.setText(str(float((int(write_feature_variable('FC0', 63, 56), 2) * float(
                resolutionA) + offset[str(resolutionA)]) / 1000) if int(write_feature_variable('FC0', 63, 56),
                                                                        2) != 0 else int(0)))
            return

        if (float(self.lineEdit_PVID_A_7.text()) > min(vid_max_value[resolutionA], self.raila_vidmax) or float(self.lineEdit_PVID_A_7.text()) < vid_min_value[resolutionA]) and float(self.lineEdit_PVID_A_7.text()) != 0:
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionA]) + " to " + str(vid_max_value[resolutionA]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_A_7.text()) > min(vid_max_value[resolutionA], self.raila_vidmax):
                self.lineEdit_PVID_A_7.setText(str(min(vid_max_value[resolutionA], self.raila_vidmax)))
            else:
                self.lineEdit_PVID_A_7.setText(str(vid_min_value[resolutionA]))

        pvid_a0 = int(
            (float(self.lineEdit_PVID_A_7.text()) * 1000 - offset[str(resolutionA)]) / int(resolutionA)) if float(
            self.lineEdit_PVID_A_7.text()) != 0 else 0

        update_database_with_temp_customer_input("FC0", 63, 56, bin(pvid_a0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a0(self):
        if self.lineEdit_PVID_B_0.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_0.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_0.setText(str(float((int(write_feature_variable('FC1', 7, 0), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 7, 0),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_0.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_0.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_0.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_0.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_0.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_0.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_0.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_0.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 7, 0, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a1(self):
        if self.lineEdit_PVID_B_1.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_1.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_1.setText(str(float((int(write_feature_variable('FC1', 15, 8), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 15, 8),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_1.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_1.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_1.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_1.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_1.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_1.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_1.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_1.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 15, 8, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a2(self):
        if self.lineEdit_PVID_B_2.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_2.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_2.setText(str(float((int(write_feature_variable('FC1', 23, 16), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 23, 16),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_2.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_2.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_2.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_2.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_2.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_2.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_2.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_2.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 23, 16, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a3(self):
        if self.lineEdit_PVID_B_3.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_3.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_3.setText(str(float((int(write_feature_variable('FC1', 31, 24), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 23, 16),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_3.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_3.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_3.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_3.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_3.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_3.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_3.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_3.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 31, 24, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a4(self):
        if self.lineEdit_PVID_B_4.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_4.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            self.lineEdit_PVID_B_4.setText(str(float((int(write_feature_variable('FC1', 39, 32), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 39, 32),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_4.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_4.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_4.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            nsleep(5)
            if float(self.lineEdit_PVID_B_4.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_4.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_4.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_4.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_4.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 39, 32, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a5(self):
        if self.lineEdit_PVID_B_5.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_5.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_5.setText(str(float((int(write_feature_variable('FC1', 47, 40), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 47, 40),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_5.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_5.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_5.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_5.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_5.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_5.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_5.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_5.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 47, 40, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a6(self):
        if self.lineEdit_PVID_B_6.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_6.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_6.setText(str(float((int(write_feature_variable('FC1', 55, 48), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 55, 48),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_6.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_6.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_6.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(vid_max_value[resolutionB]), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_6.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_6.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_6.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int(
            (float(self.lineEdit_PVID_B_6.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(
            self.lineEdit_PVID_B_6.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 55, 48, bin(pvid_b0).split("0b")[1].zfill(8))

    def svid_pvid_railB_a7(self):
        if self.lineEdit_PVID_B_7.text() == "":
            return
        try:
            float(self.lineEdit_PVID_B_7.text()).is_integer()
        except ValueError:
            print_log("Text box value should be floating point number", "ERROR")
            nsleep(5)
            self.lineEdit_PVID_B_7.setText(str(float((int(write_feature_variable('FC1', 63, 56), 2) * float(
                resolutionB) + offset[str(resolutionB)]) / 1000) if int(write_feature_variable('FC1', 63, 56),
                                                                        2) != 0 else int(0)))
            return

        if ((float(self.lineEdit_PVID_B_7.text()) > min(vid_max_value[resolutionB], self.railb_vidmax) or float(self.lineEdit_PVID_B_7.text()) < vid_min_value[resolutionB]) and float(self.lineEdit_PVID_B_7.text()) != 0):
            print_log("Input pvid is out of range", "ERROR")
            print_log("Please enter input in range " + str(vid_min_value[resolutionB]) + " to " + str(min(vid_max_value[resolutionB], self.railb_vidmax)), "INFO")
            nsleep(5)
            if float(self.lineEdit_PVID_B_7.text()) > min(vid_max_value[resolutionB], self.railb_vidmax):
                self.lineEdit_PVID_B_7.setText(str(min(vid_max_value[resolutionB], self.railb_vidmax)))
            else:
                self.lineEdit_PVID_B_7.setText(str(vid_min_value[resolutionB]))

        pvid_b0 = int((float(self.lineEdit_PVID_B_7.text()) * 1000 - offset[str(resolutionB)]) / float(resolutionB)) if float(self.lineEdit_PVID_B_7.text()) != 0 else 0

        update_database_with_temp_customer_input("FC1", 63, 56, bin(pvid_b0).split("0b")[1].zfill(8))

    def boot_onClicked(self):
        if self.radioButton_pline.isChecked() == True:
            update_database_with_temp_customer_input("DF", 10, 10, "1")
        elif self.radioButton_sline.isChecked() == True:
            update_database_with_temp_customer_input("DF", 10, 10, "0")

    def railA_svid_protocol(self):
        raila_svid = svid_protocol[str(self.comboBox_svid_protocol_RailA.currentIndex())]
        update_database_with_temp_customer_input("E80", 89, 86, bin(int(raila_svid)).split("0b")[1].zfill(4))

    def railB_svid_protocol(self):
        railb_svid = svid_protocol[str(self.comboBox_svid_protocol_RailB.currentIndex())]
        update_database_with_temp_customer_input("E81", 89, 86, bin(int(railb_svid)).split("0b")[1].zfill(4))

    def railA_resolution_sel_onClicked(self):
        if self.radioButton_internal_registerA.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(False)
            update_database_with_temp_customer_input("E10", 25, 25, bin(1).split("0b")[1].zfill(1))
        elif self.radioButton_vid_sel_pinA.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railA.setDisabled(True)
            update_database_with_temp_customer_input("E10", 25, 25, bin(0).split("0b")[1].zfill(1))

    def railB_resolution_sel_onClicked(self):
        if self.radioButton_internal_registerB.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(False)
            update_database_with_temp_customer_input("E11", 25, 25, bin(1).split("0b")[1].zfill(1))
        elif self.radioButton_vid_sel_pinB.isChecked() == True:
            self.comboBox_vid_sel_pin_resolution_railB.setDisabled(True)
            update_database_with_temp_customer_input("E11", 25, 25, bin(0).split("0b")[1].zfill(1))

    def pmbus_override_railA(self):
        if self.checkBox_PMBus_override_railA.isChecked():
            self.comboBox_vid_resolution_RailA.setDisabled(False)
            update_database_with_temp_customer_input("010", 5, 4, "00")
            resolution_calculation(1)
        else:
            self.comboBox_vid_resolution_RailA.setDisabled(True)
            update_database_with_temp_customer_input("010", 5, 4, "11")
            resolution_calculation(1)

    def pmbus_override_railB(self):
        if self.checkBox_PMBus_override_railA_2.isChecked():
            self.comboBox_vid_resolution_RailB.setDisabled(False)
            update_database_with_temp_customer_input("011", 5, 4, "00")
            resolution_calculation(1)
        else:
            self.comboBox_vid_resolution_RailB.setDisabled(True)
            update_database_with_temp_customer_input("011", 5, 4, "11")
            resolution_calculation(1)

    def railA_vid_resolution(self):
        if (self.comboBox_vid_resolution_RailA.currentText() == '5mV'):
            update_database_with_temp_customer_input("E4", 3, 3, bin(0).split("0b")[1].zfill(1))
            update_database_with_temp_customer_input("200", 7, 0, bin(int(0x1F)).split("0b")[1].zfill(8))
            resolution_calculation(1)
            self.svid_vidmax_railA()
            self.svid_vidmax_railB()
            self.svid_pvid_railB_a0()
            self.svid_pvid_railB_a1()
            self.svid_pvid_railB_a2()
            self.svid_pvid_railB_a3()
            self.svid_pvid_railB_a4()
            self.svid_pvid_railB_a5()
            self.svid_pvid_railB_a6()
            self.svid_pvid_railB_a7()

        elif (self.comboBox_vid_resolution_RailA.currentText() == '10mV'):
            update_database_with_temp_customer_input("E4", 3, 3, bin(1).split("0b")[1].zfill(1))
            update_database_with_temp_customer_input("200", 7, 0, bin(int(0x1F)).split("0b")[1].zfill(8))
            resolution_calculation(1)
            self.svid_vidmax_railA()
            self.svid_vidmax_railB()
            self.svid_pvid_railB_a0()
            self.svid_pvid_railB_a1()
            self.svid_pvid_railB_a2()
            self.svid_pvid_railB_a3()
            self.svid_pvid_railB_a4()
            self.svid_pvid_railB_a5()
            self.svid_pvid_railB_a6()
            self.svid_pvid_railB_a7()
        else:
            update_database_with_temp_customer_input("200", 7, 0, bin(int(0x1E)).split("0b")[1].zfill(8))
            resolution_calculation(1)
            self.svid_vidmax_railA()
        self.svid_pvid_a0()
        self.svid_pvid_a1()
        self.svid_pvid_a2()
        self.svid_pvid_a3()
        self.svid_pvid_a4()
        self.svid_pvid_a5()
        self.svid_pvid_a6()
        self.svid_pvid_a7()

    def railB_vid_resolution(self):
        if (self.comboBox_vid_resolution_RailB.currentText() == '5mV'):
            update_database_with_temp_customer_input("E4", 3, 3, bin(0).split("0b")[1].zfill(1))
            update_database_with_temp_customer_input("201", 7, 0, bin(int(0x1F)).split("0b")[1].zfill(8))
            resolution_calculation(1)
            self.svid_vidmax_railB()
            self.svid_vidmax_railA()
            self.svid_pvid_a0()
            self.svid_pvid_a1()
            self.svid_pvid_a2()
            self.svid_pvid_a3()
            self.svid_pvid_a4()
            self.svid_pvid_a5()
            self.svid_pvid_a6()
            self.svid_pvid_a7()
        elif (self.comboBox_vid_resolution_RailB.currentText() == '10mV'):
            update_database_with_temp_customer_input("E4", 3, 3, bin(1).split("0b")[1].zfill(1))
            update_database_with_temp_customer_input("201", 7, 0, bin(int(0x1F)).split("0b")[1].zfill(8))
            resolution_calculation(1)
            self.svid_vidmax_railB()
            self.svid_vidmax_railA()
            self.svid_pvid_a0()
            self.svid_pvid_a1()
            self.svid_pvid_a2()
            self.svid_pvid_a3()
            self.svid_pvid_a4()
            self.svid_pvid_a5()
            self.svid_pvid_a6()
            self.svid_pvid_a7()
        else:
            update_database_with_temp_customer_input("201", 7, 0, bin(int(0x1E)).split("0b")[1].zfill(8))
            resolution_calculation(1)
            self.svid_vidmax_railB()
        self.svid_pvid_railB_a0()
        self.svid_pvid_railB_a1()
        self.svid_pvid_railB_a2()
        self.svid_pvid_railB_a3()
        self.svid_pvid_railB_a4()
        self.svid_pvid_railB_a5()
        self.svid_pvid_railB_a6()
        self.svid_pvid_railB_a7()

    def railA_slew_fast(self):
        raila_slew_fast = slew_rate[str(self.comboBox_slewrate_fast_RailA.currentText().replace("mV/μs", ""))]
        update_database_with_temp_customer_input("270", 15, 0, bin(int(raila_slew_fast)).split("0b")[1].zfill(16))

    def railB_slew_fast(self):
        railb_slew_fast = slew_rate[str(self.comboBox_slewrate_fast_RailB.currentText().replace("mV/μs", ""))]
        update_database_with_temp_customer_input("271", 15, 0, bin(int(railb_slew_fast)).split("0b")[1].zfill(16))

    def svid_save_register(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                resolution_calculation(0)
                print_log("ResolutionA is " + str(resolutionA), "INFO")
                initial_vid_raila = float(
                    ((int(initialize_feature_variable('E80', 31, 24), 2) * float(resolutionA)) + offset[
                        str(resolutionA)]) / 1000) if int(initialize_feature_variable('E80', 31, 24), 2) != 0 else int(0)
                print(initial_vid_raila)
                self.lineEdit_vid_max_railA.setText(str(initial_vid_raila))
                initial_vid_railb = float((int(initialize_feature_variable('E81', 31, 24), 2) * float(resolutionB) + offset[
                    str(resolutionB)]) / 1000) if int(initialize_feature_variable('E81', 31, 24), 2) != 0 else int(0)
                self.lineEdit_vid_max_railB.setText(str(initial_vid_railb))

                initial_pvid_0 = float((int(initialize_feature_variable('FC0', 7, 0), 2) * float(resolutionA) + offset[
                    str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 7, 0), 2) != 0 else int(0)
                self.lineEdit_PVID_A_0.setText(str(initial_pvid_0))
                self.lineEdit_PVID_A_1.setText(str(float((int(initialize_feature_variable('FC0', 15, 8), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 15, 8),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_2.setText(str(float((int(initialize_feature_variable('FC0', 23, 16), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 23, 16),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_3.setText(str(float((int(initialize_feature_variable('FC0', 31, 24), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 31, 24),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_4.setText(str(float((int(initialize_feature_variable('FC0', 39, 32), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 39, 32),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_5.setText(str(float((int(initialize_feature_variable('FC0', 47, 40), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 47, 40),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_6.setText(str(float((int(initialize_feature_variable('FC0', 55, 48), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 55, 48),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_A_7.setText(str(float((int(initialize_feature_variable('FC0', 63, 56), 2) * float(
                    resolutionA) + offset[str(resolutionA)]) / 1000) if int(initialize_feature_variable('FC0', 63, 56),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_0.setText(str(float((int(initialize_feature_variable('FC1', 7, 0), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 7, 0),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_1.setText(str(float((int(initialize_feature_variable('FC1', 15, 8), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 15, 8),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_2.setText(str(float((int(initialize_feature_variable('FC1', 23, 16), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 23, 16),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_3.setText(str(float((int(initialize_feature_variable('FC1', 31, 24), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 31, 24),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_4.setText(str(float((int(initialize_feature_variable('FC1', 39, 32), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 39, 32),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_5.setText(str(float((int(initialize_feature_variable('FC1', 47, 40), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 47, 40),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_6.setText(str(float((int(initialize_feature_variable('FC1', 55, 48), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 55, 48),
                                                                            2) != 0 else int(0)))
                self.lineEdit_PVID_B_7.setText(str(float((int(initialize_feature_variable('FC1', 63, 56), 2) * float(
                    resolutionB) + offset[str(resolutionB)]) / 1000) if int(initialize_feature_variable('FC1', 63, 56),
                                                                            2) != 0 else int(0)))
                print_log("SVI2 configuration settings are saved", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def svid_discard_changes(self):
        print_log("SVID manager Settings discarded.", "INFO")
        resolution_calculation(0)
        self.main = frame_svi2()
        self.main.show()
        self.close()


class PMBus_Configuration(QMainWindow, Ui_PMBus_configuration_main_window):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, resolutionA, resolutionB, list_of_registers_used_in_this_frame,PARTNAME
        super(PMBus_Configuration, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["EF", "E4", "010", "200", "220", "210", "250", "240", "2B0", "600",
                                                "640", "330", "260", "011", "201", "221", "211", "251", "241", "2B1",
                                                "601", "641", "331", "261"]

        # Feature related initialization on GUI display
        if int(initialize_feature_variable("010", 5, 4), 2) == 0:
            self.checkBox_PMBus_override.setChecked(True)
        else:
            self.checkBox_PMBus_override.setChecked(False)

        if int(initialize_feature_variable("011", 5, 4), 2) == 0:
            self.checkBox_PMBus_overrideB.setChecked(True)
        else:
            self.checkBox_PMBus_overrideB.setChecked(False)

        if int(initialize_feature_variable("EF", 7, 7), 2) == 1:
            self.comboBox_address_selection_mode.setCurrentIndex(1)
        else:
            self.comboBox_address_selection_mode.setCurrentIndex(1)

        if initialize_feature_variable("220", 9, 9) == "1":
            c = initialize_feature_variable("220", 9, 0)
            d = ''.join([str((int(i) ^ 1)) for i in c])
            d = int(d, base=2)
            d = (d + 1) * -1
        else:
            d = int(initialize_feature_variable("220", 15, 0), 2)
        if int(initialize_feature_variable("200", 4, 0), 2) == 30:
            self.comboBox_vout_mode_RailA.setCurrentIndex(1)
            self.lineEdit_vout_max_railA.setText(str(round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125, 5)))
            if float(self.lineEdit_vout_max_railA.text()) > 1.55:
                self.lineEdit_vout_max_railA.setText("1.55")
                update_database_with_temp_customer_input("240", 15, 0, "0000010011011000")
            self.lineEdit_vout_min_railA.setText(str(round((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125, 5)))
            self.lineEdit_vout_trim_railA.setText(str(round(d * 0.00625, 5)))
            self.lineEdit_vout_command_railA.setText(str(round((int(initialize_feature_variable("210", 15, 0), 2)) * 0.00625, 5)))
            self.lineEdit_vout_margin_high_railA.setText(
                str(round((int(initialize_feature_variable("250", 15, 0), 2)) * 0.00625, 5)))
            self.lineEdit_vout_margin_low_railA.setText(
                str(round((int(initialize_feature_variable("260", 15, 0), 2)) * 0.00625, 5)))
        else:
            if int(initialize_feature_variable("E4", 3, 3), 2) == 0:
                self.comboBox_vout_mode_RailA.setCurrentIndex(0)
                if int(initialize_feature_variable("240", 15, 0), 2) == 0:
                    self.lineEdit_vout_max_railA.setText(
                        str((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_max_railA.setText(
                        str(round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                    if float(self.lineEdit_vout_max_railA.text()) > 1.52:
                        self.lineEdit_vout_max_railA.setText("1.52")
                        update_database_with_temp_customer_input("240", 15, 0, "0000001111111100")
                if int(initialize_feature_variable("2B0", 15, 0), 2) == 0:
                    self.lineEdit_vout_min_railA.setText(
                        str((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_min_railA.setText(
                        str(round((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                self.lineEdit_vout_trim_railA.setText(str(round(d * 0.005, 3)))

                if int(initialize_feature_variable("210", 15, 0), 2) == 0:
                    self.lineEdit_vout_command_railA.setText(str((int(initialize_feature_variable("210", 15, 0), 2)) * 0.005))
                else:
                    self.lineEdit_vout_command_railA.setText(
                        str(round((int(initialize_feature_variable("210", 15, 0), 2)) * 0.005 + 0.245, 3)))

                if int(initialize_feature_variable("250", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_high_railA.setText(
                        str((int(initialize_feature_variable("250", 15, 0), 2)) * 0.005))
                else:
                    self.lineEdit_vout_margin_high_railA.setText(
                        str(round((int(initialize_feature_variable("250", 15, 0), 2)) * 0.005 + 0.245, 3)))

                if int(initialize_feature_variable("260", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_low_railA.setText(
                        str((int(initialize_feature_variable("260", 15, 0), 2)) * 0.005))
                else:
                    self.lineEdit_vout_margin_low_railA.setText(
                        str(round((int(initialize_feature_variable("260", 15, 0), 2)) * 0.005 + 0.245, 3)))

            else:
                self.comboBox_vout_mode_RailA.setCurrentIndex(2)
                if int(initialize_feature_variable("240", 15, 0), 2) == 0:
                    self.lineEdit_vout_max_railA.setText(
                        str((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_max_railA.setText(
                        str(round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                    if float(self.lineEdit_vout_max_railA.text()) > 2.74:
                        self.lineEdit_vout_max_railA.setText("2.74")
                        update_database_with_temp_customer_input("240", 15, 0, "0000001111111000")
                if int(initialize_feature_variable("2B0", 15, 0), 2) == 0:
                    self.lineEdit_vout_min_railA.setText(
                        str((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_min_railA.setText(
                        str(round((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                self.lineEdit_vout_trim_railA.setText(str(round(d * 0.01, 2)))
                if int(initialize_feature_variable("210", 15, 0), 2) == 0:
                    self.lineEdit_vout_command_railA.setText(
                        str((int(initialize_feature_variable("210", 15, 0), 2)) * 0.01))
                else:
                    self.lineEdit_vout_command_railA.setText(
                        str(round((int(initialize_feature_variable("210", 15, 0), 2)) * 0.01 + 0.190, 2)))
                if int(initialize_feature_variable("250", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_high_railA.setText(
                        str((int(initialize_feature_variable("250", 15, 0), 2)) * 0.01))
                else:
                    self.lineEdit_vout_margin_high_railA.setText(
                        str(round((int(initialize_feature_variable("250", 15, 0), 2)) * 0.01 + 0.190, 2)))
                if int(initialize_feature_variable("260", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_low_railA.setText(
                        str((int(initialize_feature_variable("260", 15, 0), 2)) * 0.01))
                else:
                    self.lineEdit_vout_margin_low_railA.setText(
                        str(round((int(initialize_feature_variable("260", 15, 0), 2)) * 0.01 + 0.190, 2)))

        if initialize_feature_variable("221", 9, 9) == "1":
            c = initialize_feature_variable("221", 9, 0)
            d = ''.join([str((int(i) ^ 1)) for i in c])
            d = int(d, base=2)
            d = (d + 1) * -1
        else:
            d = int(initialize_feature_variable("221", 15, 0), 2)
        if int(initialize_feature_variable("201", 4, 0), 2) == 30:
            self.comboBox_vout_mode_RailB.setCurrentIndex(1)
            self.lineEdit_vout_max_railB.setText(str(round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125, 5)))
            if float(self.lineEdit_vout_max_railB.text()) > 1.55:
                self.lineEdit_vout_max_railB.setText("1.55")
                update_database_with_temp_customer_input("241", 15, 0, "0000010011011000")
            self.lineEdit_vout_min_railB.setText(str(round((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125, 5)))
            self.lineEdit_vout_command_railB.setText(str(round((int(initialize_feature_variable("211", 15, 0), 2)) * 0.00625, 5)))
            self.lineEdit_vout_trim_railB.setText(str(round(d * 0.00625, 5)))
            self.lineEdit_vout_margin_high_railB.setText(
                str(round((int(initialize_feature_variable("251", 15, 0), 2)) * 0.00625, 5)))
            self.lineEdit_vout_margin_low_railB.setText(
                str(round((int(initialize_feature_variable("261", 15, 0), 2)) * 0.00625, 5)))
        else:
            if int(initialize_feature_variable("E4", 3, 3), 2) == 0:
                self.comboBox_vout_mode_RailB.setCurrentIndex(0)
                if int(initialize_feature_variable("241", 15, 0), 2) == 0:
                    self.lineEdit_vout_max_railB.setText(
                        str((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_max_railB.setText(
                        str(round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                    if float(self.lineEdit_vout_max_railB.text()) > 1.52:
                        self.lineEdit_vout_max_railB.setText("1.52")
                        update_database_with_temp_customer_input("241", 15, 0, "0000001111111100")
                if int(initialize_feature_variable("2B1", 15, 0), 2) == 0:
                    self.lineEdit_vout_min_railB.setText(
                        str((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_min_railB.setText(
                        str(round((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                if int(initialize_feature_variable("211", 15, 0), 2) == 0:
                    self.lineEdit_vout_command_railB.setText(
                        str((int(initialize_feature_variable("211", 15, 0), 2)) * 0.005))
                else:
                    self.lineEdit_vout_command_railB.setText(
                        str(round((int(initialize_feature_variable("211", 15, 0), 2)) * 0.005 + 0.245, 3)))
                self.lineEdit_vout_trim_railB.setText(str(round(d * 0.005, 3)))
                if int(initialize_feature_variable("251", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_high_railB.setText(
                        str((int(initialize_feature_variable("251", 15, 0), 2)) * 0.005))
                else:
                    self.lineEdit_vout_margin_high_railB.setText(
                        str(round((int(initialize_feature_variable("251", 15, 0), 2)) * 0.005 + 0.245, 3)))
                if int(initialize_feature_variable("261", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_low_railB.setText(
                        str((int(initialize_feature_variable("261", 15, 0), 2)) * 0.005))
                else:
                    self.lineEdit_vout_margin_low_railB.setText(
                        str(round((int(initialize_feature_variable("261", 15, 0), 2)) * 0.005 + 0.245, 3)))

            else:
                self.comboBox_vout_mode_RailB.setCurrentIndex(2)
                if int(initialize_feature_variable("241", 15, 0), 2) == 0:
                    self.lineEdit_vout_max_railB.setText(
                        str((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_max_railB.setText(
                        str(round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                    if float(self.lineEdit_vout_max_railB.text()) > 2.74:
                        self.lineEdit_vout_max_railB.setText("2.74")
                        update_database_with_temp_customer_input("241", 15, 0, "0000001111111000")
                if int(initialize_feature_variable("2B1", 15, 0), 2) == 0:
                    self.lineEdit_vout_min_railB.setText(
                        str((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125))
                else:
                    self.lineEdit_vout_min_railB.setText(
                        str(round((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                if int(initialize_feature_variable("211", 15, 0), 2) == 0:
                    self.lineEdit_vout_command_railB.setText(
                        str((int(initialize_feature_variable("211", 15, 0), 2)) * 0.01))
                else:
                    self.lineEdit_vout_command_railB.setText(
                        str(round((int(initialize_feature_variable("211", 15, 0), 2)) * 0.01 + 0.190, 2)))
                self.lineEdit_vout_trim_railB.setText(str(round(d * 0.01, 3)))
                if int(initialize_feature_variable("251", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_high_railB.setText(
                        str((int(initialize_feature_variable("251", 15, 0), 2)) * 0.01))
                else:
                    self.lineEdit_vout_margin_high_railB.setText(
                        str(round((int(initialize_feature_variable("251", 15, 0), 2)) * 0.01 + 0.190, 2)))
                if int(initialize_feature_variable("261", 15, 0), 2) == 0:
                    self.lineEdit_vout_margin_low_railB.setText(
                        str((int(initialize_feature_variable("261", 15, 0), 2)) * 0.01))
                else:
                    self.lineEdit_vout_margin_low_railB.setText(
                        str(round((int(initialize_feature_variable("261", 15, 0), 2)) * 0.01 + 0.190, 2)))

        if initialize_feature_variable("600", 15, 0) == "0000000000000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(0)
        elif initialize_feature_variable("600", 15, 0) == "1011001000000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(1)
        elif initialize_feature_variable("600", 15, 0) == "1011101000000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(2)
        elif initialize_feature_variable("600", 15, 0) == "1011101100000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(3)
        elif initialize_feature_variable("600", 15, 0) == "1100001000000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(4)
        elif initialize_feature_variable("600", 15, 0) == "1100001100000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(5)
        elif initialize_feature_variable("600", 15, 0) == "1100101000000000":
            self.comboBox_ton_delay_RailA.setCurrentIndex(6)
        else:
            self.comboBox_ton_delay_RailA.setCurrentIndex(7)

        if initialize_feature_variable("601", 15, 0) == "0000000000000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(0)
        elif initialize_feature_variable("601", 15, 0) == "1011001000000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(1)
        elif initialize_feature_variable("601", 15, 0) == "1011101000000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(2)
        elif initialize_feature_variable("601", 15, 0) == "1011101100000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(3)
        elif initialize_feature_variable("601", 15, 0) == "1100001000000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(4)
        elif initialize_feature_variable("601", 15, 0) == "1100001100000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(5)
        elif initialize_feature_variable("601", 15, 0) == "1100101000000000":
            self.comboBox_ton_delay_RailB.setCurrentIndex(6)
        else:
            self.comboBox_ton_delay_RailB.setCurrentIndex(7)

        if initialize_feature_variable("640", 15, 0) == "0000000000000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(0)
        elif initialize_feature_variable("640", 15, 0) == "1011001000000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(1)
        elif initialize_feature_variable("640", 15, 0) == "1011101000000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(2)
        elif initialize_feature_variable("640", 15, 0) == "1011101100000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(3)
        elif initialize_feature_variable("640", 15, 0) == "1100001000000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(4)
        elif initialize_feature_variable("640", 15, 0) == "1100001100000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(5)
        elif initialize_feature_variable("640", 15, 0) == "1100101000000000":
            self.comboBox_toff_delay_RailA.setCurrentIndex(6)
        else:
            self.comboBox_toff_delay_RailA.setCurrentIndex(7)

        if initialize_feature_variable("641", 15, 0) == "0000000000000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(0)
        elif initialize_feature_variable("641", 15, 0) == "1011001000000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(1)
        elif initialize_feature_variable("641", 15, 0) == "1011101000000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(2)
        elif initialize_feature_variable("641", 15, 0) == "1011101100000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(3)
        elif initialize_feature_variable("641", 15, 0) == "1100001000000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(4)
        elif initialize_feature_variable("641", 15, 0) == "1100001100000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(5)
        elif initialize_feature_variable("641", 15, 0) == "1100101000000000":
            self.comboBox_toff_delay_RailB.setCurrentIndex(6)
        else:
            self.comboBox_toff_delay_RailB.setCurrentIndex(7)

        if initialize_feature_variable("331", 15, 0) == "1010101000000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(0)
        elif initialize_feature_variable("331", 15, 0) == "1011001000000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(1)
        elif initialize_feature_variable("331", 15, 0) == "1011001100000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(2)
        elif initialize_feature_variable("331", 15, 0) == "1011101000000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(3)
        elif initialize_feature_variable("331", 15, 0) == "1011101010000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(4)
        elif initialize_feature_variable("331", 15, 0) == "1011101100000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(5)
        elif initialize_feature_variable("331", 15, 0) == "1011101110000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(6)
        elif initialize_feature_variable("331", 15, 0) == "1100001000000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(7)
        elif initialize_feature_variable("331", 15, 0) == "1100001001000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(8)
        elif initialize_feature_variable("331", 15, 0) == "1100001010000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(9)
        elif initialize_feature_variable("331", 15, 0) == "1100001011000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(10)
        elif initialize_feature_variable("331", 15, 0) == "1100001100000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(11)
        elif initialize_feature_variable("331", 15, 0) == "1100001101000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(12)
        elif initialize_feature_variable("331", 15, 0) == "1100001110000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(13)
        elif initialize_feature_variable("331", 15, 0) == "1100001111000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(14)
        elif initialize_feature_variable("331", 15, 0) == "1100101000000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(15)
        elif initialize_feature_variable("331", 15, 0) == "1100101000100000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(16)
        elif initialize_feature_variable("331", 15, 0) == "1100101001000000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(17)
        elif initialize_feature_variable("331", 15, 0) == "1100101001100000":
            self.comboBox_freq_switch_RailB.setCurrentIndex(18)
        else:
            self.comboBox_freq_switch_RailB.setCurrentIndex(19)

        if initialize_feature_variable("330", 15, 0) == "1010101000000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(0)
        elif initialize_feature_variable("330", 15, 0) == "1011001000000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(1)
        elif initialize_feature_variable("330", 15, 0) == "1011001100000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(2)
        elif initialize_feature_variable("330", 15, 0) == "1011101000000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(3)
        elif initialize_feature_variable("330", 15, 0) == "1011101010000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(4)
        elif initialize_feature_variable("330", 15, 0) == "1011101100000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(5)
        elif initialize_feature_variable("330", 15, 0) == "1011101110000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(6)
        elif initialize_feature_variable("330", 15, 0) == "1100001000000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(7)
        elif initialize_feature_variable("330", 15, 0) == "1100001001000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(8)
        elif initialize_feature_variable("330", 15, 0) == "1100001010000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(9)
        elif initialize_feature_variable("330", 15, 0) == "1100001011000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(10)
        elif initialize_feature_variable("330", 15, 0) == "1100001100000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(11)
        elif initialize_feature_variable("330", 15, 0) == "1100001101000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(12)
        elif initialize_feature_variable("330", 15, 0) == "1100001110000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(13)
        elif initialize_feature_variable("330", 15, 0) == "1100001111000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(14)
        elif initialize_feature_variable("330", 15, 0) == "1100101000000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(15)
        elif initialize_feature_variable("330", 15, 0) == "1100101000100000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(16)
        elif initialize_feature_variable("330", 15, 0) == "1100101001000000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(17)
        elif initialize_feature_variable("330", 15, 0) == "1100101001100000":
            self.comboBox_freq_switch_RailA.setCurrentIndex(18)
        else:
            self.comboBox_freq_switch_RailA.setCurrentIndex(19)

        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        # Customer GUI interaction related function mapping
        self.pushButton_Discard.clicked.connect(self.Discard)
        self.pushButton_Save.clicked.connect(self.Save)
        self.checkBox_PMBus_override.stateChanged.connect(self.PMBus_override_changed)
        self.checkBox_PMBus_overrideB.stateChanged.connect(self.PMBus_override_changed_B)

        self.comboBox_address_selection_mode.activated.connect(self.address_selection_mode_changed)
        self.comboBox_vout_mode_RailA.activated.connect(self.vout_mode_change_A)
        self.lineEdit_vout_trim_railA.textChanged.connect(self.vout_trim_changed_A)
        self.lineEdit_vout_command_railA.textChanged.connect(self.vout_command_changed_A)
        self.lineEdit_vout_margin_high_railA.textChanged.connect(self.vout_margin_high_changed_A)
        self.lineEdit_vout_margin_low_railA.textChanged.connect(self.vout_margin_low_changed_A)
        self.lineEdit_vout_max_railA.textChanged.connect(self.vout_max_changed_A)
        self.lineEdit_vout_min_railA.textChanged.connect(self.vout_min_changed_A)
        self.comboBox_ton_delay_RailA.activated.connect(self.ton_delay_changed_A)
        self.comboBox_toff_delay_RailA.activated.connect(self.toff_delay_changed_A)
        self.comboBox_freq_switch_RailA.activated.connect(self.freq_switch_changed_A)

        self.lineEdit_vout_trim_railB.textChanged.connect(self.vout_trim_changed_B)
        self.comboBox_vout_mode_RailB.activated.connect(self.vout_mode_change_B)
        self.lineEdit_vout_command_railB.textChanged.connect(self.vout_command_changed_B)
        self.lineEdit_vout_margin_high_railB.textChanged.connect(self.vout_margin_high_changed_B)
        self.lineEdit_vout_margin_low_railB.textChanged.connect(self.vout_margin_low_changed_B)
        self.lineEdit_vout_max_railB.textChanged.connect(self.vout_max_changed_B)
        self.lineEdit_vout_min_railB.textChanged.connect(self.vout_min_changed_B)
        self.comboBox_ton_delay_RailB.activated.connect(self.ton_delay_changed_B)
        self.comboBox_toff_delay_RailB.activated.connect(self.toff_delay_changed_B)
        self.comboBox_freq_switch_RailB.activated.connect(self.freq_switch_changed_B)

    def PMBus_override_changed(self):
        if self.checkBox_PMBus_override.isChecked():
            update_database_with_temp_customer_input("010", 5, 4, "00")
        else:
            update_database_with_temp_customer_input("010", 5, 4, "11")

    def PMBus_override_changed_B(self):
        if self.checkBox_PMBus_overrideB.isChecked():
            update_database_with_temp_customer_input("011", 5, 4, "00")
        else:
            update_database_with_temp_customer_input("011", 5, 4, "11")

    def address_selection_mode_changed(self):
        if self.comboBox_address_selection_mode.currentIndex() == 0:
            update_database_with_temp_customer_input("EF", 7, 7, "0")
        else:
            update_database_with_temp_customer_input("EF", 7, 7, "1")

    def vout_mode_change_A(self):
        if self.comboBox_vout_mode_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("200", 4, 0, "11111")
            update_database_with_temp_customer_input("E4", 3, 3, "0")
            if float(self.lineEdit_vout_max_railA.text()) > 1.52:
                self.lineEdit_vout_max_railA.setText("1.52")
                update_database_with_temp_customer_input("240", 15, 0, "0000001111111100")
            if float(self.lineEdit_vout_trim_railA.text()) < -1.28:
                self.lineEdit_vout_trim_railA.setText("-1.28")
                update_database_with_temp_customer_input("220", 15, 0, "0000001100000000")
            elif float(self.lineEdit_vout_trim_railA.text()) > 1.275:
                self.lineEdit_vout_trim_railA.setText("1.275")
                update_database_with_temp_customer_input("220", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_min_railA.text()) > 1.52:
                self.lineEdit_vout_min_railA.setText("1.52")
                update_database_with_temp_customer_input("2B0", 15, 0, "0000001111111100")
            if float(self.lineEdit_vout_command_railA.text()) > 1.52:
                self.lineEdit_vout_command_railA.setText("1.52")
                update_database_with_temp_customer_input("210", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_margin_high_railA.text()) > 1.52:
                self.lineEdit_vout_margin_high_railA.setText("1.52")
                update_database_with_temp_customer_input("250", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_margin_low_railA.text()) > 1.52:
                self.lineEdit_vout_margin_low_railA.setText("1.52")
                update_database_with_temp_customer_input("260", 15, 0, "0000000011111111")

        elif self.comboBox_vout_mode_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("200", 4, 0, "11110")
            if float(self.lineEdit_vout_max_railA.text()) > 1.55:
                self.lineEdit_vout_max_railA.setText("1.55")
                update_database_with_temp_customer_input("240", 15, 0, "0000010011011000")
            if float(self.lineEdit_vout_trim_railA.text()) < -1.6:
                self.lineEdit_vout_trim_railA.setText("-1.6")
                update_database_with_temp_customer_input("220", 15, 0, "0000001100000000")
            elif float(self.lineEdit_vout_trim_railA.text()) > 1.59375:
                self.lineEdit_vout_trim_railA.setText("1.59375")
                update_database_with_temp_customer_input("220", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_min_railA.text()) > 1.55:
                self.lineEdit_vout_min_railA.setText("1.55")
                update_database_with_temp_customer_input("2B0", 15, 0, "0000010011011000")
            if float(self.lineEdit_vout_command_railA.text()) > 1.55:
                self.lineEdit_vout_command_railA.setText("1.55")
                update_database_with_temp_customer_input("210", 15, 0, "0000000011111000")
            if float(self.lineEdit_vout_margin_high_railA.text()) > 1.55:
                self.lineEdit_vout_margin_high_railA.setText("1.55")
                update_database_with_temp_customer_input("250", 15, 0, "0000000011111000")
            if float(self.lineEdit_vout_margin_low_railA.text()) > 1.55:
                self.lineEdit_vout_margin_low_railA.setText("1.55")
                update_database_with_temp_customer_input("260", 15, 0, "0000000011111000")

        else:
            update_database_with_temp_customer_input("200", 4, 0, "11111")
            update_database_with_temp_customer_input("E4", 3, 3, "1")
            if float(self.lineEdit_vout_max_railA.text()) > 2.74:
                self.lineEdit_vout_max_railA.setText("2.74")
                update_database_with_temp_customer_input("240", 15, 0, "0000011111111000")

        self.vout_max_changed_A()
        self.vout_min_changed_A()
        self.vout_command_changed_A()
        self.vout_margin_high_changed_A()
        self.vout_margin_low_changed_A()
        self.vout_trim_changed_A()

    def vout_mode_change_B(self):
        if self.comboBox_vout_mode_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("201", 4, 0, "11111")
            update_database_with_temp_customer_input("E4", 3, 3, "0")
            if float(self.lineEdit_vout_max_railB.text()) > 1.52:
                self.lineEdit_vout_max_railB.setText("1.52")
                update_database_with_temp_customer_input("241", 15, 0, "0000001111111100")
            if float(self.lineEdit_vout_trim_railB.text()) < -1.28:
                self.lineEdit_vout_trim_railB.setText("-1.28")
                update_database_with_temp_customer_input("221", 15, 0, "0000001100000000")
            elif float(self.lineEdit_vout_trim_railB.text()) > 1.275:
                self.lineEdit_vout_trim_railB.setText("1.275")
                update_database_with_temp_customer_input("221", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_min_railB.text()) > 1.52:
                self.lineEdit_vout_min_railB.setText("1.52")
                update_database_with_temp_customer_input("2B1", 15, 0, "0000001111111100")
            if float(self.lineEdit_vout_command_railB.text()) > 1.52:
                self.lineEdit_vout_command_railB.setText("1.52")
                update_database_with_temp_customer_input("211", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_margin_high_railB.text()) > 1.52:
                self.lineEdit_vout_margin_high_railB.setText("1.52")
                update_database_with_temp_customer_input("251", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_margin_low_railB.text()) > 1.52:
                self.lineEdit_vout_margin_low_railB.setText("1.52")
                update_database_with_temp_customer_input("261", 15, 0, "0000000011111111")

        elif self.comboBox_vout_mode_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("201", 4, 0, "11110")
            if float(self.lineEdit_vout_max_railB.text()) > 1.55:
                self.lineEdit_vout_max_railB.setText("1.55")
                update_database_with_temp_customer_input("241", 15, 0, "0000010011011000")
            if float(self.lineEdit_vout_trim_railB.text()) < -1.6:
                self.lineEdit_vout_trim_railB.setText("-1.6")
                update_database_with_temp_customer_input("221", 15, 0, "0000001100000000")
            elif float(self.lineEdit_vout_trim_railB.text()) > 1.59375:
                self.lineEdit_vout_trim_railB.setText("1.59375")
                update_database_with_temp_customer_input("221", 15, 0, "0000000011111111")
            if float(self.lineEdit_vout_min_railB.text()) > 1.55:
                self.lineEdit_vout_min_railB.setText("1.55")
                update_database_with_temp_customer_input("2B1", 15, 0, "0000010011011000")
            if float(self.lineEdit_vout_command_railB.text()) > 1.55:
                self.lineEdit_vout_command_railB.setText("1.55")
                update_database_with_temp_customer_input("211", 15, 0, "0000000011111000")
            if float(self.lineEdit_vout_margin_high_railB.text()) > 1.55:
                self.lineEdit_vout_margin_high_railB.setText("1.55")
                update_database_with_temp_customer_input("251", 15, 0, "0000000011111000")
            if float(self.lineEdit_vout_margin_low_railB.text()) > 1.55:
                self.lineEdit_vout_margin_low_railB.setText("1.55")
                update_database_with_temp_customer_input("261", 15, 0, "0000000011111000")

        else:
            print("Vout mode", self.comboBox_vout_mode_RailB.currentText())
            update_database_with_temp_customer_input("201", 4, 0, "11111")
            update_database_with_temp_customer_input("E4", 3, 3, "1")
            if float(self.lineEdit_vout_max_railB.text()) > 2.74:
                self.lineEdit_vout_max_railB.setText("2.74")
                update_database_with_temp_customer_input("241", 15, 0, "0000011111111000")

        self.vout_max_changed_B()
        self.vout_min_changed_B()
        self.vout_command_changed_B()
        self.vout_margin_high_changed_B()
        self.vout_margin_low_changed_B()
        self.vout_trim_changed_B()

    def vout_trim_changed_A(self):
        if self.lineEdit_vout_trim_railA.text():
            if "-" in self.lineEdit_vout_trim_railA.text():
                b = self.lineEdit_vout_trim_railA.text().replace("-", "")
                if len(b) == 0:
                    return
                else:
                    value = "-" + b
            else:
                value = self.lineEdit_vout_trim_railA.text()
            try:
                if isinstance(float(self.lineEdit_vout_trim_railA.text()), float):
                    if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                        value = round(float(value), 3)
                        if -1.28 <= value <= 1.275:
                            mod = round((value % 0.005), 6)
                            if mod > 0.002499:
                                value += round((0.005 - mod), 6)
                            else:
                                value = round((value - mod), 6)
                            value = int(round(value / 0.005))
                            update_database_with_temp_customer_input("220", 15, 0, bin(int(hex(int((value + (1 << 10)) %
                                                                                                   (1 << 10))).split(
                                "0x")[1],
                                                                                           16)).split("0b")[1].zfill(
                                16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_trim_railA.text() + ", give vout trim ranging (-1.28V to 1.275V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_trim_railA.setText("0.00")
                            update_database_with_temp_customer_input("220", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailA.currentIndex() == 1:
                        value = round(float(self.lineEdit_vout_trim_railA.text()), 5)
                        if -1.6 <= value <= 1.59375:
                            mod = round((value % 0.00625), 6)
                            if mod > 0.003125:
                                value += round((0.00625 - mod), 6)
                            else:
                                value = round((value - mod), 6)
                            value = int(round(value / 0.00625))
                            update_database_with_temp_customer_input("220", 15, 0, bin(int(hex(int((value + (1 << 10)) %
                                                                                                   (1 << 10))).split(
                                "0x")[1],
                                                                                           16)).split("0b")[1].zfill(
                                16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_trim_railA.text() + ", give vout trim ranging (-1.6V to 1.59375V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_trim_railA.setText("0.00")
                            update_database_with_temp_customer_input("220", 15, 0, "0000000000000000")

                    else:
                        value = round(float(self.lineEdit_vout_trim_railA.text()), 2)
                        if -2.56 <= value <= 2.55:
                            mod = round((value % 0.01), 6)
                            if mod > 0.004999:
                                value += round((0.01 - mod), 6)
                            else:
                                value = round((value - mod), 6)
                            value = int(round(value / 0.01))
                            update_database_with_temp_customer_input("220", 15, 0, bin(int(hex(int((value + (1 << 10)) %
                                                                                                   (1 << 10))).split(
                                "0x")[1],
                                                                                           16)).split("0b")[1].zfill(
                                16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_trim_railA.text() + ", give vout trim ranging (-2.56V to 2.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_trim_railA.setText("0.00")
                            update_database_with_temp_customer_input("220", 15, 0, "0000000000000000")

            except:
                print_log("Entered " + self.lineEdit_vout_trim_railA.text() + ", give valid float", "WARNING")
                time.sleep(2)
                self.lineEdit_vout_trim_railA.setText("0.00")
                return

    def vout_trim_changed_B(self):
        if self.lineEdit_vout_trim_railB.text():
            if "-" in self.lineEdit_vout_trim_railB.text():
                b = self.lineEdit_vout_trim_railB.text().replace("-", "")
                if len(b) == 0:
                    return
                else:
                    value = "-" + b
            else:
                value = self.lineEdit_vout_trim_railB.text()
            try:
                if isinstance(float(self.lineEdit_vout_trim_railB.text()), float):
                    if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                        value = round(float(value), 3)
                        if -1.28 <= value <= 1.275:
                            mod = round((value % 0.005), 6)
                            if mod > 0.002499:
                                value += round((0.005 - mod), 6)
                            else:
                                value = round((value - mod), 6)
                            value = int(round(value / 0.005))
                            update_database_with_temp_customer_input("221", 15, 0, bin(int(hex(int((value + (1 << 10)) %
                                                                                                   (1 << 10))).split(
                                "0x")[1],
                                                                                           16)).split("0b")[1].zfill(
                                16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_trim_railB.text() + ", give vout trim ranging (-1.28V to 1.275V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_trim_railB.setText("0.00")
                            update_database_with_temp_customer_input("221", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailB.currentIndex() == 1:
                        value = round(float(self.lineEdit_vout_trim_railB.text()), 5)
                        if -1.6 <= value <= 1.59375:
                            mod = round((value % 0.00625), 6)
                            if mod > 0.003125:
                                value += round((0.00625 - mod), 6)
                            else:
                                value = round((value - mod), 6)
                            value = int(round(value / 0.00625))
                            update_database_with_temp_customer_input("221", 15, 0, bin(int(hex(int((value + (1 << 10)) %
                                                                                                   (1 << 10))).split(
                                "0x")[1],
                                                                                           16)).split("0b")[1].zfill(
                                16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_trim_railB.text() + ", give vout trim ranging (-1.6V to 1.59375V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_trim_railB.setText("0.00")
                            update_database_with_temp_customer_input("221", 15, 0, "0000000000000000")

                    else:
                        value = round(float(self.lineEdit_vout_trim_railB.text()), 2)
                        if -2.56 <= value <= 2.55:
                            mod = round((value % 0.01), 6)
                            if mod > 0.004999:
                                value += round((0.01 - mod), 6)
                            else:
                                value = round((value - mod), 6)
                            value = int(round(value / 0.01))
                            update_database_with_temp_customer_input("221", 15, 0, bin(int(hex(int((value + (1 << 10)) %
                                                                                                   (1 << 10))).split(
                                "0x")[1],
                                                                                           16)).split("0b")[1].zfill(
                                16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_trim_railB.text() + ", give vout trim ranging (-2.56V to 2.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_trim_railB.setText("0.00")
                            update_database_with_temp_customer_input("221", 15, 0, "0000000000000000")

            except:
                print_log("Entered " + self.lineEdit_vout_trim_railB.text() + ", give valid float", "WARNING")
                time.sleep(2)
                self.lineEdit_vout_trim_railB.setText("0.00")
                return

    def vout_command_changed_A(self):
        if self.lineEdit_vout_command_railA.text():
            if self.lineEdit_vout_command_railA.text()[0] == ".":
                self.lineEdit_vout_command_railA.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_command_railA.text()), float):
                    val_1 = float(self.lineEdit_vout_command_railA.text())
                    if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("210", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            val_1 -= 0.245
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("210", 15, 0,
                                                                     bin(round(val_1 / 0.005)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_command_railA.text() + ", give VID ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_command_railA.setText("0.00")
                            update_database_with_temp_customer_input("210", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailA.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("210", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            val_1 -= 0.19
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("210", 15, 0,
                                                                     bin(round(val_1 / 0.01)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_command_railA.text() + ", give VID ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_command_railA.setText("0.00")
                            update_database_with_temp_customer_input("210", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("210", 15, 0,
                                                                     bin(round(val_1 / 0.00625)).split("0b")[1].zfill(
                                                                         16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_command_railA.text() + ", give VID ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_command_railA.setText("0.00")
                            update_database_with_temp_customer_input("210", 15, 0, "0000000000000000")

            except:
                print_log("Entered " + self.lineEdit_vout_command_railA.text() + ", give valid float VID", "WARNING")
                time.sleep(2)
                self.lineEdit_vout_command_railA.setText("0.00")
                return

    def vout_command_changed_B(self):
        if self.lineEdit_vout_command_railB.text():
            if self.lineEdit_vout_command_railB.text()[0] == ".":
                self.lineEdit_vout_command_railB.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_command_railB.text()), float):
                    val_1 = float(self.lineEdit_vout_command_railB.text())
                    if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("211", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            val_1 -= 0.245
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("211", 15, 0,
                                                                     bin(round(val_1 / 0.005)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_command_railB.text() + ", give VID ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_command_railB.setText("0.00")
                            update_database_with_temp_customer_input("211", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailB.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("211", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            val_1 -= 0.19
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("211", 15, 0,
                                                                     bin(round(val_1 / 0.01)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_command_railB.text() + ", give VID ranging (200mV - 2.74V or 0V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_command_railB.setText("0.00")
                            update_database_with_temp_customer_input("211", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("211", 15, 0,
                                                                     bin(round(val_1 / 0.00625)).split("0b")[1].zfill(
                                                                         16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_command_railB.text() + ", give VID ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_command_railB.setText("0.00")
                            update_database_with_temp_customer_input("211", 15, 0, "0000000000000000")

            except:
                print_log("Entered " + self.lineEdit_vout_command_railB.text() + ", give valid float VID", "WARNING")
                time.sleep(2)
                self.lineEdit_vout_command_railB.setText("0.00")
                return

    def vout_margin_high_changed_A(self):
        if self.lineEdit_vout_margin_high_railA.text():
            if self.lineEdit_vout_margin_high_railA.text()[0] == ".":
                self.lineEdit_vout_margin_high_railA.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_margin_high_railA.text()), float):
                    val_1 = float(self.lineEdit_vout_margin_high_railA.text())
                    if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("250", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            val_1 -= 0.245
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("250", 15, 0,
                                                                     bin(round(val_1 / 0.005)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_high_railA.text() + ", give Vout margin high ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_high_railA.setText("0.00")
                            update_database_with_temp_customer_input("250", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailA.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("250", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            val_1 -= 0.19
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("250", 15, 0,
                                                                     bin(round(val_1 / 0.01)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_high_railA.text() + ", give Vout margin high ranging (200mV - 2.74V or 0V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_high_railA.setText("0.00")
                            update_database_with_temp_customer_input("250", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("250", 15, 0,
                                                                     bin(round(val_1 / 0.00625)).split("0b")[1].zfill(
                                                                         16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_high_railA.text() + ", give Vout margin high ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_high_railA.setText("0.00")
                            update_database_with_temp_customer_input("250", 15, 0, "0000000000000000")

            except:
                print_log(
                    "Entered " + self.lineEdit_vout_margin_high_railA.text() + ", give valid float for Vout margin high",
                    "WARNING")
                time.sleep(2)
                self.lineEdit_vout_margin_high_railA.setText("0.00")
                return

    def vout_margin_high_changed_B(self):
        if self.lineEdit_vout_margin_high_railB.text():
            if self.lineEdit_vout_margin_high_railB.text()[0] == ".":
                self.lineEdit_vout_margin_high_railB.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_margin_high_railB.text()), float):
                    val_1 = float(self.lineEdit_vout_margin_high_railB.text())
                    if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("251", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            val_1 -= 0.245
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("251", 15, 0,
                                                                     bin(round(val_1 / 0.005)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_high_railB.text() + ", give Vout margin high ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_high_railB.setText("0.00")
                            update_database_with_temp_customer_input("251", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailB.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("251", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            val_1 -= 0.19
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("251", 15, 0,
                                                                     bin(round(val_1 / 0.01)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_high_railB.text() + ", give Vout margin high ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_high_railB.setText("0.00")
                            update_database_with_temp_customer_input("251", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("251", 15, 0,
                                                                     bin(round(val_1 / 0.00625)).split("0b")[1].zfill(
                                                                         16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_high_railB.text() + ", give Vout margin high ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_high_railB.setText("0.00")
                            update_database_with_temp_customer_input("251", 15, 0, "0000000000000000")

            except:
                print_log(
                    "Entered " + self.lineEdit_vout_margin_high_railB.text() + ", give valid float for Vout margin high",
                    "WARNING")
                time.sleep(2)
                self.lineEdit_vout_margin_high_railB.setText("0.00")
                return

    def vout_margin_low_changed_A(self):
        if self.lineEdit_vout_margin_low_railA.text():
            if self.lineEdit_vout_margin_low_railA.text()[0] == ".":
                self.lineEdit_vout_margin_low_railA.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_margin_low_railA.text()), float):
                    val_1 = float(self.lineEdit_vout_margin_low_railA.text())
                    if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("260", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            val_1 -= 0.245
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("260", 15, 0,
                                                                     bin(round(val_1 / 0.005)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_low_railA.text() + ", give Vout margin low ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_low_railA.setText("0.00")
                            update_database_with_temp_customer_input("260", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailA.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("260", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            val_1 -= 0.19
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("260", 15, 0,
                                                                     bin(round(val_1 / 0.01)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_low_railA.text() + ", give Vout margin low ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_low_railA.setText("0.00")
                            update_database_with_temp_customer_input("260", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("260", 15, 0,
                                                                     bin(round(val_1 / 0.00625)).split("0b")[1].zfill(
                                                                         16))

                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_low_railA.text() + ", give Vout margin low ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_low_railA.setText("0.00")
                            update_database_with_temp_customer_input("260", 15, 0, "0000000000000000")

            except:
                print_log(
                    "Entered " + self.lineEdit_vout_margin_low_railA.text() + ", give valid float for Vout margin low",
                    "WARNING")
                time.sleep(2)
                self.lineEdit_vout_margin_low_railA.setText("0.00")
                return

    def vout_margin_low_changed_B(self):
        if self.lineEdit_vout_margin_low_railB.text():
            if self.lineEdit_vout_margin_low_railB.text()[0] == ".":
                self.lineEdit_vout_margin_low_railB.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_margin_low_railB.text()), float):
                    val_1 = float(self.lineEdit_vout_margin_low_railB.text())
                    if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("261", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            val_1 -= 0.245
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("261", 15, 0,
                                                                     bin(round(val_1 / 0.005)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_low_railB.text() + ", give Vout margin low ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_low_railB.setText("0.00")
                            update_database_with_temp_customer_input("261", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailB.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("261", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            val_1 -= 0.19
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("261", 15, 0,
                                                                     bin(round(val_1 / 0.01)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_low_railB.text() + ", give margin low ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_low_railB.setText("0.00")
                            update_database_with_temp_customer_input("261", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            update_database_with_temp_customer_input("261", 15, 0,
                                                                     bin(round(val_1 / 0.00625)).split("0b")[1].zfill(
                                                                         16))

                        else:
                            print_log("Entered " + self.lineEdit_vout_margin_low_railB.text() + ", give Vout margin low ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_margin_low_railB.setText("0.00")
                            update_database_with_temp_customer_input("261", 15, 0, "0000000000000000")

            except:
                print_log(
                    "Entered " + self.lineEdit_vout_margin_low_railB.text() + ", give valid float for Vout margin low ",
                    "WARNING")
                time.sleep(2)
                self.lineEdit_vout_margin_low_railB.setText("0.00")
                return

    def vout_max_changed_A(self):
        if self.lineEdit_vout_max_railA.text():
            if self.lineEdit_vout_max_railA.text()[0] == ".":
                self.lineEdit_vout_max_railA.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_max_railA.text()), float):
                    val_1 = float(self.lineEdit_vout_max_railA.text())
                    if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("240", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.245) / 0.00125
                            update_database_with_temp_customer_input("240", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_max_railA.text() + ", give Vout max ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_max_railA.setText("1.52")
                            update_database_with_temp_customer_input("240", 15, 0, "0000001111111100")

                    elif self.comboBox_vout_mode_RailA.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("240", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.190) / 0.00125
                            update_database_with_temp_customer_input("240", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_max_railA.text() + ", give Vout max ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_max_railA.setText("2.74")
                            update_database_with_temp_customer_input("240", 15, 0, "0000011111111000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = val_1 / 0.00125
                            update_database_with_temp_customer_input("240", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_max_railA.text() + ", give Vout max ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_max_railA.setText("1.55")
                            update_database_with_temp_customer_input("240", 15, 0, "0000010011011000")

            except:
                print_log("Entered " + self.lineEdit_vout_max_railA.text() + ", give valid float for Vout max ",
                          "WARNING")
                time.sleep(2)
                self.lineEdit_vout_max_railA.setText("1.55")
                return

    def vout_max_changed_B(self):
        if self.lineEdit_vout_max_railB.text():
            if self.lineEdit_vout_max_railB.text()[0] == ".":
                self.lineEdit_vout_max_railB.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_max_railB.text()), float):
                    val_1 = float(self.lineEdit_vout_max_railB.text())
                    if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("241", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.245) / 0.00125
                            update_database_with_temp_customer_input("241", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_max_railB.text() + ", give Vout max ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_max_railB.setText("1.52")
                            update_database_with_temp_customer_input("241", 15, 0, "0000001111111100")

                    elif self.comboBox_vout_mode_RailB.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("241", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.190) / 0.00125
                            update_database_with_temp_customer_input("241", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_max_railB.text() + ", give Vout max ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_max_railB.setText("2.74")
                            update_database_with_temp_customer_input("241", 15, 0, "0000011111111000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = val_1 / 0.00125
                            update_database_with_temp_customer_input("241", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_max_railB.text() + ", give Vout max ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_max_railB.setText("1.55")
                            update_database_with_temp_customer_input("241", 15, 0, "0000010011011000")

            except:
                print_log("Entered " + self.lineEdit_vout_max_railB.text() + ", give valid float for Vout max ",
                          "WARNING")
                time.sleep(2)
                self.lineEdit_vout_max_railB.setText("1.55")
                return

    def vout_min_changed_A(self):
        if self.lineEdit_vout_min_railA.text():
            if self.lineEdit_vout_min_railA.text()[0] == ".":
                self.lineEdit_vout_min_railA.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_min_railA.text()), float):
                    val_1 = float(self.lineEdit_vout_min_railA.text())
                    if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("2B0", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.245) / 0.00125
                            update_database_with_temp_customer_input("2B0", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_min_railA.text() + ", give Vout min ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_min_railA.setText("0.00")
                            update_database_with_temp_customer_input("2B0", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailA.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("2B0", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.190) / 0.00125
                            update_database_with_temp_customer_input("2B0", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_min_railA.text() + ", give Vout min ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_min_railA.setText("0.00")
                            update_database_with_temp_customer_input("2B0", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = val_1 / 0.00125
                            update_database_with_temp_customer_input("2B0", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_min_railA.text() + ", give Vout min ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_min_railA.setText("0.00")
                            update_database_with_temp_customer_input("2B0", 15, 0, "0000000000000000")

            except:
                print_log("Entered " + self.lineEdit_vout_min_railA.text() + ", give valid float for Vout min ",
                          "WARNING")
                time.sleep(2)
                self.lineEdit_vout_min_railA.setText("0.00")
                return

    def vout_min_changed_B(self):
        if self.lineEdit_vout_min_railB.text():
            if self.lineEdit_vout_min_railB.text()[0] == ".":
                self.lineEdit_vout_min_railB.setText("0.")
            try:
                if isinstance(float(self.lineEdit_vout_min_railB.text()), float):
                    val_1 = float(self.lineEdit_vout_min_railB.text())
                    if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("2B1", 15, 0, "0000000000000000")
                        elif 0.25 <= val_1 <= 1.52:
                            mod = round((val_1 % 0.005), 6)
                            if mod > 0.002499:
                                val_1 += round((0.005 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.245) / 0.00125
                            update_database_with_temp_customer_input("2B1", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_min_railB.text() + ", give Vout min ranging (250mV - 1.52V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_min_railB.setText("0.00")
                            update_database_with_temp_customer_input("2B1", 15, 0, "0000000000000000")

                    elif self.comboBox_vout_mode_RailB.currentIndex() == 2:
                        if val_1 == 0:
                            update_database_with_temp_customer_input("2B1", 15, 0, "0000000000000000")
                        elif 0.2 <= val_1 <= 2.74:
                            mod = round((val_1 % 0.01), 6)
                            if mod > 0.004999:
                                val_1 += round((0.01 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = (val_1 - 0.190) / 0.00125
                            update_database_with_temp_customer_input("2B1", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_min_railB.text() + ", give Vout min ranging (200mV - 2.74V) or 0V", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_min_railB.setText("0.00")
                            update_database_with_temp_customer_input("2B1", 15, 0, "0000000000000000")

                    else:
                        if 0 <= val_1 <= 1.55:
                            mod = round((val_1 % 0.00625), 6)
                            if mod > 0.003125:
                                val_1 += round((0.00625 - mod), 6)
                            else:
                                val_1 = round((val_1 - mod), 6)
                            val_1 = val_1 / 0.00125
                            update_database_with_temp_customer_input("2B1", 15, 0,
                                                                     bin(round(val_1)).split("0b")[1].zfill(16))
                        else:
                            print_log("Entered " + self.lineEdit_vout_min_railB.text() + ", give Vout min ranging (0 - 1.55V)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_vout_min_railB.setText("0.00")
                            update_database_with_temp_customer_input("2B1", 15, 0, "0000000000000000")

            except:
                print_log("Entered " + self.lineEdit_vout_min_railB.text() + ", give valid float for Vout min ",
                          "WARNING")
                time.sleep(2)
                self.lineEdit_vout_min_railB.setText("0.00")
                return

    def ton_delay_changed_A(self):
        if self.comboBox_ton_delay_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("600", 15, 0, "0000000000000000")
        elif self.comboBox_ton_delay_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("600", 15, 0, "1011001000000000")
        elif self.comboBox_ton_delay_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("600", 15, 0, "1011101000000000")
        elif self.comboBox_ton_delay_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("600", 15, 0, "1011101100000000")
        elif self.comboBox_ton_delay_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("600", 15, 0, "1100001000000000")
        elif self.comboBox_ton_delay_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("600", 15, 0, "1100001100000000")
        elif self.comboBox_ton_delay_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("600", 15, 0, "1100101000000000")
        else:
            update_database_with_temp_customer_input("600", 15, 0, "1100101010000000")

    def ton_delay_changed_B(self):
        if self.comboBox_ton_delay_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("601", 15, 0, "0000000000000000")
        elif self.comboBox_ton_delay_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("601", 15, 0, "1011001000000000")
        elif self.comboBox_ton_delay_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("601", 15, 0, "1011101000000000")
        elif self.comboBox_ton_delay_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("601", 15, 0, "1011101100000000")
        elif self.comboBox_ton_delay_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("601", 15, 0, "1100001000000000")
        elif self.comboBox_ton_delay_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("601", 15, 0, "1100001100000000")
        elif self.comboBox_ton_delay_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("601", 15, 0, "1100101000000000")
        else:
            update_database_with_temp_customer_input("601", 15, 0, "1100101010000000")

    def toff_delay_changed_A(self):
        if self.comboBox_toff_delay_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("640", 15, 0, "0000000000000000")
        elif self.comboBox_toff_delay_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("640", 15, 0, "1011001000000000")
        elif self.comboBox_toff_delay_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("640", 15, 0, "1011101000000000")
        elif self.comboBox_toff_delay_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("640", 15, 0, "1011101100000000")
        elif self.comboBox_toff_delay_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("640", 15, 0, "1100001000000000")
        elif self.comboBox_toff_delay_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("640", 15, 0, "1100001100000000")
        elif self.comboBox_toff_delay_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("640", 15, 0, "1100101000000000")
        else:
            update_database_with_temp_customer_input("640", 15, 0, "1100101010000000")

    def toff_delay_changed_B(self):
        if self.comboBox_toff_delay_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("641", 15, 0, "0000000000000000")
        elif self.comboBox_toff_delay_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("641", 15, 0, "1011001000000000")
        elif self.comboBox_toff_delay_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("641", 15, 0, "1011101000000000")
        elif self.comboBox_toff_delay_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("641", 15, 0, "1011101100000000")
        elif self.comboBox_toff_delay_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("641", 15, 0, "1100001000000000")
        elif self.comboBox_toff_delay_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("641", 15, 0, "1100001100000000")
        elif self.comboBox_toff_delay_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("641", 15, 0, "1100101000000000")
        else:
            update_database_with_temp_customer_input("641", 15, 0, "1100101010000000")

    def freq_switch_changed_A(self):
        if self.comboBox_freq_switch_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("330", 15, 0, "1010101000000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("330", 15, 0, "1011001000000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("330", 15, 0, "1011001100000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 3:
            update_database_with_temp_customer_input("330", 15, 0, "1011101000000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 4:
            update_database_with_temp_customer_input("330", 15, 0, "1011101010000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 5:
            update_database_with_temp_customer_input("330", 15, 0, "1011101100000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 6:
            update_database_with_temp_customer_input("330", 15, 0, "1011101110000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 7:
            update_database_with_temp_customer_input("330", 15, 0, "1100001000000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 8:
            update_database_with_temp_customer_input("330", 15, 0, "1100001001000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 9:
            update_database_with_temp_customer_input("330", 15, 0, "1100001010000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 10:
            update_database_with_temp_customer_input("330", 15, 0, "1100001011000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 11:
            update_database_with_temp_customer_input("330", 15, 0, "1100001100000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 12:
            update_database_with_temp_customer_input("330", 15, 0, "1100001101000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 13:
            update_database_with_temp_customer_input("330", 15, 0, "1100001110000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 14:
            update_database_with_temp_customer_input("330", 15, 0, "1100001111000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 15:
            update_database_with_temp_customer_input("330", 15, 0, "1100101000000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 16:
            update_database_with_temp_customer_input("330", 15, 0, "1100101000100000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 17:
            update_database_with_temp_customer_input("330", 15, 0, "1100101001000000")
        elif self.comboBox_freq_switch_RailA.currentIndex() == 18:
            update_database_with_temp_customer_input("330", 15, 0, "1100101001100000")
        else:
            update_database_with_temp_customer_input("330", 15, 0, "1100101010000000")

    def freq_switch_changed_B(self):
        if self.comboBox_freq_switch_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("331", 15, 0, "1010101000000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("331", 15, 0, "1011001000000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("331", 15, 0, "1011001100000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 3:
            update_database_with_temp_customer_input("331", 15, 0, "1011101000000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 4:
            update_database_with_temp_customer_input("331", 15, 0, "1011101010000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 5:
            update_database_with_temp_customer_input("331", 15, 0, "1011101100000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 6:
            update_database_with_temp_customer_input("331", 15, 0, "1011101110000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 7:
            update_database_with_temp_customer_input("331", 15, 0, "1100001000000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 8:
            update_database_with_temp_customer_input("331", 15, 0, "1100001001000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 9:
            update_database_with_temp_customer_input("331", 15, 0, "1100001010000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 10:
            update_database_with_temp_customer_input("331", 15, 0, "1100001011000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 11:
            update_database_with_temp_customer_input("331", 15, 0, "1100001100000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 12:
            update_database_with_temp_customer_input("331", 15, 0, "1100001101000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 13:
            update_database_with_temp_customer_input("331", 15, 0, "1100001110000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 14:
            update_database_with_temp_customer_input("331", 15, 0, "1100001111000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 15:
            update_database_with_temp_customer_input("331", 15, 0, "1100101000000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 16:
            update_database_with_temp_customer_input("331", 15, 0, "1100101000100000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 17:
            update_database_with_temp_customer_input("331", 15, 0, "1100101001000000")
        elif self.comboBox_freq_switch_RailB.currentIndex() == 18:
            update_database_with_temp_customer_input("331", 15, 0, "1100101001100000")
        else:
            update_database_with_temp_customer_input("331", 15, 0, "1100101010000000")

    def Save(self):
        global list_of_registers_used_in_this_frame, resolutionB, resolutionA, initial_device_status, homeWin_obj, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)

                if initialize_feature_variable("220", 9, 9) == "1":
                    c = initialize_feature_variable("220", 9, 0)
                    d = ''.join([str((int(i) ^ 1)) for i in c])
                    d = int(d, base=2)
                    d = (d + 1) * -1
                else:
                    d = int(initialize_feature_variable("220", 15, 0), 2)
                if self.comboBox_vout_mode_RailA.currentIndex() == 0:
                    resolutionA = "5"
                    if int(initialize_feature_variable("210", 15, 0), 2) == 0:
                        self.lineEdit_vout_command_railA.setText(
                            str((int(initialize_feature_variable("210", 15, 0), 2)) * 0.005))
                    else:
                        self.lineEdit_vout_command_railA.setText(
                            str(round((int(initialize_feature_variable("210", 15, 0), 2)) * 0.005 + 0.245, 3)))
                    if int(initialize_feature_variable("250", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_high_railA.setText(
                            str((int(initialize_feature_variable("250", 15, 0), 2)) * 0.005))
                    else:
                        self.lineEdit_vout_margin_high_railA.setText(
                            str(round((int(initialize_feature_variable("250", 15, 0), 2)) * 0.005 + 0.245, 3)))
                    if int(initialize_feature_variable("260", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_low_railA.setText(
                            str((int(initialize_feature_variable("260", 15, 0), 2)) * 0.005))
                    else:
                        self.lineEdit_vout_margin_low_railA.setText(
                            str(round((int(initialize_feature_variable("260", 15, 0), 2)) * 0.005 + 0.245, 3)))
                    if int(initialize_feature_variable("240", 15, 0), 2) == 0:
                        self.lineEdit_vout_max_railA.setText(
                            str((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_max_railA.setText(
                            str(round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                    if int(initialize_feature_variable("2B0", 15, 0), 2) == 0:
                        self.lineEdit_vout_min_railA.setText(
                            str((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_min_railA.setText(
                            str(round((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                    self.lineEdit_vout_trim_railA.setText(str(round(d * 0.005, 3)))

                elif self.comboBox_vout_mode_RailA.currentIndex() == 1:
                    resolutionA = "6.25"
                    self.lineEdit_vout_command_railA.setText(
                        str(round((int(initialize_feature_variable("210", 15, 0), 2)) * 0.00625, 5)))
                    self.lineEdit_vout_margin_high_railA.setText(
                        str(round((int(initialize_feature_variable("250", 15, 0), 2)) * 0.00625, 5)))
                    self.lineEdit_vout_margin_low_railA.setText(
                        str(round((int(initialize_feature_variable("260", 15, 0), 2)) * 0.00625, 5)))
                    self.lineEdit_vout_max_railA.setText(str(round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125, 5)))
                    self.lineEdit_vout_min_railA.setText(
                        str(round((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125, 5)))
                    self.lineEdit_vout_trim_railA.setText(str(round(d * 0.00625, 5)))
                else:
                    resolutionA = "10"
                    if int(initialize_feature_variable("210", 15, 0), 2) == 0:
                        self.lineEdit_vout_command_railA.setText(
                            str((int(initialize_feature_variable("210", 15, 0), 2)) * 0.010))
                    else:
                        self.lineEdit_vout_command_railA.setText(
                            str(round((int(initialize_feature_variable("210", 15, 0), 2)) * 0.010 + 0.190, 2)))
                    if int(initialize_feature_variable("250", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_high_railA.setText(
                            str((int(initialize_feature_variable("250", 15, 0), 2)) * 0.01))
                    else:
                        self.lineEdit_vout_margin_high_railA.setText(
                            str(round((int(initialize_feature_variable("250", 15, 0), 2)) * 0.010 + 0.190, 2)))
                    if int(initialize_feature_variable("260", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_low_railA.setText(
                            str((int(initialize_feature_variable("260", 15, 0), 2)) * 0.01))
                    else:
                        self.lineEdit_vout_margin_low_railA.setText(
                            str(round((int(initialize_feature_variable("260", 15, 0), 2)) * 0.01 + 0.190, 2)))
                    if int(initialize_feature_variable("240", 15, 0), 2) == 0:
                        self.lineEdit_vout_max_railA.setText(
                            str((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_max_railA.setText(
                            str(round((int(initialize_feature_variable("240", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                    if int(initialize_feature_variable("2B0", 15, 0), 2) == 0:
                        self.lineEdit_vout_min_railA.setText(
                            str((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_min_railA.setText(
                            str(round((int(initialize_feature_variable("2B0", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                    self.lineEdit_vout_trim_railA.setText(str(round(d * 0.01, 2)))

                if initialize_feature_variable("221", 9, 9) == "1":
                    c = initialize_feature_variable("221", 9, 0)
                    d = ''.join([str((int(i) ^ 1)) for i in c])
                    d = int(d, base=2)
                    d = (d + 1) * -1
                else:
                    d = int(initialize_feature_variable("221", 15, 0), 2)
                if self.comboBox_vout_mode_RailB.currentIndex() == 0:
                    resolutionB = "5"
                    if int(initialize_feature_variable("211", 15, 0), 2) == 0:
                        self.lineEdit_vout_command_railB.setText(
                            str((int(initialize_feature_variable("211", 15, 0), 2)) * 0.005))
                    else:
                        self.lineEdit_vout_command_railB.setText(
                            str(round((int(initialize_feature_variable("211", 15, 0), 2)) * 0.005 + 0.245, 3)))
                    if int(initialize_feature_variable("251", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_high_railB.setText(
                            str((int(initialize_feature_variable("251", 15, 0), 2)) * 0.005))
                    else:
                        self.lineEdit_vout_margin_high_railB.setText(
                            str(round((int(initialize_feature_variable("251", 15, 0), 2)) * 0.005 + 0.245, 3)))
                    if int(initialize_feature_variable("261", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_low_railB.setText(
                            str((int(initialize_feature_variable("261", 15, 0), 2)) * 0.005))
                    else:
                        self.lineEdit_vout_margin_low_railB.setText(
                            str(round((int(initialize_feature_variable("261", 15, 0), 2)) * 0.005 + 0.245, 3)))
                    if int(initialize_feature_variable("241", 15, 0), 2) == 0:
                        self.lineEdit_vout_max_railB.setText(
                            str((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_max_railB.setText(
                            str(round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                    if int(initialize_feature_variable("2B1", 15, 0), 2) == 0:
                        self.lineEdit_vout_min_railB.setText(
                            str((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_min_railB.setText(
                            str(round((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125 + 0.245, 5)))
                    self.lineEdit_vout_trim_railB.setText(str(round(d * 0.005, 3)))

                elif self.comboBox_vout_mode_RailB.currentIndex() == 1:
                    resolutionB = "6.25"
                    self.lineEdit_vout_command_railB.setText(
                        str(round((int(initialize_feature_variable("211", 15, 0), 2)) * 0.00625, 5)))
                    self.lineEdit_vout_margin_high_railB.setText(
                        str(round((int(initialize_feature_variable("251", 15, 0), 2)) * 0.00625, 5)))
                    self.lineEdit_vout_margin_low_railB.setText(
                        str(round((int(initialize_feature_variable("261", 15, 0), 2)) * 0.00625, 5)))
                    self.lineEdit_vout_max_railB.setText(str(round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125, 5)))
                    self.lineEdit_vout_min_railB.setText(
                        str(round((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125, 5)))
                    self.lineEdit_vout_trim_railB.setText(str(round(d * 0.00625, 5)))
                else:
                    resolutionB = "10"
                    if int(initialize_feature_variable("211", 15, 0), 2) == 0:
                        self.lineEdit_vout_command_railB.setText(
                            str((int(initialize_feature_variable("211", 15, 0), 2)) * 0.010))
                    else:
                        self.lineEdit_vout_command_railB.setText(
                            str(round((int(initialize_feature_variable("211", 15, 0), 2)) * 0.010 + 0.190, 2)))
                    if int(initialize_feature_variable("251", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_high_railB.setText(
                            str((int(initialize_feature_variable("251", 15, 0), 2)) * 0.01))
                    else:
                        self.lineEdit_vout_margin_high_railB.setText(
                            str(round((int(initialize_feature_variable("251", 15, 0), 2)) * 0.010 + 0.190, 2)))
                    if int(initialize_feature_variable("261", 15, 0), 2) == 0:
                        self.lineEdit_vout_margin_low_railB.setText(
                            str((int(initialize_feature_variable("261", 15, 0), 2)) * 0.01))
                    else:
                        self.lineEdit_vout_margin_low_railB.setText(
                            str(round((int(initialize_feature_variable("261", 15, 0), 2)) * 0.01 + 0.190, 2)))
                    if int(initialize_feature_variable("241", 15, 0), 2) == 0:
                        self.lineEdit_vout_max_railB.setText(
                            str((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_max_railB.setText(
                            str(round((int(initialize_feature_variable("241", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                    if int(initialize_feature_variable("2B1", 15, 0), 2) == 0:
                        self.lineEdit_vout_min_railB.setText(
                            str((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125))
                    else:
                        self.lineEdit_vout_min_railB.setText(
                            str(round((int(initialize_feature_variable("2B1", 15, 0), 2)) * 0.00125 + 0.190, 5)))
                    self.lineEdit_vout_trim_railB.setText(str(round(d * 0.01, 2)))

                print_log("PMBus configuration values are saved and rounded off.", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def Discard(self):
        print_log("PMBus Configuration Settings discarded.", "INFO")
        self.main = PMBus_Configuration()
        self.main.show()
        self.close()

# Telemetry Configuration class
class frame_telemetry_sensitivity(QMainWindow, Ui_Telemetry_senstivity):

    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame,PARTNAME
        super(frame_telemetry_sensitivity, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["F2"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        ioutA = float(1000 / float(
            np.frombuffer(struct.pack("H", int(initialize_feature_variable('F2', 274, 260), 2)), dtype=np.float16)[0])) if int(initialize_feature_variable('F2', 274, 260), 2) != 0 else int(1)
        # print(ioutA)
        self.lineEdit_ioutA_scaling.setText(str(round(ioutA, 2)))
        ioutB = float(1000 / float(np.frombuffer(struct.pack("H", int(initialize_feature_variable('F2', 214, 200), 2)), dtype=np.float16)[0])) if int(initialize_feature_variable('F2', 214, 200), 2) != 0 else int(1)
        self.lineEdit_ioutB_scaling.setText(str(round(ioutB, 2)))

        tmonA = float(1000 / float(
            np.frombuffer(struct.pack("H", int(initialize_feature_variable('F2', 259, 245), 2)), dtype=np.float16)[0])) if int(initialize_feature_variable('F2', 259, 245), 2) != 0 else int(1)
        self.lineEdit_tmonA_scaling.setText(str(round(tmonA, 2)))
        tmonB = float(1000 / float(
            np.frombuffer(struct.pack("H", int(initialize_feature_variable('F2', 244, 230), 2)), dtype=np.float16)[0])) if int(initialize_feature_variable('F2', 244, 230), 2) != 0 else int(1)
        self.lineEdit_tmonB_scaling.setText(str(round(tmonB, 2)))
        imon_aux = float(1000 / float(
            np.frombuffer(struct.pack("H", int(initialize_feature_variable('F2', 304, 290), 2)), dtype=np.float16)[0])) if int(initialize_feature_variable('F2', 304, 290), 2) != 0 else int(1)
        self.lineEdit_imon_aux_scaling.setText(str(round(imon_aux, 2)))

        ## code starts
        self.lineEdit_tmonA_scaling.textEdited.connect(self.tele_tmonA_reg)
        self.lineEdit_tmonB_scaling.textEdited.connect(self.tele_tmonB_reg)
        self.lineEdit_ioutA_scaling.textEdited.connect(self.tele_ioutA_reg)
        self.lineEdit_ioutB_scaling.textEdited.connect(self.tele_ioutB_reg)
        self.lineEdit_imon_aux_scaling.textEdited.connect(self.tele_imon_aux_reg)

        self.pushButton_Save.clicked.connect(self.tele_save_register)
        self.pushButton_Discard.clicked.connect(self.tele_discard_changes)

        if resolutionA == 6.25:
            self.maxA = 248
        else:
            self.maxA = 255

        if resolutionB == 6.25:
            self.maxB = 248
        else:
            self.maxB = 255

    def tele_ioutA_reg(self):
        if self.lineEdit_ioutA_scaling.text() == "":
            return
        try:
            float(self.lineEdit_ioutA_scaling.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        iccmax_hp = bin(np.float16(1000 / float(self.lineEdit_ioutA_scaling.text())).view('H'))[2:].zfill(15) if float(self.lineEdit_ioutA_scaling.text()) != 0 else bin(15360)
        update_database_with_temp_customer_input("F2", 274, 260, str(iccmax_hp))

    def tele_ioutB_reg(self):
        if self.lineEdit_ioutB_scaling.text() == "":
            return
        try:
            float(self.lineEdit_ioutB_scaling.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        iccmaxB_hp = bin(np.float16(1000 / float(self.lineEdit_ioutB_scaling.text())).view('H'))[2:].zfill(15) if float(self.lineEdit_ioutB_scaling.text()) != 0 else bin(15360)
        update_database_with_temp_customer_input("F2", 214, 200, iccmaxB_hp)

    def tele_tmonA_reg(self):
        if self.lineEdit_tmonA_scaling.text() == "":
            return
        try:
            float(self.lineEdit_tmonA_scaling.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        tmon_hp = bin(np.float16(1000 / float(self.lineEdit_tmonA_scaling.text())).view('H'))[2:].zfill(15) if float(self.lineEdit_tmonA_scaling.text()) != 0 else bin(15360)
        update_database_with_temp_customer_input("F2", 259, 245, tmon_hp)

    def tele_tmonB_reg(self):
        if self.lineEdit_tmonB_scaling.text() == "":
            return
        try:
            float(self.lineEdit_tmonB_scaling.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        tmonB_hp = bin(np.float16(1000 / float(self.lineEdit_tmonB_scaling.text())).view('H'))[2:].zfill(15) if float(self.lineEdit_tmonB_scaling.text()) != 0 else bin(15360)
        update_database_with_temp_customer_input("F2", 244, 230, tmonB_hp)

    def tele_imon_aux_reg(self):
        if self.lineEdit_imon_aux_scaling.text() == "":
            return
        try:
            float(self.lineEdit_imon_aux_scaling.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        imon_aux_hp = bin(np.float16(1000 / float(self.lineEdit_imon_aux_scaling.text())).view('H'))[2:].zfill(15) if float(self.lineEdit_imon_aux_scaling.text()) != 0 else bin(15360)
        update_database_with_temp_customer_input("F2", 304, 290, imon_aux_hp)

    def tele_save_register(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("Telemetry sensitivity configuration has been saved", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def tele_discard_changes(self):
        print_log("Telemetry sensitivity Settings discarded.", "INFO")
        self.main = frame_telemetry_sensitivity()
        self.main.show()
        self.close()


class frame_telemetry_calibration(QMainWindow, Ui_Telemetry_calibration):

    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(frame_telemetry_calibration, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["F2", "380", "390", "381", "391"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)
        iout_gainA = float(int(initialize_feature_variable('380', 8, 0), 2) / 256.0)
        # print(iout_gainA)
        self.lineEdit_iout_gain_railA.setText(str(round(iout_gainA, 2)))
        iout_gainB = float(int(initialize_feature_variable('381', 8, 0), 2) / 256.0)
        print(iout_gainB)
        self.lineEdit_iout_gain_railB.setText(str(round(iout_gainB, 2)))
        iout_offsetA = int(initialize_feature_variable('390', 3, 0), 2)
        if initialize_feature_variable('390', 4, 4) == '1':
            offsetA = str("-") + str(iout_offsetA)
        else:
            offsetA = str(iout_offsetA)
        self.lineEdit_iout_offset_railA.setText(offsetA)
        iout_offsetB = int(initialize_feature_variable('391', 3, 0), 2)
        if initialize_feature_variable('391', 4, 4) == '1':
            offsetB = str("-") + str(iout_offsetB)
        else:
            offsetB = str(iout_offsetB)
        self.lineEdit_iout_offset_railB.setText(offsetB)

        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        ## code starts

        self.lineEdit_iout_gain_railA.textEdited.connect(self.tele_ioutA_gain)
        self.lineEdit_iout_gain_railB.textEdited.connect(self.tele_ioutB_gain)
        self.lineEdit_iout_offset_railA.textEdited.connect(self.tele_ioutA_offset)
        self.lineEdit_iout_offset_railB.textEdited.connect(self.tele_ioutB_offset)

        self.pushButton_Save.clicked.connect(self.tele_save_register)
        self.pushButton_Discard.clicked.connect(self.tele_discard_changes)

    def tele_ioutA_gain(self):
        if self.lineEdit_iout_gain_railA.text() == "":
            return
        try:
            float(self.lineEdit_iout_gain_railA.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        self.gainA = float(self.lineEdit_iout_gain_railA.text())

        if (self.gainA >= 2 or self.gainA < 0):
            print_log("Gain value should be between 0 to 2", "ERROR")
            if (self.gainA >= 2):
                self.lineEdit_iout_gain_railA.setText("2")
            elif (self.gainA < 0):
                self.lineEdit_iout_gain_railA.setText("0")
        self.gainA = float(self.lineEdit_iout_gain_railA.text())
        iouta_gain = "000000000"
        for i in range(9):
            if (self.gainA >= pow(2, (-i))):
                iouta_gain = iouta_gain[:i] + "1" + iouta_gain[i + 1:]
                self.gainA = self.gainA - pow(2, (-i))
            else:
                iouta_gain = iouta_gain[:i] + "0" + iouta_gain[i + 1:]

        # iouta_gain = bin(np.float16(float(self.lineEdit_iout_gain_railA.text())).view('H'))[2:].zfill(16)
        update_database_with_temp_customer_input("380", 8, 0, str(iouta_gain))

    def tele_ioutB_gain(self):
        if self.lineEdit_iout_gain_railB.text() == "":
            return
        try:
            float(self.lineEdit_iout_gain_railB.text()).is_integer()
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return
        self.gainB = float(self.lineEdit_iout_gain_railB.text())

        if (self.gainB >= 2 or self.gainB < 0):
            print_log("Gain value should be between 0 to 2", "ERROR")
            if (self.gainB >= 2):
                self.lineEdit_iout_gain_railB.setText("2")
            elif (self.gainB < 0):
                self.lineEdit_iout_gain_railB.setText("0")

        self.gainB = float(self.lineEdit_iout_gain_railB.text())

        ioutb_gain = "000000000"
        for i in range(9):
            if (self.gainB >= pow(2, (-i))):
                ioutb_gain = ioutb_gain[:i] + "1" + ioutb_gain[i + 1:]
                self.gainB = self.gainB - pow(2, (-i))
            else:
                ioutb_gain = ioutb_gain[:i] + "0" + ioutb_gain[i + 1:]

        # iouta_gain = bin(np.float16(float(self.lineEdit_iout_gain_railA.text())).view('H'))[2:].zfill(16)
        update_database_with_temp_customer_input("381", 8, 0, str(ioutb_gain))

    def tele_ioutA_offset(self):
        if self.lineEdit_iout_offset_railA.text() == "":
            return
        try:
            int(self.lineEdit_iout_offset_railA.text())
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return

        offsetA = int(self.lineEdit_iout_offset_railA.text())

        if (offsetA > 15):
            print_log("Enter offset between +15 and -15", "ERROR")
            self.lineEdit_iout_offset_railA.setText("15")
        elif (offsetA < -15):
            print_log("Enter offset between +15 and -15", "ERROR")
            self.lineEdit_iout_offset_railA.setText("-15")

        offsetA = int(self.lineEdit_iout_offset_railA.text())
        if offsetA < 0:
            mag_offsetA = "1" + str(bin(abs(offsetA)).replace("0b", "").zfill(4))
        else:
            mag_offsetA = "0" + str(bin(offsetA).replace("0b", "").zfill(4))
        update_database_with_temp_customer_input("390", 4, 0, mag_offsetA)

    def tele_ioutB_offset(self):
        if self.lineEdit_iout_offset_railB.text() == "":
            return
        try:
            int(self.lineEdit_iout_offset_railB.text())
        except ValueError:
            print_log("Text box value should be number", "ERROR")
            return

        offsetA = int(self.lineEdit_iout_offset_railB.text())
        if (offsetA > 15):
            print_log("Enter offset between +15 and -15", "ERROR")
            self.lineEdit_iout_offset_railB.setText("15")
        elif (offsetA < -15):
            print_log("Enter offset between +15 and -15", "ERROR")
            self.lineEdit_iout_offset_railB.setText("-15")

        offsetA = int(self.lineEdit_iout_offset_railB.text())

        if offsetA < 0:
            mag_offsetA = "1" + str(bin(abs(offsetA)).replace("0b", "").zfill(4))
        else:
            mag_offsetA = "0" + str(bin(offsetA).replace("0b", "").zfill(4))
        update_database_with_temp_customer_input("391", 4, 0, mag_offsetA)

    def tele_save_register(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("Telemetry sensitivity configuration has been saved", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def tele_discard_changes(self):
        print_log("Telemetry sensitivity Settings discarded.", "INFO")
        self.main = frame_telemetry_calibration()
        self.main.show()
        self.close()

# Fault configuration class
class Fault_configuration(QMainWindow, Ui_Fault_config):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(Fault_configuration, self).__init__(parent)
        self.setupUi(self)

        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["E4", "F2", "55", "59", "C50", "620", "C51", "621"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        if int(initialize_feature_variable('E4', 9, 9), 2) == 0:
            self.radioButton_soft_shutdown.setChecked(True)
            self.radioButton_immediate_shutdown.setChecked(False)
        elif int(initialize_feature_variable('E4', 9, 9), 2) == 1:
            self.radioButton_soft_shutdown.setChecked(False)
            self.radioButton_immediate_shutdown.setChecked(True)

        self.comboBox_trackingUVP_RailA.setCurrentIndex(int(initialize_feature_variable('C50', 27, 25), 2))
        self.comboBox_trackingUVP_RailB.setCurrentIndex(int(initialize_feature_variable('C51', 27, 25), 2))
        self.comboBox_trackingOVP_RailA.setCurrentIndex(int(initialize_feature_variable('C50', 30, 28), 2))
        self.comboBox_trackingOVP_RailB.setCurrentIndex(int(initialize_feature_variable('C51', 30, 28), 2))
        self.comboBox_PhaseOverCurrent_RailA.setCurrentIndex(int(initialize_feature_variable('C50', 48, 45), 2))
        self.comboBox_PhaseOverCurrent_RailB.setCurrentIndex(int(initialize_feature_variable('C51', 48, 45), 2))
        self.comboBox_AbsoluteOVP_RailA.setCurrentIndex(int(initialize_feature_variable('C50', 59, 57), 2))
        self.comboBox_AbsoluteOVP_RailB.setCurrentIndex(int(initialize_feature_variable('C51', 59, 57), 2))

        if str(hex(int(initialize_feature_variable('620', 15, 0), 2)).replace("0x", "")).upper() == "0000":
            self.comboBox_voutTurnOnTime_RailA.setCurrentIndex(0)
        elif str(hex(int(initialize_feature_variable('620', 15, 0), 2)).replace("0x", "")).upper() == "E320":
            self.comboBox_voutTurnOnTime_RailA.setCurrentIndex(1)
        elif str(hex(int(initialize_feature_variable('620', 15, 0), 2)).replace("0x", "")).upper() == "EA58":
            self.comboBox_voutTurnOnTime_RailA.setCurrentIndex(2)
        elif str(hex(int(initialize_feature_variable('620', 15, 0), 2)).replace("0x", "")).upper() == "EB20":
            self.comboBox_voutTurnOnTime_RailA.setCurrentIndex(3)

        if str(hex(int(initialize_feature_variable('621', 15, 0), 2)).replace("0x", "")).upper() == "0000":
            self.comboBox_voutTurnOnTime_RailB.setCurrentIndex(0)
        elif str(hex(int(initialize_feature_variable('621', 15, 0), 2)).replace("0x", "")).upper() == "E320":
            self.comboBox_voutTurnOnTime_RailB.setCurrentIndex(1)
        elif str(hex(int(initialize_feature_variable('621', 15, 0), 2)).replace("0x", "")).upper() == "EA58":
            self.comboBox_voutTurnOnTime_RailB.setCurrentIndex(2)
        elif str(hex(int(initialize_feature_variable('621', 15, 0), 2)).replace("0x", "")).upper() == "EB20":
            self.comboBox_voutTurnOnTime_RailB.setCurrentIndex(3)

        self.comboBox_temp_fault_threshold.setCurrentIndex(int(initialize_feature_variable('F2', 308, 307), 2))
        self.comboBox_iout_fault_limit.setCurrentIndex(int(initialize_feature_variable('F2', 306, 305), 2))

        # if str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "C380":
        #     self.comboBox_vin_UVP.setCurrentIndex(0)
        # elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "C3C0":
        #     self.comboBox_vin_UVP.setCurrentIndex(1)
        # elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "CA00":
        #     self.comboBox_vin_UVP.setCurrentIndex(2)
        if str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "CA20":
            self.comboBox_vin_UVP.setCurrentIndex(0)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "CA60":
            self.comboBox_vin_UVP.setCurrentIndex(1)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "CAA0":
            self.comboBox_vin_UVP.setCurrentIndex(2)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "CB20":
            self.comboBox_vin_UVP.setCurrentIndex(3)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "CBA0":
            self.comboBox_vin_UVP.setCurrentIndex(4)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "D210":
            self.comboBox_vin_UVP.setCurrentIndex(5)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "D250":
            self.comboBox_vin_UVP.setCurrentIndex(6)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "D290":
            self.comboBox_vin_UVP.setCurrentIndex(7)
        elif str(hex(int(initialize_feature_variable('59', 15, 0), 2)).replace("0x", "")).upper() == "D2D0":
            self.comboBox_vin_UVP.setCurrentIndex(8)
        else:
            print("Needs debug", "ERROR")

        if str(hex(int(initialize_feature_variable('55', 15, 0), 2)).replace("0x", "")).upper() == "DA00":
            self.comboBox_vin_OVP.setCurrentIndex(0)
        elif str(hex(int(initialize_feature_variable('55', 15, 0), 2)).replace("0x", "")).upper() == "DA10":
            self.comboBox_vin_OVP.setCurrentIndex(1)
        elif str(hex(int(initialize_feature_variable('55', 15, 0), 2)).replace("0x", "")).upper() == "DA20":
            self.comboBox_vin_OVP.setCurrentIndex(2)
        elif str(hex(int(initialize_feature_variable('55', 15, 0), 2)).replace("0x", "")).upper() == "DA30":
            self.comboBox_vin_OVP.setCurrentIndex(3)
        elif str(hex(int(initialize_feature_variable('55', 15, 0), 2)).replace("0x", "")).upper() == "DA40":
            self.comboBox_vin_OVP.setCurrentIndex(4)

        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        # Connecting to functions
        self.radioButton_soft_shutdown.toggled.connect(self.on_clicked_soft_shutdown_radio)
        self.radioButton_immediate_shutdown.toggled.connect(self.on_clicked_immediate_shutdown_radio)
        self.comboBox_temp_fault_threshold.activated.connect(self.temp_fault_threshold)
        self.comboBox_iout_fault_limit.activated.connect(self.iout_fault_limit)
        self.comboBox_trackingUVP_RailA.activated.connect(self.trackingUVP_A)
        self.comboBox_trackingUVP_RailB.activated.connect(self.trackingUVP_B)
        self.comboBox_trackingOVP_RailA.activated.connect(self.trackingOVP_A)
        self.comboBox_trackingOVP_RailB.activated.connect(self.trackingOVP_B)
        self.comboBox_PhaseOverCurrent_RailA.activated.connect(self.OCL_A)
        self.comboBox_PhaseOverCurrent_RailB.activated.connect(self.OCL_B)
        self.comboBox_AbsoluteOVP_RailA.activated.connect(self.AbsoluteOVP_A)
        self.comboBox_AbsoluteOVP_RailB.activated.connect(self.AbsoluteOVP_B)
        self.comboBox_voutTurnOnTime_RailA.activated.connect(self.voutTurnOnTime_A)
        self.comboBox_voutTurnOnTime_RailB.activated.connect(self.voutTurnOnTime_B)
        self.comboBox_vin_UVP.activated.connect(self.vin_UVP)
        self.comboBox_vin_OVP.activated.connect(self.vin_OVP)
        self.pushButton_Discard.clicked.connect(self.discard_changes)
        self.pushButton_Save.clicked.connect(self.fault_save)
        # self.comboBox_vin_OVP.activated.connect

    def on_clicked_soft_shutdown_radio(self):
        if self.radioButton_soft_shutdown.isChecked() == True:
            update_database_with_temp_customer_input("E4", 9, 9, bin(0).split("0b")[1].zfill(1))
            self.radioButton_immediate_shutdown.setChecked(False)
        elif self.radioButton_soft_shutdown.isChecked() == False:
            update_database_with_temp_customer_input("E4", 9, 9, bin(1).split("0b")[1].zfill(1))
            self.radioButton_immediate_shutdown.setChecked(True)

    def on_clicked_immediate_shutdown_radio(self):
        if self.radioButton_immediate_shutdown.isChecked() == True:
            update_database_with_temp_customer_input("E4", 9, 9, bin(1).split("0b")[1].zfill(1))
            self.radioButton_soft_shutdown.setChecked(False)
        elif self.radioButton_immediate_shutdown.isChecked() == False:
            update_database_with_temp_customer_input("E4", 9, 9, bin(0).split("0b")[1].zfill(1))
            self.radioButton_soft_shutdown.setChecked(True)

    def temp_fault_threshold(self):
        update_database_with_temp_customer_input("F2", 308, 307, bin(self.comboBox_temp_fault_threshold.currentIndex()).split("0b")[1].zfill(2))

    def iout_fault_limit(self):
        update_database_with_temp_customer_input("F2", 306, 305, bin(self.comboBox_iout_fault_limit.currentIndex()).split("0b")[1].zfill(2))

    def trackingUVP_A(self):
        update_database_with_temp_customer_input("C50", 27, 25, bin(self.comboBox_trackingUVP_RailA.currentIndex()).split("0b")[1].zfill(3))

    def trackingUVP_B(self):
        update_database_with_temp_customer_input("C51", 27, 25, bin(self.comboBox_trackingUVP_RailB.currentIndex()).split("0b")[1].zfill(3))

    def trackingOVP_A(self):
        update_database_with_temp_customer_input("C50", 30, 28, bin(self.comboBox_trackingOVP_RailA.currentIndex()).split("0b")[1].zfill(3))

    def trackingOVP_B(self):
        update_database_with_temp_customer_input("C51", 30, 28, bin(self.comboBox_trackingOVP_RailB.currentIndex()).split("0b")[1].zfill(3))

    def OCL_A(self):
        update_database_with_temp_customer_input("C50", 48, 45, bin(self.comboBox_PhaseOverCurrent_RailA.currentIndex()).split("0b")[1].zfill(4))

    def OCL_B(self):
        update_database_with_temp_customer_input("C51", 48, 45, bin(self.comboBox_PhaseOverCurrent_RailB.currentIndex()).split("0b")[1].zfill(4))

    def AbsoluteOVP_A(self):
        update_database_with_temp_customer_input("C50", 59, 57, bin(self.comboBox_AbsoluteOVP_RailA.currentIndex()).split("0b")[1].zfill(3))

    def AbsoluteOVP_B(self):
        update_database_with_temp_customer_input("C51", 59, 57, bin(self.comboBox_AbsoluteOVP_RailB.currentIndex()).split("0b")[1].zfill(3))

    def voutTurnOnTime_A(self):
        if self.comboBox_voutTurnOnTime_RailA.currentIndex() == 0:
            update_database_with_temp_customer_input("620", 15, 0, "0000000000000000")
        elif self.comboBox_voutTurnOnTime_RailA.currentIndex() == 1:
            update_database_with_temp_customer_input("620", 15, 0, "1110001100100000")
        elif self.comboBox_voutTurnOnTime_RailA.currentIndex() == 2:
            update_database_with_temp_customer_input("620", 15, 0, "1110101001011000")
        else:
            update_database_with_temp_customer_input("620", 15, 0, "1110101100100000")

    def voutTurnOnTime_B(self):
        if self.comboBox_voutTurnOnTime_RailB.currentIndex() == 0:
            update_database_with_temp_customer_input("621", 15, 0, "0000000000000000")
        elif self.comboBox_voutTurnOnTime_RailB.currentIndex() == 1:
            update_database_with_temp_customer_input("621", 15, 0, "1110001100100000")
        elif self.comboBox_voutTurnOnTime_RailB.currentIndex() == 2:
            update_database_with_temp_customer_input("621", 15, 0, "1110101001011000")
        else:
            update_database_with_temp_customer_input("621", 15, 0, "1110101100100000")

    def vin_UVP(self):
        mapping = { 0: "CA20", 1: "CA60", 2: "CAA0", 3: "CB20", 4: "CBA0", 5: "D210", 6: "D250", 7: "D290", 8: "D2D0"}
        update_database_with_temp_customer_input("59", 15, 0, bin(int(mapping[self.comboBox_vin_UVP.currentIndex()], 16)).split("0b")[1].zfill(16))

    def vin_OVP(self):
        mapping = {0: "DA00", 1: "DA10", 2: "DA20", 3: "DA30", 4: "DA40"}
        update_database_with_temp_customer_input("55", 15, 0, bin(int(mapping[self.comboBox_vin_OVP.currentIndex()], 16)).split("0b")[1].zfill(16))

    def fault_save(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("Fault Configuration Settings saved.", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def discard_changes(self):
        print_log("Fault  Configuration Settings discarded.", "INFO")
        self.main = Fault_configuration()
        self.main.show()
        self.close()


class Fault_response(QMainWindow, Ui_Fault_response):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(Fault_response, self).__init__(parent)
        self.setupUi(self)
        # List of registers associated with used feature
        list_of_registers_used_in_this_frame = ["56", "5A", "470", "410", "500", "450", "630", "471", "411", "501",
                                                "451", "631"]

        # Initialize temp update from customer field in database with final register value field at start of the frame.
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        # GUI default value look initialization
        # Non feature related initializaton on GUI_ display
        self.label_display_RailA.setText(RailA_name)
        self.label_display_RailB.setText(RailB_name)

        # Feature related initialization on GUI display
        self.comboBox_over_current_response_railA.setCurrentIndex(int(initialize_feature_variable("470", 7, 6), 2))
        self.comboBox_over_current_retry_railA.setCurrentIndex(int(initialize_feature_variable("470", 5, 3), 2))
        self.comboBox_over_current_delay_railA.setCurrentIndex(int(initialize_feature_variable("470", 2, 0), 2))
        self.comboBox_over_current_response_railB.setCurrentIndex(int(initialize_feature_variable("471", 7, 6), 2))
        self.comboBox_over_current_retry_railB.setCurrentIndex(int(initialize_feature_variable("471", 5, 3), 2))
        self.comboBox_over_current_delay_railB.setCurrentIndex(int(initialize_feature_variable("471", 2, 0), 2))

        self.comboBox_vout_over_voltage_response_railA.setCurrentIndex(int(initialize_feature_variable("410", 7, 6), 2))
        self.comboBox_vout_over_voltage_retry_railA.setCurrentIndex(int(initialize_feature_variable("410", 5, 3), 2))
        self.comboBox_vout_over_voltage_delay_railA.setCurrentIndex(int(initialize_feature_variable("410", 2, 0), 2))
        self.comboBox_vout_over_voltage_response_railB.setCurrentIndex(int(initialize_feature_variable("411", 7, 6), 2))
        self.comboBox_vout_over_voltage_retry_railB.setCurrentIndex(int(initialize_feature_variable("411", 5, 3), 2))
        self.comboBox_vout_over_voltage_delay_railB.setCurrentIndex(int(initialize_feature_variable("411", 2, 0), 2))

        self.comboBox_over_temp_response_railA.setCurrentIndex(int(initialize_feature_variable("500", 7, 6), 2))
        self.comboBox_over_temp_retry_railA.setCurrentIndex(int(initialize_feature_variable("500", 5, 3), 2))
        self.comboBox_over_temp_delay_railA.setCurrentIndex(int(initialize_feature_variable("500", 2, 0), 2))
        self.comboBox_over_temp_response_railB.setCurrentIndex(int(initialize_feature_variable("501", 7, 6), 2))
        self.comboBox_over_temp_retry_railB.setCurrentIndex(int(initialize_feature_variable("501", 5, 3), 2))
        self.comboBox_over_temp_delay_railB.setCurrentIndex(int(initialize_feature_variable("501", 2, 0), 2))

        self.comboBox_vout_under_voltage_response_railA.setCurrentIndex(int(initialize_feature_variable("450", 7, 6), 2))
        self.comboBox_vout_under_voltage_retry_railA.setCurrentIndex(int(initialize_feature_variable("450", 5, 3), 2))
        self.comboBox_vout_under_voltage_delay_railA.setCurrentIndex(int(initialize_feature_variable("450", 2, 0), 2))
        self.comboBox_vout_under_voltage_response_railB.setCurrentIndex(int(initialize_feature_variable("451", 7, 6), 2))
        self.comboBox_vout_under_voltage_retry_railB.setCurrentIndex(int(initialize_feature_variable("451", 5, 3), 2))
        self.comboBox_vout_under_voltage_delay_railB.setCurrentIndex(int(initialize_feature_variable("451", 2, 0), 2))

        self.comboBox_vin_OV_response.setCurrentIndex(int(initialize_feature_variable("56", 7, 6), 2))
        self.comboBox_vin_OV_retry.setCurrentIndex(int(initialize_feature_variable("56", 5, 3), 2))
        self.comboBox_vin_OV_delay.setCurrentIndex(int(initialize_feature_variable("56", 2, 0), 2))

        self.comboBox_vin_UV_response.setCurrentIndex(int(initialize_feature_variable("5A", 7, 6), 2))
        self.comboBox_vin_UV_retry.setCurrentIndex(int(initialize_feature_variable("5A", 5, 3), 2))
        self.comboBox_vin_UV_delay.setCurrentIndex(int(initialize_feature_variable("5A", 2, 0), 2))

        self.comboBox_vout_turn_on_time_response_railA.setCurrentIndex(int(initialize_feature_variable("630", 7, 6), 2))
        self.comboBox_vout_turn_on_time_response_railB.setCurrentIndex(int(initialize_feature_variable("631", 7, 6), 2))

        # Customer GUI interaction related function mapping
        self.comboBox_vout_turn_on_time_response_railA.activated.connect(self.vout_turn_on_time_response_changed_A)
        self.comboBox_vout_turn_on_time_response_railB.activated.connect(self.vout_turn_on_time_response_changed_B)

        self.comboBox_vin_OV_response.activated.connect(self.vin_OV_response_changed)
        self.comboBox_vin_OV_retry.activated.connect(self.vin_OV_retry_changed)
        self.comboBox_vin_OV_delay.activated.connect(self.vin_OV_delay_changed)

        self.comboBox_vin_UV_response.activated.connect(self.vin_UV_response_changed)
        self.comboBox_vin_UV_retry.activated.connect(self.vin_UV_retry_changed)
        self.comboBox_vin_UV_delay.activated.connect(self.vin_UV_delay_changed)

        self.comboBox_over_current_response_railA.activated.connect(self.over_current_response_changed_A)
        self.comboBox_over_current_retry_railA.activated.connect(self.over_current_retry_changed_A)
        self.comboBox_over_current_delay_railA.activated.connect(self.over_current_delay_changed_A)
        self.comboBox_over_current_response_railB.activated.connect(self.over_current_response_changed_B)
        self.comboBox_over_current_retry_railB.activated.connect(self.over_current_retry_changed_B)
        self.comboBox_over_current_delay_railB.activated.connect(self.over_current_delay_changed_B)

        self.comboBox_vout_over_voltage_response_railA.activated.connect(self.vout_over_voltage_response_changed_A)
        self.comboBox_vout_over_voltage_retry_railA.activated.connect(self.vout_over_voltage_retry_changed_A)
        self.comboBox_vout_over_voltage_delay_railA.activated.connect(self.vout_over_voltage_delay_changed_A)
        self.comboBox_vout_over_voltage_response_railB.activated.connect(self.vout_over_voltage_response_changed_B)
        self.comboBox_vout_over_voltage_retry_railB.activated.connect(self.vout_over_voltage_retry_changed_B)
        self.comboBox_vout_over_voltage_delay_railB.activated.connect(self.vout_over_voltage_delay_changed_B)

        self.comboBox_over_temp_response_railA.activated.connect(self.over_temp_response_changed_A)
        self.comboBox_over_temp_retry_railA.activated.connect(self.over_temp_retry_changed_A)
        self.comboBox_over_temp_delay_railA.activated.connect(self.over_temp_delay_changed_A)
        self.comboBox_over_temp_response_railB.activated.connect(self.over_temp_response_changed_B)
        self.comboBox_over_temp_retry_railB.activated.connect(self.over_temp_retry_changed_B)
        self.comboBox_over_temp_delay_railB.activated.connect(self.over_temp_delay_changed_B)

        self.comboBox_vout_under_voltage_response_railA.activated.connect(self.vout_under_voltage_response_changed_A)
        self.comboBox_vout_under_voltage_retry_railA.activated.connect(self.vout_under_voltage_retry_changed_A)
        self.comboBox_vout_under_voltage_delay_railA.activated.connect(self.vout_under_voltage_delay_changed_A)
        self.comboBox_vout_under_voltage_response_railB.activated.connect(self.vout_under_voltage_response_changed_B)
        self.comboBox_vout_under_voltage_retry_railB.activated.connect(self.vout_under_voltage_retry_changed_B)
        self.comboBox_vout_under_voltage_delay_railB.activated.connect(self.vout_under_voltage_delay_changed_B)

        self.pushButton_Discard.clicked.connect(self.Discard)
        self.pushButton_Save.clicked.connect(self.Save)

    def over_current_response_changed_A(self):
        if self.comboBox_over_current_response_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("470", 7, 6, "00")
        elif self.comboBox_over_current_response_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("470", 7, 6, "01")
        elif self.comboBox_over_current_response_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("470", 7, 6, "10")
        else:
            update_database_with_temp_customer_input("470", 7, 6, "11")

    def over_current_retry_changed_A(self):
        if self.comboBox_over_current_retry_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("470", 5, 3, "000")
        elif self.comboBox_over_current_retry_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("470", 5, 3, "001")
        elif self.comboBox_over_current_retry_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("470", 5, 3, "010")
        elif self.comboBox_over_current_retry_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("470", 5, 3, "011")
        elif self.comboBox_over_current_retry_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("470", 5, 3, "100")
        elif self.comboBox_over_current_retry_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("470", 5, 3, "101")
        elif self.comboBox_over_current_retry_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("470", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("470", 5, 3, "111")

    def over_current_delay_changed_A(self):
        if self.comboBox_over_current_delay_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("470", 2, 0, "000")
        elif self.comboBox_over_current_delay_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("470", 2, 0, "001")
        elif self.comboBox_over_current_delay_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("470", 2, 0, "010")
        elif self.comboBox_over_current_delay_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("470", 2, 0, "011")
        elif self.comboBox_over_current_delay_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("470", 2, 0, "100")
        elif self.comboBox_over_current_delay_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("470", 2, 0, "101")
        elif self.comboBox_over_current_delay_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("470", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("470", 2, 0, "111")

    def over_current_response_changed_B(self):
        if self.comboBox_over_current_response_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("471", 7, 6, "00")
        elif self.comboBox_over_current_response_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("471", 7, 6, "01")
        elif self.comboBox_over_current_response_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("471", 7, 6, "10")
        else:
            update_database_with_temp_customer_input("471", 7, 6, "11")

    def over_current_retry_changed_B(self):
        if self.comboBox_over_current_retry_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("471", 5, 3, "000")
        elif self.comboBox_over_current_retry_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("471", 5, 3, "001")
        elif self.comboBox_over_current_retry_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("471", 5, 3, "010")
        elif self.comboBox_over_current_retry_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("471", 5, 3, "011")
        elif self.comboBox_over_current_retry_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("471", 5, 3, "100")
        elif self.comboBox_over_current_retry_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("471", 5, 3, "101")
        elif self.comboBox_over_current_retry_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("471", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("471", 5, 3, "111")

    def over_current_delay_changed_B(self):
        if self.comboBox_over_current_delay_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("471", 2, 0, "000")
        elif self.comboBox_over_current_delay_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("471", 2, 0, "001")
        elif self.comboBox_over_current_delay_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("471", 2, 0, "010")
        elif self.comboBox_over_current_delay_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("471", 2, 0, "011")
        elif self.comboBox_over_current_delay_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("471", 2, 0, "100")
        elif self.comboBox_over_current_delay_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("471", 2, 0, "101")
        elif self.comboBox_over_current_delay_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("471", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("471", 2, 0, "111")

    def vout_over_voltage_response_changed_A(self):
        if self.comboBox_vout_over_voltage_response_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("410", 7, 6, "00")
        elif self.comboBox_vout_over_voltage_response_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("410", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("410", 7, 6, "10")

    def vout_over_voltage_retry_changed_A(self):
        if self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("410", 5, 3, "000")
        elif self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("410", 5, 3, "001")
        elif self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("410", 5, 3, "010")
        elif self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("410", 5, 3, "011")
        elif self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("410", 5, 3, "100")
        elif self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("410", 5, 3, "101")
        elif self.comboBox_vout_over_voltage_retry_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("410", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("410", 5, 3, "111")

    def vout_over_voltage_delay_changed_A(self):
        if self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("410", 2, 0, "000")
        elif self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("410", 2, 0, "001")
        elif self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("410", 2, 0, "010")
        elif self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("410", 2, 0, "011")
        elif self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("410", 2, 0, "100")
        elif self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("410", 2, 0, "101")
        elif self.comboBox_vout_over_voltage_delay_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("410", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("410", 2, 0, "111")

    def vout_over_voltage_response_changed_B(self):
        if self.comboBox_vout_over_voltage_response_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("411", 7, 6, "00")
        elif self.comboBox_vout_over_voltage_response_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("411", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("411", 7, 6, "10")

    def vout_over_voltage_retry_changed_B(self):
        if self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("411", 5, 3, "000")
        elif self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("411", 5, 3, "001")
        elif self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("411", 5, 3, "010")
        elif self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("411", 5, 3, "011")
        elif self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("411", 5, 3, "100")
        elif self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("411", 5, 3, "101")
        elif self.comboBox_vout_over_voltage_retry_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("411", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("411", 5, 3, "111")

    def vout_over_voltage_delay_changed_B(self):
        if self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("411", 2, 0, "000")
        elif self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("411", 2, 0, "001")
        elif self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("411", 2, 0, "010")
        elif self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("411", 2, 0, "011")
        elif self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("411", 2, 0, "100")
        elif self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("411", 2, 0, "101")
        elif self.comboBox_vout_over_voltage_delay_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("411", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("411", 2, 0, "111")

    def over_temp_response_changed_A(self):
        if self.comboBox_over_temp_response_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("500", 7, 6, "00")
        elif self.comboBox_over_temp_response_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("500", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("500", 7, 6, "10")

    def over_temp_retry_changed_A(self):
        if self.comboBox_over_temp_retry_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("500", 5, 3, "000")
        elif self.comboBox_over_temp_retry_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("500", 5, 3, "001")
        elif self.comboBox_over_temp_retry_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("500", 5, 3, "010")
        elif self.comboBox_over_temp_retry_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("500", 5, 3, "011")
        elif self.comboBox_over_temp_retry_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("500", 5, 3, "100")
        elif self.comboBox_over_temp_retry_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("500", 5, 3, "101")
        elif self.comboBox_over_temp_retry_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("500", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("500", 5, 3, "111")

    def over_temp_delay_changed_A(self):
        if self.comboBox_over_temp_delay_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("500", 2, 0, "000")
        elif self.comboBox_over_temp_delay_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("500", 2, 0, "001")
        elif self.comboBox_over_temp_delay_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("500", 2, 0, "010")
        elif self.comboBox_over_temp_delay_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("500", 2, 0, "011")
        elif self.comboBox_over_temp_delay_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("500", 2, 0, "100")
        elif self.comboBox_over_temp_delay_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("500", 2, 0, "101")
        elif self.comboBox_over_temp_delay_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("500", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("500", 2, 0, "111")

    def over_temp_response_changed_B(self):
        if self.comboBox_over_temp_response_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("501", 7, 6, "00")
        elif self.comboBox_over_temp_response_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("501", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("501", 7, 6, "10")

    def over_temp_retry_changed_B(self):
        if self.comboBox_over_temp_retry_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("501", 5, 3, "000")
        elif self.comboBox_over_temp_retry_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("501", 5, 3, "001")
        elif self.comboBox_over_temp_retry_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("501", 5, 3, "010")
        elif self.comboBox_over_temp_retry_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("501", 5, 3, "011")
        elif self.comboBox_over_temp_retry_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("501", 5, 3, "100")
        elif self.comboBox_over_temp_retry_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("501", 5, 3, "101")
        elif self.comboBox_over_temp_retry_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("501", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("501", 5, 3, "111")

    def over_temp_delay_changed_B(self):
        if self.comboBox_over_temp_delay_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("501", 2, 0, "000")
        elif self.comboBox_over_temp_delay_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("501", 2, 0, "001")
        elif self.comboBox_over_temp_delay_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("501", 2, 0, "010")
        elif self.comboBox_over_temp_delay_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("501", 2, 0, "011")
        elif self.comboBox_over_temp_delay_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("501", 2, 0, "100")
        elif self.comboBox_over_temp_delay_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("501", 2, 0, "101")
        elif self.comboBox_over_temp_delay_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("501", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("501", 2, 0, "111")

    def vout_under_voltage_response_changed_A(self):
        if self.comboBox_vout_under_voltage_response_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("450", 7, 6, "00")
        elif self.comboBox_vout_under_voltage_response_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("450", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("450", 7, 6, "10")

    def vout_under_voltage_retry_changed_A(self):
        if self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("450", 5, 3, "000")
        elif self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("450", 5, 3, "001")
        elif self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("450", 5, 3, "010")
        elif self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("450", 5, 3, "011")
        elif self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("450", 5, 3, "100")
        elif self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("450", 5, 3, "101")
        elif self.comboBox_vout_under_voltage_retry_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("450", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("450", 5, 3, "111")

    def vout_under_voltage_delay_changed_A(self):
        if self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("450", 2, 0, "000")
        elif self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("450", 2, 0, "001")
        elif self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 2:
            update_database_with_temp_customer_input("450", 2, 0, "010")
        elif self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 3:
            update_database_with_temp_customer_input("450", 2, 0, "011")
        elif self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 4:
            update_database_with_temp_customer_input("450", 2, 0, "100")
        elif self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 5:
            update_database_with_temp_customer_input("450", 2, 0, "101")
        elif self.comboBox_vout_under_voltage_delay_railA.currentIndex() == 6:
            update_database_with_temp_customer_input("410", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("450", 2, 0, "111")

    def vout_under_voltage_response_changed_B(self):
        if self.comboBox_vout_under_voltage_response_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("451", 7, 6, "00")
        elif self.comboBox_vout_under_voltage_response_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("451", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("451", 7, 6, "10")

    def vout_under_voltage_retry_changed_B(self):
        if self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("451", 5, 3, "000")
        elif self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("451", 5, 3, "001")
        elif self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("451", 5, 3, "010")
        elif self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("451", 5, 3, "011")
        elif self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("451", 5, 3, "100")
        elif self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("451", 5, 3, "101")
        elif self.comboBox_vout_under_voltage_retry_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("451", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("451", 5, 3, "111")

    def vout_under_voltage_delay_changed_B(self):
        if self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("451", 2, 0, "000")
        elif self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("451", 2, 0, "001")
        elif self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 2:
            update_database_with_temp_customer_input("451", 2, 0, "010")
        elif self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 3:
            update_database_with_temp_customer_input("451", 2, 0, "011")
        elif self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 4:
            update_database_with_temp_customer_input("451", 2, 0, "100")
        elif self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 5:
            update_database_with_temp_customer_input("451", 2, 0, "101")
        elif self.comboBox_vout_under_voltage_delay_railB.currentIndex() == 6:
            update_database_with_temp_customer_input("451", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("451", 2, 0, "111")

    def vin_OV_response_changed(self):
        if self.comboBox_vin_OV_response.currentIndex() == 0:
            update_database_with_temp_customer_input("56", 7, 6, "00")
        elif self.comboBox_vin_OV_response.currentIndex() == 1:
            update_database_with_temp_customer_input("56", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("56", 7, 6, "10")

    def vin_OV_retry_changed(self):
        if self.comboBox_vin_OV_retry.currentIndex() == 0:
            update_database_with_temp_customer_input("56", 5, 3, "000")
        elif self.comboBox_vin_OV_retry.currentIndex() == 1:
            update_database_with_temp_customer_input("56", 5, 3, "001")
        elif self.comboBox_vin_OV_retry.currentIndex() == 2:
            update_database_with_temp_customer_input("56", 5, 3, "010")
        elif self.comboBox_vin_OV_retry.currentIndex() == 3:
            update_database_with_temp_customer_input("56", 5, 3, "011")
        elif self.comboBox_vin_OV_retry.currentIndex() == 4:
            update_database_with_temp_customer_input("56", 5, 3, "100")
        elif self.comboBox_vin_OV_retry.currentIndex() == 5:
            update_database_with_temp_customer_input("56", 5, 3, "101")
        elif self.comboBox_vin_OV_retry.currentIndex() == 6:
            update_database_with_temp_customer_input("56", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("56", 5, 3, "111")

    def vin_OV_delay_changed(self):
        if self.comboBox_vin_OV_delay.currentIndex() == 0:
            update_database_with_temp_customer_input("56", 2, 0, "000")
        elif self.comboBox_vin_OV_delay.currentIndex() == 1:
            update_database_with_temp_customer_input("56", 2, 0, "001")
        elif self.comboBox_vin_OV_delay.currentIndex() == 2:
            update_database_with_temp_customer_input("56", 2, 0, "010")
        elif self.comboBox_vin_OV_delay.currentIndex() == 3:
            update_database_with_temp_customer_input("56", 2, 0, "011")
        elif self.comboBox_vin_OV_delay.currentIndex() == 4:
            update_database_with_temp_customer_input("56", 2, 0, "100")
        elif self.comboBox_vin_OV_delay.currentIndex() == 5:
            update_database_with_temp_customer_input("56", 2, 0, "101")
        elif self.comboBox_vin_OV_delay.currentIndex() == 6:
            update_database_with_temp_customer_input("56", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("56", 2, 0, "111")

    def vin_UV_response_changed(self):
        if self.comboBox_vin_UV_response.currentIndex() == 0:
            update_database_with_temp_customer_input("5A", 7, 6, "00")
        elif self.comboBox_vin_UV_response.currentIndex() == 1:
            update_database_with_temp_customer_input("5A", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("5A", 7, 6, "10")

    def vin_UV_retry_changed(self):
        if self.comboBox_vin_UV_retry.currentIndex() == 0:
            update_database_with_temp_customer_input("5A", 5, 3, "000")
        elif self.comboBox_vin_UV_retry.currentIndex() == 1:
            update_database_with_temp_customer_input("5A", 5, 3, "001")
        elif self.comboBox_vin_UV_retry.currentIndex() == 2:
            update_database_with_temp_customer_input("5A", 5, 3, "010")
        elif self.comboBox_vin_UV_retry.currentIndex() == 3:
            update_database_with_temp_customer_input("5A", 5, 3, "011")
        elif self.comboBox_vin_UV_retry.currentIndex() == 4:
            update_database_with_temp_customer_input("5A", 5, 3, "100")
        elif self.comboBox_vin_UV_retry.currentIndex() == 5:
            update_database_with_temp_customer_input("5A", 5, 3, "101")
        elif self.comboBox_vin_UV_retry.currentIndex() == 6:
            update_database_with_temp_customer_input("5A", 5, 3, "110")
        else:
            update_database_with_temp_customer_input("5A", 5, 3, "111")

    def vin_UV_delay_changed(self):
        if self.comboBox_vin_UV_delay.currentIndex() == 0:
            update_database_with_temp_customer_input("5A", 2, 0, "000")
        elif self.comboBox_vin_UV_delay.currentIndex() == 1:
            update_database_with_temp_customer_input("5A", 2, 0, "001")
        elif self.comboBox_vin_UV_delay.currentIndex() == 2:
            update_database_with_temp_customer_input("5A", 2, 0, "010")
        elif self.comboBox_vin_UV_delay.currentIndex() == 3:
            update_database_with_temp_customer_input("5A", 2, 0, "011")
        elif self.comboBox_vin_UV_delay.currentIndex() == 4:
            update_database_with_temp_customer_input("5A", 2, 0, "100")
        elif self.comboBox_vin_UV_delay.currentIndex() == 5:
            update_database_with_temp_customer_input("5A", 2, 0, "101")
        elif self.comboBox_vin_UV_delay.currentIndex() == 6:
            update_database_with_temp_customer_input("5A", 2, 0, "110")
        else:
            update_database_with_temp_customer_input("5A", 2, 0, "111")

    def vout_turn_on_time_response_changed_A(self):
        if self.comboBox_vout_turn_on_time_response_railA.currentIndex() == 0:
            update_database_with_temp_customer_input("630", 7, 6, "00")
        elif self.comboBox_vout_turn_on_time_response_railA.currentIndex() == 1:
            update_database_with_temp_customer_input("630", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("630", 7, 6, "10")

    def vout_turn_on_time_response_changed_B(self):
        if self.comboBox_vout_turn_on_time_response_railB.currentIndex() == 0:
            update_database_with_temp_customer_input("631", 7, 6, "00")
        elif self.comboBox_vout_turn_on_time_response_railB.currentIndex() == 1:
            update_database_with_temp_customer_input("631", 7, 6, "01")
        else:
            update_database_with_temp_customer_input("631", 7, 6, "10")

    def Save(self):
        global list_of_registers_used_in_this_frame, initial_device_status, homeWin_obj, parallel_thread, stop_thread
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                for i in list_of_registers_used_in_this_frame:
                    if register_database[i]['Final_register_value'] != register_database[i]['Temp_update_from_customer']:
                        # update the final register
                        register_database[i]['Final_register_value'] = register_database[i]['Temp_update_from_customer']
                        register_database[i]['Customer_interaction'] = "YES"
                        write_PMBUS_entry_in_command_xlsx(i)
                print_log("Fault Response settings saved.", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

    def Discard(self):
        print_log("Fault Response settings discarded.", "INFO")
        self.main = Fault_response()
        self.main.show()
        self.close()

# Custom command class
class PMBus_custom_commands(QMainWindow, Ui_PMBus_custom_commands):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame, register_database, PAGE
        super(PMBus_custom_commands, self).__init__(parent)
        self.setupUi(self)
        # self.thread = MyThread()
        # self.thread.any_signal.connect(lambda : self.label_ACK.setText(str(count)))
        # self.thread.start()

        # List of registers associated with used feature

        # Initialize temp update from customer field in database with final register value field at start of the frame.

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        self.cc=""
        self.not_paged = []
        for i in register_database:
            if len(i) == 2:
                self.not_paged.append(i)

        # Feature related initialization on GUI display
        if initial_device_status == 1:
            self.label_device_address.setText("--")
        else:
            self.label_device_address.setText(''.join(format(int(initialize_feature_variable("EF", 6, 0), 2), '02x')))
        self.label_command_type.setText("")
        self.label_command_type.setStyleSheet("color: rgb(255, 255, 255);\n"
                                              "background-color: rgb(160, 188, 173);\n"
                                              "border-radius:2px;")

        # page control details
        self.radioButton_railA.setText(RailA_name)
        self.radioButton_railB.setText(RailB_name)
        if PAGE == 0:
            self.radioButton_railA.setChecked(True)
            self.paged_command_data = "0x00"
        else:
            self.radioButton_railB.setChecked(True)
            self.paged_command_data = "0x01"

        self.temp_paged_command_data = ""

        self.radioButton_railB_3.setChecked(True)
        self.lineEdit_device_command.setText("")
        self.lineEdit_Payload.setText("")
        self.lineEdit_Payload.setDisabled(True)
        self.label_size.setText("")

        self.command_code = ""
        self.access = ""
        self.command_size = 1
        self.payload = ""
        self.command_type = self.label_command_type.text()
        self.temp_PMBus_custom_command = []

        self.lineEdit_Received_DATA.setText("--")
        self.lineEdit_Received_DATA.setDisabled(True)
        self.lineEdit_Received_DATA.setStyleSheet("QLineEdit{background-color: rgb(160, 188, 173);}")

        # Customer GUI interaction related function mapping
        self.lineEdit_Payload.textChanged.connect(self.payload_changed)
        self.radioButton_railA.toggled.connect(self.PMBus_addr_railA)
        self.radioButton_railB.toggled.connect(self.PMBus_addr_railB)
        self.pushButton_Discard.clicked.connect(self.PMBus_Send)
        self.lineEdit_device_command.textChanged.connect(self.Command_code_changed)
        self.radioButton_railA_3.toggled.connect(self.Read_command)
        self.radioButton_railB_3.toggled.connect(self.Write_command)

    def PMBus_addr_railA(self):
        global PAGE
        if self.radioButton_railA.isChecked():
            self.paged_command_data = "0x00"
            PAGE = 0

    def PMBus_addr_railB(self):
        global PAGE
        if self.radioButton_railB.isChecked():
            self.paged_command_data = "0x01"
            PAGE = 1

    def Read_command(self):
        if self.radioButton_railA_3.isChecked():
            if self.cc != "1B1":
                self.lineEdit_Payload.setText("--")
                self.lineEdit_Payload.setDisabled(True)
                self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(160, 188, 173);}")
            else:
                self.lineEdit_Payload.setText("")
                self.lineEdit_Payload.setDisabled(False)
                self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.Command_code_changed()
            self.lineEdit_Received_DATA.setText("")
            self.lineEdit_Received_DATA.setDisabled(False)
            self.lineEdit_Received_DATA.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")

    def Write_command(self):
        if self.radioButton_railB_3.isChecked():
            self.lineEdit_Payload.setText("")
            self.lineEdit_Payload.setDisabled(False)
            self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.Command_code_changed()
            self.lineEdit_Payload.setText("")
            self.lineEdit_Payload.setDisabled(False)
            self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
            self.lineEdit_Received_DATA.setText("--")
            self.lineEdit_Received_DATA.setDisabled(True)
            self.lineEdit_Received_DATA.setStyleSheet("QLineEdit{background-color: rgb(160, 188, 173);}")

    def Command_code_changed(self):
        global register_database,PAGE
        if self.lineEdit_device_command.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_device_command.text()) and int(
                    self.lineEdit_device_command.text(), base=16) <= 255 and len(
                    self.lineEdit_device_command.text()) < 3:
                # Remove the operation command register from register database temporarily
                dict1 = register_database.pop('020')
                dict2 = register_database.pop('021')
                self.cc = self.lineEdit_device_command.text().upper()
                if self.cc not in self.not_paged:
                    if PAGE == 0:
                        self.cc += "0"
                    else:
                        self.cc += "1"
                if self.cc in register_database.keys():
                    self.radioButton_railA_3.setCheckable(True)
                    self.command_code = "0x" + self.lineEdit_device_command.text()
                    self.command_size = int((register_database[self.cc]["size_in_bits"]) / 8)
                    self.label_size.setText(str(self.command_size))
                    self.lineEdit_Payload.setReadOnly(False)
                    self.lineEdit_Payload.setText("")
                    self.lineEdit_Payload.setDisabled(False)
                    self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                    if self.cc == "1B1" or self.cc == "1B0":
                        self.command_size = 2
                        self.label_size.setText(str(self.command_size))
                    if self.radioButton_railA_3.isChecked():
                        self.label_command_type.setText(register_database[self.cc]["Read_command_type"])
                        self.label_command_type.setStyleSheet("color: rgb(255, 255, 255);\n"
                                                              "background-color: rgb(160, 188, 173);\n"
                                                              "border-radius:2px;")
                        self.command_type = self.label_command_type.text()
                        if self.cc != "1B1" or self.cc != "1B0":
                            self.lineEdit_Payload.setText("--")
                            self.lineEdit_Payload.setDisabled(True)
                            self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(160, 188, 173);}")
                            self.payload = "NA"
                        else:
                            if self.radioButton_railA_3.isChecked():
                                self.label_size.setText("1")
                                self.command_size = 1
                            self.lineEdit_Payload.setText("")
                            self.lineEdit_Payload.setDisabled(False)
                            self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(255, 255, 255);}")
                            self.payload = ""
                    else:
                        self.label_command_type.setText(register_database[self.cc]["Write_command_type"])
                        self.label_command_type.setStyleSheet("color: rgb(255, 255, 255);\n"
                                                              "background-color: rgb(160, 188, 173);\n"
                                                              "border-radius:2px;")
                        self.command_type = self.label_command_type.text()

                elif self.lineEdit_device_command.text() == "03" or self.lineEdit_device_command.text() == "11" \
                        or self.lineEdit_device_command.text() == "12":
                    self.radioButton_railB_3.setChecked(True)
                    self.radioButton_railA_3.setCheckable(False)
                    self.label_size.setText("")
                    self.command_size = 1
                    self.payload = "NA"
                    self.label_command_type.setText("Send Byte")
                    self.label_command_type.setStyleSheet("color: rgb(255, 255, 255);\n"
                                                          "background-color: rgb(160, 188, 173);\n"
                                                          "border-radius:2px;")
                    self.command_type = "Send Byte"
                    self.command_code = "0x" + self.lineEdit_device_command.text()
                    self.lineEdit_Payload.setText("--")
                    self.lineEdit_Payload.setDisabled(True)
                    self.lineEdit_Payload.setStyleSheet("QLineEdit{background-color: rgb(160, 188, 173);}")

                else:
                    print_log("Command code [ 0x"+self.lineEdit_device_command.text()+" ] is not supported", "ERROR")
                    self.lineEdit_Payload.setReadOnly(True)
                    self.label_size.setText("")
                    self.label_command_type.setText("Not supported")
                    self.label_command_type.setStyleSheet("color: rgb(255, 0, 0);\n"
                                                          "background-color: rgb(160, 188, 173);\n"
                                                          "border-radius:2px;")

                # Adding the operation command register back to register database
                register_database['020'] = dict1
                register_database['021'] = dict2

            else:
                print_log(
                    "Entered " + self.lineEdit_device_command.text() + ", give command code ranging (0x00 - 0xFF)",
                    "ERROR")
                time.sleep(2)
                self.label_command_type.setText("")
                self.lineEdit_device_command.setText("")
                self.command_code = ""
                self.label_size.setText("")
                self.command_size = ""

    def payload_changed(self):
        if self.radioButton_railB_3.isChecked():
            self.access = "W"
            if self.lineEdit_Payload.text():
                if self.cc != "1B1" or self.cc !="1B0":
                    if all(ch in string.hexdigits for ch in self.lineEdit_Payload.text()) and int(
                            self.lineEdit_Payload.text(), base=16) <= (
                            2 ** (register_database[self.cc]["size_in_bits"])) - 1 \
                            and len(self.lineEdit_Payload.text()) <= (
                            register_database[self.cc]["size_in_bits"] / 4):
                        self.payload = "0x" + self.lineEdit_Payload.text()
                        self.command_size = int(register_database[self.cc]["size_in_bits"] / 8)

                    else:
                        if self.lineEdit_Payload.text() != "--":
                            print_log("Entered " + self.lineEdit_Payload.text() + ", give payload ranging (0x" +
                                      "0".zfill(int(register_database[self.cc]["size_in_bits"] / 4)) + "- 0x" +
                                      "F".ljust(int(register_database[self.cc]["size_in_bits"] / 4), 'F') + "). Expected "+str(int(register_database[self.cc]["size_in_bits"] / 4))+ " hex digits but received "+str(len(self.lineEdit_Payload.text()))+".",
                                      "ERROR")
                            time.sleep(2)
                            self.lineEdit_Payload.setText("")
                            self.payload = ""
                            self.command_size = ""
                        else:
                            self.command_size = 1

                elif self.cc == "1B1" or self.cc == "1B0":
                    if all(ch in string.hexdigits for ch in self.lineEdit_Payload.text()) and int(
                            self.lineEdit_Payload.text(), base=16) <= 65535 and len(self.lineEdit_Payload.text()) <= 4:
                        self.payload = "0x" + self.lineEdit_Payload.text()
                        self.command_size = 2
                    else:
                        if self.lineEdit_Payload.text() != "--":
                            print_log("Entered " + self.lineEdit_Payload.text() +
                                      ", give payload ranging (0x0000 - 0xFFFF)", "ERROR")
                            time.sleep(2)
                            self.lineEdit_Payload.setText("")
                            self.payload = ""
                            self.command_size = ""
                        else:
                            self.command_size = 1

        else:
            self.access = "R"
            if self.cc != "1B1" or  self.cc != "1B0":
                self.payload = "NA"
            else:
                if self.lineEdit_Payload.text():
                    if all(ch in string.hexdigits for ch in self.lineEdit_Payload.text()) and int(
                            self.lineEdit_Payload.text(), base=16) <= 255 and len(self.lineEdit_Payload.text()) <= 2:
                        self.payload = "0x" + self.lineEdit_Payload.text()
                    else:
                        print_log("Entered " + self.lineEdit_Payload.text() +
                                  ", give payload ranging (0x00 - 0xFF)", "ERROR")
                        time.sleep(2)
                        self.lineEdit_Payload.setText("")

    def PMBus_Send(self):
        global next_row_pointer_command_xlsx, stop_thread, parallel_thread, initial_device_status, PMBus_send, PAGE, homeWin_obj
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                PMBus_send = True
                if PAGE == 0:
                    self.radioButton_railA.setChecked(True)
                    self.paged_command_data = "0x00"
                else:
                    self.radioButton_railB.setChecked(True)
                    self.paged_command_data = "0x01"

                # special command page handling
                if ((self.cc == "331") or (self.cc == "621") or (self.cc == "601") or (self.cc == "641") or
                       (self.cc == "271")) and int(self.paged_command_data, 16) and self.radioButton_railA_3.isChecked():
                    # print(self.cc, self.paged_command_data)
                    self.paged_command_data = "0xff"
                else:
                    if PAGE == 0:
                        self.paged_command_data = "0x00"
                    else:
                        self.paged_command_data = "0x01"

                # We need to write the PAGE i.e. self.paged_command_data

                workbook = openpyxl.load_workbook("command.xlsx")
                worksheet = workbook.active
                worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
                worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = "0x00"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = "W"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = "1"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = self.paged_command_data
                worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
                worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
                worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = "Write Byte"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
                worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
                workbook.save("command.xlsx")
                next_row_pointer_command_xlsx += 1

                if self.command_code != "" and self.payload != "":
                    if self.radioButton_railB_3.isChecked():
                        if self.lineEdit_Payload.text() != "--":
                            if len(self.lineEdit_Payload.text()) == int(int(self.command_size) * 2):
                                # self.temp_PMBus_custom_command = [self.command_code, self.access, self.command_size, self.payload,
                                #                                   self.command_type, self.paged_command_data]
                                workbook = openpyxl.load_workbook("command.xlsx")
                                worksheet = workbook.active
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = self.command_code
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = self.access
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = str(self.command_size)
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = self.payload
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = self.command_type
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
                                workbook.save("command.xlsx")
                                next_row_pointer_command_xlsx += 1
                                # print(self.command_code, self.access, self.command_size, self.payload, self.command_type)
                                print_log("PMBUS command has been written", "INFO")

                                # Update the register database final register value & temp register value when written
                                # through PMBus custom command
                                data = bin(int(self.payload[2:], 16))[2:].zfill(len(self.payload[2:]) * 4)
                                if self.cc not in self.not_paged:
                                    register_database[self.cc[:2] + str(PAGE)]["Temp_update_from_customer"] = data
                                    register_database[self.cc[:2] + str(PAGE)]["Final_register_value"] = data
                                else:
                                    register_database[self.cc]["Temp_update_from_customer"] = data
                                    register_database[self.cc]["Final_register_value"] = data

                            else:
                                print_log("Entered " + self.lineEdit_Payload.text() + ", Expected payload is "
                                          + str(int(self.command_size * 2)) + " hex digits but received "
                                          + str(len(self.lineEdit_Payload.text())) + " hex digits", "ERROR")
                                self.lineEdit_Payload.setText("")
                        else:
                            # self.temp_PMBus_custom_command = [self.command_code, self.access, self.command_size, self.payload,
                            #                                   self.command_type, self.paged_command_data]
                            workbook = openpyxl.load_workbook("command.xlsx")
                            worksheet = workbook.active
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = self.command_code
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = self.access
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = str(self.command_size)
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = self.payload
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = self.command_type
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
                            workbook.save("command.xlsx")
                            next_row_pointer_command_xlsx += 1
                            # print(self.command_code, self.access, self.command_size, self.payload, self.command_type)
                            print_log("PMBUS command has been written", "INFO")
                    else:
                        # self.temp_PMBus_custom_command = [self.command_code, self.access, self.command_size, self.payload,
                        #                                   self.command_type, self.paged_command_data]
                        if self.cc == "1B1":
                            if len(self.lineEdit_Payload.text()) == int(int(self.command_size) * 2):
                                workbook = openpyxl.load_workbook("command.xlsx")
                                worksheet = workbook.active
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = self.command_code
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = self.access
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = str(self.command_size)
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = self.payload
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = self.command_type
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
                                worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
                                workbook.save("command.xlsx")
                                next_row_pointer_command_xlsx += 1
                                # print(self.command_code, self.access, self.command_size, self.payload, self.command_type)
                                print_log("PMBUS command has been written", "INFO")
                            else:
                                # print_log("Entered " + self.lineEdit_Payload.text() + ", Give a valid payload of "
                                #           + str(self.command_size) + " byte", "ERROR")
                                print_log("Entered " + self.lineEdit_Payload.text() + ", Expected payload is "
                                          + str(int(self.command_size * 2)) + " hex digits but received "
                                          + str(len(self.lineEdit_Payload.text())) + " hex digits", "ERROR")
                                self.lineEdit_Payload.setText("")
                        else:
                            # if ((self.cc == "331") or (self.cc == "621") or (self.cc == "601") or (self.cc == "641") or (
                            #         self.cc == "271")) and int(self.paged_command_data, 16):
                            workbook = openpyxl.load_workbook("command.xlsx")
                            worksheet = workbook.active
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "PMBUS"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = self.command_code
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = self.access
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = str(self.command_size)
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = self.payload
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = PMBus_freq
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = PMBus_parity
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = "NA"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=10).value = "NA"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = self.command_type
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
                            worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
                            workbook.save("command.xlsx")
                            next_row_pointer_command_xlsx += 1
                            # print(self.command_code, self.access, self.command_size, self.payload, self.command_type)
                            print_log("PMBUS command has been written", "INFO")
                time.sleep(0.1)
                stop_thread = False
                parallel_thread.start()
                time.sleep(0.1)
                # PMBus_send = False
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")

class SVID_custom_commands(QMainWindow, Ui_SVID_custom_commands):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, list_of_registers_used_in_this_frame
        super(SVID_custom_commands, self).__init__(parent)
        self.setupUi(self)
        # self.thread = MyThread()
        # List of registers associated with used feature

        # Initialize temp update from customer field in database with final register value field at start of the frame.

        # Used feature: global variable initialization # Do not assign if any global feature variable is used for reading in a frame.
        # GUI default value look initialization
        # Non feature related initializaton on GUI_ display

        # Feature related initialization on GUI display
        self.radioButton_railA.setText(RailA_name)
        self.radioButton_railB.setText(RailB_name)
        self.radioButton_railA.setChecked(True)
        self.svid_address = "0x" + format(int(initialize_feature_variable("E6", 7, 4), 2), '02x').zfill(2)
        self.radioButton_railB.setChecked(False)
        self.comboBox_SVID_command.setCurrentIndex(0)
        self.SVID_command = "0x" + format(self.comboBox_SVID_command.currentIndex() + 1, '02x').zfill(2)
        self.lineEdit_Payload.setText("00")
        self.payload = "0x00"
        self.lineEdit_Received_DATA.setReadOnly(True)
        self.lineEdit_Received_DATA.setStyleSheet("QLineEdit{background-color: rgb(160, 188, 173);}")
        self.temp_svid_data = []
        self.label_ACK.setText("")

        # Customer GUI interaction related function mapping
        self.radioButton_railA.toggled.connect(self.SVID_addr_rail0)
        self.radioButton_railB.toggled.connect(self.SVID_addr_rail1)
        self.comboBox_SVID_command.activated.connect(self.SVID_command_changed)
        self.lineEdit_Payload.textChanged.connect(self.Payload_changed)
        self.pushButton_Discard.clicked.connect(self.SVID_Send)

    def SVID_addr_rail0(self):
        if self.radioButton_railA.isChecked():
            self.svid_address = "0x" + format(int(initialize_feature_variable("E6", 7, 4), 2), '02x').zfill(2)

    def SVID_addr_rail1(self):
        if self.radioButton_railB.isChecked():
            self.svid_address = "0x" + format(int(initialize_feature_variable("E6", 3, 0), 2), '02x').zfill(2)

    def SVID_command_changed(self):
        if self.comboBox_SVID_command.currentIndex() == 7:
            self.SVID_command = "0x09"
        else:
            self.SVID_command = "0x" + format(self.comboBox_SVID_command.currentIndex() + 1, '02x').zfill(2)

    def Payload_changed(self):
        if self.lineEdit_Payload.text():
            if all(ch in string.hexdigits for ch in self.lineEdit_Payload.text()) and int(
                    self.lineEdit_Payload.text(), base=16) <= 255 and len(self.lineEdit_Payload.text()) <= 2:
                self.payload = "0x" + self.lineEdit_Payload.text().zfill(2)
            else:
                print_log("Entered " + self.lineEdit_Payload.text() + ", give payload ranging (0x00 - 0xFF)", "ERROR")
                time.sleep(2)
                self.lineEdit_Payload.setText("")

    def SVID_Send(self):
        global next_row_pointer_command_xlsx, stop_thread, parallel_thread, initial_device_status, homeWin_obj
        if initial_device_status == 0:
            if homeWin_obj.pushButton_Enable_VR.isChecked():
                stop_thread = True
                time.sleep(1)
                # if self.temp_svid_data != [self.SVID_command, self.svid_address, self.payload]:
                #     self.temp_svid_data = [self.SVID_command, self.svid_address, self.payload]
                if len(self.lineEdit_Payload.text()) == 2:
                    workbook = openpyxl.load_workbook("command.xlsx")
                    worksheet = workbook.active
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=1).value = str(next_row_pointer_command_xlsx - 1)
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=2).value = "SVID"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=3).value = self.SVID_command
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=4).value = "WR"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=5).value = "1"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=6).value = self.payload
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=7).value = "25MHz"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=8).value = "NA"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=9).value = self.svid_address
                    worksheet.cell(row=next_row_pointer_command_xlsx,
                                   column=10).value = self.comboBox_SVID_command.currentText()
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=11).value = "NA"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=12).value = "NA"
                    worksheet.cell(row=next_row_pointer_command_xlsx, column=13).value = "NA"
                    workbook.save("command.xlsx")
                    next_row_pointer_command_xlsx += 1
                    print_log("SVID command has been written", "INFO")
                else:
                    print_log("Entered " + self.lineEdit_Payload.text() + ", Give a valid payload of 1 byte", "ERROR")
                stop_thread = False
                parallel_thread.start()
            else:
                print_log("Enable VR to write into the device", "WARNING")
        else:
            print_log("TI dongle or V3p3 is not connected, check the connections", "ERROR")


class SVI2_custom_command(QMainWindow, Ui_Form):
    def __init__(self, parent=None):
        super().__init__()
        super(SVI2_custom_command, self).__init__(parent)
        self.setupUi(self)
        self.lineEdit_vid.setText('0')
        self.pushButton.clicked.connect(self.send_command)
        self.pushButton_tele.clicked.connect(self.receive_telemetry)
        self.radioButton_votf.setDisabled(True)

        workbook = openpyxl.open('telemetry_data.xlsx')
        worksheet = workbook.active
        header = ["Vout1", "Vout2", "Iout1", "Iout2"]
        worksheet.append(header)
        workbook.save("telemetry_data.xlsx")

    def send_command(self):
        global stop_thread, parallel_thread
        stop_thread = True
        time.sleep(1)
        loadline = self.comboBox_ll_trim.currentIndex()
        offset = self.comboBox_offset.currentIndex()
        vdd1_s = self.comboBox_vdd1.currentIndex()
        vdd2_s = self.comboBox_vdd2.currentIndex()
        psi0_l_s = self.comboBox_psi0.currentIndex()
        psi1_l_s = self.comboBox_psi1.currentIndex()
        tfn_s = self.comboBox_tfn.currentIndex()
        clk_s = self.lineEdit_clk.text()
        if (float(self.lineEdit_vid.text()) > 1.55):
            QMessageBox.about(self, "Error", "Enter Valid Voltage value between 0V and 1.55V")
        else:
            vid = 248 - int(float(self.lineEdit_vid.text()) * 1000.0 / 6.25)
            votf = exec_command(vdd1_s, vdd2_s, vid, psi0_l_s, psi1_l_s, tfn_s, loadline, offset, clk_s)

        if (votf == 1):
            self.radioButton_votf.setChecked(True)
        else:
            self.radioButton_votf.setChecked(False)
        stop_thread = False
        parallel_thread.start()

    def receive_telemetry(self):
        global parallel_thread, stop_thread
        stop_thread = True
        time.sleep(1)
        if (self.comboBox_tfn.currentIndex() == 1 and self.comboBox_vdd1.currentIndex() == 1 and self.comboBox_vdd2.currentIndex() == 0):
            QMessageBox.about(self, "Error", "Telemetry is disabled")
        else:
            decode_telemetry()
        stop_thread = False
        parallel_thread.start()

# Home window class
class HomeWindow(QMainWindow, Ui_Home):
    def __init__(self, parent=None):
        global RailA_name, RailB_name, PARTNAME, RailA_phase_count_arg, RailB_phase_count_arg, PMBUS_ADDR, resolutionA, resolutionB, device_status, parallel_thread, homeWin_obj, queue
        global VR_Enabled, list_of_registers_used_in_this_frame, initial_device_status, svid_address
        # loading GUI frame and initializing it.
        super(HomeWindow, self).__init__(parent)
        self.setupUi(self)

        # GUI look initialization
        # Device setting-> remove Eco phase configuration
        self.label_Device_Status_2.setVisible(False)
        self.comboBox_Eco_phase.setVisible(False)               # eco phase feature is disabled in device configuration settings

        # hiding IMON AUX from telemetry display
        self.frame_EFFICIENCY.hide()# IMON_AUX requires reading SVID register, hence removed.
        self.frame_VIN.hide() # PSYS value is not getting updated from SVID register to PMBUS register, hence removing monitor

        #Readjusting telemetry monitors.
        self.frame_IMONA.setGeometry(QtCore.QRect(40, 300, 150, 150))
        self.frame_IMONB.setGeometry(QtCore.QRect(230, 300, 150, 150))
        self.frame_VOUTA.setGeometry(QtCore.QRect(40, 100, 150, 150))
        self.frame_VOUTB.setGeometry(QtCore.QRect(230, 100, 150, 150))

        # self.frame_31.setGeometry(QtCore.QRect(10, 10, 150, 150))
        # Readjusting PSYS monitor
        # self.frame_VIN.setGeometry(QtCore.QRect(140, 360, 150, 150))

        self.label.setText(PARTNAME)
        self.label_RAILA.setText(RailA_name)
        self.label_RAILB.setText(RailB_name)

        RailA_phase_count_arg = str(int(initialize_feature_variable("DF", 8, 4), 2))
        RailB_phase_count_arg = str(int(initialize_feature_variable("DF", 3, 0), 2))
        self.label_phase_A.setText(RailA_phase_count_arg)
        self.label_phase_B.setText(RailB_phase_count_arg)

        self.label_PMBus_addressA.setText(PMBUS_ADDR + "h")
        self.label_PMBus_addressB.setText(PMBUS_ADDR + "h")
        self.label_boot_vol_A.setText(str(round(float((int(initialize_feature_variable('E10', 23, 13), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13), 2) != 0 else int(0)) + "V")
        self.label_boot_vol_B.setText(str(round(float((int(initialize_feature_variable('E11', 23, 13), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13), 2) != 0 else int(0)) + "V")

        # Device Configuration feature initialization
        list_of_registers_used_in_this_frame = ["DF", "C9", "C50", "C51"]
        initialize_Temp_update_from_customer(list_of_registers_used_in_this_frame)

        if int(initialize_feature_variable("DF", 9, 9), 2) == 1:
            self.pushButton_Low_power_std_mode.setChecked(True)

            self.pushButton_Low_power_std_mode.setStyleSheet("QPushButton{\n"
                                                             "border-image: url(GUI_IMAGE/but1.png);\n"
                                                             "}\n"
                                                             "\n"
                                                             "\n"
                                                             "QPushButton::hover {\n"
                                                             "border-image: url(GUI_IMAGE/but1_hover.png);\n"
                                                             "    }\n"
                                                             "")
        else:
            self.pushButton_Low_power_std_mode.setChecked(False)
            self.pushButton_Low_power_std_mode.setStyleSheet("QPushButton{\n"
                                                             "border-image: url(GUI_IMAGE/but0.png);\n"
                                                             "}\n"
                                                             "\n"
                                                             "\n"
                                                             "QPushButton::hover {\n"
                                                             "border-image: url(GUI_IMAGE/but0_hover.png);\n"
                                                             "    }\n"
                                                             "")

        # Eco phase configuration mode # we are reading railA and updating it for RailB too.
        if int(initialize_feature_variable("C9", 31, 31), 2) == 0:
            if int(initialize_feature_variable("C50", 85, 85), 2) == 0:
                self.comboBox_Eco_phase.setCurrentIndex(0)
            else:
                self.comboBox_Eco_phase.setCurrentIndex(1)
        else:
            if int(initialize_feature_variable("C50", 85, 85), 2) == 0:
                self.comboBox_Eco_phase.setCurrentIndex(2)
            else:
                self.comboBox_Eco_phase.setCurrentIndex(3)

            # ENable VR case handling
        if device_status == "ON":
            self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                    "border-image: url(GUI_IMAGE/but1.png);\n"
                                                    "}\n"
                                                    "\n"
                                                    "\n"
                                                    "QPushButton::hover {\n"
                                                    "border-image: url(GUI_IMAGE/but1_hover"
                                                    ".png);\n "
                                                    "    }\n"
                                                    "")

            VR_Enabled = "ON"

        elif device_status == "OFF":
            self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                    "border-image: url(GUI_IMAGE/but0.png);\n"
                                                    "}\n"
                                                    "\n"
                                                    "\n"
                                                    "QPushButton::hover {\n"
                                                    "border-image: url(GUI_IMAGE/but0_hover"
                                                    ".png);\n "
                                                    "    }\n"
                                                    "")
            VR_Enabled = "OFF"
            self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_off_button.png"))

            # Part based label changing
        if PARTNAME == "AMP4592":
            self.pushButton_SVID_Configuration.setText("SVID Configuration")
            self.pushButton_Custom_protocol_config.setText("SVID")
            self.label_prot_A.setText("SVID:")
            self.label_prot_B.setText("SVID:")
            self.label_prot_address_A.setText(svid_address + "h")
            self.label_prot_address_B.setText(str(int(svid_address) + 1) + "h")
        elif PARTNAME == "AMP4692":
            self.pushButton_SVID_Configuration.setText("SVI2 Configuration")
            self.pushButton_Custom_protocol_config.setText("SVI2")
            self.label_prot_A.setText("SVI2:")
            self.label_prot_B.setText("SVI2:")
            self.label_prot_address_A.setText("--")
            self.label_prot_address_B.setText("--")
        elif PARTNAME == "AMP4792":
            self.pushButton_SVID_Configuration.setText("SVI3 Configuration")
            self.pushButton_Custom_protocol_config.setText("SVI3")
        elif PARTNAME == "AMP4291":
            self.pushButton_SVID_Configuration.setText("OVR Configuration")
            self.pushButton_Custom_protocol_config.setText("OVR")
        else:
            self.pushButton_SVID_Configuration.setText("Unknown")

        # PMBUS connected handling
        # if PMBus_connected == "YES":
        #     self.pushButton_PMBus_Enable.setStyleSheet("QPushButton{\n"
        #                                                "border-image: url(GUI_IMAGE/but1.png);\n"
        #                                                "}\n"
        #                                                "\n"
        #                                                "\n"
        #                                                "QPushButton::hover {\n"
        #                                                "border-image: url(GUI_IMAGE/but1_hover.png);\n"
        #                                                "    }\n"
        #                                                "")
        #     self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_on_button.png"))
        #
        # elif PMBus_connected == "NO":
        #     self.pushButton_PMBus_Enable.setStyleSheet("QPushButton{\n"
        #                                                "border-image: url(GUI_IMAGE/but0.png);\n"
        #                                                "}\n"
        #                                                "\n"
        #                                                "\n"
        #                                                "QPushButton::hover {\n"
        #                                                "border-image: url(GUI_IMAGE/but0_hover.png);\n"
        #                                                "    }\n"
        #                                                "")
        #     self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_off_button.png"))

        # Resetting the telemetry display values to 0, when refresh is selected
        self.label_slew_A.setText(str(0) + " V")
        self.label_slew_B.setText(str(0) + " V")
        self.label_absolute7.setText(
            "0" + "/" + str(int(initialize_feature_variable("E9", 33, 26), 2)) + " W")
        self.label_per7.setText(
            "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
            "vertical-align:super;\">%</span></p>")
        self.label_absolute8.setText(
            "0" + "/" + str(int(initialize_feature_variable("E9", 61, 54), 2)) + " W")
        self.label_per8.setText(
            "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
            "vertical-align:super;\">%</span></p>")
        self.label_absolute3.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
        self.label_per3.setText(
            "<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
            "font-size:20pt;\">V</span></p></body></html>")
        self.label_absolute5.setText(
            "Max:" + str(int(initialize_feature_variable("E80", 55, 48), 2)) + " A")
        self.label_per5.setText(
            "<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
            "font-size:18pt;\">A</span></p></body></html>")
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setBold(False)
        font.setWeight(50)
        font.setPointSize(20)
        self.label_temp_A.setFont(font)
        self.label_temp_A.setText(str(0))
        self.label_temp_B.setFont(font)
        self.label_temp_B.setText(str(0))
        self.label_absolute4.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
        self.label_per4.setText(
            "<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
            "font-size:20pt;\">V</span></p></body></html>")
        self.label_absolute6.setText(
            "Max:" + str(int(initialize_feature_variable("E81", 55, 48), 2)) + " A")
        self.label_per6.setText(
            "<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
            "font-size:18pt;\">A</span></p></body></html>")

        # Controller ON_OFF Config
        queue = deque()  # queuing to print in the print log function
        # self.dio = DIO()
        # self.dio.ni845xOpen()  # opening the TI dongle to check for his addr
        # self.dio.ni845xClose()
        # V3p3_not_present = 1
        try:
            self.device_handler = usb_to_gpio.USB_TO_GPIO()
        except Exception as err:
            print("oh")
            print(err)

        if self.device_handler.adapter_status == 0:  # returns 0 if not connected
            TI_dongle_not_present = 1
            queue.appendleft("TI Dongle is not connected, and hence PMBus will not be accessible.%ERROR")
            device_status = "OFF"
            self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_off_button.png"))
            self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
            self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
            self.pushButton_Enable_VR.setChecked(False)
            self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                    "border-image: url(GUI_IMAGE/but0.png);\n"
                                                    "}\n"
                                                    "\n"
                                                    "\n"
                                                    "QPushButton::hover {\n"
                                                    "border-image: url(GUI_IMAGE/but0_hover"
                                                    ".png);\n "
                                                    "    }\n"
                                                    "")
            # self.pushButton_Enable_VR.setDisabled(True)
        else:
            TI_dongle_not_present = 0
            if is_PMBus_connected() == "YES":
                value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                                                   pullup_enable=1, clockrate=100, writesize=1, writedata=[0x02],
                                                   noofbytestoread=1)
                device_status = "ON"
                value = format(value[0], "08b")
                if (value[3:6] == "000") or (value[3:6] == "001") or (value[3:6] == "010") or (value[3:6] == "011") or (value[3:6] == "100"):
                    VR_Enabled = "ON"
                    # power_up_device()
                    self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_on_button.png"))
                    self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_green_icon.png);")
                    self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_green_icon.png);")
                    self.pushButton_Enable_VR.setChecked(True)
                    self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                            "border-image: url(GUI_IMAGE/but1.png);\n"
                                                            "}\n"
                                                            "\n"
                                                            "\n"
                                                            "QPushButton::hover {\n"
                                                            "border-image: url(GUI_IMAGE/but1_hover"
                                                            ".png);\n "
                                                            "    }\n"
                                                            "")
                    self.pushButton_Enable_VR.setDisabled(True)
                    queue.appendleft("Controller has been configured to turn on with V3p3 supply.%INFO")

                elif (value[3:6] == "101") or (value[3:6] == "111") or (value[3:6] == "110"):
                    # VR_Enabled = "OFF"
                    # power_down_device()
                    if VR_Enabled == "OFF":
                        self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_off_button.png"))
                        self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
                        self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
                        self.pushButton_Enable_VR.setDisabled(False)
                        self.pushButton_Enable_VR.setChecked(False)
                        self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                                "border-image: url(GUI_IMAGE/but0.png);\n"
                                                                "}\n"
                                                                "\n"
                                                                "\n"
                                                                "QPushButton::hover {\n"
                                                                "border-image: url(GUI_IMAGE/but0_hover"
                                                                ".png);\n "
                                                                "    }\n"
                                                                "")
                        queue.appendleft("V3p3 supply is present. "
                                         "Toggle Enable VR to turn on the chip.%INFO")

                        # Resetting the telemetry display values to 0, when enable is low
                        self.label_slew_A.setText(str(0) + " V")
                        self.label_slew_B.setText(str(0) + " V")
                        self.label_absolute7.setText(
                            "0" + "/" + str(int(initialize_feature_variable("E9", 33, 26), 2)) + " W")
                        self.label_per7.setText(
                            "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                            "vertical-align:super;\">%</span></p>")
                        self.label_absolute8.setText(
                            "0" + "/" + str(int(initialize_feature_variable("E9", 61, 54), 2)) + " W")
                        self.label_per8.setText(
                            "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                            "vertical-align:super;\">%</span></p>")
                        self.label_absolute3.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
                        self.label_per3.setText(
                            "<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
                            "font-size:20pt;\">V</span></p></body></html>")
                        self.label_absolute5.setText(
                            "Max:" + str(int(initialize_feature_variable("E80", 55, 48), 2)) + " A")
                        self.label_per5.setText(
                            "<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
                            "font-size:18pt;\">A</span></p></body></html>")
                        font = QtGui.QFont()
                        font.setFamily("Arial")
                        font.setBold(False)
                        font.setWeight(50)
                        font.setPointSize(20)
                        self.label_temp_A.setFont(font)
                        self.label_temp_A.setText(str(0))
                        self.label_temp_B.setFont(font)
                        self.label_temp_B.setText(str(0))
                        self.label_absolute4.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
                        self.label_per4.setText(
                            "<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
                            "font-size:20pt;\">V</span></p></body></html>")
                        self.label_absolute6.setText(
                            "Max:" + str(int(initialize_feature_variable("E81", 55, 48), 2)) + " A")
                        self.label_per6.setText(
                            "<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
                            "font-size:18pt;\">A</span></p></body></html>")
                    else:
                        self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_on_button.png"))
                        self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_green_icon.png);")
                        self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_green_icon.png);")
                        self.pushButton_Enable_VR.setDisabled(False)
                        self.pushButton_Enable_VR.setChecked(True)
                        self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                                "border-image: url(GUI_IMAGE/but1.png);\n"
                                                                "}\n"
                                                                "\n"
                                                                "\n"
                                                                "QPushButton::hover {\n"
                                                                "border-image: url(GUI_IMAGE/but1_hover"
                                                                ".png);\n "
                                                                "    }\n"
                                                                "")

            else:
                queue.appendleft("PMBus address entered is wrong or check the V3p3 supply connection.%ERROR")
                device_status = "OFF"
                self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_off_button.png"))
                self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
                self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
                self.pushButton_Enable_VR.setChecked(False)
                self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                        "border-image: url(GUI_IMAGE/but0.png);\n"
                                                        "}\n"
                                                        "\n"
                                                        "\n"
                                                        "QPushButton::hover {\n"
                                                        "border-image: url(GUI_IMAGE/but0_hover"
                                                        ".png);\n "
                                                        "    }\n"
                                                        "")
                # self.pushButton_Enable_VR.setDisabled(True)

        initial_device_status = TI_dongle_not_present or (device_status == "OFF")

        # overriding to 0 to get access to save the commands in Command.xlsx
        # initial_device_status = 0
        if initial_device_status == 1:
            # print_log("Check the PMBus address and V3p3 supply connection", "ERROR")
            pass

        # if parallel_thread.V3p3_comparator() == 0:          # checking v3p3 status using a DIO line, either 3.3V pullup on dongle or device
        #     V3p3_not_present = 1
        #     queue.appendleft("V3p3 is not connected, check the connection%ERROR")
        #     device_status = "OFF"
        #     self.pushButton_Enable_VR.setChecked(False)
        #     self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
        #                                             "border-image: url(GUI_IMAGE/but0.png);\n"
        #                                             "}\n"
        #                                             "\n"
        #                                             "\n"
        #                                             "QPushButton::hover {\n"
        #                                             "border-image: url(GUI_IMAGE/but0_hover"
        #                                             ".png);\n "
        #                                             "    }\n"
        #                                             "")
        #     self.pushButton_Enable_VR.setDisabled(True)
        # else:
        #     V3p3_not_present = 0
        #     if is_PMBus_connected() == "YES":
        #         device_status = "ON"
        #         value = parallel_thread.read_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR,16),
        #                                            pullup_enable=1, clockrate=100, writesize=1, writedata=[0x02],
        #                                            noofbytestoread=1)
        #         value = format(value[0], "08b")
        #         if value[3:6] == "000" or "001" or "010" or "011" or "100":
        #             self.pushButton_Enable_VR.setChecked(True)
        #             self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
        #                                                     "border-image: url(GUI_IMAGE/but1.png);\n"
        #                                                     "}\n"
        #                                                     "\n"
        #                                                     "\n"
        #                                                     "QPushButton::hover {\n"
        #                                                     "border-image: url(GUI_IMAGE/but1_hover"
        #                                                     ".png);\n "
        #                                                     "    }\n"
        #                                                     "")
        #             self.pushButton_Enable_VR.setDisabled(True)
        #             queue.appendleft("V3p3 is connected to the controller%INFO")
        #
        #         elif value[3:6] == "101":
        #             self.pushButton_Enable_VR.setDisabled(False)
        #             self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
        #                                                     "border-image: url(GUI_IMAGE/but0.png);\n"
        #                                                     "}\n"
        #                                                     "\n"
        #                                                     "\n"
        #                                                     "QPushButton::hover {\n"
        #                                                     "border-image: url(GUI_IMAGE/but0_hover"
        #                                                     ".png);\n "
        #                                                     "    }\n"
        #                                                     "")
        #             queue.appendleft("V3p3 is connected to the controller. "
        #                              "Toggle Enable VR to turn on the chip%INFO")
        #     else:
        #         queue.appendleft("PMBus address entered is wrong%ERROR")
        #         device_status = "OFF"
        #         self.pushButton_Enable_VR.setChecked(False)
        #         self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
        #                                                 "border-image: url(GUI_IMAGE/but0.png);\n"
        #                                                 "}\n"
        #                                                 "\n"
        #                                                 "\n"
        #                                                 "QPushButton::hover {\n"
        #                                                 "border-image: url(GUI_IMAGE/but0_hover"
        #                                                 ".png);\n "
        #                                                 "    }\n"
        #                                                 "")
        #         self.pushButton_Enable_VR.setDisabled(True)

        # # setting up main functions

        if PARTNAME == "AMP4592":
            self.pushButton_SVID_Configuration.setText("SVID Configuration")
            self.pushButton_Custom_protocol_config.setText("SVID")
            self.label_prot_A.setText("SVID:")
            self.label_prot_B.setText("SVID:")
            self.label_prot_address_A.setText(svid_address + "h")
            self.label_prot_address_B.setText(str(int(svid_address) + 1) + "h")

            #Remove SVID custom command.
            self.pushButton_Custom_protocol_config.setVisible(False)
            self.pushButton_PMBUS_custom_command.setGeometry(QtCore.QRect(250, 230, 200, 60))

            if initial_device_status == 1:
                self.label_PMBus_addressA.setText( "--")
                self.label_PMBus_addressB.setText( "--")
                self.label_prot_address_A.setText( "--")
                self.label_prot_address_B.setText( "--")

        elif PARTNAME == "AMP4692":
            self.pushButton_SVID_Configuration.setText("SVI2 Configuration")
            self.pushButton_Custom_protocol_config.setText("SVI2")
            self.label_prot_A.setText("SVI2:")
            self.label_prot_B.setText("SVI2:")
            self.label_prot_address_A.setText("--")
            self.label_prot_address_B.setText("--")
            if initial_device_status == 1:
                self.label_PMBus_addressA.setText( "--")
                self.label_PMBus_addressB.setText( "--")
        elif PARTNAME == "AMP4792":
            self.pushButton_SVID_Configuration.setText("SVI3 Configuration")
            self.pushButton_Custom_protocol_config.setText("SVI3")
            if initial_device_status == 1:
                self.label_PMBus_addressA.setText( "--")
                self.label_PMBus_addressB.setText( "--")
        elif PARTNAME == "AMP4291":
            self.pushButton_SVID_Configuration.setText("OVR Configuration")
            self.pushButton_Custom_protocol_config.setText("OVR")
            if initial_device_status == 1:
                self.label_PMBus_addressA.setText( "--")
                self.label_PMBus_addressB.setText( "--")
        else:
            self.pushButton_SVID_Configuration.setText("Unknown")

        self.pushButton_Burn_MTP.clicked.connect(self.Burn_MTP)
        self.pushButton_Refresh.clicked.connect(self.Refresh)
        self.pushButton_Help.clicked.connect(self.help)
        self.pushButton_Home.clicked.connect(self.Home_button)
        self.pushButton_Fault.clicked.connect(lambda: self.stackedWidget_monitor_fault.setCurrentIndex(1))
        self.pushButton_back_to_monitor.clicked.connect(lambda: self.stackedWidget_monitor_fault.setCurrentIndex(0))
        self.pushButton_Enable_VR.clicked.connect(self.enable_device)

        self.pushButton_Load_command.clicked.connect(self.Load_command_xlsx)
        self.pushButton_Save_command.clicked.connect(self.save_command_xlsx)

        self.pushButton_Device_Configuration.clicked.connect(lambda: self.stackedWidget_main.setCurrentIndex(1))
        self.pushButton_Telemetry.clicked.connect(lambda: self.stackedWidget_main.setCurrentIndex(2))
        self.pushButton_Telemetry.clicked.connect(lambda: self.stackedWidget_main.setCurrentIndex(2))
        self.pushButton_Faults.clicked.connect(lambda: self.stackedWidget_main.setCurrentIndex(3))
        self.pushButton_Custom_Commands.clicked.connect(lambda: self.stackedWidget_main.setCurrentIndex(4))

        ## Setting up Device configuration functions
        self.pushButton_phase_configuration.clicked.connect(self.device_configuration_phase_configuration)
        self.pushButton_autonomous_phase_manager.clicked.connect(self.device_configuration_autonomous_phase_manager)
        self.pushButton_boot_voltage.clicked.connect(self.device_configuration_Boot_voltage_configuration)
        self.pushButton_PMBus_address.clicked.connect(self.device_configuration_PMBus_address)
        self.pushButton_transient_control.clicked.connect(self.device_configuration_transient_configuration)
        self.pushButton_phase_thermal_balance.clicked.connect(self.device_configuration_phase_thermal_balance)
        self.pushButton_SVID_Configuration.clicked.connect(self.device_configuration_SVID_configuration)
        self.pushButton_PMBus_Configuration.clicked.connect(self.device_configuration_PMBus_Configuration)
        self.pushButton_Low_power_std_mode.clicked.connect(self.device_configuration_low_power_std_mode)
        self.comboBox_Eco_phase.activated.connect(self.device_configuration_eco_phase_mode)

        # Setting up Telemetry configuration functions.
        self.pushButton_Senstivity.clicked.connect(self.telemetry_configuration_senstivity_configuration)
        self.pushButton_Calibration.clicked.connect(self.telemetry_configuration_calibration_configuration)

        # Setting up Fault configuration functions
        self.pushButton_Fault_configuration.clicked.connect(self.Fault_configuration_fn)
        self.pushButton_Fault_Response.clicked.connect(self.Fault_response_fn)

        # Setting up Custom command functions
        self.pushButton_PMBUS_custom_command.clicked.connect(self.Custom_command_PMBUS)
        self.pushButton_Custom_protocol_config.clicked.connect(self.Custom_command_protocol_specific)

    def Refresh(self):
        global homeWin_obj, initial_device_status, parallel_thread, stop_thread
        stop_thread = True
        time.sleep(2)

        homeWin_obj = HomeWindow()
        self.main = homeWin_obj
        while (queue):
            v3p3_status_print = queue.pop()
            v3p3_status_print = v3p3_status_print.split("%")
            print_log(v3p3_status_print[0], v3p3_status_print[1])
        self.main.show()
        self.close()
        if initial_device_status == 0:
            master_command_xlsx_user_input()
            if VR_Enabled != "OFF":
                stop_thread = False
                parallel_thread.start()

    def Burn_MTP(self):
        global parallel_thread, PMBUS_ACK_status, initial_device_status, register_database, stop_thread
        if initial_device_status == 0:
            stop_thread = True
            time.sleep(1)
            testmode_entry()
            number_of_times_mtp_burn = otp_test()
            testmode_entry()
            wrong_combo_popup = QMessageBox()
            wrong_combo_popup.setWindowTitle("Burn MTP")
            wrong_combo_popup.setIcon(QMessageBox.Information)
            wrong_combo_popup.setText("MTP burn count = " + str(number_of_times_mtp_burn) + ". Are you sure you want to Burn MTP?")
            wrong_combo_popup.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            wrong_combo_popup.setDefaultButton(QMessageBox.No)
            result = wrong_combo_popup.exec_()
            if result == QMessageBox.Yes:
                usb_replace_bit(reg_address=0xDD, reg_size=2, high_index=7, low_index=5, reg_value='000', command_format="Non-block")
                usb_replace_bit(reg_address=0xDD, reg_size=2, high_index=3, low_index=3, reg_value='1', command_format="Non-block")
                time.sleep(0.5)
                usb_replace_bit(reg_address=0xDD, reg_size=2, high_index=3, low_index=3, reg_value='0', command_format="Non-block")
                # parallel_thread.usb_to_gpio_send_byte(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                #                             pullup_enable=1, clockrate=100, writesize=1,
                #                             writedata=[0x11])           # SEND BYTE
                # parallel_thread.write_PMBus(voltagelevel=33, address_size=0, address=int(PMBUS_ADDR, 16),
                #                             pullup_enable=1, clockrate=100, writesize=2,
                #                             writedata=[0, 0])
                testmode_exit()
                if PMBUS_ACK_status == 0:
                    print_log("MTP burn complete", "INFO")
                    # update the "Updated_device_MTP_value" key of register database with device read value. And check
                    # if device value is same as "Final_register_value" then MTP is burnt successfully for that command
                    # code, else print_log() the list of unsuccessfully burnt command codes.
                    # MTP_burnt_verification()
                    stop_thread = False
                    time.sleep(0.1)
                    parallel_thread.start()
                else:
                    print_log("MTP burn unsuccessful, check the PMBus address and V3p3 supply connection", "ERROR")
        else:
            print_log("Check the PMBus address and V3p3 supply connection", "ERROR")

    def help(self):
        print_log("Please share the log file present inside folder \"GUI_LOG\".", "INFO")

    def Home_button(self):
        self.stackedWidget_main.setCurrentIndex(0)
        # GUI look initialization
        self.label.setText(PARTNAME)
        self.label_RAILA.setText(RailA_name)
        self.label_RAILB.setText(RailB_name)
        RailA_phase_count_arg = str(int(initialize_feature_variable("DF", 8, 4), 2))
        RailB_phase_count_arg = str(int(initialize_feature_variable("DF", 3, 0), 2))
        self.label_phase_A.setText(RailA_phase_count_arg)
        self.label_phase_B.setText(RailB_phase_count_arg)
        if initial_device_status == 1:
            self.label_PMBus_addressA.setText("--")
            self.label_PMBus_addressB.setText("--")
        else:
            self.label_PMBus_addressA.setText(PMBUS_ADDR + "h")
            self.label_PMBus_addressB.setText(PMBUS_ADDR + "h")
        self.label_boot_vol_A.setText(str(round(float((int(initialize_feature_variable('E10', 23, 13), 2) * float(resolutionA) + offset[str(resolutionA)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13), 2) != 0 else int(0)) + "V")
        self.label_boot_vol_B.setText(str(round(float((int(initialize_feature_variable('E11', 23, 13), 2) * float(resolutionB) + offset[str(resolutionB)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13), 2) != 0 else int(0)) + "V")

    def enable_device(self):
        global device_status, VR_Enabled, stop_thread, parallel_thread, initial_device_status
        if initial_device_status == 1:
            print_log("Please check the TI dongle and V3p3 supply connection", "ERROR")
            return
        if VR_Enabled == "OFF":
            self.pushButton_Enable_VR.setChecked(True)
            self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                    "border-image: url(GUI_IMAGE/but1.png);\n"
                                                    "}\n"
                                                    "\n"
                                                    "\n"
                                                    "QPushButton::hover {\n"
                                                    "border-image: url(GUI_IMAGE/but1_hover"
                                                    ".png);\n "
                                                    "    }\n"
                                                    "")
            time.sleep(1)
            print_log("VR is enabled now", "INFO")
            power_up_device()
            VR_Enabled = "ON"
            self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_on_button.png"))
            self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_green_icon.png);")
            self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_green_icon.png);")
            self.label_per7.setText(
                    "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                    "vertical-align:super;\">%</span></p>")
            self.label_absolute8.setText("0" + "/" + str(int(initialize_feature_variable("E9", 61, 54), 2)) + " W")
            self.label_per8.setText(
                    "<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                    "vertical-align:super;\">%</span></p>")



            # start the thread when Enable VR is high
            stop_thread = False
            time.sleep(0.1)
            parallel_thread.start()

        else:
            # stop the thread when Enable VR is low
            stop_thread = True
            time.sleep(1)

            # Resetting the telemetry display values to 0, when enable is low
            self.label_slew_A.setText(str(0) + " V")
            self.label_slew_B.setText(str(0) + " V")
            self.label_absolute7.setText("0" + "/" + str(int(initialize_feature_variable("E9", 33, 26), 2)) + " W")
            self.label_per7.setText("<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                                    "vertical-align:super;\">%</span></p>")
            self.label_absolute8.setText("0" + "/" + str(int(initialize_feature_variable("E9", 61, 54), 2)) + " W")
            self.label_per8.setText("<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                                    "vertical-align:super;\">%</span></p>")
            self.label_absolute3.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
            self.label_per3.setText("<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
                                    "font-size:20pt;\">V</span></p></body></html>")
            self.label_absolute5.setText("Max:" + str(int(initialize_feature_variable("E80", 55, 48), 2)) + " A")
            self.label_per5.setText("<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
                                    "font-size:18pt;\">A</span></p></body></html>")
            font = QtGui.QFont()
            font.setFamily("Arial")
            font.setBold(False)
            font.setWeight(50)
            font.setPointSize(20)
            self.label_temp_A.setFont(font)
            self.label_temp_A.setText(str(0))
            self.label_temp_B.setFont(font)
            self.label_temp_B.setText(str(0))
            self.label_absolute4.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
            self.label_per4.setText("<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
                                    "font-size:20pt;\">V</span></p></body></html>")
            self.label_absolute6.setText("Max:" + str(int(initialize_feature_variable("E81", 55, 48), 2)) + " A")
            self.label_per6.setText("<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
                                    "font-size:18pt;\">A</span></p></body></html>")

            print_log("VR is powering down now", "INFO")
            power_down_device()
            VR_Enabled = "OFF"
            self.pushButton_Enable_VR.setChecked(False)
            self.pushButton_Enable_VR.setStyleSheet("QPushButton{\n"
                                                    "border-image: url(GUI_IMAGE/but0.png);\n"
                                                    "}\n"
                                                    "\n"
                                                    "\n"
                                                    "QPushButton::hover {\n"
                                                    "border-image: url(GUI_IMAGE/but0_hover"
                                                    ".png);\n "
                                                    "    }\n"
                                                    "")
            # self.pushButton_PMBus_Enable.setStyleSheet("QPushButton{\n"
            #                                            "border-image: url(GUI_IMAGE/but0.png);\n"
            #                                            "}\n"
            #                                            "\n"
            #                                            "\n"
            #                                            "QPushButton::hover {\n"
            #                                            "border-image: url(GUI_IMAGE/but0_hover.png);\n"
            #                                            "    }\n"
            #                                            "")
            self.label_Device_status_image.setPixmap(QtGui.QPixmap("GUI_IMAGE/Status_off_button.png"))
            self.label_RailA_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")
            self.label_RailB_on_img.setStyleSheet("image: url(GUI_IMAGE/power_grey_icon.png);")

    def Load_command_xlsx(self):
        global initial_device_status, stop_thread, load_settings_done
        # command_xlsx_file_name = QFileDialog.getOpenFileName(self, 'Load File', '', '*.xlsx')
        # if len(command_xlsx_file_name[0]) and initial_device_status == 0:
        #     # Replace the master_command.xlsx
        #     workbook = openpyxl.load_workbook(command_xlsx_file_name[0])
        #     workbook.save('master_command.xlsx')
        #     stop_thread = True
        #     update_command_xlsx_from_master_command_xlsx()  # call the parallel thread when all
        #     # master_command xlsx has been copied.
        #     stop_thread = False
        #
        #     update_register_database_from_master_command_xlsx()
        #     parallel_thread.start()
        #     print_log("File loaded located at: " + command_xlsx_file_name[0], "INFO")

        if initial_device_status == 0:
            wrong_combo_popup = QMessageBox()
            wrong_combo_popup.setWindowTitle("Load device settings")
            wrong_combo_popup.setIcon(QMessageBox.Information)
            wrong_combo_popup.setText("Are you sure you want to load the device settings?")
            wrong_combo_popup.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            wrong_combo_popup.setDefaultButton(QMessageBox.No)
            result = wrong_combo_popup.exec_()
            if result == QMessageBox.Yes:
                stop_thread = True
                time.sleep(1)
                # Resetting the telemetry display values to 0, when enable is low
                self.label_slew_A.setText(str(0) + " V")
                self.label_slew_B.setText(str(0) + " V")
                self.label_absolute7.setText("0" + "/" + str(int(initialize_feature_variable("E9", 33, 26), 2)) + " W")
                self.label_per7.setText("<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                                        "vertical-align:super;\">%</span></p>")
                self.label_absolute8.setText("0" + "/" + str(int(initialize_feature_variable("E9", 61, 54), 2)) + " W")
                self.label_per8.setText("<p><span style=\" font-size:28pt;\">0</span><span style=\" font-size:18pt; "
                                        "vertical-align:super;\">%</span></p>")
                self.label_absolute3.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
                self.label_per3.setText("<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
                                        "font-size:20pt;\">V</span></p></body></html>")
                self.label_absolute5.setText("Max:" + str(int(initialize_feature_variable("E80", 55, 48), 2)) + " A")
                self.label_per5.setText("<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
                                        "font-size:18pt;\">A</span></p></body></html>")
                font = QtGui.QFont()
                font.setFamily("Arial")
                font.setBold(False)
                font.setWeight(50)
                font.setPointSize(20)
                self.label_temp_A.setFont(font)
                self.label_temp_A.setText(str(0))
                self.label_temp_B.setFont(font)
                self.label_temp_B.setText(str(0))
                self.label_absolute4.setText("Max:" + str(vid_max_value[resolutionA]) + " V")
                self.label_per4.setText("<html><head/><body><p><span style=\" font-size:28pt;\">0</span><span style=\" "
                                        "font-size:20pt;\">V</span></p></body></html>")
                self.label_absolute6.setText("Max:" + str(int(initialize_feature_variable("E81", 55, 48), 2)) + " A")
                self.label_per6.setText("<html><head/><body><p><span style=\" font-size:21pt;\">0</span><span style=\" "
                                        "font-size:18pt;\">A</span></p></body></html>")

                # loading the excel starts
                command_xlsx_file_name = "Setting.xlsx"
                if len(command_xlsx_file_name) and initial_device_status == 0:
                    # Replace the master_command.xlsx
                    workbook = openpyxl.load_workbook(command_xlsx_file_name)
                    workbook.save('master_command.xlsx')

                    wrong_combo_popup = QMessageBox()
                    wrong_combo_popup.setWindowTitle("Loading previous settings")
                    wrong_combo_popup.setText("Loading the previously saved settings, Please wait till Telemetry monitor starts changing...")
                    wrong_combo_popup.exec_()
                    print_log("MTP settings will be loaded from file: " + command_xlsx_file_name, "INFO")
                    # global variable to be used to print-log after excel contents are written into the device
                    load_settings_done = "YES"

                    update_command_xlsx_from_master_command_xlsx()  # call the parallel thread when all master_command xlsx has been copied
                    stop_thread = False

                    # Updating the home page labels when load condition is selected
                    RailA_phase_count_arg = read_bits(reg_address=0xDF, reg_size=2, high_index=8, low_index=4,
                                                      command_format="Non-block")
                    RailB_phase_count_arg = read_bits(reg_address=0xDF, reg_size=2, high_index=3, low_index=0,
                                                      command_format="Non-block")
                    self.label_phase_A.setText(str(int(RailA_phase_count_arg, 2)))
                    self.label_phase_B.setText(str(int(RailB_phase_count_arg, 2)))
                    self.label_PMBus_addressA.setText(PMBUS_ADDR + "h")
                    self.label_PMBus_addressB.setText(PMBUS_ADDR + "h")
                    self.label_prot_address_A.setText(
                        str(hex(int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
                    self.label_prot_address_B.setText(
                        str(hex(1 + int(initialize_feature_variable('E6', 7, 4), 2)).replace("0x", "")) + "h")
                    self.label_boot_vol_A.setText(str(round(float((int(initialize_feature_variable('E10', 23, 13), 2) * float(
                        resolutionA) + offset[str(resolutionA)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13),
                                                                                2) != 0 else int(0)) + "V")
                    self.label_boot_vol_B.setText(str(round(float((int(initialize_feature_variable('E11', 23, 13), 2) * float(
                        resolutionB) + offset[str(resolutionB)]) / 1000), 4) if int(initialize_feature_variable('E10', 23, 13),
                                                                                2) != 0 else int(0)) + "V")

                    update_register_database_from_master_command_xlsx()
                    stop_thread = False
                    time.sleep(0.1)
                    parallel_thread.start()
        else:
            print_log("Please check the device connection and PMBus Address.", "ERROR")

    def save_command_xlsx(self):
        global initial_device_status
        # command_xlsx_file_name = QFileDialog.getSaveFileName(self, 'Save File', '', '*.xlsx')
        # if len(command_xlsx_file_name[0]) and initial_device_status == 0:
        #     workbook = openpyxl.load_workbook('master_command.xlsx')
        #     workbook.save(command_xlsx_file_name[0])
        #     print_log("File saved at: " + command_xlsx_file_name[0], "INFO")
        if initial_device_status == 0:
            wrong_combo_popup = QMessageBox()
            wrong_combo_popup.setWindowTitle("Save device settings")
            wrong_combo_popup.setIcon(QMessageBox.Information)
            wrong_combo_popup.setText("Are you sure you want to save the device settings?")
            wrong_combo_popup.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            wrong_combo_popup.setDefaultButton(QMessageBox.No)
            result = wrong_combo_popup.exec_()
            if result == QMessageBox.Yes:
                # creating Setting.xlsx for saving the device settings
                new_xlsx = openpyxl.Workbook()
                file = 'Setting.xlsx'
                new_xlsx.save(file)

                command_xlsx_file_name = "Setting.xlsx"
                if len(command_xlsx_file_name) and initial_device_status == 0:
                    workbook = openpyxl.load_workbook('master_command.xlsx')
                    workbook.save(command_xlsx_file_name)

                print_log("MTP settings have been saved in file: " + command_xlsx_file_name, "INFO")
        else:
            print_log("Please check the device connection and PMBus Address", "ERROR")

    def device_configuration_phase_configuration(self):
        self.main = frame_phase_configuration()
        self.main.show()

    def device_configuration_autonomous_phase_manager(self):
        self.main = Phase_add_drop()
        self.main.show()

    def device_configuration_Boot_voltage_configuration(self):
        self.main = Boot_voltage_configuration()
        self.main.show()

    def device_configuration_PMBus_address(self):
        self.main = PMBus_Address()
        self.main.show()

    def device_configuration_transient_configuration(self):
        self.main = Transient_Configuration()
        self.main.show()

    def device_configuration_phase_thermal_balance(self):
        self.main = Phase_thermal_balance()
        self.main.show()

    def device_configuration_SVID_configuration(self):
        global PARTNAME
        if PARTNAME == "AMP4592":
            self.main = frame_svid()
            self.main.show()
        elif PARTNAME == "AMP4692":
            self.main = frame_svi2()
            self.main.show()
        elif PARTNAME == "AMP4792":
            print_log("This function has not be coded yet", "WARNING")
        elif PARTNAME == "AMP4291":
            print_log("This function has not be coded yet", "WARNING")
        else:
            print_log("This function has not be coded yet ", "ERROR")

    def device_configuration_PMBus_Configuration(self):
        self.main = PMBus_Configuration()
        self.main.show()

    def device_configuration_low_power_std_mode(self):
        # as soon as this is clicked, it will toggle the button.
        global initial_device_status
        if initial_device_status == 1:
            print_log("Please check PMBus connectivity.","WARN")
            return
        if self.pushButton_Low_power_std_mode.isChecked() == False:

            self.pushButton_Low_power_std_mode.setStyleSheet("QPushButton{\n"
                                                             "border-image: url(GUI_IMAGE/but0.png);\n"
                                                             "}\n"
                                                             "\n"
                                                             "\n"
                                                             "QPushButton::hover {\n"
                                                             "border-image: url(GUI_IMAGE/but0_hover.png);\n"
                                                             "    }\n"
                                                             "")
            update_database_with_temp_customer_input("DF", 9, 9, "0")
            register_database["DF"]['Final_register_value'] = register_database["DF"]['Temp_update_from_customer']
            register_database["DF"]['Customer_interaction'] = "YES"
            write_PMBUS_entry_in_command_xlsx("DF")
            print_log("Low power stand by mode is off.", "INFO")
        else:
            self.pushButton_Low_power_std_mode.setStyleSheet("QPushButton{\n"
                                                             "border-image: url(GUI_IMAGE/but1.png);\n"
                                                             "}\n"
                                                             "\n"
                                                             "\n"
                                                             "QPushButton::hover {\n"
                                                             "border-image: url(GUI_IMAGE/but1_hover.png);\n"
                                                             "    }\n"
                                                             "")
            update_database_with_temp_customer_input("DF", 9, 9, "1")
            register_database["DF"]['Final_register_value'] = register_database["DF"]['Temp_update_from_customer']
            register_database["DF"]['Customer_interaction'] = "YES"
            write_PMBUS_entry_in_command_xlsx("DF")
            print_log("Low power stand by mode is On.", "INFO")

    def device_configuration_eco_phase_mode(self):
        global initial_device_status
        if initial_device_status == 1:
            print_log("Please check PMBus connectivity.", "WARN")
            return

        if self.comboBox_Eco_phase.currentIndex() == 0:
            update_database_with_temp_customer_input("C9", 31, 31, "0")
            update_database_with_temp_customer_input("C50", 85, 85, "0")
            update_database_with_temp_customer_input("C51", 85, 85, "0")
            print_log("Current mode is active in PS3 state.", "INFO")

        elif self.comboBox_Eco_phase.currentIndex() == 1:

            update_database_with_temp_customer_input("C9", 31, 31, "0")
            update_database_with_temp_customer_input("C50", 85, 85, "1")
            update_database_with_temp_customer_input("C51", 85, 85, "1")

            print_log("Voltage mode is active in PS3 state.", "INFO")

        elif self.comboBox_Eco_phase.currentIndex() == 2:
            update_database_with_temp_customer_input("C9", 31, 31, "1")
            update_database_with_temp_customer_input("C50", 85, 85, "0")
            update_database_with_temp_customer_input("C51", 85, 85, "0")
            print_log("Auto Switch mode is active.", "INFO")

        elif self.comboBox_Eco_phase.currentIndex() == 3:
            update_database_with_temp_customer_input("C9", 31, 31, "1")
            update_database_with_temp_customer_input("C50", 85, 85, "1")
            update_database_with_temp_customer_input("C51", 85, 85, "1")
            print_log("Manual mode is active.", "INFO")

        register_database["C9"]['Final_register_value'] = register_database["C9"]['Temp_update_from_customer']
        register_database["C50"]['Final_register_value'] = register_database["C50"]['Temp_update_from_customer']
        register_database["C51"]['Final_register_value'] = register_database["C51"]['Temp_update_from_customer']

        register_database["C9"]['Customer_interaction'] = "YES"
        register_database["C50"]['Customer_interaction'] = "YES"
        register_database["C51"]['Customer_interaction'] = "YES"

        write_PMBUS_entry_in_command_xlsx("C9")
        write_PMBUS_entry_in_command_xlsx("C50")
        write_PMBUS_entry_in_command_xlsx("C51")

    # Telemetry Functions
    def telemetry_configuration_senstivity_configuration(self):
        self.main = frame_telemetry_sensitivity()
        self.main.show()

    def telemetry_configuration_calibration_configuration(self):
        self.main = frame_telemetry_calibration()
        self.main.show()

    # Fault Functions
    def Fault_configuration_fn(self):
        self.main = Fault_configuration()
        self.main.show()

    def Fault_response_fn(self):
        self.main = Fault_response()
        self.main.show()

    # Custom command Functions
    def Custom_command_PMBUS(self):
        global PMBus_custom_obj
        PMBus_custom_obj = PMBus_custom_commands()
        PMBus_custom_obj.show()

    def Custom_command_protocol_specific(self):
        global PARTNAME, SVID_custom_obj, SVI2_custom_obj
        if PARTNAME == "AMP4592":
            SVID_custom_obj = SVID_custom_commands()
            SVID_custom_obj.show()
        elif PARTNAME == "AMP4692":
            SVI2_custom_obj = SVI2_custom_command()
            SVI2_custom_obj.show()
        else:
            print_log("Function not coded yet", "ERROR")


class StartUpWindow(QMainWindow, Ui_start_up):

    def __init__(self, parent=None):
        # load and initalize the gui frame.
        super(StartUpWindow, self).__init__(parent)
        self.setupUi(self)

        self.pushButton_proceed.clicked.connect(self.button_proceed_fn)
        ## code to disalbe dropdown options
        self.device_combo_box.setCurrentIndex(1)
        self.device_combo_box.setEnabled(False)
        self.label_ver.setText("Version: 0.95")

    def button_proceed_fn(self):
        global PARTNAME, homeWin_obj, queue, PMBUS_ADDR, initial_device_status, stop_thread, parallel_thread

        PARTNAME = self.device_combo_box.currentText()

        if PARTNAME == "Select Device":
            wrong_combo_popup = QMessageBox()
            wrong_combo_popup.setWindowTitle("Error")
            wrong_combo_popup.setText("Please choose one of the device from the list !")
            wrong_combo_popup.setIcon(QMessageBox.Critical)
            wrong_combo_popup.exec_()
        else:
            wrong_combo_popup = QMessageBox()
            wrong_combo_popup.setWindowTitle("Information")
            wrong_combo_popup.setIcon(QMessageBox.Information)
            wrong_combo_popup.setText("Default PMBus address is 75h !\n Do you want to edit the address?")
            wrong_combo_popup.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            wrong_combo_popup.setDefaultButton(QMessageBox.No)
            result = wrong_combo_popup.exec_()

            if result == QMessageBox.No:
                print("Default address remains 75h.", "INFO")
                PMBUS_ADDR = "75"
            elif result == QMessageBox.Yes:
                print("Please enter chip`s updated PMBus address.", "INFO")
                correct_input = False
                hex_code_set = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9", "A", "B", "C", "D", "E", "F", "a", "b", "c", "d", "e", "f"]
                while (correct_input == False):
                    text, okPressed = QInputDialog.getText(self, "PMBus Address", "Please enter PMBus Address in hex format within (00-7f)", QLineEdit.Normal, "")
                    text = text.replace(" ", "")  # remove whitespaces

                    if okPressed and text != "" and len(text.strip()) == 2 and text.strip()[0] in hex_code_set and \
                            text.strip()[1] in hex_code_set and int(text, 16) <= 127:
                        correct_input = True
                        PMBUS_ADDR = text
                    elif okPressed == False:
                        PMBUS_ADDR = "75"
                        correct_input = True
                    else:
                        correct_input = False

            PARTNAME = self.device_combo_box.currentText()

            # homeWin_obj.label.setText(PARTNAME)

            homeWin_obj = HomeWindow()  # initializing it here as we have the print_log function using this object

            self.main = homeWin_obj
            print_log("Chosen device is:" + PARTNAME + ".", "INFO")
            while (queue):
                v3p3_status_print = queue.pop()
                v3p3_status_print = v3p3_status_print.split("%")
                print_log(v3p3_status_print[0], v3p3_status_print[1])
            self.main.show()
            self.close()
            if initial_device_status == 0:
                raila_phase_count = read_bits(reg_address=0xDF, reg_size=2, high_index=8, low_index=4, command_format="Non-block")
                railb_phase_count = read_bits(reg_address=0xDF, reg_size=2, high_index=3, low_index=0, command_format="Non-block")
                homeWin_obj.label_phase_A.setText(str(int(raila_phase_count, 2)))
                homeWin_obj.label_phase_B.setText(str(int(railb_phase_count, 2)))
                master_command_xlsx_user_input()


if __name__ == '__main__':
    # parallel_thread = MyThread()
    # otp_test()

    # global homeWin_obj
    global SVID_custom_obj, PMBus_custom_obj
    # Creating incremental log file
    GUI_log_file_name = 'gui_log.txt'
    if os.path.isfile("GUI_LOG\\" + GUI_log_file_name):
        while os.path.isfile("GUI_LOG\\" + GUI_log_file_name):
            mod_file_name = GUI_log_file_name.replace("gui_log.txt", "")
            mod_file_name = mod_file_name.replace("_", "")
            if mod_file_name == "":
                GUI_log_file_name = "1_gui_log.txt"
            else:
                GUI_log_file_name = str(int(mod_file_name) + 1) + '_gui_log.txt'
    else:
        GUI_log_file_name = "gui_log.txt"

    log_gui_interaction("Inside main()")
    app = QApplication(sys.argv)
    myWin = StartUpWindow()

    parallel_thread = MyThread()
    myWin.show()
    SVID_custom_obj = SVID_custom_commands()
    create_new_command_xlsx()  # it will create a fresh command.xlsx
    PMBus_custom_obj = PMBus_custom_commands()

    sys.exit(app.exec_())
